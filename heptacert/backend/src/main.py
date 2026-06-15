import csv
import base64
import hashlib
import ipaddress
import hmac
import io
import json
import math
import os
import secrets
import logging
import time
import zipfile
import asyncio
import re
import textwrap
from datetime import datetime, timedelta, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.header import Header as EmailHeader
from email.utils import formataddr
from enum import Enum
from html import escape
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Dict, Optional, List, Tuple, Literal
from urllib.parse import urlencode, quote, urlparse
from pydantic import field_validator
import uuid as _uuid_module
import aiosmtplib
import httpx
from itsdangerous import URLSafeTimedSerializer, SignatureExpired, BadSignature
from cryptography.fernet import Fernet, InvalidToken
from cryptography.hazmat.primitives import hashes, serialization
from cryptography.hazmat.primitives.serialization import pkcs7
from cryptography import x509
import pandas as pd
import pyotp
from fastapi import FastAPI, Body, Depends, HTTPException, UploadFile, File, Query, Request, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse, RedirectResponse, JSONResponse, Response
from .moderation import moderate_public_text
try:
    from PIL import Image as PILImage
except ImportError:
    PILImage = None  # type: ignore
try:
    import qrcode
    import qrcode.image.pil
except ImportError:
    qrcode = None  # type: ignore
import jwt
from jwt import InvalidTokenError as JWTError
from passlib.context import CryptContext
from pydantic import BaseModel, ConfigDict, EmailStr, Field, field_validator
from pydantic_settings import BaseSettings
from slowapi import Limiter
from slowapi.errors import RateLimitExceeded
from datetime import date as date_type
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
from apscheduler.triggers.date import DateTrigger
from sqlalchemy import (
    Boolean, String, Integer, BigInteger, DateTime, ForeignKey, Text,
    Enum as SAEnum, UniqueConstraint, Index, select, func, distinct, update, delete, or_, and_,
    Date as sa_Date, Time as sa_Time, literal, union_all
)
from sqlalchemy import cast
import sqlalchemy as sa
from sqlalchemy import JSON as _JSON
from sqlalchemy.dialects.postgresql import JSONB as _PgJSONB, INET as _PgINET, insert as _pg_insert

# Use native PostgreSQL JSONB/INET on PostgreSQL, fall back to JSON/String on SQLite
JSONB = _JSON().with_variant(_PgJSONB(), "postgresql")
INET = String(45).with_variant(_PgINET(), "postgresql")
BIGINT_PK = BigInteger().with_variant(Integer, "sqlite")
from sqlalchemy.ext.asyncio import create_async_engine, async_sessionmaker, AsyncSession
from sqlalchemy.orm import DeclarativeBase, Mapped, mapped_column, relationship, selectinload

from .event_features import (
    FEATURE_DEFAULTS,
    is_certificate_enabled,
    is_checkin_enabled,
    is_cpd_enabled,
    is_gamification_enabled,
    is_public_registration_enabled,
    is_quiz_enabled,
    is_raffles_enabled,
    is_ticketing_enabled,
    normalize_event_type,
    normalize_feature_bool,
)
from .generator import TemplateConfig, render_certificate_pdf, render_certificate_png_watermarked, new_certificate_uuid
from .enums import Role, CertStatus, TxType, OrderStatus, AttendeeSource  # enums.py'a tasindi (god-dosya bolme)
from .services import *  # noqa: F401,F403  (servis/auth fonk. services.py'a tasindi)
from .utils import *  # noqa: F401,F403  (saf yardimcilar utils.py'a tasindi)

logger = logging.getLogger("heptacert")

ALLOWED_RICH_TEXT_STYLES = {
    "color",
    "font-size",
    "font-family",
    "font-weight",
    "font-style",
    "text-decoration",
}
FONT_SIZE_MAP = {
    "1": "12px",
    "2": "14px",
    "3": "16px",
    "4": "18px",
    "5": "24px",
    "6": "30px",
    "7": "36px",
}


def _sanitize_rich_text_style(style_text: str) -> str:
    sanitized: List[str] = []
    for chunk in style_text.split(";"):
        if ":" not in chunk:
            continue
        raw_key, raw_value = chunk.split(":", 1)
        key = raw_key.strip().lower()
        value = raw_value.strip()
        if key not in ALLOWED_RICH_TEXT_STYLES or not value:
            continue
        if key == "color" and not re.fullmatch(r"#(:[0-9a-fA-F]{3}|[0-9a-fA-F]{6})|rgba\([\d\s.,%]+\)", value):
            continue
        if key == "font-size" and not re.fullmatch(r"\d{1,3}(:px|em|rem|%)", value):
            continue
        if key == "font-family" and not re.fullmatch(r"[A-Za-z0-9 ,\"'_-]{1,120}", value):
            continue
        if key == "font-weight" and value.lower() not in {"normal", "bold", "600", "700", "800"}:
            continue
        if key == "font-style" and value.lower() not in {"normal", "italic"}:
            continue
        if key == "text-decoration" and value.lower() not in {"none", "underline", "line-through"}:
            continue
        sanitized.append(f"{key}: {escape(value, quote=True)}")
    return "; ".join(sanitized)


class _RichTextSanitizer(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.parts: List[str] = []
        self.open_tags: List[str] = []

    def handle_starttag(self, tag: str, attrs: List[tuple[str, Optional[str]]]) -> None:
        canonical = self._canonical_tag(tag)
        if canonical is None:
            return
        if canonical == "br":
            self.parts.append("<br>")
            return
        attr_html = self._sanitize_attrs(tag, attrs, canonical)
        self.parts.append(f"<{canonical}{attr_html}>")
        self.open_tags.append(canonical)

    def handle_endtag(self, tag: str) -> None:
        canonical = self._canonical_tag(tag)
        if canonical in (None, "br"):
            return
        if self.open_tags and self.open_tags[-1] == canonical:
            self.parts.append(f"</{canonical}>")
            self.open_tags.pop()

    def handle_data(self, data: str) -> None:
        if data:
            self.parts.append(escape(data))

    def handle_entityref(self, name: str) -> None:
        self.parts.append(f"&{name};")

    def handle_charref(self, name: str) -> None:
        self.parts.append(f"&#{name};")

    def get_html(self) -> str:
        while self.open_tags:
            self.parts.append(f"</{self.open_tags.pop()}>")
        return "".join(self.parts)

    def _canonical_tag(self, tag: str) -> Optional[str]:
        tag_name = tag.lower()
        mapping = {
            "b": "strong",
            "strong": "strong",
            "i": "em",
            "em": "em",
            "u": "u",
            "p": "p",
            "div": "p",
            "br": "br",
            "ul": "ul",
            "ol": "ol",
            "li": "li",
            "span": "span",
            "font": "span",
        }
        return mapping.get(tag_name)

    def _sanitize_attrs(
        self,
        source_tag: str,
        attrs: List[tuple[str, Optional[str]]],
        canonical: str,
    ) -> str:
        styles: List[str] = []
        for key, raw_value in attrs:
            attr_key = (key or "").lower()
            value = (raw_value or "").strip()
            if not value:
                continue
            if attr_key == "style":
                cleaned_style = _sanitize_rich_text_style(value)
                if cleaned_style:
                    styles.append(cleaned_style)
            if source_tag.lower() == "font":
                if attr_key == "size" and value in FONT_SIZE_MAP:
                    styles.append(f"font-size: {FONT_SIZE_MAP[value]}")
                if attr_key == "face" and re.fullmatch(r"[A-Za-z0-9 ,\"'_-]{1,120}", value):
                    styles.append(f"font-family: {escape(value, quote=True)}")
                if attr_key == "color" and re.fullmatch(r"#(:[0-9a-fA-F]{3}|[0-9a-fA-F]{6})", value):
                    styles.append(f"color: {value}")
        if canonical not in {"p", "span"} or not styles:
            return ""
        style_attr = _sanitize_rich_text_style("; ".join(styles))
        if not style_attr:
            return ""
        return f' style="{style_attr}"'


def sanitize_event_description_html(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    cleaned = value.strip()
    if not cleaned:
        return None
    if "<" not in cleaned and ">" not in cleaned:
        return escape(cleaned).replace("\r\n", "\n").replace("\r", "\n").replace("\n", "<br>") or None
    parser = _RichTextSanitizer()
    parser.feed(cleaned)
    parser.close()
    html = parser.get_html().strip()
    return html or None


from .config import Settings, settings  # Settings + settings config.py'a tasindi (god-dosya bolme Adim 1)
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
_startup_time: float = time.time()

from .db import Base, engine, SessionLocal, get_db  # db.py'a tasindi (god-dosya bolme)
bulk_cert_job_lock = asyncio.Lock()
superadmin_bulk_email_tasks_lock = asyncio.Lock()
superadmin_bulk_email_tasks: Dict[int, asyncio.Task] = {}

# Lightweight in-process caches for high-traffic QR check-in reads
CHECKIN_CONTEXT_TTL_SECONDS = 20
EVENT_TOTAL_SESSIONS_TTL_SECONDS = 30
_checkin_context_cache: Dict[str, tuple[float, Dict[str, Any]]] = {}
_event_total_sessions_cache: Dict[int, tuple[float, int]] = {}
REGISTRATION_DEVICE_COOKIE = "heptacert_reg_device"


# Base -> db.py (yukarida import edildi)


# Role, CertStatus, TxType -> enums.py (yukarida import edildi)


from .models import *  # noqa: F401,F403  (modeller models.py'a tasindi)


































































async def _get_event_capacities(event_id: int, db: AsyncSession) -> Dict[str, List[Dict[str, Any]]]:
    """Return capacities for event as mapping field_id -> list of {label, capacity, remaining}.
    If DB row not present for an option, derive from event.config if present.
    """
    out: Dict[str, List[Dict[str, Any]]] = {}
    # Load DB rows
    res = await db.execute(select(RegistrationOptionCapacity).where(RegistrationOptionCapacity.event_id == event_id))
    rows = res.scalars().all()
    by_field: Dict[str, List[RegistrationOptionCapacity]] = {}
    for r in rows:
        by_field.setdefault(r.field_id, []).append(r)

    # Also load event config options as fallback
    ev_res = await db.execute(select(Event).where(Event.id == event_id))
    ev = ev_res.scalar_one_or_none()
    registration_fields = _get_event_registration_fields(ev) if ev else []

    for field in registration_fields:
        if field.get("type") != "select":
            continue
        fid = str(field.get("id") or "").strip()
        opts = []
        db_rows = {r.option_label: r for r in by_field.get(fid, [])}
        for o in field.get("options") or []:
            if isinstance(o, dict):
                label = str(o.get("label") or "").strip()
                cap = o.get("capacity")
            else:
                label = str(o or "").strip()
                cap = None
            if label == "":
                continue
            row = db_rows.get(label)
            if row:
                remaining = None if row.capacity is None else max(0, int(row.capacity) - int(row.reserved_count or 0))
                opts.append({"label": label, "capacity": row.capacity, "remaining": remaining})
            else:
                remaining = None if cap is None else int(cap)
                opts.append({"label": label, "capacity": cap, "remaining": remaining})
        if opts:
            out[fid] = opts
    return out


def _collect_registration_option_reservations(
    event: "Event",
    registration_answers: Dict[str, Any],
) -> List[Tuple[int, str, str, Optional[int]]]:
    reservations_to_attempt: List[Tuple[int, str, str, Optional[int]]] = []
    registration_fields = _get_event_registration_fields(event)
    for field in registration_fields:
        if field.get("type") != "select":
            continue
        field_id = str(field.get("id") or "").strip()
        if not field_id:
            continue
        selected = registration_answers.get(field_id)
        if not selected:
            continue
        selected_values = selected if isinstance(selected, (list, tuple)) else [selected]
        options = field.get("options") or []
        for sel in selected_values:
            sel_label = str(sel or "").strip()
            if not sel_label:
                continue
            opt_obj: Optional[dict] = None
            for option in options:
                if isinstance(option, dict) and str(option.get("label") or "").strip() == sel_label:
                    opt_obj = option
                    break
                if isinstance(option, str) and option == sel_label:
                    opt_obj = {"label": option, "capacity": None}
                    break
            if not opt_obj:
                continue
            capacity_value = opt_obj.get("capacity")
            try:
                capacity_int = int(capacity_value) if capacity_value is not None and str(capacity_value).strip() != "" else None
            except Exception:
                capacity_int = None
            reservations_to_attempt.append((event.id, field_id, sel_label, capacity_int))
    return reservations_to_attempt


async def _sync_registration_option_capacities(db: AsyncSession, event: "Event") -> None:
    registration_fields = _get_event_registration_fields(event)
    select_fields = [field for field in registration_fields if field.get("type") == "select"]
    if not select_fields:
        await db.execute(delete(RegistrationOptionCapacity).where(RegistrationOptionCapacity.event_id == event.id))
        return

    current_specs: Dict[Tuple[str, str], Optional[int]] = {}
    for field in select_fields:
        field_id = str(field.get("id") or "").strip()
        if not field_id:
            continue
        for option in field.get("options") or []:
            if isinstance(option, dict):
                label = str(option.get("label") or "").strip()
                capacity_value = option.get("capacity")
            else:
                label = str(option or "").strip()
                capacity_value = None
            if not label:
                continue
            try:
                current_specs[(field_id, label)] = int(capacity_value) if capacity_value is not None and str(capacity_value).strip() != "" else None
            except Exception:
                current_specs[(field_id, label)] = None

    if not current_specs:
        await db.execute(delete(RegistrationOptionCapacity).where(RegistrationOptionCapacity.event_id == event.id))
        return

    verified_counts: Dict[Tuple[str, str], int] = {}
    attendee_rows = await db.execute(
        select(Attendee.registration_answers).where(
            Attendee.event_id == event.id,
            Attendee.email_verified.is_(True),
        )
    )
    for raw_answers in attendee_rows.scalars().all():
        answers = raw_answers if isinstance(raw_answers, dict) else {}
        for field in select_fields:
            field_id = str(field.get("id") or "").strip()
            if not field_id:
                continue
            selected_value = answers.get(field_id)
            if not selected_value:
                continue
            selected_values = selected_value if isinstance(selected_value, (list, tuple)) else [selected_value]
            for selected in selected_values:
                label = str(selected or "").strip()
                key = (field_id, label)
                if label and key in current_specs:
                    verified_counts[key] = verified_counts.get(key, 0) + 1

    existing_rows_res = await db.execute(
        select(RegistrationOptionCapacity).where(RegistrationOptionCapacity.event_id == event.id)
    )
    existing_rows = {
        (row.field_id, row.option_label): row
        for row in existing_rows_res.scalars().all()
    }

    current_keys = set(current_specs.keys())
    for key, capacity in current_specs.items():
        field_id, label = key
        used_count = verified_counts.get(key, 0)
        row = existing_rows.get(key)
        if row:
            row.capacity = capacity
            row.reserved_count = used_count
        else:
            db.add(
                RegistrationOptionCapacity(
                    event_id=event.id,
                    field_id=field_id,
                    option_label=label,
                    capacity=capacity,
                    reserved_count=used_count,
                )
            )

    for key, row in existing_rows.items():
        if key not in current_keys:
            await db.delete(row)


# Ã¢â€â‚¬Ã¢â€â‚¬ Payment DB models (created by migration 002) Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬

# OrderStatus -> enums.py (yukarida import edildi)






# Ã¢â€â‚¬Ã¢â€â‚¬ Enterprise DB models (created by migration 003) Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬




































# Ã¢â€â‚¬Ã¢â€â‚¬ Email System Models (created by migration 008) Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬

























# Ã¢â€â‚¬Ã¢â€â‚¬ Attendaonce management models (migration 003) Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬

# AttendeeSource -> enums.py (yukarida import edildi)












# Ã¢â€â‚¬Ã¢â€â‚¬ Gamification: Badge Rules & Participant Badges Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬









# Keep the correctly spelled public model name without changing the existing table mapping.
AttendanceRecord = AttendaonceRecord






# Ã¢â€â‚¬Ã¢â€â‚¬ Certificate Tiers Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬







# Ã¢â€â‚¬Ã¢â€â‚¬ Survey System Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬





# Ã¢â€â‚¬Ã¢â€â‚¬ Sponsor Slots Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬



from .schemas import *  # noqa: F401,F403  (semalar schemas.py'a tasindi)














































from .event_team import (  # event_team.py'a tasindi (god-dosya bolme)
    EVENT_TEAM_ROLES, EVENT_TEAM_STATUSES, EVENT_TEAM_ROLE_PERMISSIONS,
    EVENT_TEAM_PERMISSION_LABELS, _normalize_event_team_permissions,
    _effective_event_team_permissions,
)









































































# Ã¢â€â‚¬Ã¢â€â‚¬ Enterprise Pydantic models Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬













# Ã¢â€â‚¬Ã¢â€â‚¬ Gamification & Survey Request/Response Models Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬

































































# Ã¢â€â‚¬Ã¢â€â‚¬ Email System Schemas Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬











































DEFAULT_PRICING: List[dict] = [
    {
        "id": "starter",
        "name_tr": "BaÃ…Å¸langÃ„Â±ÃƒÂ§",
        "name_en": "Starter",
        "price_monthly": 0,
        "price_annual": 0,
        "hc_quota": 50,
        "features_tr": [
            "50 HC hoÅŸ geldin bonusu (tek seferlik)",
            "QR kod doÃ„Å¸rulama",
            "Sertifika arÃ…Å¸ivi (1 yÃ„Â±l)",
            "Temel ÅŸablon editÃ¶rÃ¼",
            "HeptaCert watermark",
        ],
        "features_en": [
            "50 HC welcome bonus (one-time)",
            "QR code verification",
            "Certificate archive (1 year)",
            "Basic template editor",
            "HeptaCert watermark",
        ],
        "is_free": True,
        "is_enterprise": False,
    },
    {
        "id": "pro",
        "name_tr": "Profesyonel",
        "name_en": "Professional",
        "price_monthly": 499,
        "price_annual": 399,
        "hc_quota": 500,
        "features_tr": [
            "AylÃ„Â±k 500 HC",
            "SÃ„Â±nÃ„Â±rsÃ„Â±z etkinlik",
            "Excel toplu basÃ„Â±m",
            "Sertifika arÃ…Å¸ivi (3 yÃ„Â±l)",
            "Etkinlik kayÃ„Â±t ve check-in sistemi",
            "QR ile yoklama takibi",
            "Ã–oncelikli destek",
        ],
        "features_en": [
            "500 HC per month",
            "Unlimited events",
            "Excel bulk generation",
            "Certificate archive (3 years)",
            "Event registration & check-in system",
            "QR attendaonce tracking",
            "Priority support",
        ],
        "is_free": False,
        "is_enterprise": False,
    },
    {
        "id": "growth",
        "name_tr": "BÃ¼yÃ¼me",
        "name_en": "Growth",
        "price_monthly": 1299,
        "price_annual": 1099,
        "hc_quota": 2000,
        "features_tr": [
            "AylÃ„Â±k 2.000 HC",
            "SÃ„Â±nÃ„Â±rsÃ„Â±z etkinlik",
            "Excel toplu basÃ„Â±m",
            "Sertifika arÃ…Å¸ivi (3 yÃ„Â±l)",
            "Etkinlik kayÃ„Â±t ve check-in sistemi",
            "QR ile yoklama takibi",
            "API eriÅŸimi (tam)",
            "Ãƒâ€“zel alan adÃ„Â± doÃ„Å¸rulama",
            "Marka watermark kaldÃ„Â±rma",
            "Otomatik email sistemi (bulk mail + ÅŸablonlar)",
            "5-7 hazÃ„Â±r sertifika Ã…Å¸ablonu",
            "Custom event aÃƒÂ§Ã„Â±klamasÃ„Â± ve banneri",
            "Webhook API desteÃ„Å¸i",
            "Advaonced analytics dashboard",
            "Custom form alanlarÃ„Â±",
            "KatÃ„Â±lÃ„Â±mcÃ„Â± self-service sertifika indirme",
        ],
        "features_en": [
            "2,000 HC per month",
            "Unlimited events",
            "Excel bulk generation",
            "Certificate archive (3 years)",
            "Event registration & check-in system",
            "QR attendaonce tracking",
            "Full API access",
            "Custom domain verification",
            "Remove branding watermark",
            "Automated email system (bulk mail + templates)",
            "5-7 pre-built certificate templates",
            "Custom event description & banner",
            "Webhook API support",
            "Advaonced analytics dashboard",
            "Custom form fields",
            "Attendee self-service certificate download",
        ],
        "is_free": False,
        "is_enterprise": False,
    },
    {
        "id": "enterprise",
        "name_tr": "Kurumsal",
        "name_en": "Enterprise",
        "price_monthly": None,
        "price_annual": None,
        "hc_quota": None,
        "features_tr": [
            "SÃ„Â±nÃ„Â±rsÃ„Â±z HC kotasÃ„Â±",
            "Ãƒâ€“zel SLA anlaÃ…Å¸masÃ„Â±",
            "API entegrasyonu",
            "Ãƒâ€“zel alan adÃ„Â± desteÃ„Å¸i",
            "Etkinlik kayÃ„Â±t ve check-in sistemi",
            "QR ile yoklama takibi",
            "Toplu sertifika Ã¼retimi",
            "7/24 kurumsal destek",
        ],
        "features_en": [
            "Unlimited HC quota",
            "Custom SLA agreement",
            "API integration",
            "Custom domain support",
            "Event registration & check-in system",
            "QR attendaonce tracking",
            "Bulk certificate generation",
            "24/7 enterprise support",
        ],
        "is_free": False,
        "is_enterprise": True,
    },
]



#helpers
# 1 coin = 10 unit (0.1 coin)
COIN_UNIT = 10
MB_PER_COIN_MONTH = 100.0
MIN_MONTHLY_UNITS = 2   # 0.2 coin
STEP_UNITS = 1          # 0.1 coin

def monthly_hosting_units(asset_size_bytes: int) -> int:
    mb = asset_size_bytes / (1024 * 1024)
    units_mult = max(1, math.ceil(mb / MB_PER_COIN_MONTH))
    raw = units_mult * STEP_UNITS
    return max(MIN_MONTHLY_UNITS, raw)

def hosting_units(term: str, asset_size_bytes: int) -> int:
    m = monthly_hosting_units(asset_size_bytes)
    if term == "monthly":
        return m
    return m * 10  # yearly: 10 ay Ã¼cret





ISSUE_UNITS_PER_CERT = 10




def certificate_to_out(cert: "Certificate", *, include_locked_pdf: bool = False) -> CertificateOut:
    asset_size_bytes = int(getattr(cert, "asset_size_bytes", 0) or 0)
    hosting_term = getattr(cert, "hosting_term", None) or "yearly"
    hosting_cost = hosting_units(hosting_term, asset_size_bytes)
    pdf_url = cert.pdf_url if include_locked_pdf or cert.status == CertStatus.active else None
    return CertificateOut(
        id=cert.id,
        uuid=cert.uuid,
        public_id=cert.public_id,
        student_name=cert.student_name,
        event_id=cert.event_id,
        status=cert.status,
        issued_at=getattr(cert, "issued_at", None),
        hosting_term=hosting_term,
        hosting_ends_at=getattr(cert, "hosting_ends_at", None),
        days_remaining=_certificate_days_remaining(cert),
        asset_size_bytes=asset_size_bytes,
        issue_cost_units=ISSUE_UNITS_PER_CERT,
        hosting_cost_units=hosting_cost,
        total_cost_units=ISSUE_UNITS_PER_CERT + hosting_cost,
        monthly_cost_units=hosting_units("monthly", asset_size_bytes),
        yearly_cost_units=hosting_units("yearly", asset_size_bytes),
        auto_renew_enabled=bool(getattr(cert, "auto_renew_enabled", False)),
        pdf_url=pdf_url,
        png_url=_certificate_png_public_url(cert.event_id, cert.uuid),
    )

#helpers

def hash_password(pw: str) -> str:
    return pwd_context.hash(pw)


def verify_password(pw: str, pw_hash: str) -> bool:
    return pwd_context.verify(pw, pw_hash)


# Ã¢â€â‚¬Ã¢â€â‚¬ Email token helpers Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬
_email_signer: Optional[URLSafeTimedSerializer] = None

def get_signer() -> URLSafeTimedSerializer:
    global _email_signer
    if _email_signer is None:
        _email_signer = URLSafeTimedSerializer(settings.email_token_secret)
    return _email_signer


def make_email_token(payload: dict) -> str:
    return get_signer().dumps(payload)


def verify_email_token(token: str, max_age: int = 86400) -> dict:
    """Raises on expired / invalid."""
    return get_signer().loads(token, max_age=max_age)


def make_survey_access_token(*, attendee_id: int, event_id: int, email: str) -> str:
    return make_email_token(
        {
            "action": "survey_access",
            "attendee_id": attendee_id,
            "event_id": event_id,
            "email": email.lower(),
        }
    )


def verify_survey_access_token(
    token: str,
    *,
    event_id: int | str,
    event_public_id: Optional[str] = None,
    max_age: int = 60 * 60 * 24 * 365,
) -> dict:
    payload = verify_email_token(token, max_age=max_age)
    if payload.get("action") != "survey_access":
        raise BadSignature("invalid survey token action")
    payload_event_id = str(payload.get("event_id") or "").strip()
    allowed_event_ids = {str(event_id)}
    if event_public_id:
        allowed_event_ids.add(str(event_public_id))
    if payload_event_id not in allowed_event_ids:
        raise BadSignature("survey token event mismatch")
    return payload


def build_public_survey_url(*, event_id: str, attendee_id: int, email: str) -> str:
    survey_token = make_survey_access_token(
        attendee_id=attendee_id,
        event_id=event_id,
        email=email,
    )
    return f"{settings.frontend_base_url.rstrip('/')}/events/{event_id}/survey?token={survey_token}"


def build_public_status_url(*, event_id: str, attendee_id: int, email: str) -> str:
    survey_token = make_survey_access_token(
        attendee_id=attendee_id,
        event_id=event_id,
        email=email,
    )
    return f"{settings.frontend_base_url.rstrip('/')}/events/{event_id}/status?token={survey_token}"










def _decrypt_smtp_password(stored: Optional[str]) -> Optional[str]:
    if not stored:
        return None
    value = str(stored)
    if not value.startswith("enc:v1:"):
        # Backward compatibility for legacy plaintext rows.
        return value
    token = value[len("enc:v1:"):]
    try:
        return _smtp_password_cipher().decrypt(token.encode("utf-8")).decode("utf-8")
    except InvalidToken:
        logger.error("SMTP password decryption failed due to invalid token")
        return None




def _decrypt_secret(stored: Optional[str]) -> Optional[str]:
    return _decrypt_smtp_password(stored)


GOOGLE_SHEETS_SCOPES = [
    "openid",
    "email",
    "profile",
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.file",
]

GOOGLE_SHEETS_REQUIRED_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
]

GOOGLE_CALENDAR_SCOPES = [
    "openid",
    "email",
    "profile",
    "https://www.googleapis.com/auth/calendar.events",
]

GOOGLE_CALENDAR_REQUIRED_SCOPES = [
    "https://www.googleapis.com/auth/calendar.events",
]






def _google_sheets_missing_scopes(scopes: Any) -> List[str]:
    present = set(_normalize_google_scopes(scopes))
    required = set(GOOGLE_SHEETS_REQUIRED_SCOPES)
    return [scope for scope in GOOGLE_SHEETS_REQUIRED_SCOPES if scope in required and scope not in present]


def _google_calendar_missing_scopes(scopes: Any) -> List[str]:
    present = set(_normalize_google_scopes(scopes))
    return [scope for scope in GOOGLE_CALENDAR_REQUIRED_SCOPES if scope not in present]












async def _get_google_access_token_for_scopes(db: AsyncSession, user_id: int, required_scopes: List[str], label: str) -> str:
    integration = await _get_user_google_integration(db, user_id)
    if not integration or not integration.refresh_token:
        raise HTTPException(status_code=409, detail=f"Google {label} connection is not ready.")
    present_scopes = set(_normalize_google_scopes(integration.scopes))
    missing = [scope for scope in required_scopes if scope not in present_scopes]
    if missing:
        raise HTTPException(status_code=409, detail=f"Google {label} authorization is missing required scopes.")

    now = datetime.now(timezone.utc)
    expires_at = ensure_utc(integration.token_expires_at)
    current_access_token = _decrypt_secret(integration.access_token)
    if current_access_token and expires_at and expires_at > now + timedelta(seconds=60):
        return current_access_token

    refresh_token = _decrypt_secret(integration.refresh_token)
    if not refresh_token:
        raise HTTPException(status_code=409, detail="Google refresh token could not be read.")

    async with httpx.AsyncClient(timeout=15.0) as client:
        token_res = await client.post(
            "https://oauth2.googleapis.com/token",
            data={
                "client_id": settings.google_oauth_client_id,
                "client_secret": settings.google_oauth_client_secret,
                "refresh_token": refresh_token,
                "grant_type": "refresh_token",
            },
        )
    if token_res.status_code >= 400:
        raise HTTPException(status_code=409, detail=f"Google {label} authorization has expired. Please connect again.")

    token_data = token_res.json()
    access_token = str(token_data.get("access_token") or "")
    if not access_token:
        raise HTTPException(status_code=409, detail="Google access token missing.")
    integration.access_token = _encrypt_secret(access_token)
    integration.token_expires_at = now + timedelta(seconds=int(token_data.get("expires_in") or 3600))
    if token_data.get("scope"):
        integration.scopes = _normalize_google_scopes(token_data.get("scope"))
    db.add(integration)
    await db.commit()
    return access_token


async def _get_google_access_token_for_sheets(db: AsyncSession, user_id: int) -> str:
    return await _get_google_access_token_for_scopes(db, user_id, GOOGLE_SHEETS_REQUIRED_SCOPES, "Sheets")


async def _get_google_access_token_for_calendar(db: AsyncSession, user_id: int) -> str:
    return await _get_google_access_token_for_scopes(db, user_id, GOOGLE_CALENDAR_REQUIRED_SCOPES, "Calendar")












def _google_sheets_header_for_event(event: "Event") -> List[str]:
    header = [
        "registered_at",
        "event_name",
        "attendee_id",
        "name",
        "email",
        "source",
        "email_verified",
    ]
    for field in _get_event_registration_fields(event):
        label = str(field.get("label") or field.get("id") or "").strip()
        if label:
            header.append(label)
    return header


def _google_sheets_row_for_attendee(event: "Event", attendee: "Attendee") -> List[str]:
    answers = attendee.registration_answers or {}
    row = [
        _format_export_datetime(attendee.registered_at),
        event.name,
        str(attendee.id),
        attendee.name,
        attendee.email,
        attendee.source,
        "yes" if attendee.email_verified else "no",
    ]
    for field in _get_event_registration_fields(event):
        field_id = str(field.get("id") or "")
        row.append(_sheet_safe_value(answers.get(field_id)))
    return row


async def _write_event_attendees_to_google_sheet(
    db: AsyncSession,
    event: "Event",
    *,
    create_if_missing: bool = False,
) -> Dict[str, Any]:
    sheets_config = _get_event_google_sheets_config(event)
    if not sheets_config.get("enabled") and not create_if_missing:
        return sheets_config

    access_token = await _get_google_access_token_for_sheets(db, event.admin_id)
    sheet_name = str(sheets_config.get("sheet_name") or "Registrations")
    spreadsheet_id = str(sheets_config.get("spreadsheet_id") or "")
    spreadsheet_url = str(sheets_config.get("spreadsheet_url") or "")

    if create_if_missing or not spreadsheet_id:
        created = await _google_json_request(
            access_token,
            "POST",
            "https://sheets.googleapis.com/v4/spreadsheets",
            json_body={"properties": {"title": f"HeptaCert - {event.name} - Registrations"}},
        )
        spreadsheet_id = str(created.get("spreadsheetId") or "")
        spreadsheet_url = str(created.get("spreadsheetUrl") or f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}")
        sheet_name = _google_sheets_default_sheet_name(created)
        if not spreadsheet_id:
            raise HTTPException(status_code=502, detail="Google Sheet could not be created.")

    attendees_res = await db.execute(
        select(Attendee).where(Attendee.event_id == event.id).order_by(Attendee.id.asc())
    )
    attendees = attendees_res.scalars().all()
    values = [_google_sheets_header_for_event(event)]
    values.extend(_google_sheets_row_for_attendee(event, attendee) for attendee in attendees)

    encoded_range = _google_sheets_a1_range(sheet_name)
    await _google_json_request(
        access_token,
        "POST",
        f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}/values/{encoded_range}:clear",
        json_body={},
    )
    await _google_json_request(
        access_token,
        "PUT",
        f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}/values/{encoded_range}?valueInputOption=USER_ENTERED",
        json_body={"majorDimension": "ROWS", "values": values},
    )

    next_sheets_config = {
        "enabled": True,
        "spreadsheet_id": spreadsheet_id,
        "spreadsheet_url": spreadsheet_url,
        "sheet_name": sheet_name,
        "header": values[0],
        "last_synced_at": datetime.now(timezone.utc).isoformat(),
    }
    _set_event_google_sheets_config(event, next_sheets_config)
    db.add(event)
    await db.commit()
    await db.refresh(event)
    return next_sheets_config


async def _append_attendee_to_google_sheet_if_enabled(db: AsyncSession, event: "Event", attendee: "Attendee") -> None:
    sheets_config = _get_event_google_sheets_config(event)
    if not sheets_config.get("enabled") or not sheets_config.get("spreadsheet_id"):
        return
    try:
        access_token = await _get_google_access_token_for_sheets(db, event.admin_id)
        sheet_name = str(sheets_config.get("sheet_name") or "Registrations")
        spreadsheet_id = str(sheets_config.get("spreadsheet_id"))
        encoded_range = _google_sheets_a1_range(sheet_name)
        await _google_json_request(
            access_token,
            "POST",
            f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}/values/{encoded_range}:append?valueInputOption=USER_ENTERED&insertDataOption=INSERT_ROWS",
            json_body={"majorDimension": "ROWS", "values": [_google_sheets_row_for_attendee(event, attendee)]},
        )
    except Exception as exc:
        logger.warning("Google Sheets attendee append failed for event_id=%s attendee_id=%s: %s", event.id, attendee.id, exc)


async def _sync_google_sheet_if_enabled(db: AsyncSession, event: "Event") -> None:
    if not _get_event_google_sheets_config(event).get("enabled"):
        return
    try:
        await _write_event_attendees_to_google_sheet(db, event)
    except Exception as exc:
        logger.warning("Google Sheets event sync failed for event_id=%s: %s", event.id, exc)


MS365_EXCEL_SCOPES = ["openid", "profile", "email", "offline_access", "User.Read", "Files.ReadWrite"]
MS365_EXCEL_REQUIRED_SCOPES = ["Files.ReadWrite"]






def _ms365_excel_missing_scopes(scopes: Any) -> List[str]:
    present = {scope.lower() for scope in _normalize_ms365_scopes(scopes)}
    return [scope for scope in MS365_EXCEL_REQUIRED_SCOPES if scope.lower() not in present]












async def _get_ms365_access_token_for_excel(db: AsyncSession, user_id: int) -> str:
    integration = await _get_user_ms365_integration(db, user_id)
    if not integration or not integration.refresh_token:
        raise HTTPException(status_code=409, detail="Microsoft Excel connection is not ready.")
    now = datetime.now(timezone.utc)
    expires_at = ensure_utc(integration.token_expires_at)
    current_access_token = _decrypt_secret(integration.access_token)
    if current_access_token and expires_at and expires_at > now + timedelta(seconds=60):
        return current_access_token
    refresh_token = _decrypt_secret(integration.refresh_token)
    if not refresh_token:
        raise HTTPException(status_code=409, detail="Microsoft refresh token could not be read.")
    async with httpx.AsyncClient(timeout=15.0) as client:
        token_res = await client.post(
            "https://login.microsoftonline.com/common/oauth2/v2.0/token",
            data={
                "client_id": settings.ms365_oauth_client_id,
                "client_secret": settings.ms365_oauth_client_secret,
                "refresh_token": refresh_token,
                "grant_type": "refresh_token",
                "scope": " ".join(MS365_EXCEL_SCOPES),
            },
        )
    if token_res.status_code >= 400:
        raise HTTPException(status_code=409, detail="Microsoft Excel authorization has expired. Please connect again.")
    token_data = token_res.json()
    access_token = str(token_data.get("access_token") or "")
    if not access_token:
        raise HTTPException(status_code=409, detail="Microsoft access token missing.")
    integration.access_token = _encrypt_secret(access_token)
    if token_data.get("refresh_token"):
        integration.refresh_token = _encrypt_secret(str(token_data.get("refresh_token")))
    integration.token_expires_at = now + timedelta(seconds=int(token_data.get("expires_in") or 3600))
    if token_data.get("scope"):
        integration.scopes = _normalize_ms365_scopes(token_data.get("scope"))
    db.add(integration)
    await db.commit()
    return access_token






def _event_attendees_xlsx_bytes(event: "Event", attendees: List["Attendee"]) -> bytes:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registrations"
    ws.append(_google_sheets_header_for_event(event))
    for attendee in attendees:
        ws.append(_google_sheets_row_for_attendee(event, attendee))
    for column_cells in ws.columns:
        max_length = max((len(str(cell.value or "")) for cell in column_cells), default=10)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 48)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


async def _write_event_attendees_to_ms365_excel(db: AsyncSession, event: "Event", *, create_if_missing: bool = False) -> Dict[str, Any]:
    excel_config = _get_event_ms365_excel_config(event)
    if not excel_config.get("enabled") and not create_if_missing:
        return excel_config
    access_token = await _get_ms365_access_token_for_excel(db, event.admin_id)
    attendees_res = await db.execute(select(Attendee).where(Attendee.event_id == event.id).order_by(Attendee.id.asc()))
    attendees = attendees_res.scalars().all()
    workbook_bytes = _event_attendees_xlsx_bytes(event, attendees)
    workbook_id = str(excel_config.get("workbook_id") or "")
    filename = _safe_ms365_filename(str(excel_config.get("workbook_name") or f"HeptaCert - {event.name} - Registrations.xlsx"))
    if not workbook_id and not create_if_missing:
        raise HTTPException(status_code=409, detail="No Microsoft Excel workbook is connected for this event.")
    item = await _ms365_upload_workbook(access_token, workbook_bytes, workbook_id=workbook_id, filename=filename)
    workbook_id = str(item.get("id") or workbook_id)
    workbook_url = str(item.get("webUrl") or excel_config.get("workbook_url") or "")
    workbook_name = str(item.get("name") or filename)
    if not workbook_id:
        raise HTTPException(status_code=502, detail="Microsoft Excel workbook could not be created.")
    next_excel_config = {
        "enabled": True,
        "workbook_id": workbook_id,
        "workbook_url": workbook_url,
        "workbook_name": workbook_name,
        "sheet_name": "Registrations",
        "header": _google_sheets_header_for_event(event),
        "last_synced_at": datetime.now(timezone.utc).isoformat(),
    }
    _set_event_ms365_excel_config(event, next_excel_config)
    db.add(event)
    await db.commit()
    await db.refresh(event)
    return next_excel_config


async def _sync_ms365_excel_if_enabled(db: AsyncSession, event: "Event") -> None:
    if not _get_event_ms365_excel_config(event).get("enabled"):
        return
    try:
        await _write_event_attendees_to_ms365_excel(db, event)
    except Exception as exc:
        logger.warning("Microsoft Excel sync failed for event_id=%s: %s", event.id, exc)


async def send_attendee_verification_email(*, attendee: "Attendee", event: "Event") -> None:
    token = attendee.email_verification_token
    if not token:
        return

    verify_link = build_attendee_verify_url(event_id=_get_public_event_identifier(event), token=token)
    await send_email_async(
        to=attendee.email,
        subject=f"{event.name} etkinliği için e-posta adresinizi doğrulayın",
        html_body=f"""
        <p>Merhaba {attendee.name},</p>
        <p>{event.name} etkinlik kaydınızı tamamlamak için e-posta adresinizi doğrulamanız gerekiyor.</p>
        <p><a href="{verify_link}">{verify_link}</a></p>
        <p>Bu bağlantıyı doğrulamadan check-in yapamaz ve cekilişlere dahil olamazsınız.</p>
        <p>Bağlantı 24 saat gecerlidir.</p>
        """,
        sender_user_id=event.admin_id,
    )







async def send_email_async(
    to: str,
    subject: str,
    html_body: str,
    template_vars: Optional[Dict[str, Any]] = None,
    attachments: Optional[List[tuple[str, bytes, str]]] = None,  # [(filename, bytes, mimetype),...]
    raise_on_error: bool = False,
    sender_user_id: Optional[int] = None,
) -> None:
    """
    Send email asynchronously.
    
    Args:
        to: Recipient email address
        subject: Email subject
        html_body: HTML body or Jinja2 template string
        template_vars: Variables to render in Jinja2 template
        attachments: Optional list of (filename, bytes_content, mimetype) tuples
    """
    smtp_host = settings.smtp_host
    smtp_port = settings.smtp_port
    smtp_user = settings.smtp_user or None
    smtp_password = settings.smtp_password or None
    smtp_from_email = settings.smtp_from
    smtp_from_name = ""
    smtp_reply_to: Optional[str] = None
    smtp_auto_cc: Optional[str] = None
    smtp_use_tls = True
    using_user_smtp = False
    list_unsubscribe_url: Optional[str] = None

    # Keep a copy for fallback if per-user SMTP is broken.
    global_smtp_host = settings.smtp_host
    global_smtp_port = settings.smtp_port
    global_smtp_user = settings.smtp_user or None
    global_smtp_password = settings.smtp_password or None
    global_smtp_from = settings.smtp_from

    if sender_user_id is not None:
        async with SessionLocal() as db_mail:
            res = await db_mail.execute(select(UserEmailConfig).where(UserEmailConfig.user_id == sender_user_id))
            user_config = res.scalar_one_or_none()
            if user_config and user_config.smtp_enabled:
                has_required_core = bool(
                    user_config.smtp_host
                    and user_config.smtp_port
                    and (user_config.from_email or user_config.smtp_user)
                )
                decrypted_user_password = _decrypt_smtp_password(user_config.smtp_password)
                needs_auth = bool(user_config.smtp_user)
                has_required_auth = (not needs_auth) or bool(decrypted_user_password)

                if has_required_core and has_required_auth:
                    smtp_host = user_config.smtp_host
                    smtp_port = int(user_config.smtp_port)
                    smtp_user = user_config.smtp_user or None
                    smtp_password = decrypted_user_password
                    smtp_from_email = user_config.from_email or user_config.smtp_user or settings.smtp_from
                    smtp_from_name = (user_config.from_name or "").strip()
                    smtp_reply_to = (user_config.reply_to or "").strip() or None
                    smtp_auto_cc = (user_config.auto_cc or "").strip() or None
                    smtp_use_tls = bool(user_config.smtp_use_tls)
                    using_user_smtp = True
                else:
                    logger.warning(
                        "User SMTP config invalid for user_id=%s; falling back to global SMTP",
                        sender_user_id,
                    )

    if not smtp_host:
        logger.warning(
            "[EMAIL - no SMTP configured] To: %s | Subject: %s\nBody: %s",
            to, subject, html_body
        )
        if raise_on_error:
            raise RuntimeError("SMTP is not configured")
        return
    
    # Render Jinja2 template if template_vars provided
    if template_vars:
        try:
            from jinja2 import Template
            template = Template(html_body)
            html_body = template.render(**template_vars)
            subject = Template(subject).render(**template_vars)
        except Exception as exc:
            logger.error("Jinja2 template rendering failed: %s", exc)
            return

    # Attempt to resolve recipient as an Attendee/PublicMember to support unsubscribe links.
    try:
        normalized_to = (to or "").strip().lower()
        is_digest_message = (
            "{{ unsubscribe_url }}" in html_body
            or "{{unsubscribe_url}}" in html_body
            or "topluluk güncellemeleri" in html_body.lower()
            or "topluluk gÃ¼ncellemeleri" in html_body.lower()
            or "topluluk guncellemeleri" in html_body.lower()
            or "topluluk guoncellemeleri" in html_body.lower()
        )
        async with SessionLocal() as db_check:
            a_res = await db_check.execute(
                select(Attendee)
                .where(func.lower(func.trim(Attendee.email)) == normalized_to)
                .order_by(Attendee.id.desc())
            )
            attendee = a_res.scalars().first()
            member_res = await db_check.execute(
                select(PublicMember)
                .where(func.lower(func.trim(PublicMember.email)) == normalized_to)
                .order_by(PublicMember.id.desc())
            )
            public_member = member_res.scalars().first()

            if is_digest_message and public_member:
                if not bool(getattr(public_member, "digest_opt_in", True)):
                    logger.info("Skipping digest email to opted-out public member %s", normalized_to)
                    return

                token = hashlib.sha256(f"public_member:{public_member.id}:{public_member.email}".encode()).hexdigest()[:16]
                list_unsubscribe_url = f"{settings.public_base_url}/api/public/members/{public_member.id}/unsubscribe-digest?token={token}"
            elif attendee:
                # If attendee previously unsubscribed, skip sending
                if attendee.unsubscribed_at:
                    logger.info("Skipping email to unsubscribed attendee %s", normalized_to)
                    return

                token = hashlib.sha256(f"{attendee.id}:{attendee.email}".encode()).hexdigest()[:16]
                list_unsubscribe_url = f"{settings.public_base_url}/api/public/attendees/{attendee.id}/unsubscribe?token={token}"

            if list_unsubscribe_url:
                had_unsubscribe_placeholder = "{{ unsubscribe_url }}" in html_body or "{{unsubscribe_url}}" in html_body
                html_body = html_body.replace("{{ unsubscribe_url }}", list_unsubscribe_url).replace(
                    "{{unsubscribe_url}}", list_unsubscribe_url
                )

                # Append a minimal unsubscribe footer if template does not contain an unsubscribe placeholder
                if not had_unsubscribe_placeholder and "unsubscribe" not in html_body.lower():
                    html_body = html_body + (
                        f"<hr><p style=\"font-size:12px;color:#666\">E-posta almak istemiyorsanız, "
                        f"<a href=\"{list_unsubscribe_url}\">buradan abonelikten çıkabilirsiniz</a>.</p>"
                    )
    except Exception:
        logger.exception("Failed to resolve attendee for unsubscribe handling")
    
    msg = MIMEMultipart("alternative")
    msg["Subject"] = str(EmailHeader(subject, charset="utf-8"))
    msg["From"] = (
        formataddr((smtp_from_name, smtp_from_email))
        if smtp_from_name
        else smtp_from_email
    )
    msg["To"] = to
    if smtp_reply_to:
        msg["Reply-To"] = smtp_reply_to
    if smtp_auto_cc:
        msg["Cc"] = smtp_auto_cc
    # Add unsubscribe header (RFC 2369) with mailto and URL if available
    list_unsub_header = f"<mailto:{smtp_reply_to or smtp_from_email}subject=unsubscribe>"
    if list_unsubscribe_url:
        list_unsub_header = f"{list_unsub_header}, <{list_unsubscribe_url}>"
    msg["List-Unsubscribe"] = list_unsub_header
    if list_unsubscribe_url:
        msg["List-Unsubscribe-Post"] = "List-Unsubscribe=One-Click"
    
    msg.attach(MIMEText(html_body, "html", "utf-8"))
    
    # Attach files if provided
    if attachments:
        from email.mime.base import MIMEBase
        from email.mime.image import MIMEImage
        from email import encoders
        
        for filename, file_bytes, mimetype in attachments:
            maintype, subtype = mimetype.split("/", 1)
            if maintype == "text":
                attachment = MIMEText(file_bytes.decode("utf-8", errors="replace"), _subtype=subtype, _charset="utf-8")
            elif maintype == "image":
                from email.mime.image import MIMEImage
                attachment = MIMEImage(file_bytes, _subtype=subtype)
            elif maintype == "application":
                attachment = MIMEBase(maintype, subtype)
                attachment.set_payload(file_bytes)
                encoders.encode_base64(attachment)
            else:
                attachment = MIMEBase(maintype, subtype)
                attachment.set_payload(file_bytes)
                encoders.encode_base64(attachment)
            
            attachment.add_header("Content-Disposition", "attachment", filename=filename)
            msg.attach(attachment)
    
    use_implicit_tls = bool(smtp_use_tls and int(smtp_port or 0) == 465)
    use_starttls = bool(smtp_use_tls and not use_implicit_tls)

    try:
        await aiosmtplib.send(
            msg,
            hostname=smtp_host,
            port=smtp_port,
            username=smtp_user,
            password=smtp_password,
            start_tls=use_starttls,
            use_tls=use_implicit_tls,
            timeout=20,
        )
        logger.info("Email sent successfully to %s", to)
    except Exception as exc:
        logger.error("SMTP send failed to %s: %s", to, exc)

        # If per-user SMTP fails, try global SMTP oonce to avoid system-wide email outage.
        if using_user_smtp and global_smtp_host:
            try:
                fallback_from = global_smtp_from or smtp_from_email
                msg.replace_header("From", fallback_from)
                msg.replace_header("List-Unsubscribe", f"<mailto:{smtp_reply_to or fallback_from}subject=unsubscribe>")

                fallback_implicit_tls = bool(int(global_smtp_port or 0) == 465)
                fallback_starttls = not fallback_implicit_tls

                await aiosmtplib.send(
                    msg,
                    hostname=global_smtp_host,
                    port=global_smtp_port,
                    username=global_smtp_user,
                    password=global_smtp_password,
                    start_tls=fallback_starttls,
                    use_tls=fallback_implicit_tls,
                    timeout=20,
                )
                logger.info("Email sent successfully to %s using global SMTP fallback", to)
                return
            except Exception as fallback_exc:
                logger.error("Global SMTP fallback also failed to %s: %s", to, fallback_exc)

        if raise_on_error:
            raise






async def send_certificate_delivery_email_task(
    *,
    event_id: int,
    cert_uuid: str,
    recipient_name: str,
    recipient_email: str,
) -> None:
    if not recipient_email:
        return

    async with SessionLocal() as db_email:
        event_res = await db_email.execute(select(Event).where(Event.id == event_id))
        event = event_res.scalar_one_or_none()
        if not event or not event.auto_email_on_cert or not event.cert_email_template_id:
            return

        template_res = await db_email.execute(
            select(EmailTemplate).where(EmailTemplate.id == event.cert_email_template_id)
        )
        template = template_res.scalar_one_or_none()
        if not template:
            logger.warning("Certificate delivery email skipped: template=%s not found", event.cert_email_template_id)
            return

        cert_res = await db_email.execute(
            select(Certificate).where(
                Certificate.uuid == cert_uuid,
                Certificate.event_id == event_id,
                Certificate.deleted_at.is_(None),
            )
        )
        cert = cert_res.scalar_one_or_none()
        if not cert:
            logger.warning("Certificate delivery email skipped: cert=%s not found", cert_uuid)
            return

        pdf_rel_path = f"pdfs/event_{event_id}/{cert_uuid}.pdf"
        pdf_abs_path = Path(settings.local_storage_dir) / pdf_rel_path
        png_rel_path = _certificate_png_rel_path(event_id, cert_uuid)
        png_abs_path = Path(settings.local_storage_dir) / png_rel_path

        attachments: List[tuple[str, bytes, str]] = []
        safe_public_id = cert.public_id or cert.uuid
        if pdf_abs_path.exists():
            attachments.append((f"certificate-{safe_public_id}.pdf", pdf_abs_path.read_bytes(), "application/pdf"))
        if png_abs_path.exists():
            attachments.append((f"certificate-{safe_public_id}.png", png_abs_path.read_bytes(), "image/png"))

        verify_url = build_certificate_verify_url(cert.uuid)
        pdf_url = cert.pdf_url if cert.status == CertStatus.active else None
        png_url = _certificate_png_public_url(event_id, cert_uuid)
        event_public_id = _get_public_event_identifier(event)
        event_link = f"{settings.frontend_base_url.rstrip('/')}/events/{event_public_id}/register"
        linkedin_share_link = build_linkedin_share_url(
            verify_url,
            f"{event.name} sertifikamı HeptaCert üzerinden doğrulanabilir şekilde paylaşıyorum.",
        )
        template_vars = {
            "recipient_name": recipient_name,
            "recipient_email": recipient_email,
            "event_name": event.name,
            "event_date": event.event_date.isoformat() if event.event_date else "",
            "event_location": event.event_location or "",
            "certificate_link": verify_url,
            "certificate_verify_url": verify_url,
            "certificate_pdf_url": pdf_url or verify_url,
            "certificate_png_url": png_url or pdf_url or verify_url,
            "certificate_public_id": cert.public_id or "",
            "certificate_uuid": cert.uuid,
            "event_link": event_link,
            "registration_link": event_link,
            "linkedin_share_link": linkedin_share_link,
            "linkedin_share_url": linkedin_share_link,
            "wallet_link": f"{settings.frontend_base_url.rstrip('/')}/profile",
            "logo_url": f"{settings.public_base_url}/static/images/email-logo.png",
        }

        try:
            await send_email_async(
                to=recipient_email,
                subject=template.subject_tr,
                html_body=template.body_html,
                template_vars=template_vars,
                attachments=attachments or None,
                raise_on_error=False,
                sender_user_id=event.admin_id,
            )
        except Exception as exc:
            logger.error("Automatic certificate email failed for %s: %s", recipient_email, exc)




async def trigger_webhooks(
    user_id: int,
    event_type: str,
    payload: Dict[str, Any],
    db_session: Optional[Any] = None,
) -> None:
    """
    Trigger all active webhook subscriptions for a given event type.
    
    Args:
        user_id: User ID to filter webhooks
        event_type: Event type (e.g., "email.sent", "email.failed", "email.opened")
        payload: Data to send in webhook payload
        db_session: Optional database session (creates one if not provided)
    """
    import hmac
    import hashlib
    import httpx
    
    # Get webhooks from database
    from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
    
    async with SessionLocal() as db:
        res = await db.execute(
            select(WebhookSubscription).where(
                (WebhookSubscription.user_id == user_id) &
                (WebhookSubscription.event_type == event_type) &
                (WebhookSubscription.is_active == True)
            )
        )
        webhooks = res.scalars().all()

    from .notification_integrations_api import trigger_notification_integrations_for_user
    async with SessionLocal() as db_notifications:
        await trigger_notification_integrations_for_user(db_notifications, user_id, event_type, payload)
    
    if not webhooks:
        return
    
    # Send to each webhook
    async with httpx.AsyncClient(timeout=10.0) as client:
        for webhook in webhooks:
            try:
                WebhookEndpointIn(url=webhook.url, events=[])
                # Create HMAC signature
                signature = ""
                if webhook.secret:
                    payload_json = json.dumps(payload)
                    signature = hmac.new(
                        webhook.secret.encode(),
                        payload_json.encode(),
                        hashlib.sha256
                    ).hexdigest()
                else:
                    payload_json = json.dumps(payload)
                
                # Send POST request
                headers = {
                    "Content-Type": "application/json",
                    "X-Webhook-Event": event_type,
                    "X-Webhook-Timestamp": datetime.now(timezone.utc).isoformat(),
                }
                
                if signature:
                    headers["X-Webhook-Signature"] = f"sha256={signature}"
                
                resp = await client.post(
                    webhook.url,
                    json=payload,
                    headers=headers,
                )
                response_body = resp.text if resp.status_code < 300 else None
                    
                # Log the delivery
                await log_webhook_delivery(
                    webhook_id=webhook.id,
                    event_type=event_type,
                    payload=payload,
                    http_status=resp.status_code,
                    error_message=response_body if resp.status_code >= 400 else None,
                )
                    
                if resp.status_code >= 400:
                    logger.warning(
                        "Webhook delivery failed for webhook %d: HTTP %d",
                        webhook.id, resp.status_code
                    )
                else:
                    logger.info(
                        "Webhook delivered successfully for webhook %d",
                        webhook.id
                    )
            
            except asyncio.TimeoutError:
                await log_webhook_delivery(
                    webhook_id=webhook.id,
                    event_type=event_type,
                    payload=payload,
                    http_status=None,
                    error_message="Request timeout (10s)",
                )
                logger.warning("Webhook request timeout for webhook %d", webhook.id)
            
            except Exception as e:
                await log_webhook_delivery(
                    webhook_id=webhook.id,
                    event_type=event_type,
                    payload=payload,
                    http_status=None,
                    error_message=str(e),
                )
                logger.error("Webhook trigger failed for webhook %d: %s", webhook.id, e)




# Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬








# get_db -> db.py (yukarida import edildi)






from fastapi import Header as FastAPIHeader



















async def require_paid_plan(
    request: Request,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Paid event features use the owner plan; collaborators require owner Enterprise."""
    if me.role == Role.superadmin:
        return me

    billing_user_id = me.id
    allowed_plans = {"pro", "growth", "enterprise"}
    event_id_raw = request.path_params.get("event_id")
    if event_id_raw is not None:
        try:
            event_id = int(event_id_raw)
        except (TypeError, ValueError):
            raise HTTPException(status_code=404, detail="Event not found")
        event_owner_res = await db.execute(select(Event.admin_id).where(Event.id == event_id))
        event_owner_id = event_owner_res.scalar_one_or_none()
        if event_owner_id is None:
            raise HTTPException(status_code=404, detail="Event not found")
        billing_user_id = int(event_owner_id)
        owner = await db.get(User, billing_user_id)
        if owner and owner.role == Role.superadmin:
            return me
        if billing_user_id != me.id:
            allowed_plans = {"enterprise"}

    res = await db.execute(
        select(Subscription)
        .where(Subscription.user_id == billing_user_id, Subscription.is_active == True)
        .order_by(Subscription.expires_at.desc())
        .limit(1)
    )
    sub = res.scalar_one_or_none()
    if not _subscription_is_active_plan(sub, allowed_plans):
        if event_id_raw is not None and allowed_plans == {"enterprise"}:
            raise HTTPException(
                status_code=403,
                detail="Organizasyon calisanlari ve ekip uyeleri icin etkinlik sahibinin Enterprise planda olmasi gerekir.",
            )
        raise HTTPException(
            status_code=403,
            detail="Bu ozellik sadece Pro, Growth ve Enterprise planlarında kullanılabilir.",
        )
    now = datetime.now(timezone.utc)
    expires_at = ensure_utc(sub.expires_at)
    if expires_at and expires_at < now:
        raise HTTPException(
            status_code=403,
            detail="Aboneliğiniz sona ermiş. Lutfen planınızı yenileyin.",
        )
    return me


async def require_email_system_access(
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Email system features (bulk mail, templates, etc) require Growth or Enterprise plan. Superadmins bypass."""
    if me.role == Role.superadmin:
        return me
    res = await db.execute(
        select(Subscription)
        .where(Subscription.user_id == me.id, Subscription.is_active == True)
        .order_by(Subscription.expires_at.desc())
        .limit(1)
    )
    sub = res.scalar_one_or_none()
    if not sub or sub.plan_id not in ("growth", "enterprise"):
        raise HTTPException(
            status_code=403,
            detail="Oto-mail sistemi Growth ve Enterprise planlarında kullanılabilir.",
        )
    now = datetime.now(timezone.utc)
    expires_at = ensure_utc(sub.expires_at)
    if expires_at and expires_at < now:
        raise HTTPException(
            status_code=403,
            detail="Aboneliğiniz sona ermiş. Lutfen planınızı yenileyin.",
        )
    return me




async def require_event_owner_premium_for_teams(
    event_id: int,
    db: AsyncSession = Depends(get_db),
    me: CurrentUser = Depends(get_current_user),
):
    """Verify event owner has Enterprise plan for team/collaborator features."""
    if me.role == Role.superadmin:
        return True
    
    res = await db.execute(select(Event).where(Event.id == event_id))
    event = res.scalar_one_or_none()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    
    if event.admin_id == me.id:
        # Event owner - check their subscription
        has_premium = await _check_event_owner_has_premium_for_teams(event_id, db)
        if not has_premium:
            raise HTTPException(
                status_code=403,
                detail="Ekip ve calisan ozellikleri sadece Enterprise planda kullanilabilir.",
            )
        return True
    else:
        # Team member - check the EVENT OWNER's subscription, not theirs
        has_premium = await _check_event_owner_has_premium_for_teams(event_id, db)
        if not has_premium:
            raise HTTPException(
                status_code=403,
                detail="Etkinlik sahibinin ekip ve calisan ozellikleri icin Enterprise planda olmasi gerekir.",
            )
        return True












app = FastAPI(title="HeptaCert API", version="2.0.0", docs_url=None, redoc_url=None, openapi_url=None)

# Prefer X-Forwarded-For only when the immediate peer is explicitly configured.
from .ratelimit import (  # ratelimit.py'a tasindi (routers-prep)
    limiter, rate_limit_storage_uri, _rate_limit_key, _client_ip_for_rate_limit,
    _is_trusted_proxy_peer, _heptacert_rate_limit_handler,
)




def _get_registration_device_id(request: Request) -> tuple[str, bool]:
    existing = (request.cookies.get(REGISTRATION_DEVICE_COOKIE) or "").strip()
    if existing and len(existing) <= 128:
        return existing, False
    return secrets.token_urlsafe(24), True






app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _heptacert_rate_limit_handler)
logger.info("Rate limiter storage: %s", "redis" if rate_limit_storage_uri.startswith("redis") else "memory")

# Include domains router (custom domains / Caddy ask endpoint)
try:
    from . import domains as _domains_module  # ensure model is importable
    from . import domains_api as _domains_api  # router
    app.include_router(_domains_api.router)
except Exception:
    # Import errors at startup shouldn't break the app; log and continue.
    logger.debug("domains_api not loaded at startup (will try on demand)")

origins = [o.strip() for o in settings.cors_origins.split(",")] if settings.cors_origins else ["*"]
# When wildcard, allow_credentials must be False (browser blocks credentials+wildcard per CORS spec).
# JWT auth uses Authorization header Ã¢â‚¬â€ no cookies Ã¢â‚¬â€ so credentials=False is fine.
if origins == ["*"]:
    app.add_middleware(
        CORSMiddleware,
        allow_origins=["*"],
        allow_credentials=False,
        allow_methods=["*"],
        allow_headers=["*"],
    )
else:
    app.add_middleware(
        CORSMiddleware,
        allow_origins=origins,
        allow_origin_regex=settings.cors_allow_origin_regex or None,
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
    )


# Ã¢â€â‚¬Ã¢â€â‚¬ Audit log middleware Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬
_AUDIT_SKIP_PREFIXES = (
    "/api/auth/", "/api/billing/webhook/", "/api/files/",
    "/api/verify/", "/api/pricing/", "/api/stats", "/api/billing/status",
    "/api/waitlist",
    "/api/public/attendees/",
    "/api/public/members/",
    "/api/admin/google/sheets/callback",
    "/api/admin/microsoft/excel/callback",
    "/api/admin/microsoft/calendar/callback",
    "/docs", "/openapi", "/redoc",
    "/api/openapi.json",
    "/api/openapi-actions.json",
)


@app.middleware("http")
async def security_headers_middleware(request: Request, call_next):
    response = await call_next(request)
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "DENY")
    response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    response.headers.setdefault("Permissions-Policy", "camera=(self), microphone=(), geolocation=()")
    if settings.public_base_url.lower().startswith("https://"):
        response.headers.setdefault("Strict-Transport-Security", "max-age=31536000; includeSubDomains")
    return response


@app.middleware("http")
async def organization_middleware(request: Request, call_next):
    """Resolve organization by Host header (if set) and attach lightweight info to request.state.organization.

    This avoids repeating the same lookup in many endpoints and centralizes host-based branding.
    """
    request.state.organization = None
    host = request.headers.get("host", "")
    if host:
        domain = host.split(":")[0]
        try:
            async with SessionLocal() as session:
                res = await session.execute(select(Organization).where(Organization.custom_domain == domain))
                org = res.scalar_one_or_none()
                if org:
                    request.state.organization = {
                        "id": org.id,
                        "org_name": org.org_name,
                        "brand_logo": org.brand_logo,
                        "brand_color": org.brand_color,
                        "settings": getattr(org, "settings", {}) or {},
                    }
        except Exception:
            # Fail open: do not block requests if DB lookup fails Ã¢â‚¬â€ logging only.
            logger.debug("organization_middleware: lookup failed for host %s", host)

    # If a human opens an API URL in the browser without Authorization, redirect
    # them to the frontend admin login. This avoids confusing raw JSON errors
    # when someone pastes an API URL into the address bar. Do NOT affect XHR
    # requests (which typically accept application/json).
    try:
        path = request.url.path or ""
        auth_hdr = request.headers.get("authorization") or request.headers.get("Authorization")
        accept = request.headers.get("accept", "")
        # Avoid redirecting public human-facing API endpoints (files, verify, etc.)
        if (
            request.method == "GET"
            and path.startswith("/api/")
            and not auth_hdr
            and "text/html" in accept
            and not any(path.startswith(p) for p in _AUDIT_SKIP_PREFIXES)
        ):
            return RedirectResponse(url=f"{settings.frontend_base_url.rstrip('/')}/admin/login", status_code=302)
    except Exception:
        pass

    response = await call_next(request)
    method = request.method
    path = request.url.path

    if method in ("POST", "PATCH", "PUT", "DELETE") and not any(
        path.startswith(p) for p in _AUDIT_SKIP_PREFIXES
    ):
        # Extract user from JWT without raising
        user_id: Optional[int] = None
        auth = request.headers.get("Authorization", "")
        if auth.lower().startswith("bearer "):
            token = auth.split(" ", 1)[1]
            try:
                if not token.startswith("hc_live_"):
                    payload = jwt.decode(token, settings.jwt_secret, algorithms=["HS256"])
                    user_id = int(payload.get("sub", 0)) or None
            except Exception:
                pass

        # Determine resource type/id from path
        parts = [p for p in path.split("/") if p]
        resource_type: Optional[str] = None
        resource_id: Optional[str] = None
        if len(parts) >= 3:
            resource_type = parts[2]  # e.g. "events", "certificates"
        if len(parts) >= 4:
            resource_id = parts[3]

        ip = _client_ip_for_rate_limit(request)

        async with SessionLocal() as db:
            try:
                db.add(AuditLog(
                    user_id=user_id,
                    action=f"{method} {path}",
                    resource_type=resource_type,
                    resource_id=resource_id,
                    ip_address=ip,
                    user_agent=request.headers.get("User-Agent", "")[:512],
                    extra={"status_code": response.status_code},
                ))
                await db.commit()
            except Exception as exc:
                logger.debug("Audit log write failed (non-critical): %s", exc)
                await db.rollback()

    return response


@app.get("/api/events/{event_id}/capacities")
async def public_event_capacities(event_id: str, db: AsyncSession = Depends(get_db)):
    ev = await _resolve_public_event(db, event_id)
    if not ev:
        raise HTTPException(status_code=404, detail="Event not found")
    caps = await _get_event_capacities(ev.id, db)
    return caps


@app.on_event("startup")
async def startup():
    ensure_dirs()

    async with SessionLocal() as db:
        res = await db.execute(select(User).where(User.email == str(settings.bootstrap_superadmin_email)))
        exists = res.scalar_one_or_none()
        if not exists:
            u = User(
                email=str(settings.bootstrap_superadmin_email),
                password_hash=hash_password(settings.bootstrap_superadmin_password),
                role=Role.superadmin,
                heptacoin_balaonce=0,
                is_verified=True,
            )
            db.add(u)
            await db.commit()

    # Fix any stored URLs that still refereonce old ports/hosts
    # (e.g. localhost:8000 from before port change to 8765)
    old_prefixes = ["http://localhost:8000", "http://localhost:3000"]
    new_base = settings.public_base_url  # e.g. http://localhost:8765
    async with SessionLocal() as db:
        from sqlalchemy import text as sa_text
        for old in old_prefixes:
            if old != new_base:
                await db.execute(sa_text(
                    "UPDATE events SET template_image_url = replace(template_image_url, :old, :new) "
                    "WHERE template_image_url LIKE :pattern"
                ), {"old": old, "new": new_base, "pattern": f"{old}%"})
                await db.execute(sa_text(
                    "UPDATE certificates SET pdf_url = replace(pdf_url, :old, :new) "
                    "WHERE pdf_url LIKE :pattern"
                ), {"old": old, "new": new_base, "pattern": f"{old}%"})
        await db.commit()
        logger.info("Startup URL migration complete.")

    # Seed default certificate templates
    async with SessionLocal() as db:
        cert_temp_count = await db.execute(select(func.count(CertificateTemplate.id)))
        if cert_temp_count.scalar() == 0:
            templates = [
                CertificateTemplate(
                    name="Minimalist",
                    template_image_url=f"{settings.public_base_url}/api/files/cert-templates/minimalist.svg",
                    config={
                        "isim_x": 300, "isim_y": 400,
                        "font_size": 48, "font_color": "#333333",
                        "name_text_align": "center", "name_font_weight": "bold",
                        "qr_x": 50, "qr_y": 650, "qr_size": 200,
                        "cert_id_x": 50, "cert_id_y": 50, "cert_id_font_size": 18,
                        "show_hologram": False,
                    },
                    is_default=True,
                    order_index=0,
                ),
                CertificateTemplate(
                    name="Profesyonel",
                    template_image_url=f"{settings.public_base_url}/api/files/cert-templates/professional.svg",
                    config={
                        "isim_x": 400, "isim_y": 380,
                        "font_size": 56, "font_color": "#1a4d99",
                        "name_text_align": "center", "name_font_weight": "bold", "name_font_style": "italic",
                        "qr_x": 80, "qr_y": 680, "qr_size": 220,
                        "cert_id_x": 80, "cert_id_y": 80, "cert_id_font_size": 20,
                        "show_hologram": True,
                    },
                    is_default=True,
                    order_index=1,
                ),
                CertificateTemplate(
                    name="Renkli",
                    template_image_url=f"{settings.public_base_url}/api/files/cert-templates/colorful.svg",
                    config={
                        "isim_x": 350, "isim_y": 420,
                        "font_size": 52, "font_color": "#ff6b6b",
                        "name_text_align": "center", "name_font_weight": "bold",
                        "qr_x": 100, "qr_y": 700, "qr_size": 200,
                        "cert_id_x": 60, "cert_id_y": 60, "cert_id_font_size": 22,
                        "show_hologram": True,
                    },
                    is_default=True,
                    order_index=2,
                ),
                CertificateTemplate(
                    name="Kurumsal",
                    template_image_url=f"{settings.public_base_url}/api/files/cert-templates/corporate.svg",
                    config={
                        "isim_x": 450, "isim_y": 400,
                        "font_size": 48, "font_color": "#2c3e50",
                        "name_text_align": "center", "name_font_weight": "normal",
                        "qr_x": 50, "qr_y": 680, "qr_size": 180,
                        "cert_id_x": 50, "cert_id_y": 50, "cert_id_font_size": 20,
                        "show_hologram": True,
                    },
                    is_default=True,
                    order_index=3,
                ),
                CertificateTemplate(
                    name="Modern",
                    template_image_url=f"{settings.public_base_url}/api/files/cert-templates/modern.svg",
                    config={
                        "isim_x": 400, "isim_y": 380,
                        "font_size": 54, "font_color": "#6366f1",
                        "name_text_align": "center", "name_font_weight": "bold",
                        "qr_x": 120, "qr_y": 700, "qr_size": 200,
                        "cert_id_x": 70, "cert_id_y": 70, "cert_id_font_size": 19,
                        "show_hologram": False,
                    },
                    is_default=True,
                    order_index=4,
                ),
                CertificateTemplate(
                    name="Elegant",
                    template_image_url=f"{settings.public_base_url}/api/files/cert-templates/elegant.svg",
                    config={
                        "isim_x": 425, "isim_y": 400,
                        "font_size": 50, "font_color": "#8b7355",
                        "name_text_align": "center", "name_font_weight": "bold", "name_font_style": "italic",
                        "qr_x": 75, "qr_y": 710, "qr_size": 190,
                        "cert_id_x": 60, "cert_id_y": 60, "cert_id_font_size": 21,
                        "show_hologram": True,
                    },
                    is_default=True,
                    order_index=5,
                ),
                CertificateTemplate(
                    name="Akademik",
                    template_image_url=f"{settings.public_base_url}/api/files/cert-templates/academic.svg",
                    config={
                        "isim_x": 400, "isim_y": 420,
                        "font_size": 56, "font_color": "#003d5c",
                        "name_text_align": "center", "name_font_weight": "bold",
                        "qr_x": 60, "qr_y": 700, "qr_size": 210,
                        "cert_id_x": 80, "cert_id_y": 80, "cert_id_font_size": 22,
                        "show_hologram": True,
                    },
                    is_default=True,
                    order_index=6,
                ),
            ]
            for tmpl in templates:
                db.add(tmpl)
            await db.commit()
            logger.info("Seeded %d default certificate templates", len(templates))

    # Seed default email templates
    async with SessionLocal() as db:
        email_temp_count = await db.execute(
            select(func.count(EmailTemplate.id)).where(EmailTemplate.template_type == "system")
        )
        if email_temp_count.scalar() == 0:
            # Get a superadmin to assign as creator (use first superadmin or create one)
            superadmin_res = await db.execute(
                select(User).where(User.role == Role.superadmin).limit(1)
            )
            superadmin = superadmin_res.scalar_one_or_none()
            if superadmin:
                email_templates = [
                    EmailTemplate(
                        event_id=None,
                        created_by=superadmin.id,
                        name="Sertifika Teslim - TR",
                        subject_tr="Sertifikanız Hazır! | {{event_name}}",
                        subject_en="Your Certificate is Ready! | {{event_name}}",
                        body_html="""
<h2>Merhaba {{recipient_name}},</h2>
<p>Tebrikler! {{event_name}} etkinliğine katılım icin sertifikanız hazır.</p>

<div style="margin: 20px 0; padding: 15px; background: #f0f9ff; border-radius: 5px;">
    <p><a href="{{certificate_link}}" style="display: inline-block; padding: 10px 20px; background: #3b82f6; color: white; text-decoration: none; border-radius: 5px;">Sertifikayı İndir</a></p>
</div>

<p><strong>QR Kod ile Doğrulama:</strong></p>
<p>Sertifikanız QR kodu tarafından korunmaktadır ve resmi olarak doğrulanabilir.</p>

<br>
<p>Sorularınız icin <a href="mailto:contact@heptapusgroup.com">contact@heptapusgroup.com</a> adresine yazabilirsiniz.</p>

<p>Saygılarımızla,<br>HeptaCert Ekibi</p>
                        """,
                        template_type="system",
                        is_default=True,
                    ),
                    EmailTemplate(
                        event_id=None,
                        created_by=superadmin.id,
                        name="Kayıt Onayı - TR",
                        subject_tr="Kaydınız Başarıyla Alındı | {{event_name}}",
                        subject_en="Your Registration is Confirmed | {{event_name}}",
                        body_html="""
<h2>Merhaba {{recipient_name}},</h2>
<p>{{event_name}} etkinliğine kaydınız başarıyla tamamlanmıştır.</p>

<p><strong>Etkinlik Detayları:</strong></p>
<ul>
    <li><strong>Tarih:</strong> {{event_date}}</li>
    <li><strong>Yer:</strong> {{event_location}}</li>
</ul>

<p>Etkinlik hakkında daha fazla bilgi icin lutfen <a href="{{event_link}}">buraya tıklayın</a>.</p>

<br>
<p>Sorularınız icin <a href="mailto:contact@heptapusgroup.com">contact@heptapusgroup.com</a> adresine yazabilirsiniz.</p>

<p>Saygılarımızla,<br>HeptaCert Ekibi</p>
                        """,
                        template_type="system",
                        is_default=True,
                    ),
                ]
                for tmpl in email_templates:
                    db.add(tmpl)
                await db.commit()
                logger.info("Seeded %d default email templates", len(email_templates))

    # Repair older local databases where the default Turkish templates were seeded with mojibake.
    async with SessionLocal() as db:
        system_templates_res = await db.execute(
            select(EmailTemplate)
            .where(EmailTemplate.template_type == "system", EmailTemplate.is_default == True)
            .order_by(EmailTemplate.created_at.asc(), EmailTemplate.id.asc())
        )
        system_templates = system_templates_res.scalars().all()
        default_template_updates = [
            {
                "name": "Sertifika Teslim - TR",
                "subject_tr": "Sertifikanız Hazır! | {{event_name}}",
                "subject_en": "Your Certificate is Ready! | {{event_name}}",
                "body_html": """
<h2>Merhaba {{recipient_name}},</h2>
<p>Tebrikler! {{event_name}} etkinliğine katılımınız için sertifikanız hazır.</p>

<div style="margin: 20px 0; padding: 15px; background: #f0f9ff; border-radius: 5px;">
    <p><a href="{{certificate_link}}" style="display: inline-block; padding: 10px 20px; background: #3b82f6; color: white; text-decoration: none; border-radius: 5px;">Sertifikayı İndir</a></p>
</div>

<p><strong>QR Kod ile Doğrulama:</strong></p>
<p>Sertifikanız QR kodu ile korunur ve resmi olarak doğrulanabilir.</p>

<br>
<p>Sorularınız için <a href="mailto:contact@heptapusgroup.com">contact@heptapusgroup.com</a> adresine yazabilirsiniz.</p>

<p>Saygılarımızla,<br>HeptaCert Ekibi</p>
                """,
            },
            {
                "name": "Kayıt Onayı - TR",
                "subject_tr": "Kaydınız Başarıyla Alındı | {{event_name}}",
                "subject_en": "Your Registration is Confirmed | {{event_name}}",
                "body_html": """
<h2>Merhaba {{recipient_name}},</h2>
<p>{{event_name}} etkinliğine kaydınız başarıyla tamamlanmıştır.</p>

<p><strong>Etkinlik Detayları:</strong></p>
<ul>
    <li><strong>Tarih:</strong> {{event_date}}</li>
    <li><strong>Yer:</strong> {{event_location}}</li>
</ul>

<p>Etkinlik hakkında daha fazla bilgi için lütfen <a href="{{event_link}}">buraya tıklayın</a>.</p>

<br>
<p>Sorularınız için <a href="mailto:contact@heptapusgroup.com">contact@heptapusgroup.com</a> adresine yazabilirsiniz.</p>

<p>Saygılarımızla,<br>HeptaCert Ekibi</p>
                """,
            },
        ]
        repaired = 0
        for template, update_data in zip(system_templates[:2], default_template_updates):
            if template.name != update_data["name"] or "Ä" in (template.body_html or "") or "Ä" in (template.subject_tr or ""):
                template.name = update_data["name"]
                template.subject_tr = update_data["subject_tr"]
                template.subject_en = update_data["subject_en"]
                template.body_html = update_data["body_html"]
                db.add(template)
                repaired += 1
        if repaired:
            await db.commit()
            logger.info("Repaired %d default email templates", repaired)

    # Start background scheduler for expiring cert notifications
    try:
        from apscheduler.schedulers.asyncio import AsyncIOScheduler
        scheduler = AsyncIOScheduler(timezone="UTC")

        async def _notify_expiring_certs():
            """Email admins about certs expiring in 7 days or 1 day."""
            now = datetime.now(timezone.utc)
            thresholds = [now + timedelta(days=7), now + timedelta(days=1)]
            async with SessionLocal() as db:
                for threshold in thresholds:
                    window_start = threshold - timedelta(hours=1)
                    window_end   = threshold + timedelta(hours=1)
                    res = await db.execute(
                        select(Certificate, Event, User)
                        .join(Event, Certificate.event_id == Event.id)
                        .join(User, Event.admin_id == User.id)
                        .where(
                            Certificate.status == CertStatus.active,
                            Certificate.deleted_at.is_(None),
                            Certificate.hosting_ends_at >= window_start,
                            Certificate.hosting_ends_at <= window_end,
                        )
                    )
                    rows = res.all()
                    # Group by admin
                    by_admin: Dict[int, list] = {}
                    for cert, ev, admin in rows:
                        by_admin.setdefault(admin.id, {"email": admin.email, "certs": []})
                        by_admin[admin.id]["certs"].append({
                            "name": cert.student_name,
                            "event": ev.name,
                            "ends": cert.hosting_ends_at.strftime("%d.%m.%Y"),
                        })
                    for _, data in by_admin.items():
                        days_left = int((threshold - now).total_seconds() / 86400)
                        rows_html = "".join(
                            f"<tr><td>{escape(c['name'])}</td><td>{escape(c['event'])}</td><td>{escape(c['ends'])}</td></tr>"
                            for c in data["certs"]
                        )
                        html = f"""
                        <h2>Ã¢Å¡Â Ã¯Â¸Â BarÃ„Â±ndÃ„Â±rma SÃƒÂ¼resi Doluyor Ã¢â‚¬â€ {days_left} GÃƒÂ¼n</h2>
                        <p>AÃ…Å¸aÃ„Å¸Ã„Â±daki sertifikalarÃ„Â±n barÃ„Â±ndÃ„Â±rma sÃƒÂ¼resi yakÃ„Â±nda dolacak. Yenilemek iÃƒÂ§in panele giriÃ…Å¸ yapÃ„Â±n.</p>
                        <table border="1" cellpadding="6" style="border-collapse:collapse">
                        <tr><th>KatÃ„Â±lÃ„Â±mcÃ„Â±</th><th>Etkinlik</th><th>BitiÃ…Å¸ Tarihi</th></tr>
                        {rows_html}
                        </table>
                        <p><a href="{settings.frontend_base_url}/admin/events">Panele Git â†’</a></p>
                        """
                        await send_email_async(
                            data["email"],
                            f"Ã¢Å¡Â Ã¯Â¸Â HeptaCert: {len(data['certs'])} sertifikanÃ„Â±n barÃ„Â±ndÃ„Â±rma sÃƒÂ¼resi {days_left} gÃƒÂ¼nde doluyor",
                            html,
                        )

        async def _auto_renew_certificates():
            """Renew due certificate hosting for certificates with auto-renew enabled."""
            now = datetime.now(timezone.utc)
            async with SessionLocal() as db_auto:
                res = await db_auto.execute(
                    select(Certificate, Event, User)
                    .join(Event, Certificate.event_id == Event.id)
                    .join(User, Event.admin_id == User.id)
                    .where(
                        Certificate.auto_renew_enabled == True,
                        Certificate.deleted_at.is_(None),
                        Certificate.status != CertStatus.revoked,
                        Certificate.hosting_ends_at.is_not(None),
                        Certificate.hosting_ends_at <= now,
                    )
                    .limit(200)
                )
                for cert, ev, admin in res.all():
                    cost = hosting_units(getattr(cert, "hosting_term", None) or "yearly", int(cert.asset_size_bytes or 0))
                    if admin.heptacoin_balaonce < cost:
                        cert.status = CertStatus.expired
                        logger.warning(
                            "Certificate auto-renew skipped for cert %s: insufficient HC balance on user %s",
                            cert.id,
                            admin.id,
                        )
                        continue

                    admin.heptacoin_balaonce -= cost
                    cert.hosting_ends_at = compute_hosting_ends(getattr(cert, "hosting_term", None) or "yearly")
                    cert.status = CertStatus.active
                    db_auto.add(Transaction(
                        user_id=admin.id,
                        amount=cost,
                        type=TxType.spend,
                        description=f"Certificate hosting auto-renew: {cert.public_id or cert.uuid}",
                    ))
                    logger.info("Certificate auto-renewed: cert=%s event=%s cost=%s", cert.id, ev.id, cost)
                await db_auto.commit()

        async def _monthly_hc_renewal():
            """Credit monthly HC quota to all active paid subscribers."""
            now_r = datetime.now(timezone.utc)
            cutoff = now_r - timedelta(days=30)
            async with SessionLocal() as db_r:
                # Single query: join Subscription + User to avoid N+1
                res_subs = await db_r.execute(
                    select(Subscription, User)
                    .join(User, User.id == Subscription.user_id)
                    .where(
                        Subscription.is_active == True,
                        Subscription.plan_id.in_(["pro", "growth", "enterprise"]),
                        Subscription.expires_at > now_r,
                        User.deleted_at.is_(None),
                    )
                )
                rows = res_subs.all()
                for sub_r2, usr in rows:
                    if sub_r2.last_hc_credited_at and sub_r2.last_hc_credited_at > cutoff:
                        continue
                    quota = _get_hc_quota(sub_r2.plan_id)
                    if not quota:
                        continue
                    usr.heptacoin_balaonce += quota
                    db_r.add(Transaction(
                        user_id=usr.id, amount=quota, type=TxType.credit,
                        description=f"Aylık HC yenileme: {sub_r2.plan_id}",
                    ))
                    sub_r2.last_hc_credited_at = now_r
                    logger.info("Monthly HC renewal: user %s +%d HC (%s)", usr.email, quota, sub_r2.plan_id)
                await db_r.commit()

        async def _process_bulk_emails():
            """Process pending bulk email jobs every 5 minutes."""
            from jinja2 import Template
            
            async with SessionLocal() as db_bulk:
                # Get all pending jobs
                res_jobs = await db_bulk.execute(
                    select(BulkEmailJob)
                    .where(BulkEmailJob.status.in_(["pending", "sending", "scheduled"]))
                    .order_by(BulkEmailJob.created_at.asc())
                    .with_for_update(skip_locked=True)
                    .limit(10)
                )
                jobs = res_jobs.scalars().all()
                
                for job in jobs:
                    try:
                        now_job = datetime.now(timezone.utc)
                        if job.status == "scheduled":
                            is_due = False
                            if job.scheduled_at is not None:
                                scheduled_at = ensure_utc(job.scheduled_at)
                                is_due = bool(scheduled_at and scheduled_at <= now_job)
                            elif job.cron_expression:
                                try:
                                    marker = ensure_utc(job.completed_at or job.started_at or job.created_at) or now_job
                                    trigger = CronTrigger.from_crontab(job.cron_expression, timezone="UTC")
                                    next_fire = trigger.get_next_fire_time(marker, marker)
                                    is_due = bool(next_fire and next_fire <= now_job)
                                except Exception as exc:
                                    job.status = "failed"
                                    job.error_message = f"Invalid cron expression: {exc}"
                                    job.completed_at = now_job
                                    db_bulk.add(job)
                                    await db_bulk.commit()
                                    continue
                            if not is_due:
                                continue

                        # Update job status to sending
                        job.status = "sending"
                        if not job.started_at:
                            job.started_at = now_job
                        job.error_message = None
                        db_bulk.add(job)
                        await db_bulk.commit()
                        
                        # Get event and template
                        ev_res = await db_bulk.execute(
                            select(Event).where(Event.id == job.event_id)
                        )
                        event = ev_res.scalar_one_or_none()
                        if not event:
                            job.status = "failed"
                            job.error_message = "Event not found"
                            db_bulk.add(job)
                            await db_bulk.commit()
                            continue
                        
                        t_res = await db_bulk.execute(
                            select(EmailTemplate).where(EmailTemplate.id == job.email_template_id)
                        )
                        template = t_res.scalar_one_or_none()
                        if not template:
                            job.status = "failed"
                            job.error_message = "Email template not found"
                            db_bulk.add(job)
                            await db_bulk.commit()
                            continue
                        
                        # Get recipients (all attendees or only certified attendees)
                        att_res = await db_bulk.execute(
                            select(Attendee).where(
                                Attendee.event_id == job.event_id,
                                Attendee.email.is_not(None),
                                func.trim(Attendee.email) != "",
                                Attendee.unsubscribed_at.is_(None),
                            )
                        )
                        all_attendees = att_res.scalars().all()

                        cert_res = await db_bulk.execute(
                            select(Certificate)
                            .where(
                                Certificate.event_id == job.event_id,
                                Certificate.status == CertStatus.active,
                                Certificate.deleted_at.is_(None),
                            )
                            .order_by(Certificate.created_at.desc())
                        )
                        cert_rows = cert_res.scalars().all()
                        cert_uuid_by_name: Dict[str, str] = {}
                        for cert in cert_rows:
                            name_key = (cert.student_name or "").strip().lower()
                            if name_key and name_key not in cert_uuid_by_name:
                                cert_uuid_by_name[name_key] = cert.uuid

                        recipient_type = job.recipient_type or "attendees"
                        if recipient_type.startswith("segment:"):
                            from .audience_segments_api import get_segment_attendees

                            segment_key = recipient_type.split(":", 1)[1]
                            attendees = await get_segment_attendees(db_bulk, event, segment_key)
                            attendees = [
                                attendee
                                for attendee in attendees
                                if attendee.email and attendee.email.strip() and attendee.unsubscribed_at is None
                            ]
                        elif recipient_type == "certified":
                            attendees = [
                                attendee
                                for attendee in all_attendees
                                if (attendee.name or "").strip().lower() in cert_uuid_by_name
                            ]
                        else:
                            attendees = all_attendees
                        target_count = len(attendees)
                        job.recipients_count = target_count

                        processed_res = await db_bulk.execute(
                            select(EmailDeliveryLog.attendee_id).where(
                                EmailDeliveryLog.bulk_job_id == job.id,
                                EmailDeliveryLog.status.in_(["sent", "failed"]),
                            )
                        )
                        processed_attendee_ids = {int(item) for item in processed_res.scalars().all() if item}
                        sent_count_res = await db_bulk.execute(
                            select(func.count(EmailDeliveryLog.id)).where(
                                EmailDeliveryLog.bulk_job_id == job.id,
                                EmailDeliveryLog.status == "sent",
                            )
                        )
                        failed_count_res = await db_bulk.execute(
                            select(func.count(EmailDeliveryLog.id)).where(
                                EmailDeliveryLog.bulk_job_id == job.id,
                                EmailDeliveryLog.status == "failed",
                            )
                        )
                        sent = int(sent_count_res.scalar_one() or 0)
                        failed = int(failed_count_res.scalar_one() or 0)
                        attendees = [attendee for attendee in attendees if attendee.id not in processed_attendee_ids]

                        if not attendees and target_count == 0:
                            job.status = "completed"
                            job.sent_count = 0
                            db_bulk.add(job)
                            await db_bulk.commit()
                            continue
                        if not attendees:
                            sent_count_res = await db_bulk.execute(
                                select(func.count(EmailDeliveryLog.id)).where(
                                    EmailDeliveryLog.bulk_job_id == job.id,
                                    EmailDeliveryLog.status == "sent",
                                )
                            )
                            failed_count_res = await db_bulk.execute(
                                select(func.count(EmailDeliveryLog.id)).where(
                                    EmailDeliveryLog.bulk_job_id == job.id,
                                    EmailDeliveryLog.status == "failed",
                                )
                            )
                            job.sent_count = int(sent_count_res.scalar_one() or 0)
                            job.failed_count = int(failed_count_res.scalar_one() or 0)
                            job.status = "scheduled" if job.cron_expression else "completed"
                            job.completed_at = datetime.now(timezone.utc)
                            db_bulk.add(job)
                            await db_bulk.commit()
                            continue
                        
                        # Process in batches for rate limiting
                        batch_size = max(1, int(settings.email_batch_size or 10))
                        attendees = attendees[:batch_size]
                        
                        for i in range(0, len(attendees), batch_size):
                            batch = attendees[i:i+batch_size]
                            
                            for attendee in batch:
                                try:
                                    # Render template with variables
                                    name_key = (attendee.name or "").strip().lower()
                                    public_event_id = _get_public_event_identifier(event)
                                    certificate_link = (
                                        build_certificate_verify_url(cert_uuid_by_name[name_key])
                                        if name_key in cert_uuid_by_name
                                        else f"{settings.frontend_base_url.rstrip('/')}/events/{public_event_id}/register"
                                    )
                                    linkedin_share_link = build_linkedin_share_url(
                                        certificate_link,
                                        f"{event.name} sertifikamı HeptaCert üzerinden doğrulanabilir şekilde paylaşıyorum.",
                                    )
                                    template_vars = {
                                        "recipient_name": attendee.name,
                                        "recipient_email": attendee.email,
                                        "event_name": event.name,
                                        "event_date": event.event_date.isoformat() if event.event_date else "TBD",
                                        "event_location": event.event_location or "Online",
                                        "certificate_link": certificate_link,
                                        "certificate_verify_url": certificate_link,
                                        "certificate_uuid": cert_uuid_by_name.get(name_key, ""),
                                        "event_link": f"{settings.frontend_base_url.rstrip('/')}/events/{public_event_id}/register",
                                        "registration_link": f"{settings.frontend_base_url.rstrip('/')}/events/{public_event_id}/register",
                                        "linkedin_share_link": linkedin_share_link,
                                        "linkedin_share_url": linkedin_share_link,
                                        "wallet_link": f"{settings.frontend_base_url.rstrip('/')}/profile",
                                        "survey_link": build_public_survey_url(
                                            event_id=public_event_id,
                                            attendee_id=attendee.id,
                                            email=attendee.email,
                                        ),
                                    }
                                    
                                    # Render subject and body
                                    subj = Template(template.subject_tr).render(**template_vars)
                                    body = Template(template.body_html).render(**template_vars)

                                    # Create delivery log first to get its ID for tracking pixel
                                    delivery_log = EmailDeliveryLog(
                                        bulk_job_id=job.id,
                                        attendee_id=attendee.id,
                                        recipient_email=attendee.email,
                                        status="pending",
                                    )
                                    db_bulk.add(delivery_log)
                                    await db_bulk.flush()  # assigns delivery_log.id

                                    # Inject tracking pixel if enabled in sender's email config
                                    _email_cfg_res = await db_bulk.execute(
                                        select(UserEmailConfig).where(UserEmailConfig.user_id == event.admin_id)
                                    )
                                    _email_cfg = _email_cfg_res.scalar_one_or_none()
                                    if _email_cfg and _email_cfg.enable_tracking_pixel:
                                        pixel_token = make_email_token({"log_id": delivery_log.id, "action": "open"})
                                        pixel_url = f"{settings.public_base_url}/api/public/track/open/{delivery_log.id}?t={pixel_token}"
                                        body = body.rstrip() + f'\n<img src="{escape(pixel_url)}" width="1" height="1" style="display:none" alt="" />'

                                    # Send mail
                                    await send_email_async(
                                        to=attendee.email,
                                        subject=subj,
                                        html_body=body,
                                        raise_on_error=True,
                                        sender_user_id=event.admin_id,
                                    )

                                    delivery_log.status = "sent"
                                    await db_bulk.commit()
                                    
                                    # Trigger webhook
                                    await trigger_webhooks(
                                        user_id=event.admin_id,
                                        event_type="email.sent",
                                        payload={
                                            "event_id": event.id,
                                            "recipient_email": attendee.email,
                                            "recipient_name": attendee.name,
                                            "sent_at": datetime.now(timezone.utc).isoformat(),
                                        },
                                    )
                                    
                                    sent += 1
                                except Exception as e:
                                    # Update or create delivery log as failed
                                    try:
                                        delivery_log.status = "failed"
                                        delivery_log.reason = str(e)[:500]
                                    except Exception:
                                        delivery_log = EmailDeliveryLog(
                                            bulk_job_id=job.id,
                                            attendee_id=attendee.id,
                                            recipient_email=attendee.email,
                                            status="failed",
                                            reason=str(e)[:500],
                                        )
                                        db_bulk.add(delivery_log)
                                    await db_bulk.commit()
                                    
                                    # Trigger failure webhook
                                    await trigger_webhooks(
                                        user_id=event.admin_id,
                                        event_type="email.failed",
                                        payload={
                                            "event_id": event.id,
                                            "recipient_email": attendee.email,
                                            "recipient_name": attendee.name,
                                            "reason": str(e),
                                            "failed_at": datetime.now(timezone.utc).isoformat(),
                                        },
                                    )
                                    
                                    failed += 1
                                    logger.error("Failed to send email to %s: %s", attendee.email, e)
                            
                            # Update job progress
                            job.sent_count = sent
                            job.failed_count = failed
                            db_bulk.add(job)
                            await db_bulk.commit()
                            
                        # Mark completed only when all targets have a delivery log. Otherwise the next scheduler
                        # cycle will continue with the next chunk.
                        if sent + failed >= target_count:
                            job.status = "scheduled" if job.cron_expression else "completed"
                            job.completed_at = datetime.now(timezone.utc)
                        else:
                            job.status = "sending"
                            job.completed_at = None
                        job.sent_count = sent
                        job.failed_count = failed
                        db_bulk.add(job)
                        await db_bulk.commit()
                        logger.info("Bulk email job %d progressed: %d/%d sent, %d failed", job.id, sent, target_count, failed)
                        
                    except Exception as e:
                        logger.error("Bulk email job %d failed: %s", job.id, e)
                        job.status = "failed"
                        job.error_message = str(e)
                        job.completed_at = datetime.now(timezone.utc)
                        db_bulk.add(job)
                        await db_bulk.commit()

        async def _process_automation_dispatches():
            from .automation_api import process_automation_dispatches_once

            stats = await process_automation_dispatches_once()
            if stats.get("sent") or stats.get("failed"):
                logger.info("Automation dispatch cycle: %s", stats)

        async def _process_training_renewal_notifications():
            from .training_api import process_training_renewal_notifications_once

            stats = await process_training_renewal_notifications_once()
            if stats.get("sent") or stats.get("failed"):
                logger.info("Training renewal notification cycle: %s", stats)

        async def _process_segment_export_jobs():
            from .audience_segments_api import process_segment_export_jobs_once

            stats = await process_segment_export_jobs_once()
            if stats.get("completed") or stats.get("failed"):
                logger.info("Segment export job cycle: %s", stats)

        async def _process_document_export_jobs():
            from .document_export_jobs import process_document_export_jobs_once

            stats = await process_document_export_jobs_once()
            if stats.get("processed") or stats.get("failed"):
                logger.info("Document export job cycle: %s", stats)

        async def _auto_calculate_badges():
            """Auto-calculate badges for events with gamification enabled (every 30 min)."""
            try:
                async with SessionLocal() as db_bg:
                    active_events_res = await db_bg.execute(
                        select(Event.id).where(
                            Event.gamification_enabled.is_(True),
                            Event.deleted_at.is_(None),
                        ).limit(50)
                    )
                    event_ids = active_events_res.scalars().all()
                    total_created = 0
                    for eid in event_ids:
                        rule_res = await db_bg.execute(
                            select(BadgeRule).where(BadgeRule.event_id == eid, BadgeRule.enabled.is_(True))
                        )
                        badge_rule = rule_res.scalar_one_or_none()
                        if not badge_rule or not badge_rule.badge_definitions:
                            continue
                        att_res = await db_bg.execute(select(Attendee).where(Attendee.event_id == eid))
                        attendees = att_res.scalars().all()
                        sessions_res = await db_bg.execute(
                            select(func.count()).select_from(EventSession).where(EventSession.event_id == eid)
                        )
                        total_sessions = sessions_res.scalar() or 0
                        rank_map = {att.id: idx + 1 for idx, att in enumerate(sorted(attendees, key=lambda a: a.registered_at))}
                        for attendee in attendees:
                            ar_res = await db_bg.execute(
                                select(func.count()).select_from(AttendaonceRecord).where(AttendaonceRecord.attendee_id == attendee.id)
                            )
                            sessions_attended = ar_res.scalar() or 0
                            for badge_def in badge_rule.badge_definitions or []:
                                badge_type = badge_def.get("type", "")
                                pb_res = await db_bg.execute(
                                    select(ParticipantBadge).where(
                                        ParticipantBadge.event_id == eid,
                                        ParticipantBadge.attendee_id == attendee.id,
                                        ParticipantBadge.badge_type == badge_type,
                                    )
                                )
                                if pb_res.scalar_one_or_none():
                                    continue
                                criteria = badge_def.get("criteria") or {}
                                passed = True
                                for key, threshold in criteria.items():
                                    if key == "min_sessions" and not (sessions_attended >= int(threshold)):
                                        passed = False; break
                                    elif key in ("attendance_rate", "attendaonce_rate"):
                                        rate = (sessions_attended / total_sessions * 100) if total_sessions > 0 else 0
                                        if not (rate >= float(threshold)):
                                            passed = False; break
                                    elif key == "registered_rank_max" and not (rank_map.get(attendee.id, 9999) <= int(threshold)):
                                        passed = False; break
                                    elif key == "survey_completed" and bool(threshold) and attendee.survey_completed_at is None:
                                        passed = False; break
                                    elif key == "can_download_cert" and bool(threshold) and not attendee.can_download_cert:
                                        passed = False; break
                                if not passed:
                                    continue
                                db_bg.add(ParticipantBadge(
                                    event_id=eid, attendee_id=attendee.id, badge_type=badge_type,
                                    awarded_at=datetime.now(timezone.utc), is_automatic=True,
                                    badge_metadata={
                                        "name": badge_def.get("name") or badge_type.replace("_", " ").title(),
                                        "icon": badge_def.get("icon"),
                                        "color": badge_def.get("color") or "#6366f1",
                                    },
                                ))
                                total_created += 1
                    if total_created:
                        await db_bg.commit()
                        logger.info("Auto badge calculation: %d badges created across %d events", total_created, len(event_ids))
            except Exception as exc:
                logger.warning("Auto badge calculation failed: %s", exc)

        async def _notify_lms_due_dates():
            """Daily: send due-date reminders for LMS assignments 3 days away."""
            try:
                from .lms_extended_models import CourseCalendarEvent
                from .lms_models import CourseEnrollment, TrainingCourse
                now = datetime.now(timezone.utc)
                window_start = now + timedelta(hours=68)
                window_end = now + timedelta(hours=76)
                async with SessionLocal() as db_sch:
                    events_res = await db_sch.execute(
                        select(CourseCalendarEvent).where(
                            CourseCalendarEvent.event_type == "due_date",
                            CourseCalendarEvent.starts_at >= window_start,
                            CourseCalendarEvent.starts_at < window_end,
                        )
                    )
                    cal_events = events_res.scalars().all()
                    for cal in cal_events:
                        course_res = await db_sch.execute(
                            select(TrainingCourse).where(TrainingCourse.id == cal.course_id)
                        )
                        course = course_res.scalar_one_or_none()
                        if not course:
                            continue
                        enrollments_res = await db_sch.execute(
                            select(CourseEnrollment, PublicMember)
                            .join(PublicMember, PublicMember.id == CourseEnrollment.member_id)
                            .where(
                                CourseEnrollment.course_id == cal.course_id,
                                CourseEnrollment.completed_at.is_(None),
                            )
                        )
                        for enrollment, member in enrollments_res.all():
                            if not member.email:
                                continue
                            due_str = cal.starts_at.strftime("%d/%m/%Y %H:%M") if cal.starts_at else ""
                            subject = f"Ödev hatırlatma: {cal.title} — 3 gün kaldı"
                            html_body = f"""
<div style="font-family:Inter,Arial,sans-serif;line-height:1.6;color:#1f2937">
  <h2 style="margin:0 0 12px">Ödev Son Tarihi Yaklaşıyor</h2>
  <p><strong>{cal.title}</strong> ödevi için son tarih <strong>{due_str}</strong>.</p>
  <p>Kurs: <strong>{course.title}</strong></p>
  <p>Ödevi zamanında teslim etmeyi unutmayın!</p>
</div>"""
                            try:
                                await send_email_async(
                                    to_email=member.email,
                                    subject=subject,
                                    html_body=html_body,
                                )
                            except Exception as mail_exc:
                                logger.warning("LMS due-date email failed for %s: %s", member.email, mail_exc)
            except Exception as exc:
                logger.warning("LMS due-date notification job failed: %s", exc)

        if settings.enable_scheduler:
            scheduler.add_job(_notify_expiring_certs, "cron", hour=2, minute=0)
            scheduler.add_job(_auto_renew_certificates, "interval", hours=1)
            scheduler.add_job(_monthly_hc_renewal, "cron", hour=3, minute=30)
            scheduler.add_job(_process_system_digest_emails, "cron", minute=0)
            scheduler.add_job(_process_bulk_emails, "interval", minutes=5)  # Every 5 minutes
            scheduler.add_job(_process_automation_dispatches, "interval", minutes=5)
            scheduler.add_job(_process_training_renewal_notifications, "cron", hour=4, minute=15)
            scheduler.add_job(_process_segment_export_jobs, "interval", seconds=10)
            scheduler.add_job(_process_document_export_jobs, "interval", seconds=10)
            scheduler.add_job(_process_bulk_certificate_jobs, "interval", seconds=3)
            scheduler.add_job(_auto_calculate_badges, "interval", minutes=30)
            scheduler.add_job(_notify_lms_due_dates, "cron", hour=7, minute=0)
            scheduler.start()
            logger.info("APScheduler started Ã¢â‚¬â€ cert notifications + monthly HC renewal + system digest + bulk email processing + bulk certificate queue")
        else:
            logger.info("APScheduler skipped because ENABLE_SCHEDULER=false")
    except Exception as e:
        logger.warning("APScheduler init failed (non-fatal): %s", e)


def _get_hc_quota(plan_id: str) -> Optional[int]:
    """Return the monthly HC quota for a plan from DEFAULT_PRICING."""
    tier = next((t for t in DEFAULT_PRICING if t.get("id") == plan_id), None)
    return tier.get("hc_quota") if tier else None




_SAFE_RASTER_CONTENT_TYPES = {
    "image/jpeg": ".jpg",
    "image/png": ".png",
    "image/webp": ".webp",
}


async def _read_safe_raster_upload(file: UploadFile, max_size: int = 10 * 1024 * 1024) -> Tuple[bytes, str]:
    content_type = str(file.content_type or "").lower().strip()
    ext = _SAFE_RASTER_CONTENT_TYPES.get(content_type)
    if not ext:
        raise bad_request("Only PNG, JPEG or WEBP image uploads are allowed")
    data = await file.read()
    if not data:
        raise bad_request("Image file is empty")
    if len(data) > max_size:
        raise HTTPException(status_code=413, detail="Image exceeds size limit")
    if PILImage is None:
        raise HTTPException(status_code=503, detail="Image validation is unavailable")
    try:
        image = PILImage.open(io.BytesIO(data))
        image.verify()
        if str(image.format or "").upper() not in {"PNG", "JPEG", "WEBP"}:
            raise ValueError("Unsupported raster format")
    except Exception as exc:
        raise bad_request("Invalid image file") from exc
    return data, ext


REGISTRATION_FIELD_TYPES = {"text", "textarea", "number", "tel", "select", "date", "file"}
EVENT_VISIBILITY_VALUES = {"private", "unlisted", "public"}
CERT_TEMPLATE_CONFIG_KEYS = {
    "isim_x",
    "isim_y",
    "qr_x",
    "qr_y",
    "qr_size",
    "font_size",
    "font_color",
    "name_text_align",
    "name_font_weight",
    "name_font_style",
    "cert_id_x",
    "cert_id_y",
    "cert_id_font_size",
    "cert_id_color",
    "show_hologram",
}


def _validate_registration_fields_for_write(raw_fields: Any, *, existing_fields: Optional[List[Dict[str, Any]]] = None) -> List[Dict[str, Any]]:
    existing_fields = existing_fields or []
    if raw_fields is None:
        return existing_fields
    if not isinstance(raw_fields, list):
        raise bad_request("registration_fields must be a list.")
    if existing_fields and not raw_fields:
        raise bad_request("registration_fields cannot be cleared accidentally. Provide valid fields or remove them explicitly in a dedicated flow.")

    def _ctx(index: int, item: Any) -> str:
        if not isinstance(item, dict):
            return f"registration_fields[{index}]"
        field_id = str(item.get("id") or "").strip()
        label = str(item.get("label") or "").strip()
        parts = [f"registration_fields[{index}]"]
        if field_id:
            parts.append(f"id={field_id}")
        if label:
            parts.append(f'label="{label}"')
        return " ".join(parts)

    normalized: List[Dict[str, Any]] = []
    seen_ids: set[str] = set()
    field_types_by_id: Dict[str, str] = {}
    select_options_by_id: Dict[str, set[str]] = {}
    conditional_specs: List[Tuple[int, str, str, str]] = []

    for index, item in enumerate(raw_fields):
        if not isinstance(item, dict):
            raise bad_request(f"{_ctx(index, item)} must be an object.")

        field_id = str(item.get("id") or "").strip()
        if not field_id:
            raise bad_request(f"{_ctx(index, item)}.id is required.")
        if len(field_id) > 64:
            raise bad_request(f"{_ctx(index, item)}.id is too long.")
        if field_id in seen_ids:
            raise bad_request(f"{_ctx(index, item)}.id is duplicated.")

        label = str(item.get("label") or "").strip()
        if not label:
            raise bad_request(f"{_ctx(index, item)}.label is required.")
        if len(label) > 120:
            raise bad_request(f"{_ctx(index, item)}.label is too long.")

        field_type = str(item.get("type") or "text").strip().lower()
        if field_type not in REGISTRATION_FIELD_TYPES:
            raise bad_request(f"{_ctx(index, item)}.type is invalid.")

        placeholder_raw = item.get("placeholder")
        placeholder = None
        if placeholder_raw is not None:
            placeholder = str(placeholder_raw).strip()
            if len(placeholder) > 200:
                raise bad_request(f"{_ctx(index, item)}.placeholder is too long.")
            if not placeholder:
                placeholder = None

        helper_text_raw = item.get("helper_text")
        helper_text = None
        if helper_text_raw is not None:
            helper_text_input = str(helper_text_raw).strip()
            if len(helper_text_input) > 5000:
                raise bad_request(f"{_ctx(index, item)}.helper_text is too long.")
            helper_text = sanitize_event_description_html(helper_text_input)

        raw_options = item.get("options")
        options: List[Dict[str, Any]] = []
        selection_mode = "single"

        if field_type == "select":
            if not isinstance(raw_options, list) or not raw_options:
                raise bad_request(f"{_ctx(index, item)}.options is required for select fields.")
            seen_labels: set[str] = set()
            for opt_index, opt in enumerate(raw_options):
                if isinstance(opt, dict):
                    opt_label = str(opt.get("label") or "").strip()
                    capacity_raw = opt.get("capacity")
                else:
                    opt_label = str(opt or "").strip()
                    capacity_raw = None
                if not opt_label:
                    raise bad_request(f"{_ctx(index, item)}.options[{opt_index}].label is required.")
                if len(opt_label) > 120:
                    raise bad_request(f"{_ctx(index, item)}.options[{opt_index}].label is too long.")
                if opt_label in seen_labels:
                    raise bad_request(f"{_ctx(index, item)}.options[{opt_index}].label is duplicated.")
                if capacity_raw in (None, ""):
                    capacity_val = None
                else:
                    try:
                        capacity_val = int(capacity_raw)
                    except Exception as exc:
                        raise bad_request(f"{_ctx(index, item)}.options[{opt_index}].capacity is invalid.") from exc
                options.append({"label": opt_label, "capacity": capacity_val})
                seen_labels.add(opt_label)
            if len(options) > 30:
                raise bad_request(f"{_ctx(index, item)}.options cannot exceed 30 items.")
            raw_selection_mode = str(item.get("selection_mode") or "single").strip().lower()
            if raw_selection_mode not in {"single", "multiple"}:
                raise bad_request(f"{_ctx(index, item)}.selection_mode is invalid.")
            selection_mode = raw_selection_mode
        elif raw_options not in (None, []):
            raise bad_request(f"{_ctx(index, item)}.options is only valid for select fields.")

        required_when_field_id = str(item.get("required_when_field_id") or "").strip()
        required_when_equals = str(item.get("required_when_equals") or "").strip()
        if bool(required_when_field_id) ^ bool(required_when_equals):
            raise bad_request(f"{_ctx(index, item)} conditional fields must include both required_when_field_id and required_when_equals.")
        if required_when_field_id and len(required_when_field_id) > 64:
            raise bad_request(f"{_ctx(index, item)}.required_when_field_id is too long.")
        if required_when_equals and len(required_when_equals) > 120:
            raise bad_request(f"{_ctx(index, item)}.required_when_equals is too long.")

        normalized_item: Dict[str, Any] = {
            "id": field_id,
            "label": label,
            "type": field_type,
            "required": bool(item.get("required")),
            "placeholder": placeholder,
            "helper_text": helper_text,
        }
        if field_type == "select":
            normalized_item["options"] = options
            normalized_item["selection_mode"] = selection_mode
            select_options_by_id[field_id] = {option["label"] for option in options}
        if required_when_field_id and required_when_equals:
            normalized_item["required_when_field_id"] = required_when_field_id
            normalized_item["required_when_equals"] = required_when_equals
            conditional_specs.append((index, field_id, required_when_field_id, required_when_equals))

        normalized.append(normalized_item)
        seen_ids.add(field_id)
        field_types_by_id[field_id] = field_type

    valid_ids = set(field_types_by_id)
    for index, field_id, cond_id, cond_value in conditional_specs:
        if cond_id not in valid_ids:
            raise bad_request(f"registration_fields[{index}] id={field_id} refereonces an unknown required_when_field_id: {cond_id}.")
        if cond_id == field_id:
            raise bad_request(f"registration_fields[{index}] id={field_id} cannot depend on itself.")
        if field_types_by_id.get(cond_id) != "select":
            raise bad_request(f"registration_fields[{index}] id={field_id} required_when_field_id must refereonce a select field: {cond_id}.")
        if cond_value not in select_options_by_id.get(cond_id, set()):
            raise bad_request(f"registration_fields[{index}] id={field_id} required_when_equals must match one of the refereonced select options on {cond_id}: {cond_value}.")

    return normalized


def _normalize_registration_fields(raw_fields: Any) -> List[Dict[str, Any]]:
    if not isinstance(raw_fields, list):
        return []

    normalized: List[Dict[str, Any]] = []
    seen_ids: set[str] = set()

    for index, item in enumerate(raw_fields):
        if not isinstance(item, dict):
            continue

        field_id = str(item.get("id") or "").strip()[:64]
        if not field_id:
            field_id = f"field_{index + 1}"
        if field_id in seen_ids:
            continue

        label = str(item.get("label") or "").strip()[:120]
        if not label:
            continue

        field_type = str(item.get("type") or "text").strip().lower()
        if field_type not in REGISTRATION_FIELD_TYPES:
            field_type = "text"

        placeholder = str(item.get("placeholder") or "").strip()[:200] or None
        helper_text = sanitize_event_description_html(str(item.get("helper_text") or "").strip()[:5000])
        required = bool(item.get("required"))
        required_when_field_id = str(item.get("required_when_field_id") or "").strip()[:64]
        required_when_equals = str(item.get("required_when_equals") or "").strip()[:120]

        options: List[Dict[str, Any]] = []
        selection_mode = "single"
        raw_options = item.get("options")
        if isinstance(raw_options, list):
            for opt in raw_options:
                if isinstance(opt, dict):
                    lbl = str(opt.get("label") or "").strip()[:120]
                    if not lbl:
                        continue
                    cap = opt.get("capacity")
                    try:
                        cap_val = int(cap) if cap is not None and str(cap).strip() != "" else None
                    except Exception:
                        cap_val = None
                    options.append({"label": lbl, "capacity": cap_val})
                else:
                    s = str(opt or "").strip()[:120]
                    if s:
                        options.append({"label": s, "capacity": None})
        if field_type == "select":
            # dedupe by label keeping first occurreonce
            seen = set()
            deduped: List[Dict[str, Any]] = []
            for o in options:
                if o["label"] in seen:
                    continue
                seen.add(o["label"])
                deduped.append(o)
            options = deduped[:30]
            if not options:
                field_type = "text"
            else:
                raw_selection_mode = str(item.get("selection_mode") or "single").strip().lower()
                selection_mode = "multiple" if raw_selection_mode == "multiple" else "single"

        normalized_item: Dict[str, Any] = {
            "id": field_id,
            "label": label,
            "type": field_type,
            "required": required,
            "placeholder": placeholder,
            "helper_text": helper_text,
        }
        if field_type == "select":
            normalized_item["options"] = options
            normalized_item["selection_mode"] = selection_mode
        if required_when_field_id and required_when_equals:
            normalized_item["required_when_field_id"] = required_when_field_id
            normalized_item["required_when_equals"] = required_when_equals

        normalized.append(normalized_item)
        seen_ids.add(field_id)

    valid_ids = {item["id"] for item in normalized}
    for item in normalized:
        cond_id = str(item.get("required_when_field_id") or "").strip()
        cond_value = str(item.get("required_when_equals") or "").strip()
        if not cond_id or not cond_value or cond_id not in valid_ids or cond_id == item["id"]:
            item.pop("required_when_field_id", None)
            item.pop("required_when_equals", None)

    return normalized


def _get_event_registration_fields(event: Event) -> List[Dict[str, Any]]:
    config = event.config or {}
    return _normalize_registration_fields(config.get("registration_fields"))


def _normalize_event_visibility(raw_visibility: Any) -> str:
    value = str(raw_visibility or "private").strip().lower()
    if value not in EVENT_VISIBILITY_VALUES:
        return "private"
    return value


def _merge_certificate_template_config(event_config: Optional[Dict[str, Any]], template_config: Any) -> Dict[str, Any]:
    next_config = dict(event_config or {})
    if not isinstance(template_config, dict):
        return next_config
    for key in CERT_TEMPLATE_CONFIG_KEYS:
        if key in template_config:
            next_config[key] = template_config[key]
    return next_config


def _get_event_visibility(event: Event) -> str:
    config = event.config or {}
    return _normalize_event_visibility(config.get("visibility"))














def _get_event_organizer_privacy_notice_text(event: Event) -> str:
    config = event.config or {}
    custom = str(config.get("organizer_privacy_notice_text") or "").strip()
    if custom:
        return sanitize_event_description_html(custom) or ""
    return sanitize_event_description_html(
        "ORGANIZATOR AYDINLATMA METNI\n\n"
        "Bu etkinlik kaydi kapsaminda paylastiginiz ad-soyad, iletisim bilgileri, TC kimlik no, pasaport no, "
        "ogrenci no, dogum tarihi, adres ve benzeri etkinlige ozel veriler organizator tarafindan kayit, katilim "
        "takibi, sertifika ve etkinlik yonetimi amaclariyla islenebilir. Bu alanlarin hangi hukuki sebebe dayandigi, "
        "saklama suresi ve paylasim kapsamı organizatorun sorumlulugundadir."
    ) or ""


































async def _get_checkin_context_by_token(checkin_token: str, db: AsyncSession) -> Optional[Dict[str, Any]]:
    now_ts = time.time()
    cached = _checkin_context_cache.get(checkin_token)
    if cached and now_ts < cached[0]:
        return cached[1]

    res = await db.execute(
        select(
            EventSession.id,
            EventSession.name,
            EventSession.session_date,
            EventSession.session_start,
            EventSession.session_location,
            EventSession.is_active,
            Event.id.label("event_id"),
            Event.public_id.label("event_public_id"),
            Event.name.label("event_name"),
            Event.event_date,
            Event.min_sessions_required,
            Event.checkin_enabled,
        )
        .join(Event, Event.id == EventSession.event_id)
        .where(EventSession.checkin_token == checkin_token)
        .limit(1)
    )
    row = res.first()
    if not row:
        return None

    ctx = {
        "session_id": row.id,
        "session_name": row.name,
        "session_date": row.session_date,
        "session_start": row.session_start,
        "session_location": row.session_location,
        "is_active": bool(row.is_active),
        "event_id": row.event_id,
        "event_public_id": row.event_public_id or str(row.event_id),
        "event_name": row.event_name,
        "event_date": row.event_date,
        "min_sessions_required": int(row.min_sessions_required or 0),
        "checkin_enabled": is_checkin_enabled(row),
    }
    _checkin_context_cache[checkin_token] = (now_ts + CHECKIN_CONTEXT_TTL_SECONDS, ctx)
    return ctx


async def _get_event_total_sessions_cached(event_id: int, db: AsyncSession) -> int:
    now_ts = time.time()
    cached = _event_total_sessions_cache.get(event_id)
    if cached and now_ts < cached[0]:
        return cached[1]

    total_sess_res = await db.execute(
        select(func.count()).where(EventSession.event_id == event_id)
    )
    total_sessions = int(total_sess_res.scalar_one() or 0)
    _event_total_sessions_cache[event_id] = (now_ts + EVENT_TOTAL_SESSIONS_TTL_SECONDS, total_sessions)
    return total_sessions








async def _process_one_bulk_certificate_job(job_id: int) -> None:
    ISSUE_UNITS_PER_CERT = 10

    async with SessionLocal() as db_job:
        res_job = await db_job.execute(
            select(BulkCertificateJob).where(BulkCertificateJob.id == job_id).with_for_update()
        )
        job = res_job.scalar_one_or_none()
        if not job:
            return

        if job.status in ["completed", "failed", "cancelled"]:
            return

        if job.status == "pending":
            job.status = "processing"
            job.started_at = datetime.now(timezone.utc)

        names: List[str] = [str(x).strip() for x in (job.names or []) if str(x).strip()]
        total = len(names)
        job.total_count = total

        if job.current_index >= total:
            if not job.zip_file_path and job.generated_files:
                zip_rel_path = f"zips/event_{job.event_id}/bulk_job_{job.id}.zip"
                zip_abs_path = Path(settings.local_storage_dir) / zip_rel_path
                zip_abs_path.parent.mkdir(parents=True, exist_ok=True)
                with zipfile.ZipFile(zip_abs_path, "w", zipfile.ZIP_DEFLATED) as zf:
                    for item in job.generated_files or []:
                        rel_pdf = item.get("rel_pdf_path")
                        filename = item.get("file_name")
                        if not rel_pdf or not filename:
                            continue
                        pdf_abs = Path(settings.local_storage_dir) / rel_pdf
                        if pdf_abs.exists():
                            zf.write(pdf_abs, filename)
                job.zip_file_path = zip_rel_path
            job.status = "completed"
            job.completed_at = datetime.now(timezone.utc)
            db_job.add(job)
            await db_job.commit()
            return

        q_event = select(Event).where(Event.id == job.event_id)
        if job.created_by:
            q_event = q_event.where(Event.admin_id == job.created_by)
        res_ev = await db_job.execute(q_event)
        ev = res_ev.scalar_one_or_none()
        if not ev:
            job.status = "failed"
            job.error_message = "Event not found or access denied"
            job.completed_at = datetime.now(timezone.utc)
            db_job.add(job)
            await db_job.commit()
            return

        try:
            cfg = editor_config_to_template_config(ev.config or {})
        except Exception as e:
            job.status = "failed"
            job.error_message = f"Invalid event config: {e}"
            job.completed_at = datetime.now(timezone.utc)
            db_job.add(job)
            await db_job.commit()
            return

        res_user = await db_job.execute(select(User).where(User.id == job.created_by))
        user = res_user.scalar_one_or_none()
        if not user:
            job.status = "failed"
            job.error_message = "Job owner not found"
            job.completed_at = datetime.now(timezone.utc)
            db_job.add(job)
            await db_job.commit()
            return

        template_path = local_path_from_url(ev.template_image_url)
        if not template_path.exists():
            job.status = "failed"
            job.error_message = "Template image not found"
            job.completed_at = datetime.now(timezone.utc)
            db_job.add(job)
            await db_job.commit()
            return
        template_bytes = template_path.read_bytes()

        org_res = await db_job.execute(select(Organization).where(Organization.user_id == job.created_by))
        org = org_res.scalar_one_or_none()
        brand_logo_bytes: Optional[bytes] = None
        if org and org.brand_logo:
            try:
                logo_path = local_path_from_url(org.brand_logo)
                if logo_path.exists():
                    brand_logo_bytes = logo_path.read_bytes()
            except Exception:
                pass

        start_idx = max(0, int(job.current_index or 0))
        end_idx = min(total, start_idx + max(1, int(job.chunk_size or 10)))

        generated_files = list(job.generated_files or [])

        for idx in range(start_idx, end_idx):
            student_name = names[idx]

            if user.heptacoin_balaonce < ISSUE_UNITS_PER_CERT:
                job.status = "failed"
                job.error_message = f"Insufficient HeptaCoin at index={idx}"
                job.current_index = idx
                job.completed_at = datetime.now(timezone.utc)
                db_job.add(job)
                await db_job.commit()
                return

            cert_check = await db_job.execute(
                select(Certificate).where(
                    Certificate.event_id == ev.id,
                    Certificate.student_name == student_name,
                    Certificate.deleted_at.is_(None),
                )
            )
            if cert_check.scalar_one_or_none():
                job.already_exists_count += 1
                job.current_index = idx + 1
                continue

            try:
                cert_uuid = new_certificate_uuid()

                ev_lock_res = await db_job.execute(select(Event).where(Event.id == ev.id).with_for_update())
                ev_locked = ev_lock_res.scalar_one()
                ev_locked.cert_seq += 1
                public_id = f"EV{ev_locked.id}-{ev_locked.cert_seq:06d}"
                verify_url = build_certificate_verify_url(cert_uuid)

                pdf_bytes = await asyncio.to_thread(
                    render_certificate_pdf,
                    template_image_bytes=template_bytes,
                    student_name=student_name,
                    verify_url=verify_url,
                    config=cfg,
                    public_id=public_id,
                    brand_logo_bytes=brand_logo_bytes,
                )

                rel_pdf_path = f"pdfs/event_{ev.id}/{cert_uuid}.pdf"
                abs_pdf_path = Path(settings.local_storage_dir) / rel_pdf_path
                abs_pdf_path.parent.mkdir(parents=True, exist_ok=True)
                abs_pdf_path.write_bytes(pdf_bytes)
                try:
                    png_bytes = await asyncio.to_thread(
                        render_certificate_png_watermarked,
                        template_image_bytes=template_bytes,
                        student_name=student_name,
                        verify_url=verify_url,
                        config=cfg,
                        public_id=public_id,
                        brand_logo_bytes=brand_logo_bytes,
                    )
                    rel_png_path = _certificate_png_rel_path(ev.id, cert_uuid)
                    abs_png_path = Path(settings.local_storage_dir) / rel_png_path
                    abs_png_path.parent.mkdir(parents=True, exist_ok=True)
                    abs_png_path.write_bytes(png_bytes)
                except Exception as png_ex:
                    logger.warning("Bulk certificate PNG render failed for job=%s idx=%s: %s", job.id, idx, png_ex)

                asset_size_bytes = abs_pdf_path.stat().st_size
                hosting_term = "yearly"
                hosting_spend = hosting_units(hosting_term, asset_size_bytes)
                spend_units = ISSUE_UNITS_PER_CERT + hosting_spend

                if user.heptacoin_balaonce < spend_units:
                    job.status = "failed"
                    job.error_message = f"Insufficient HeptaCoin at index={idx}"
                    job.current_index = idx
                    job.completed_at = datetime.now(timezone.utc)
                    db_job.add(job)
                    await db_job.commit()
                    return

                pdf_url = build_public_pdf_url(rel_pdf_path)
                hosting_ends_at = compute_hosting_ends(hosting_term)

                cert = Certificate(
                    uuid=cert_uuid,
                    public_id=public_id,
                    student_name=student_name,
                    event_id=ev.id,
                    pdf_url=pdf_url,
                    status=CertStatus.active,
                    hosting_term=hosting_term,
                    hosting_ends_at=hosting_ends_at,
                    asset_size_bytes=asset_size_bytes,
                )
                db_job.add(cert)
                await db_job.flush()
                try:
                    await _maybe_log_cpd(db_job, ev.id, cert.id, student_name=student_name)
                except Exception:
                    pass

                user.heptacoin_balaonce -= spend_units
                db_job.add(Transaction(user_id=user.id, amount=spend_units, type=TxType.spend))

                job.created_count += 1
                job.spent_heptacoin += spend_units
                generated_files.append({
                    "rel_pdf_path": rel_pdf_path,
                    "file_name": _safe_cert_filename(student_name, public_id),
                })
                job.generated_files = generated_files
                job.current_index = idx + 1

                attendee_res = await db_job.execute(
                    select(Attendee)
                    .where(
                        Attendee.event_id == ev.id,
                        func.lower(Attendee.name) == (student_name or "").strip().lower(),
                    )
                    .order_by(Attendee.id.asc())
                    .limit(1)
                )
                attendee_for_email = attendee_res.scalar_one_or_none()
                if attendee_for_email and attendee_for_email.email:
                    from .crm_snapshot_hooks import refresh_crm_snapshot_for_attendee
                    await refresh_crm_snapshot_for_attendee(db_job, attendee_for_email)
                    await send_certificate_delivery_email_task(
                        event_id=ev.id,
                        cert_uuid=cert_uuid,
                        recipient_name=attendee_for_email.name,
                        recipient_email=attendee_for_email.email,
                    )
            except Exception as ex:
                logger.error("Bulk certificate job %s failed at idx=%s: %s", job.id, idx, ex)
                job.failed_count += 1
                job.current_index = idx + 1

        if job.current_index >= total:
            zip_rel_path = f"zips/event_{job.event_id}/bulk_job_{job.id}.zip"
            zip_abs_path = Path(settings.local_storage_dir) / zip_rel_path
            zip_abs_path.parent.mkdir(parents=True, exist_ok=True)
            with zipfile.ZipFile(zip_abs_path, "w", zipfile.ZIP_DEFLATED) as zf:
                for item in generated_files:
                    rel_pdf = item.get("rel_pdf_path")
                    filename = item.get("file_name")
                    if not rel_pdf or not filename:
                        continue
                    pdf_abs = Path(settings.local_storage_dir) / rel_pdf
                    if pdf_abs.exists():
                        zf.write(pdf_abs, filename)
            job.zip_file_path = zip_rel_path
            job.status = "completed"
            job.completed_at = datetime.now(timezone.utc)

        db_job.add(job)
        await db_job.commit()


async def _process_bulk_certificate_jobs() -> None:
    if bulk_cert_job_lock.locked():
        return

    async with bulk_cert_job_lock:
        async with SessionLocal() as db_q:
            res = await db_q.execute(
                select(BulkCertificateJob.id)
                .where(BulkCertificateJob.status.in_(["pending", "processing"]))
                .order_by(BulkCertificateJob.created_at.asc())
                .limit(1)
            )
            job_ids = [r[0] for r in res.all()]

        for job_id in job_ids:
            try:
                await _process_one_bulk_certificate_job(job_id)
            except Exception as e:
                logger.error("Bulk certificate queue processor failed for job=%s: %s", job_id, e)


# Ã¢â€â‚¬Ã¢â€â‚¬ Badge Management Endpoints Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬

async def _can_manage_organization_event(db: AsyncSession, me: CurrentUser, owner_user_id: int) -> bool:
    if me.role == Role.superadmin or owner_user_id == me.id:
        return True
    from .organization_access_api import user_can_manage_owner_organization
    return await user_can_manage_owner_organization(db, me, owner_user_id, "events:manage")


@app.post("/api/admin/events/{event_id}/badge-rules", response_model=BadgeRulesOut)
async def create_or_update_badge_rules(
    event_id: int,
    rules_in: BadgeRulesIn,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Create or update badge rules for an event. Admin only."""
    # Check authorization
    e_res = await db.execute(select(Event).where(Event.id == event_id))
    event = e_res.scalar_one_or_none()
    if not event:
        raise HTTPException(status_code=404, detail="Etkinlik bulunamadÃ„Â±")
    
    if not await _can_manage_organization_event(db, current_user, event.admin_id):
        raise HTTPException(status_code=403, detail="Unauthorized")

    if not is_gamification_enabled(event):
        raise HTTPException(status_code=403, detail="Oyunlaştırma bu etkinlikte kapalı")

    # Check if rules already exist
    br_res = await db.execute(
        select(BadgeRule).where(BadgeRule.event_id == event_id)
    )
    badge_rule = br_res.scalar_one_or_none()

    if badge_rule:
        badge_rule.badge_definitions = [d.model_dump() for d in rules_in.badge_definitions]
        badge_rule.enabled = rules_in.enabled
        badge_rule.updated_at = datetime.utcnow()
    else:
        badge_rule = BadgeRule(
            event_id=event_id,
            badge_definitions=[d.model_dump() for d in rules_in.badge_definitions],
            enabled=rules_in.enabled,
            created_by=current_user.id,
            updated_at=datetime.utcnow(),
        )
        db.add(badge_rule)

    await db.commit()
    await db.refresh(badge_rule)
    return badge_rule


@app.get("/api/admin/events/{event_id}/badge-rules", response_model=Optional[BadgeRulesOut])
async def get_badge_rules(
    event_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Get badge rules for an event."""
    e_res = await db.execute(select(Event).where(Event.id == event_id))
    event = e_res.scalar_one_or_none()
    if not event:
        raise HTTPException(status_code=404, detail="Etkinlik bulunamadÃ„Â±")

    if not await _can_manage_organization_event(db, current_user, event.admin_id):
        raise HTTPException(status_code=403, detail="Yetkisiz eriÅŸim")

    if not is_gamification_enabled(event):
        raise HTTPException(status_code=403, detail="Oyunlaştırma bu etkinlikte kapalı")

    br_res = await db.execute(
        select(BadgeRule).where(BadgeRule.event_id == event_id)
    )
    badge_rule = br_res.scalar_one_or_none()
    return badge_rule


@app.post("/api/admin/events/{event_id}/badges", response_model=ParticipantBadgeOut)
async def award_badge_manually(
    event_id: int,
    badge_in: AwardBadgeIn,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Manually award a badge to an attendee."""
    # Check authorization
    e_res = await db.execute(select(Event).where(Event.id == event_id))
    event = e_res.scalar_one_or_none()
    if not event:
        raise HTTPException(status_code=404, detail="Etkinlik bulunamadÃ„Â±")

    if not await _can_manage_organization_event(db, current_user, event.admin_id):
        raise HTTPException(status_code=403, detail="Yetkisiz eriÅŸim")

    if not is_gamification_enabled(event):
        raise HTTPException(status_code=403, detail="Oyunlaştırma bu etkinlikte kapalı")

    # Check attendee exists
    att_res = await db.execute(
        select(Attendee).where(
            Attendee.id == badge_in.attendee_id,
            Attendee.event_id == event_id,
        )
    )
    attendee = att_res.scalar_one_or_none()
    if not attendee:
        raise HTTPException(status_code=404, detail="KatÃ„Â±lÃ„Â±mcÃ„Â± bulunamadÃ„Â±")

    # Check if badge already exists
    pb_res = await db.execute(
        select(ParticipantBadge).where(
            ParticipantBadge.event_id == event_id,
            ParticipantBadge.attendee_id == badge_in.attendee_id,
            ParticipantBadge.badge_type == badge_in.badge_type,
        )
    )
    existing_badge = pb_res.scalar_one_or_none()
    if existing_badge:
        raise HTTPException(status_code=409, detail="Bu rozet zaten veriliÅŸ")

    # Create badge
    new_badge = ParticipantBadge(
        event_id=event_id,
        attendee_id=badge_in.attendee_id,
        badge_type=badge_in.badge_type,
        criteria_met=badge_in.criteria_met,
        awarded_by=current_user.id,
        awarded_at=datetime.utcnow(),
        is_automatic=False,
        badge_metadata=badge_in.badge_metadata,
    )
    db.add(new_badge)
    await db.commit()
    await db.refresh(new_badge)
    return new_badge




@app.get("/api/admin/events/{event_id}/badges")
async def list_badges(
    event_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List all awarded badges for an event."""
    e_res = await db.execute(select(Event).where(Event.id == event_id))
    event = e_res.scalar_one_or_none()
    if not event:
        raise HTTPException(status_code=404, detail="Etkinlik bulunamadÃ„Â±")

    if not await _can_manage_organization_event(db, current_user, event.admin_id):
        raise HTTPException(status_code=403, detail="Yetkisiz eriÅŸim")

    if not is_gamification_enabled(event):
        raise HTTPException(status_code=403, detail="Oyunlaştırma bu etkinlikte kapalı")

    pb_res = await db.execute(
        select(ParticipantBadge)
        .where(ParticipantBadge.event_id == event_id)
        .order_by(ParticipantBadge.awarded_at.desc())
    )
    badges = pb_res.scalars().all()

    by_type: Dict[str, int] = {}
    automatic_count = 0
    manual_count = 0
    for badge in badges:
        by_type[badge.badge_type] = by_type.get(badge.badge_type, 0) + 1
        if badge.is_automatic:
            automatic_count += 1
        else:
            manual_count += 1

    badge_items = await _build_participant_badge_items(db, event_id, badges)

    return {
        "total_badges": len(badges),
        "badges": badge_items,
        "badge_summary": {
            "by_type": by_type,
            "automatic_vs_manual": {"automatic": automatic_count, "manual": manual_count},
        },
    }


@app.get("/api/events/{event_id}/attendees/{attendee_id}/badges")
async def list_public_attendee_badges(
    event_id: str,
    attendee_id: int,
    email: str = Query(..., min_length=3),
    db: AsyncSession = Depends(get_db),
):
    """List awarded badges for a public attendee view."""
    event = await _resolve_public_event(db, event_id)
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    if not is_gamification_enabled(event):
        raise HTTPException(status_code=403, detail="Oyunlaştırma bu etkinlikte kapalı")

    attendee_res = await db.execute(
        select(Attendee).where(
            Attendee.id == attendee_id,
            Attendee.event_id == event.id,
        )
    )
    attendee = attendee_res.scalar_one_or_none()
    if not attendee or attendee.email.lower() != email.lower():
        raise HTTPException(status_code=404, detail="Attendee not found")

    pb_res = await db.execute(
        select(ParticipantBadge)
        .where(
            ParticipantBadge.event_id == event.id,
            ParticipantBadge.attendee_id == attendee_id,
        )
        .order_by(ParticipantBadge.awarded_at.desc())
    )
    badges = pb_res.scalars().all()
    badge_items = await _build_participant_badge_items(db, event.id, badges)

    return {
        "total_badges": len(badge_items),
        "badges": badge_items,
    }


@app.get("/api/events/{event_id}/survey-access")
async def resolve_public_survey_access(
    event_id: str,
    token: str = Query(..., min_length=8),
    db: AsyncSession = Depends(get_db),
):
    """Resolve a signed public survey token to attendee context."""
    event = await _resolve_public_event(db, event_id)
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    try:
        payload = verify_survey_access_token(
            token,
            event_id=event.id,
            event_public_id=_get_public_event_identifier(event),
        )
    except SignatureExpired:
        raise HTTPException(status_code=410, detail="Survey access link expired")
    except BadSignature:
        raise HTTPException(status_code=400, detail="Invalid survey access link")

    attendee_id = int(payload.get("attendee_id") or 0)
    attendee_res = await db.execute(
        select(Attendee).where(
            Attendee.id == attendee_id,
            Attendee.event_id == event.id,
        )
    )
    attendee = attendee_res.scalar_one_or_none()
    if not attendee or attendee.email.lower() != str(payload.get("email") or "").lower():
        raise HTTPException(status_code=404, detail="Attendee not found")

    return {
        "attendee_id": attendee.id,
        "attendee_name": attendee.name,
        "attendee_email": attendee.email,
        "survey_token": token,
    }


@app.get("/api/events/{event_id}/participant-status", response_model=PublicParticipantStatusOut)
async def get_public_participant_status(
    event_id: str,
    token: str = Query(..., min_length=8),
    db: AsyncSession = Depends(get_db),
):
    event = await _resolve_public_event(db, event_id)
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    try:
        payload = verify_survey_access_token(
            token,
            event_id=event.id,
            event_public_id=_get_public_event_identifier(event),
        )
    except SignatureExpired:
        raise HTTPException(status_code=410, detail="Survey access link expired")
    except BadSignature:
        raise HTTPException(status_code=400, detail="Invalid survey access link")

    attendee_id = int(payload.get("attendee_id") or 0)
    attendee_res = await db.execute(
        select(Attendee).where(
            Attendee.id == attendee_id,
            Attendee.event_id == event.id,
        )
    )
    attendee = attendee_res.scalar_one_or_none()
    if not attendee or attendee.email.lower() != str(payload.get("email") or "").lower():
        raise HTTPException(status_code=404, detail="Attendee not found")

    return await build_public_participant_status(db, event=event, attendee=attendee)


@app.get("/api/events/{event_id}/participant-status/me", response_model=PublicParticipantStatusOut)
async def get_my_public_participant_status(
    event_id: str,
    member: CurrentPublicMember = Depends(get_current_public_member),
    db: AsyncSession = Depends(get_db),
):
    event = await _resolve_public_event(db, event_id)
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    legacy_res = await db.execute(
        select(Attendee).where(
            Attendee.event_id == event.id,
            func.lower(Attendee.email) == member.email.lower(),
            Attendee.public_member_id.is_(None),
        )
    )
    legacy_attendees = legacy_res.scalars().all()
    if legacy_attendees:
        for legacy_attendee in legacy_attendees:
            legacy_attendee.public_member_id = member.id
        await db.commit()

    attendee_res = await db.execute(
        select(Attendee)
        .where(
            Attendee.event_id == event.id,
            Attendee.public_member_id == member.id,
        )
        .order_by(Attendee.registered_at.desc())
        .limit(1)
    )
    attendee = attendee_res.scalar_one_or_none()
    if not attendee:
        raise HTTPException(status_code=404, detail="Attendee not found")

    return await build_public_participant_status(db, event=event, attendee=attendee)


@app.post("/api/admin/events/{event_id}/badges/calculate")
async def trigger_automatic_badge_calculation(
    event_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Trigger automatic badge calculation for all attendees in an event."""
    e_res = await db.execute(select(Event).where(Event.id == event_id))
    event = e_res.scalar_one_or_none()
    if not event:
        raise HTTPException(status_code=404, detail="Etkinlik bulunamadÃ„Â±")

    if not await _can_manage_organization_event(db, current_user, event.admin_id):
        raise HTTPException(status_code=403, detail="Yetkisiz eriÅŸim")

    if not is_gamification_enabled(event):
        raise HTTPException(status_code=403, detail="Oyunlaştırma bu etkinlikte kapalı")

    # Get badge rules
    br_res = await db.execute(
        select(BadgeRule).where(BadgeRule.event_id == event_id)
    )
    badge_rule = br_res.scalar_one_or_none()
    if not badge_rule:
        raise HTTPException(status_code=404, detail="Bu etkinlik iÃƒÂ§in rozet kurallarÃ„Â± belirlenmemiÃ…Å¸")

    if not badge_rule.enabled:
        raise HTTPException(status_code=400, detail="Rozet sistemi bu etkinlik iÃƒÂ§in devre dÃ„Â±Ã…Å¸Ã„Â±")
    if not badge_rule.badge_definitions:
        raise HTTPException(status_code=400, detail="Hesaplama iÃƒÂ§in en az bir rozet tanÃ„Â±mÃ„Â± gerekli")

    # Get all attendees
    att_res = await db.execute(
        select(Attendee).where(Attendee.event_id == event_id)
    )
    attendees = att_res.scalars().all()

    # Pre-compute total session count for attendaonce_rate
    sessions_res = await db.execute(
        select(func.count()).select_from(EventSession).where(EventSession.event_id == event_id)
    )
    total_sessions = sessions_res.scalar() or 0

    # Pre-compute registration ranks (ordered by registered_at asc)
    rank_map: dict[int, int] = {
        att.id: idx + 1
        for idx, att in enumerate(sorted(attendees, key=lambda a: a.registered_at))
    }

    created_count = 0
    for attendee in attendees:
        # Load attendaonce records for this attendee oonce
        ar_res = await db.execute(
            select(func.count()).select_from(AttendaonceRecord).where(
                AttendaonceRecord.attendee_id == attendee.id
            )
        )
        sessions_attended = ar_res.scalar() or 0

        # For each badge definition, check if criteria are met
        for badge_def in badge_rule.badge_definitions or []:
            badge_type = badge_def.get("type", "")

            # Check if badge already exists
            pb_res = await db.execute(
                select(ParticipantBadge).where(
                    ParticipantBadge.event_id == event_id,
                    ParticipantBadge.attendee_id == attendee.id,
                    ParticipantBadge.badge_type == badge_type,
                )
            )
            if pb_res.scalar_one_or_none():
                continue  # Already has this badge

            # Evaluate each criterion
            criteria = badge_def.get("criteria") or {}
            passed = True
            criteria_met: dict = {}

            for key, threshold in criteria.items():
                if key == "min_sessions":
                    ok = sessions_attended >= int(threshold)
                    criteria_met[key] = {"required": threshold, "actual": sessions_attended, "passed": ok}
                    if not ok:
                        passed = False

                elif key in ("attendance_rate", "attendaonce_rate"):  # support both old typo and corrected key
                    rate = (sessions_attended / total_sessions * 100) if total_sessions > 0 else 0
                    ok = rate >= float(threshold)
                    criteria_met["attendance_rate"] = {"required": threshold, "actual": round(rate, 1), "passed": ok}
                    if not ok:
                        passed = False

                elif key == "registered_rank_max":
                    rank = rank_map.get(attendee.id, 9999)
                    ok = rank <= int(threshold)
                    criteria_met[key] = {"required": threshold, "actual": rank, "passed": ok}
                    if not ok:
                        passed = False

                elif key == "survey_completed":
                    required = bool(threshold)
                    actual = attendee.survey_completed_at is not None
                    ok = actual == required if required else True
                    criteria_met[key] = {"required": required, "actual": actual, "passed": ok}
                    if not ok:
                        passed = False

                elif key == "can_download_cert":
                    required = bool(threshold)
                    ok = attendee.can_download_cert == required if required else True
                    criteria_met[key] = {"required": required, "actual": attendee.can_download_cert, "passed": ok}
                    if not ok:
                        passed = False

            if not passed:
                continue

            badge = ParticipantBadge(
                event_id=event_id,
                attendee_id=attendee.id,
                badge_type=badge_type,
                criteria_met=criteria_met,
                awarded_at=datetime.now(timezone.utc),
                is_automatic=True,
                badge_metadata={
                    "name": badge_def.get("name") or badge_type.replace("_", " ").title(),
                    "icon": badge_def.get("icon"),
                    "color": badge_def.get("color") or "#6366f1",
                    "description": badge_def.get("description"),
                    "calculation_rule_version": "1.1",
                    "calculated_at": datetime.now(timezone.utc).isoformat(),
                },
            )
            db.add(badge)
            created_count += 1

    await db.commit()

    return {
        "status": "success",
        "message": f"{created_count} rozet hesaplandÃ„Â± ve verildi",
        "badges_created": created_count,
    }


# Ã¢â€â‚¬Ã¢â€â‚¬ Certificate Tier Endpoints Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬



@app.post("/api/admin/events/{event_id}/certificates/assign-tiers")
async def assign_certificate_tiers(
    event_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Assign certificate tiers to all attendees based on the defined rules."""
    # Check authorization
    e_res = await db.execute(select(Event).where(Event.id == event_id))
    event = e_res.scalar_one_or_none()
    if not event:
        raise HTTPException(status_code=404, detail="Etkinlik bulunamadÃ„Â±")

    if not await _can_manage_organization_event(db, current_user, event.admin_id):
        raise HTTPException(status_code=403, detail="Yetkisiz eriÅŸim")

    # Get tier rules
    ctr_res = await db.execute(
        select(CertificateTierRule).where(CertificateTierRule.event_id == event_id)
    )
    tier_rule = ctr_res.scalar_one_or_none()
    if not tier_rule:
        raise HTTPException(status_code=404, detail="Bu etkinlik iÃƒÂ§in sertifika seviyesi kurallarÃ„Â± belirlenmemiÃ…Å¸")

    # Get all certificates that don't have a tier yet
    certs_res = await db.execute(
        select(Certificate)
        .where(
            Certificate.event_id == event_id,
            Certificate.status == CertStatus.active,
            Certificate.certificate_tier.is_(None),
        )
        .order_by(Certificate.created_at)
    )
    certificates = certs_res.scalars().all()

    assigned_count = 0
    for cert in certificates:
        # For simplicity, assign first matching tier
        # In production, implement complex condition evaluation
        for tier_def in tier_rule.tier_definitions or []:
            tier_name = tier_def.get("tier_name", "Unknown")
            cert.certificate_tier = tier_name
            assigned_count += 1
            break

    await db.commit()

    return {
        "status": "success",
        "message": f"{assigned_count} sertifikaya seviye atandÃ„Â±",
        "certificates_assigned": assigned_count,
    }


@app.get("/api/admin/events/{event_id}/certificates/tier-summary")
async def get_tier_summary(
    event_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Get summary of certificate tier distribution for an event."""
    e_res = await db.execute(select(Event).where(Event.id == event_id))
    event = e_res.scalar_one_or_none()
    if not event:
        raise HTTPException(status_code=404, detail="Etkinlik bulunamadÃ„Â±")

    if not await _can_manage_organization_event(db, current_user, event.admin_id):
        raise HTTPException(status_code=403, detail="Yetkisiz eriÅŸim")

    # Get tier distribution
    certs_res = await db.execute(
        select(Certificate.certificate_tier, func.count(Certificate.id).label('count'))
        .where(Certificate.event_id == event_id, Certificate.status == CertStatus.active)
        .group_by(Certificate.certificate_tier)
    )
    tier_counts = certs_res.all()

    tier_summary = {
        "total_certificates": 0,
        "by_tier": {},
        "tier_percentages": {},
    }

    total = sum(count for _, count in tier_counts)
    tier_summary["total_certificates"] = total

    for tier_name, count in tier_counts:
        tier_key = tier_name or "Unassigned"
        tier_summary["by_tier"][tier_key] = count
        if total > 0:
            tier_summary["tier_percentages"][tier_key] = round((count / total) * 100, 2)

    return tier_summary


# Ã¢â€â‚¬Ã¢â€â‚¬ Survey Integration Endpoints Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬



@app.post("/api/surveys/{event_id}/submit", response_model=SurveyResponseOut)
async def submit_builtin_survey(
    event_id: str,
    survey_resp_in: SurveyResponseIn,
    attendee_id_header_snake: Optional[int] = FastAPIHeader(default=None, alias="attendee_id"),
    attendee_id_header_kebab: Optional[int] = FastAPIHeader(default=None, alias="attendee-id"),
    db: AsyncSession = Depends(get_db),
):
    """Submit a built-in survey response. Attendee endpoint."""
    event = await _resolve_public_event(db, event_id)
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    event_db_id = event.id
    attendee_id = survey_resp_in.attendee_id or attendee_id_header_snake or attendee_id_header_kebab
    token_email: Optional[str] = None
    if survey_resp_in.survey_token:
        try:
            token_payload = verify_survey_access_token(
                survey_resp_in.survey_token,
                event_id=event_db_id,
                event_public_id=_get_public_event_identifier(event),
            )
        except SignatureExpired:
            raise HTTPException(status_code=410, detail="Survey access link expired")
        except BadSignature:
            raise HTTPException(status_code=400, detail="Invalid survey access link")

        attendee_id = int(token_payload.get("attendee_id") or 0)
        token_email = str(token_payload.get("email") or "").lower()

    if not attendee_id:
        raise HTTPException(status_code=422, detail="attendee_id zorunludur")

    att_res = await db.execute(
        select(Attendee).where(
            Attendee.id == attendee_id,
            Attendee.event_id == event_db_id,
        )
    )
    attendee = att_res.scalar_one_or_none()
    if not attendee:
        raise HTTPException(status_code=404, detail="Katlmc bulunamad")
    if token_email and attendee.email.lower() != token_email:
        raise HTTPException(status_code=404, detail="KatÃ„Â±lÃ„Â±mcÃ„Â± bulunamadÃ„Â±")

    es_res = await db.execute(select(EventSurvey).where(EventSurvey.event_id == event_db_id))
    event_survey = es_res.scalar_one_or_none()
    if not event_survey:
        raise HTTPException(status_code=400, detail="Bu etkinlik icin anket yaplandrlmam")

    if survey_resp_in.survey_type != "builtin":
        raise HTTPException(status_code=400, detail="Bu endpoint yalnzca yerleik anket icindir")

    if event_survey.survey_type not in {"builtin", "both"}:
        raise HTTPException(status_code=400, detail="Bu etkinlik yerleik anket kabul etmiyor")

    builtin_questions = event_survey.builtin_questions or []
    if not builtin_questions:
        raise HTTPException(status_code=400, detail="Yerleik anket sorular tanmlanmam")

    answers = survey_resp_in.answers or {}
    for question in builtin_questions:
        question_id = str(question.get("id") or "").strip()
        if not question_id:
            continue

        raw_answer = answers.get(question_id)
        normalized_answer = str(raw_answer).strip() if raw_answer is not None else ""
        if question.get("required") and not normalized_answer:
            raise HTTPException(status_code=400, detail=f"'{question_id}' sorusu zorunludur")

        if question.get("type") == "multiple_choice" and normalized_answer:
            options = [str(option) for option in (question.get("options") or [])]
            if options and normalized_answer not in options:
                raise HTTPException(status_code=400, detail=f"'{question_id}' sorusu icin gecersiz secenek")

    sr_res = await db.execute(
        select(SurveyResponse).where(
            SurveyResponse.event_id == event_db_id,
            SurveyResponse.attendee_id == attendee_id,
            SurveyResponse.survey_type == survey_resp_in.survey_type,
        )
    )
    existing = sr_res.scalar_one_or_none()
    if existing:
        raise HTTPException(status_code=409, detail="Bu anket zaten gnderilmi")

    now = datetime.utcnow()
    survey_response = SurveyResponse(
        event_id=event_db_id,
        attendee_id=attendee_id,
        survey_type=survey_resp_in.survey_type,
        answers=answers,
        external_response_id=survey_resp_in.external_response_id,
        completed_at=now,
        completion_proof={"submitted_at": now.isoformat()},
    )
    db.add(survey_response)

    attendee.survey_completed_at = now
    attendee.can_download_cert = True
    from .crm_snapshot_hooks import refresh_crm_snapshot_for_attendee
    await refresh_crm_snapshot_for_attendee(db, attendee)

    await db.commit()
    await db.refresh(survey_response)
    return survey_response


@app.post("/api/surveys/external/webhook")
async def handle_external_survey_webhook(
    request: Request,
    db: AsyncSession = Depends(get_db),
):
    """Handle webhook from external survey providers (Typeform, Qualtrics, etc.)."""
    try:
        payload = await request.json()
    except Exception as e:
        raise HTTPException(status_code=400, detail="Invalid JSON payload")

    # Extract webhook key from headers
    webhook_key = request.headers.get("X-Webhook-Key", "")
    event_id = request.query_params.get("event_id")
    attendee_id = request.query_params.get("attendee_id")

    if not all([webhook_key, event_id, attendee_id]):
        raise HTTPException(status_code=400, detail="Missing required parameters")

    # Verify webhook key
    es_res = await db.execute(
        select(EventSurvey).where(EventSurvey.event_id == int(event_id))
    )
    event_survey = es_res.scalar_one_or_none()
    if not event_survey or event_survey.external_webhook_key != webhook_key:
        raise HTTPException(status_code=401, detail="Webhook key verification failed")

    if event_survey.survey_type not in {"external", "both"}:
        raise HTTPException(status_code=400, detail="Bu etkinlik harici anket kabul etmiyor")

    # Check if response already exists
    sr_res = await db.execute(
        select(SurveyResponse).where(
            SurveyResponse.event_id == int(event_id),
            SurveyResponse.attendee_id == int(attendee_id),
            SurveyResponse.survey_type == "external",
        )
    )
    existing = sr_res.scalar_one_or_none()
    if existing:
        raise HTTPException(status_code=409, detail="Survey response already recorded")

    # Create response
    survey_response = SurveyResponse(
        event_id=int(event_id),
        attendee_id=int(attendee_id),
        survey_type="external",
        answers=None,
        external_response_id=payload.get("response_id"),
        completed_at=datetime.utcnow(),
        completion_proof=payload,
    )
    db.add(survey_response)

    # Update attendee
    att_res = await db.execute(
        select(Attendee).where(
            Attendee.id == int(attendee_id),
            Attendee.event_id == int(event_id),
        )
    )
    attendee = att_res.scalar_one_or_none()
    if not attendee:
        raise HTTPException(status_code=404, detail="Attendee not found")

    attendee.survey_completed_at = datetime.utcnow()
    attendee.can_download_cert = True
    from .crm_snapshot_hooks import refresh_crm_snapshot_for_attendee
    await refresh_crm_snapshot_for_attendee(db, attendee)

    await db.commit()

    return {
        "status": "success",
        "message": "Survey response recorded",
    }


@app.get("/api/admin/events/{event_id}/surveys/responses")
async def get_survey_responses(
    event_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Get all survey responses for an event. Admin only."""
    e_res = await db.execute(select(Event).where(Event.id == event_id))
    event = e_res.scalar_one_or_none()
    if not event:
        raise HTTPException(status_code=404, detail="Etkinlik bulunamadÃ„Â±")

    if not await _can_manage_organization_event(db, current_user, event.admin_id):
        raise HTTPException(status_code=403, detail="Yetkisiz eriÅŸim")

    sr_res = await db.execute(
        select(SurveyResponse)
        .where(SurveyResponse.event_id == event_id)
        .order_by(SurveyResponse.completed_at.desc())
    )
    responses = sr_res.scalars().all()

    att_res = await db.execute(select(Attendee).where(Attendee.event_id == event_id))
    attendees = att_res.scalars().all()
    attendee_map = {attendee.id: attendee for attendee in attendees}
    completed_attendee_ids = {response.attendee_id for response in responses}

    return {
        "total_responses": len(responses),
        "responses": [
            SurveyResponseOut(
                id=response.id,
                event_id=response.event_id,
                attendee_id=response.attendee_id,
                attendee_name=attendee_map.get(response.attendee_id).name if attendee_map.get(response.attendee_id) else None,
                attendee_email=attendee_map.get(response.attendee_id).email if attendee_map.get(response.attendee_id) else None,
                survey_type=response.survey_type,
                answers=response.answers,
                external_response_id=response.external_response_id,
                completed_at=response.completed_at,
                completion_proof=response.completion_proof,
            )
            for response in responses
        ],
        "response_rate": {
            "completed": len(completed_attendee_ids),
            "pending": max(len(attendees) - len(completed_attendee_ids), 0),
        },
    }


# Ã¢â€â‚¬Ã¢â€â‚¬ Sponsor Management Endpoints Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬




@app.get("/api/health")
@limiter.exempt
async def health_check():
    return {"status": "ok"}


@app.get("/api/openapi.json", include_in_schema=False)
@limiter.exempt
async def openapi_schema():
    schema = app.openapi()
    schema.setdefault("servers", [{"url": "https://heptacert.com"}])
    return schema


@app.get("/api/openapi-actions.json", include_in_schema=False)
@limiter.exempt
async def openapi_actions_schema():
    """Compressed OpenAPI schema for ChatGPT Actions — all admin endpoints, stripped for size."""
    import json as _json
    import re as _re

    _STRIP_OP = {"description", "x-openapi-router-controller"}
    _STRIP_SCHEMA = {"title", "description", "example", "examples"}

    def _slim_schema(obj):
        if isinstance(obj, dict):
            return {
                k: _slim_schema(v)
                for k, v in obj.items()
                if k not in _STRIP_SCHEMA
            }
        if isinstance(obj, list):
            return [_slim_schema(i) for i in obj]
        return obj

    def _slim_op(op: dict, path: str, method: str) -> dict:
        slim = {k: v for k, v in op.items() if k not in _STRIP_OP}
        # Keep only 200/201 response without schema detail (saves ~40% size)
        if "responses" in slim:
            compact_resp = {}
            for code, resp in slim["responses"].items():
                compact_resp[code] = {"description": resp.get("description", "")}
            slim["responses"] = compact_resp
        # Slim request body schemas
        if "requestBody" in slim:
            slim["requestBody"] = _slim_schema(slim["requestBody"])
        # Slim parameters — drop Authorization header (ChatGPT manages auth separately)
        if "parameters" in slim:
            slim["parameters"] = [
                {k: v for k, v in p.items() if k not in ("description", "example")}
                for p in slim["parameters"]
                if not (p.get("in") == "header" and p.get("name", "").lower() == "authorization")
            ]
        if "operationId" not in slim:
            slim["operationId"] = f"{method}_{path.replace('/', '_').strip('_')}"
        return slim

    # Exactly 30 operations for ChatGPT Actions limit
    ALLOWED: set[tuple[str, str]] = {
        ("/api/health",                                                          "get"),
        ("/api/admin/events",                                                    "get"),
        ("/api/admin/events",                                                    "post"),
        ("/api/admin/events/{event_id}",                                         "get"),
        ("/api/admin/events/{event_id}",                                         "patch"),
        ("/api/admin/events/{event_id}",                                         "delete"),
        ("/api/admin/events/{event_id}/attendees",                               "get"),
        ("/api/admin/events/{event_id}/attendees",                               "post"),
        ("/api/admin/events/{event_id}/attendees/{attendee_id}",                 "patch"),
        ("/api/admin/events/{event_id}/attendees/{attendee_id}",                 "delete"),
        ("/api/admin/events/{event_id}/certificates",                            "get"),
        ("/api/admin/events/{event_id}/certificates",                            "post"),
        ("/api/admin/certificates/{cert_id}/revoke",                             "post"),
        ("/api/admin/events/{event_id}/sessions",                                "get"),
        ("/api/admin/events/{event_id}/sessions",                                "post"),
        ("/api/admin/events/{event_id}/sessions/{session_id}",                   "patch"),
        ("/api/admin/events/{event_id}/sessions/{session_id}",                   "delete"),
        ("/api/admin/events/{event_id}/checkin-lookup",                          "get"),
        ("/api/admin/events/{event_id}/sessions/{session_id}/checkin",           "post"),
        ("/api/admin/events/{event_id}/attendance",                              "get"),
        ("/api/admin/events/{event_id}/automations",                             "get"),
        ("/api/admin/events/{event_id}/automations",                             "post"),
        ("/api/admin/events/{event_id}/stats",                                   "get"),
        ("/api/admin/webhooks",                                                  "get"),
        ("/api/admin/webhooks",                                                  "post"),
        ("/api/admin/webhooks/{webhook_id}",                                     "delete"),
        ("/api/admin/api-keys",                                                  "get"),
        ("/api/admin/api-keys",                                                  "post"),
        ("/api/admin/api-keys/{key_id}",                                         "delete"),
        ("/api/admin/events/{event_id}/certificates/tier-summary",               "get"),
    }

    full = app.openapi()
    slim_paths: dict = {}
    for path, path_item in full.get("paths", {}).items():
        methods = {
            method: _slim_op(op, path, method)
            for method, op in path_item.items()
            if (path, method) in ALLOWED
        }
        if methods:
            slim_paths[path] = methods

    # Collect only $refs actually used
    refs_txt = _json.dumps(slim_paths)
    used_refs: set = set(_re.findall(r'"#/components/schemas/(\w+)"', refs_txt))
    all_comp = full.get("components", {}).get("schemas", {})

    def _collect(name: str, seen: set):
        if name in seen or name not in all_comp:
            return
        seen.add(name)
        for n in _re.findall(r'"#/components/schemas/(\w+)"', _json.dumps(all_comp[name])):
            _collect(n, seen)

    all_used: set = set()
    for r in used_refs:
        _collect(r, all_used)

    return {
        "openapi": "3.1.0",
        "info": {"title": "HeptaCert API", "version": "2.0.0"},
        "servers": [{"url": "https://heptacert.com"}],
        "paths": slim_paths,
        "components": {
            "schemas": {k: _slim_schema(v) for k, v in all_comp.items() if k in all_used},
        },
    }




@app.post("/api/auth/login")
@limiter.limit("5/minute")
async def login(request: Request, data: LoginIn, db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(User).where(User.email == str(data.email)))
    user = res.scalar_one_or_none()
    if not user or not verify_password(data.password, user.password_hash):
        await write_audit_log(
            db,
            user_id=user.id if user else None,
            action="security.login_failed",
            resource_type="auth",
            resource_id=str(data.email),
            ip_address=_client_ip_for_rate_limit(request),
            user_agent=request.headers.get("User-Agent"),
            extra={"email": str(data.email)},
        )
        await db.commit()
        raise HTTPException(status_code=401, detail="E-posta veya şifre hatalı.")
    if user.deleted_at is not None:
        raise HTTPException(status_code=401, detail="Bu hesap silinmiştir. Destek için iletişime geçin.")
    if not user.is_verified:
        await write_audit_log(
            db,
            user_id=user.id,
            action="security.login_unverified_blocked",
            resource_type="auth",
            resource_id=str(user.id),
            ip_address=_client_ip_for_rate_limit(request),
            user_agent=request.headers.get("User-Agent"),
            extra={"email": user.email},
        )
        await db.commit()
        raise HTTPException(status_code=403, detail="E-posta adresinizi doğrulamanız gerekiyor. Lütfen gelen kutunuzu kontrol edin.")

    owns_admin_workspace = user.role in (Role.admin, Role.superadmin)

    # Organization employees need admin-panel access even if they registered as a public member first.
    if user.role not in (Role.admin, Role.superadmin):
        try:
            from .organization_access_api import OrganizationMember

            normalized_email = (user.email or "").strip().lower()
            org_member_res = await db.execute(
                select(OrganizationMember.id).where(
                    OrganizationMember.status == "active",
                    or_(
                        OrganizationMember.user_id == user.id,
                        func.lower(func.trim(OrganizationMember.email)) == normalized_email,
                    ),
                ).limit(1)
            )
            event_member_res = await db.execute(
                select(EventTeamMember.id).where(
                    EventTeamMember.status == "active",
                    or_(
                        EventTeamMember.user_id == user.id,
                        func.lower(func.trim(EventTeamMember.email)) == normalized_email,
                    ),
                ).limit(1)
            )
            if org_member_res.scalar_one_or_none() is not None or event_member_res.scalar_one_or_none() is not None:
                user.role = Role.admin
                await db.flush()
        except Exception:
            logger.exception("Failed to synchronize organization employee role for user %s", user.id)

    if owns_admin_workspace:
        await _get_or_create_admin_organization(db, user.id)

    # Check if 2FA is enabled for this user
    totp_res = await db.execute(select(TotpSecret).where(TotpSecret.user_id == user.id, TotpSecret.enabled.is_(True)))
    totp = totp_res.scalar_one_or_none()
    if totp:
        partial = create_partial_token(user_id=user.id)
        await write_audit_log(
            db,
            user_id=user.id,
            action="security.login_2fa_required",
            resource_type="auth",
            resource_id=str(user.id),
            ip_address=_client_ip_for_rate_limit(request),
            user_agent=request.headers.get("User-Agent"),
        )
        await db.commit()
        return LoginWith2FAOut(requires_2fa=True, partial_token=partial)
    # Enforce host-scoped organization access: if request resolved to an organization
    # (via organization_middleware), only allow login from users that either
    # - own the organization (Organization.user_id), or
    # - have an email matching the organization's custom domain.
    # Future: support an explicit allowlist table for per-org approved emails.
    try:
        org_info = getattr(request.state, "organization", None)
        if org_info and org_info.get("id"):
            # load full Organization to get its owner user_id and custom_domain
            org_res = await db.execute(select(Organization).where(Organization.id == int(org_info.get("id"))))
            org = org_res.scalar_one_or_none()
            if org:
                # owner bypass
                if user.id == org.user_id:
                    pass
                else:
                    # compare email domain OR check explicit allowlist entries
                    email_val = (user.email or "").strip().lower()
                    email_domain = email_val.split("@")[-1] if "@" in email_val else ""
                    org_domain = (org.custom_domain or "").lower()
                    allowed = False
                    if email_domain and org_domain and email_domain == org_domain:
                        allowed = True
                    else:
                        # consult allowlist table
                        try:
                            allow_res = await db.execute(
                                select(OrganizationAllowlist).where(
                                    OrganizationAllowlist.org_id == org.id,
                                    OrganizationAllowlist.email == email_val,
                                )
                            )
                            allow_entry = allow_res.scalar_one_or_none()
                            if allow_entry:
                                allowed = True
                        except Exception:
                            # DB problem when checking allowlist Ã¢â‚¬â€ fail closed
                            raise HTTPException(status_code=500, detail="Alan adÃ„Â± eriÃ…Å¸imi doÃ„Å¸rulanamadÃ„Â±.")

                    if not allowed:
                        raise HTTPException(status_code=403, detail="Bu alan adÃ„Â± iÃƒÂ§in yalnÃ„Â±zca yetkili hesaplar giriÃ…Å¸ yapabilir.")
    except HTTPException:
        raise
    except Exception:
        # If DB check fails for some reason, fail closed to be safe.
        raise HTTPException(status_code=500, detail="Alan adÃ„Â± eriÃ…Å¸imi doÃ„Å¸rulanamadÃ„Â±.")

    await write_audit_log(
        db,
        user_id=user.id,
        action="security.login_success",
        resource_type="auth",
        resource_id=str(user.id),
        ip_address=_client_ip_for_rate_limit(request),
        user_agent=request.headers.get("User-Agent"),
    )
    await db.commit()
    return LoginWith2FAOut(
        requires_2fa=False,
        access_token=create_access_token(user_id=user.id, role=user.role),
    )


@app.post("/api/auth/2fa/validate")
@limiter.limit("10/minute")
async def validate_login_2fa(request: Request, data: TotpValidateIn, db: AsyncSession = Depends(get_db)):
    try:
        payload = jwt.decode(data.partial_token, settings.jwt_secret, algorithms=["HS256"])
        if not payload.get("partial"):
            raise HTTPException(status_code=401, detail="Gecersiz 2FA oturumu")
        user_id = int(payload.get("sub"))
    except HTTPException:
        raise
    except (JWTError, ValueError, TypeError):
        raise HTTPException(status_code=401, detail="2FA oturumu gecersiz veya suresi dolmus")

    user_res = await db.execute(select(User).where(User.id == user_id))
    user = user_res.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=401, detail="Kullanici bulunamadi")

    totp_res = await db.execute(
        select(TotpSecret).where(TotpSecret.user_id == user.id, TotpSecret.enabled.is_(True))
    )
    totp_secret = totp_res.scalar_one_or_none()
    if not totp_secret:
        raise HTTPException(status_code=400, detail="Bu hesapta 2FA etkin degil")

    code = (data.code or "").strip()
    if not pyotp.TOTP(totp_secret.secret).verify(code, valid_window=1):
        # Try backup code fallback
        import hashlib as _hashlib
        code_hash = _hashlib.sha256(code.upper().replace("-", "").encode()).hexdigest()
        backup_res = await db.execute(
            select(TotpBackupCode).where(
                TotpBackupCode.user_id == user.id,
                TotpBackupCode.code_hash == code_hash,
                TotpBackupCode.used_at.is_(None),
            )
        )
        backup = backup_res.scalar_one_or_none()
        if not backup:
            raise HTTPException(status_code=401, detail="Geçersiz doğrulama kodu")
        backup.used_at = datetime.now(timezone.utc)
        await write_audit_log(
            db,
            user_id=user.id,
            action="security.2fa_backup_code_used",
            resource_type="auth",
            resource_id=str(user.id),
            ip_address=request.client.host if request.client else None,
            user_agent=request.headers.get("user-agent"),
        )
        await db.commit()

    return {
        "access_token": create_access_token(user_id=user.id, role=user.role),
        "token_type": "bearer",
    }


@app.post("/api/auth/register", status_code=201)
@limiter.limit("3/minute")
async def register(request: Request, data: RegisterIn, db: AsyncSession = Depends(get_db)):
    if not data.terms_accepted:
        raise bad_request("Kayıt icin kullanım koşullarını kabul etmelisiniz.")

    email = str(data.email).strip().lower()
    res = await db.execute(select(User).where(func.lower(User.email) == email))
    existing_user = res.scalar_one_or_none()
    if existing_user and existing_user.deleted_at is None:
        raise bad_request("Bu e-posta adresi zaten kayıtlı.")

    token = make_email_token({"email": email, "action": "verify"})
    if existing_user:
        existing_user.email = email
        existing_user.password_hash = hash_password(data.password)
        existing_user.role = Role.admin
        existing_user.deleted_at = None
        existing_user.is_verified = False
        existing_user.verification_token = token
        existing_user.password_reset_token = None
        existing_user.magic_link_token = None
        existing_user.heptacoin_balaonce = existing_user.heptacoin_balaonce or 100
        user = existing_user
        from .cache import cache
        await cache.delete(f"user:{user.id}")
    else:
        user = User(
            email=email,
            password_hash=hash_password(data.password),
            role=Role.admin,
            heptacoin_balaonce=100,  # 100 HC welcome bonus
            is_verified=False,
            verification_token=token,
        )
        db.add(user)
    await db.commit()

    verify_link = f"{settings.frontend_base_url.rstrip('/')}/verify-email?token={token}"
    await send_email_async(
        to=email,
        subject="HeptaCert - E-posta Adresinizi Doğrulayın",
        html_body=f"""
        <p>Merhaba,</p>
        <p>HeptaCert'e hoş geldiniz! Hesabınızı aktif etmek icin aşağıdaki bağlantıya tıklayın:</p>
        <p><a href="{verify_link}">{verify_link}</a></p>
        <p>Bu bağlantı 24 saat gecerlidir.</p>
        """,
    )
    return {"detail": "Kayıt başarılı. Aktivasyon e-postası gonderildi."}


@app.get("/api/auth/verify-email")
async def verify_email_endpoint(token: str = Query(...), db: AsyncSession = Depends(get_db)):
    try:
        payload = verify_email_token(token, max_age=86400)
    except SignatureExpired:
        raise bad_request("Doğrulama bağlantısının suresi dolmuş. Lutfen yeniden kayıt olun.")
    except (BadSignature, Exception):
        raise bad_request("Gecersiz doğrulama bağlantısı.")

    if payload.get("action") != "verify":
        raise bad_request("Gecersiz token turu.")

    email = payload.get("email")
    res = await db.execute(select(User).where(User.email == email))
    user = res.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı.")
    if (not user.is_verified) and (not user.verification_token or not hmac.compare_digest(str(user.verification_token), token)):
        raise bad_request("Geçersiz doğrulama bağlantısı.")
    if user.is_verified:
        if user.role in (Role.admin, Role.superadmin):
            await _get_or_create_admin_organization(db, user.id)
            await db.commit()
        return {"detail": "Hesabınız zaten doğrulanmış."}

    user.is_verified = True
    user.verification_token = None
    if user.role in (Role.admin, Role.superadmin):
        await _get_or_create_admin_organization(db, user.id)
    await db.commit()
    return {"detail": "E-posta başarıyla doğrulandı. Giriş yapabilirsiniz."}


@app.post("/api/auth/resend-verification")
@limiter.limit("3/hour")
async def resend_verification_email(request: Request, data: ForgotPasswordIn, db: AsyncSession = Depends(get_db)):
    email = str(data.email).strip().lower()
    res = await db.execute(select(User).where(func.lower(User.email) == email))
    user = res.scalar_one_or_none()
    if user and not user.is_verified:
        token = make_email_token({"email": user.email, "action": "verify"})
        user.verification_token = token
        db.add(user)
        await db.commit()

        verify_link = f"{settings.frontend_base_url.rstrip('/')}/verify-email?token={token}"
        await send_email_async(
            to=user.email,
            subject="HeptaCert - E-posta adresinizi doğrulayın",
            html_body=f"""
            <p>Merhaba,</p>
            <p>HeptaCert hesabınızı aktif etmek için aşağıdaki bağlantıya tıklayın:</p>
            <p><a href="{verify_link}">{verify_link}</a></p>
            <p>Bu bağlantı 24 saat geçerlidir. Maili göremiyorsanız spam klasörünü de kontrol edin.</p>
            """,
        )

    return {"detail": "Eğer doğrulanmamış bir hesap varsa doğrulama e-postası yeniden gönderildi."}






OAUTH_BRIDGE_COOKIE = "heptacert_oauth_bridge"
OAUTH_BRIDGE_PATH = "/api/auth/oauth/bridge/exchange"


def _oauth_bridge_redirect(redirect_target: str, *, token: str, mode: str) -> RedirectResponse:
    bridge_value = make_email_token({"action": "oauth_bridge", "mode": mode, "token": token})
    response = RedirectResponse(redirect_target)
    response.set_cookie(
        key=OAUTH_BRIDGE_COOKIE,
        value=bridge_value,
        max_age=60,
        httponly=True,
        secure=settings.public_base_url.lower().startswith("https://"),
        samesite="lax",
        path=OAUTH_BRIDGE_PATH,
    )
    return response


@app.post("/api/auth/oauth/bridge/exchange")
async def oauth_bridge_exchange(request: Request):
    raw_bridge = request.cookies.get(OAUTH_BRIDGE_COOKIE)
    if not raw_bridge:
        raise HTTPException(status_code=401, detail="OAuth bridge session missing.")
    try:
        payload = verify_email_token(raw_bridge, max_age=60)
    except Exception:
        raise HTTPException(status_code=401, detail="OAuth bridge session invalid or expired.")
    if payload.get("action") != "oauth_bridge" or not payload.get("token"):
        raise HTTPException(status_code=401, detail="OAuth bridge session invalid.")
    response = JSONResponse({
        "mode": "admin" if payload.get("mode") == "admin" else "member",
        "access_token": str(payload["token"]),
    })
    response.delete_cookie(OAUTH_BRIDGE_COOKIE, path=OAUTH_BRIDGE_PATH)
    return response




@app.get("/api/auth/google/start")
async def google_oauth_start(
    mode: str = Query(default="member", pattern="^(member|admin)$"),
    next: Optional[str] = Query(default=None),
):
    if not settings.google_oauth_client_id or not settings.google_oauth_client_secret:
        raise HTTPException(status_code=503, detail="Google OAuth is not configured.")

    fallback_next = "/admin/events" if mode == "admin" else "/events"
    state = make_email_token({
        "action": "google_oauth",
        "mode": mode,
        "next": _normalize_oauth_next(next, fallback_next),
    })
    params = {
        "client_id": settings.google_oauth_client_id,
        "redirect_uri": _google_oauth_redirect_uri(),
        "response_type": "code",
        "scope": "openid email profile",
        "state": state,
        "access_type": "online",
        "prompt": "select_account",
    }
    return RedirectResponse(f"https://accounts.google.com/o/oauth2/v2/auth?{urlencode(params)}")


@app.get("/api/auth/google/callback")
async def google_oauth_callback(
    code: str = Query(...),
    state: str = Query(...),
    db: AsyncSession = Depends(get_db),
):
    if not settings.google_oauth_client_id or not settings.google_oauth_client_secret:
        raise HTTPException(status_code=503, detail="Google OAuth is not configured.")
    try:
        state_payload = verify_email_token(state, max_age=300)
    except Exception:
        raise bad_request("Google OAuth state is invalid or expired.")
    if state_payload.get("action") != "google_oauth":
        raise bad_request("Invalid Google OAuth state.")

    mode = "admin" if state_payload.get("mode") == "admin" else "member"
    fallback_next = "/admin/events" if mode == "admin" else "/events"
    next_url = _normalize_oauth_next(str(state_payload.get("next") or ""), fallback_next)

    import httpx
    async with httpx.AsyncClient(timeout=10.0) as client:
        token_res = await client.post(
            "https://oauth2.googleapis.com/token",
            data={
                "code": code,
                "client_id": settings.google_oauth_client_id,
                "client_secret": settings.google_oauth_client_secret,
                "redirect_uri": _google_oauth_redirect_uri(),
                "grant_type": "authorization_code",
            },
        )
        if token_res.status_code >= 400:
            raise HTTPException(status_code=400, detail="Google token exchange failed.")
        access_token = token_res.json().get("access_token")
        if not access_token:
            raise HTTPException(status_code=400, detail="Google access token missing.")
        userinfo_res = await client.get(
            "https://openidconnect.googleapis.com/v1/userinfo",
            headers={"Authorization": f"Bearer {access_token}"},
        )
        if userinfo_res.status_code >= 400:
            raise HTTPException(status_code=400, detail="Google profile could not be read.")
        profile = userinfo_res.json()

    email = str(profile.get("email") or "").strip().lower()
    if not email or not profile.get("email_verified"):
        raise HTTPException(status_code=400, detail="Google email must be verified.")
    display_name = str(profile.get("name") or email.split("@")[0]).strip()[:120] or email
    avatar_url = str(profile.get("picture") or "").strip() or None

    if mode == "admin":
        res = await db.execute(select(User).where(User.email == email))
        user = res.scalar_one_or_none()
        if not user:
            user = User(
                email=email,
                password_hash=hash_password(secrets.token_urlsafe(32)),
                role=Role.admin,
                heptacoin_balaonce=100,
                is_verified=True,
                verification_token=None,
            )
            db.add(user)
            await db.commit()
            await db.refresh(user)
        elif not user.is_verified:
            user.is_verified = True
            user.verification_token = None
            await db.commit()
        token = create_access_token(user_id=user.id, role=user.role)
    else:
        res = await db.execute(select(PublicMember).where(PublicMember.email == email))
        member = res.scalar_one_or_none()
        if not member:
            member = PublicMember(
                public_id=await _generate_public_member_public_id(db),
                email=email,
                display_name=display_name,
                avatar_url=avatar_url,
                password_hash=hash_password(secrets.token_urlsafe(32)),
                is_verified=True,
                verification_token=None,
            )
            db.add(member)
            await db.commit()
            await db.refresh(member)
        else:
            changed = False
            if not member.is_verified:
                member.is_verified = True
                member.verification_token = None
                changed = True
            if avatar_url and not member.avatar_url:
                member.avatar_url = avatar_url
                changed = True
            if changed:
                await db.commit()
        token = create_public_member_access_token(member_id=member.id)

    params = urlencode({"mode": mode, "bridge": "1", "next": next_url})
    return _oauth_bridge_redirect(
        f"{settings.frontend_base_url.rstrip('/')}/auth/google/callback?{params}",
        token=token,
        mode=mode,
    )


@app.get(
    "/api/admin/google/sheets/status",
    response_model=GoogleSheetsStatusOut,
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def google_sheets_status(
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    configured = bool(settings.google_oauth_client_id and settings.google_oauth_client_secret)
    integration = await _get_user_google_integration(db, me.id)
    scopes = _normalize_google_scopes(integration.scopes if integration else [])
    missing_scopes = _google_sheets_missing_scopes(scopes) if integration else GOOGLE_SHEETS_SCOPES
    return GoogleSheetsStatusOut(
        configured=configured,
        connected=bool(configured and integration and integration.refresh_token and not missing_scopes),
        google_email=integration.google_email if integration else None,
        scopes=scopes,
        missing_scopes=missing_scopes,
    )


@app.get(
    "/api/admin/google/sheets/start",
    response_model=GoogleSheetsAuthStartOut,
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def google_sheets_auth_start(
    next: Optional[str] = Query(default="/admin/events"),
    frontend_origin: Optional[str] = Query(default=None),
    event_id: Optional[int] = Query(default=None),
    me: CurrentUser = Depends(get_current_user),
):
    if not settings.google_oauth_client_id or not settings.google_oauth_client_secret:
        raise HTTPException(status_code=503, detail="Google OAuth is not configured.")
    state = make_email_token({
        "action": "google_sheets_oauth",
        "user_id": me.id,
        "next": _normalize_oauth_next(next, "/admin/events"),
        "frontend_origin": _normalize_oauth_frontend_origin(frontend_origin),
        "event_id": event_id,
    })
    params = {
        "client_id": settings.google_oauth_client_id,
        "redirect_uri": _google_sheets_redirect_uri(),
        "response_type": "code",
        "scope": " ".join(GOOGLE_SHEETS_SCOPES),
        "state": state,
        "access_type": "offline",
        "prompt": "consent select_account",
        "include_granted_scopes": "true",
    }
    return GoogleSheetsAuthStartOut(
        authorization_url=f"https://accounts.google.com/o/oauth2/v2/auth{urlencode(params)}"
    )


@app.get("/api/admin/google/sheets/callback")
async def google_sheets_auth_callback(
    code: str = Query(...),
    state: str = Query(...),
    db: AsyncSession = Depends(get_db),
):
    if not settings.google_oauth_client_id or not settings.google_oauth_client_secret:
        raise HTTPException(status_code=503, detail="Google OAuth is not configured.")
    try:
        state_payload = verify_email_token(state, max_age=300)
    except Exception:
        raise bad_request("Google Sheets OAuth state is invalid or expired.")
    oauth_action = state_payload.get("action")
    if oauth_action not in {"google_sheets_oauth", "google_calendar_oauth"}:
        raise bad_request("Invalid Google OAuth state.")

    user_id = int(state_payload.get("user_id") or 0)
    if user_id <= 0:
        raise bad_request("Invalid Google Sheets OAuth user.")

    token_data = await _google_exchange_code_for_tokens(code, _google_sheets_redirect_uri())
    access_token = str(token_data.get("access_token") or "")
    if not access_token:
        raise HTTPException(status_code=400, detail="Google access token missing.")
    profile = await _google_get_profile(access_token)
    email = str(profile.get("email") or "").strip().lower() or None

    integration = await _get_user_google_integration(db, user_id)
    if not integration:
        integration = UserGoogleIntegration(user_id=user_id)
        db.add(integration)
    integration.google_email = email
    integration.access_token = _encrypt_secret(access_token)
    if token_data.get("refresh_token"):
        integration.refresh_token = _encrypt_secret(str(token_data.get("refresh_token")))
    integration.token_expires_at = datetime.now(timezone.utc) + timedelta(seconds=int(token_data.get("expires_in") or 3600))
    integration.scopes = _normalize_google_scopes(token_data.get("scope"))
    db.add(integration)
    await db.commit()

    next_url = _normalize_oauth_next(str(state_payload.get("next") or ""), "/admin/events")
    frontend_origin = _normalize_oauth_frontend_origin(str(state_payload.get("frontend_origin") or ""))
    sheet_status = "connected"
    event_id = int(state_payload.get("event_id") or 0)
    if oauth_action == "google_sheets_oauth" and event_id > 0:
        event_res = await db.execute(select(Event).where(Event.id == event_id, Event.admin_id == user_id))
        event = event_res.scalar_one_or_none()
        if event:
            try:
                await _write_event_attendees_to_google_sheet(db, event, create_if_missing=True)
                sheet_status = "created"
            except Exception:
                logger.exception("Google Sheets OAuth connected but automatic sheet creation failed for event_id=%s", event_id)
                sheet_status = "connected"
    user_res = await db.execute(select(User).where(User.id == user_id))
    user = user_res.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="Google Sheets OAuth user not found.")
    bridge_payload = {"oauth_bridge": "1"}
    if oauth_action == "google_sheets_oauth":
        bridge_payload["google_sheets"] = sheet_status
    if oauth_action == "google_calendar_oauth":
        bridge_payload["google_calendar"] = "connected"
    bridge_params = urlencode(bridge_payload)
    separator = "&" if "?" in next_url else "?"
    redirect_target = f"{frontend_origin}{next_url}{separator}{bridge_params}"
    logger.info(
        "Google Sheets OAuth connected for user_id=%s; redirecting to frontend_origin=%s next=%s",
        user_id,
        frontend_origin,
        next_url,
    )
    return _oauth_bridge_redirect(
        redirect_target,
        token=create_access_token(user_id=user.id, role=user.role),
        mode="admin",
    )


@app.get(
    "/api/admin/microsoft/excel/status",
    response_model=MicrosoftExcelStatusOut,
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def microsoft_excel_status(
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    configured = bool(settings.ms365_oauth_client_id and settings.ms365_oauth_client_secret)
    integration = await _get_user_ms365_integration(db, me.id)
    scopes = _normalize_ms365_scopes(integration.scopes if integration else [])
    missing_scopes = _ms365_excel_missing_scopes(scopes) if integration else MS365_EXCEL_REQUIRED_SCOPES
    return MicrosoftExcelStatusOut(
        configured=configured,
        connected=bool(configured and integration and integration.refresh_token and not missing_scopes),
        microsoft_email=integration.microsoft_email if integration else None,
        scopes=scopes,
        missing_scopes=missing_scopes,
    )


@app.get(
    "/api/admin/microsoft/excel/start",
    response_model=MicrosoftExcelAuthStartOut,
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def microsoft_excel_auth_start(
    next: Optional[str] = Query(default="/admin/events"),
    frontend_origin: Optional[str] = Query(default=None),
    event_id: Optional[int] = Query(default=None),
    me: CurrentUser = Depends(get_current_user),
):
    if not settings.ms365_oauth_client_id or not settings.ms365_oauth_client_secret:
        raise HTTPException(status_code=503, detail="Microsoft OAuth is not configured.")
    state = make_email_token({
        "action": "ms365_excel_oauth",
        "user_id": me.id,
        "next": _normalize_oauth_next(next, "/admin/events"),
        "frontend_origin": _normalize_oauth_frontend_origin(frontend_origin),
        "event_id": event_id,
    })
    params = {
        "client_id": settings.ms365_oauth_client_id,
        "redirect_uri": _ms365_excel_redirect_uri(),
        "response_type": "code",
        "scope": " ".join(MS365_EXCEL_SCOPES),
        "state": state,
        "response_mode": "query",
        "prompt": "select_account",
    }
    return MicrosoftExcelAuthStartOut(
        authorization_url=f"https://login.microsoftonline.com/common/oauth2/v2.0/authorize?{urlencode(params)}"
    )


@app.get("/api/admin/microsoft/excel/callback")
async def microsoft_excel_auth_callback(
    code: str = Query(...),
    state: str = Query(...),
    db: AsyncSession = Depends(get_db),
):
    if not settings.ms365_oauth_client_id or not settings.ms365_oauth_client_secret:
        raise HTTPException(status_code=503, detail="Microsoft OAuth is not configured.")
    try:
        state_payload = verify_email_token(state, max_age=300)
    except Exception:
        raise bad_request("Microsoft Excel OAuth state is invalid or expired.")
    if state_payload.get("action") != "ms365_excel_oauth":
        raise bad_request("Invalid Microsoft Excel OAuth state.")

    user_id = int(state_payload.get("user_id") or 0)
    if user_id <= 0:
        raise bad_request("Invalid Microsoft Excel OAuth user.")

    token_data = await _ms365_exchange_code_for_tokens(code, _ms365_excel_redirect_uri())
    access_token = str(token_data.get("access_token") or "")
    if not access_token:
        raise HTTPException(status_code=400, detail="Microsoft access token missing.")
    profile = await _ms365_get_profile(access_token)
    email = str(profile.get("mail") or profile.get("userPrincipalName") or "").strip().lower() or None

    integration = await _get_user_ms365_integration(db, user_id)
    if not integration:
        integration = UserMicrosoftIntegration(user_id=user_id)
        db.add(integration)
    integration.microsoft_email = email
    integration.access_token = _encrypt_secret(access_token)
    if token_data.get("refresh_token"):
        integration.refresh_token = _encrypt_secret(str(token_data.get("refresh_token")))
    integration.token_expires_at = datetime.now(timezone.utc) + timedelta(seconds=int(token_data.get("expires_in") or 3600))
    integration.scopes = _normalize_ms365_scopes(token_data.get("scope"))
    db.add(integration)
    await db.commit()

    next_url = _normalize_oauth_next(str(state_payload.get("next") or ""), "/admin/events")
    frontend_origin = _normalize_oauth_frontend_origin(str(state_payload.get("frontend_origin") or ""))
    excel_status = "connected"
    event_id = int(state_payload.get("event_id") or 0)
    if event_id > 0:
        event_res = await db.execute(select(Event).where(Event.id == event_id, Event.admin_id == user_id))
        event = event_res.scalar_one_or_none()
        if event:
            try:
                await _write_event_attendees_to_ms365_excel(db, event, create_if_missing=True)
                excel_status = "created"
            except Exception:
                logger.exception("Microsoft Excel OAuth connected but automatic workbook creation failed for event_id=%s", event_id)
                excel_status = "connected"
    user_res = await db.execute(select(User).where(User.id == user_id))
    user = user_res.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="Microsoft Excel OAuth user not found.")
    bridge_params = urlencode({
        "ms365_excel": excel_status,
        "oauth_bridge": "1",
    })
    separator = "&" if "?" in next_url else "?"
    redirect_target = f"{frontend_origin}{next_url}{separator}{bridge_params}"
    logger.info(
        "Microsoft Excel OAuth connected for user_id=%s; redirecting to frontend_origin=%s next=%s",
        user_id,
        frontend_origin,
        next_url,
    )
    return _oauth_bridge_redirect(
        redirect_target,
        token=create_access_token(user_id=user.id, role=user.role),
        mode="admin",
    )


@app.post("/api/public/auth/register", status_code=201)
@limiter.limit("3/minute")
async def public_member_register(request: Request, data: PublicMemberRegisterIn, db: AsyncSession = Depends(get_db)):
    if not data.terms_accepted:
        raise bad_request("You must accept the terms of use to register.")

    email = str(data.email).strip().lower()
    display_name = data.display_name.strip()
    if not display_name:
        raise bad_request("Display name is required.")

    res = await db.execute(select(PublicMember).where(PublicMember.email == email))
    if res.scalar_one_or_none():
        raise bad_request("This email address is already registered.")

    token = make_email_token({"email": email, "action": "public_member_verify"})
    member = PublicMember(
        public_id=await _generate_public_member_public_id(db),
        email=email,
        display_name=display_name,
        password_hash=hash_password(data.password),
        is_verified=False,
        verification_token=token,
    )
    db.add(member)
    await db.commit()

    verify_link = f"{settings.frontend_base_url.rstrip('/')}/member/verify-email?token={token}"
    await send_email_async(
        to=email,
        subject="HeptaCert - Verify your member account",
        html_body=f"""
        <p>Hello {display_name},</p>
        <p>Verify your HeptaCert member account to browse public events and join community features.</p>
        <p><a href="{verify_link}">{verify_link}</a></p>
        <p>This link is valid for 24 hours.</p>
        """,
    )
    return {"detail": "Member registration successful. Verification email sent."}


@app.post("/api/public/auth/login", response_model=PublicMemberTokenOut)
@limiter.limit("5/minute")
async def public_member_login(request: Request, data: PublicMemberLoginIn, db: AsyncSession = Depends(get_db)):
    email = str(data.email).strip().lower()
    res = await db.execute(select(PublicMember).where(PublicMember.email == email))
    member = res.scalar_one_or_none()
    if not member or not verify_password(data.password, member.password_hash):
        raise HTTPException(status_code=401, detail="E-posta veya şifre hatalı.")
    if member.deleted_at is not None:
        raise HTTPException(status_code=401, detail="Bu hesap silinmiştir. Destek için iletişime geçin.")
    if not member.is_verified:
        raise HTTPException(status_code=403, detail="E-posta adresinizi doğrulamanız gerekiyor. Lütfen gelen kutunuzu kontrol edin.")

    member_out = PublicMemberMeOut(
        id=member.id,
        public_id=member.public_id,
        email=member.email,
        display_name=member.display_name,
        bio=member.bio,
        avatar_url=member.avatar_url,
        headline=member.headline,
        location=member.location,
        website_url=member.website_url,
        created_at=member.created_at,
    )
    return PublicMemberTokenOut(
        access_token=create_public_member_access_token(member_id=member.id),
        member=member_out,
    )


@app.get("/api/public/auth/verify-email")
async def verify_public_member_email(token: str = Query(...), db: AsyncSession = Depends(get_db)):
    try:
        payload = verify_email_token(token, max_age=86400)
    except SignatureExpired:
        raise bad_request("Verification link expired. Please register again.")
    except BadSignature:
        raise bad_request("Invalid verification link.")

    if payload.get("action") != "public_member_verify":
        raise bad_request("Invalid token action.")

    email = str(payload.get("email") or "").strip().lower()
    res = await db.execute(select(PublicMember).where(PublicMember.email == email))
    member = res.scalar_one_or_none()
    if not member:
        raise HTTPException(status_code=404, detail="Member not found.")
    if not member.verification_token or not hmac.compare_digest(str(member.verification_token), token):
        raise bad_request("Invalid verification link.")
    if member.is_verified:
        return {"detail": "Your member account is already verified."}

    member.is_verified = True
    member.verification_token = None
    await db.commit()
    return {"detail": "Member email verified successfully. You can now sign in."}


@app.post("/api/public/auth/resend-verification")
@limiter.limit("3/hour")
async def public_member_resend_verification(request: Request, data: ForgotPasswordIn, db: AsyncSession = Depends(get_db)):
    email = str(data.email).strip().lower()
    res = await db.execute(select(PublicMember).where(PublicMember.email == email))
    member = res.scalar_one_or_none()
    if member and not member.is_verified:
        token = make_email_token({"email": email, "action": "public_member_verify"})
        member.verification_token = token
        db.add(member)
        await db.commit()

        verify_link = f"{settings.frontend_base_url.rstrip('/')}/member/verify-email?token={token}"
        await send_email_async(
            to=email,
            subject="HeptaCert - Üye hesabınızı doğrulayın",
            html_body=f"""
            <p>Merhaba {member.display_name},</p>
            <p>HeptaCert üye hesabınızı doğrulamak için aşağıdaki bağlantıya tıklayın:</p>
            <p><a href="{verify_link}">{verify_link}</a></p>
            <p>Bu bağlantı 24 saat geçerlidir. Maili göremiyorsanız spam klasörünü de kontrol edin.</p>
            """,
        )

    return {"detail": "Eğer doğrulanmamış bir üye hesabı varsa doğrulama e-postası yeniden gönderildi."}


@app.post("/api/public/auth/forgot-password")
@limiter.limit("3/minute")
async def public_member_forgot_password(request: Request, data: ForgotPasswordIn, db: AsyncSession = Depends(get_db)):
    email = str(data.email).strip().lower()
    res = await db.execute(select(PublicMember).where(PublicMember.email == email))
    member = res.scalar_one_or_none()
    if member and member.is_verified:
        token = make_email_token({"email": email, "action": "public_member_reset"})
        member.password_reset_token = token
        await db.commit()

        reset_link = f"{settings.frontend_base_url}/reset-password?token={token}&mode=member"
        await send_email_async(
            to=email,
            subject="HeptaCert - Reset your member password",
            html_body=f"""
            <p>Hello {member.display_name},</p>
            <p>Click the link below to reset your member account password:</p>
            <p><a href="{reset_link}">{reset_link}</a></p>
            <p>This link is valid for 1 hour.</p>
            """,
        )
    return {"detail": "If the account exists, password reset instructions were sent to the email address."}


@app.post("/api/public/auth/reset-password")
@limiter.limit("5/minute")
async def public_member_reset_password(request: Request, data: ResetPasswordIn, db: AsyncSession = Depends(get_db)):
    try:
        payload = verify_email_token(data.token, max_age=3600)
    except SignatureExpired:
        raise bad_request("Password reset link expired.")
    except (BadSignature, Exception):
        raise bad_request("Invalid password reset link.")

    if payload.get("action") != "public_member_reset":
        raise bad_request("Invalid token action.")

    email = str(payload.get("email") or "").strip().lower()
    res = await db.execute(select(PublicMember).where(PublicMember.email == email))
    member = res.scalar_one_or_none()
    if not member:
        raise HTTPException(status_code=404, detail="Member not found.")
    if not member.password_reset_token or not hmac.compare_digest(str(member.password_reset_token), data.token):
        raise bad_request("Invalid password reset link.")

    member.password_hash = hash_password(data.new_password)
    member.password_reset_token = None
    await db.commit()
    return {"detail": "Password reset successful."}


@app.post("/api/auth/forgot-password")
@limiter.limit("3/minute")
async def forgot_password(request: Request, data: ForgotPasswordIn, db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(User).where(User.email == str(data.email)))
    user = res.scalar_one_or_none()
    # Always return 200 to avoid email enumeration
    if user and user.is_verified:
        token = make_email_token({"email": str(data.email), "action": "reset"})
        user.password_reset_token = token
        await db.commit()

        reset_link = f"{settings.frontend_base_url}/reset-password?token={token}"
        await send_email_async(
            to=str(data.email),
            subject="HeptaCert - Şifre Sıfırlama",
            html_body=f"""
            <p>Şifrenizi sıfırlamak için aşağıdaki bağlantıya tıklayın:</p>
            <p><a href="{reset_link}">{reset_link}</a></p>
            <p>Bu bağlantı 1 saat gecerlidir.</p>
            """,
        )
    return {"detail": "Şifre sıfırlama talimatları e-posta adresinize gönderildi."}


@app.post("/api/auth/reset-password")
@limiter.limit("5/minute")
async def reset_password(request: Request, data: ResetPasswordIn, db: AsyncSession = Depends(get_db)):
    try:
        payload = verify_email_token(data.token, max_age=3600)
    except SignatureExpired:
        raise bad_request("Şifre sıfırlama bağlantısının süresi dolmuş.")
    except (BadSignature, Exception):
        raise bad_request("Gecersiz sıfırlama bağlantısı.")

    if payload.get("action") != "reset":
        raise bad_request("Gecersiz token turu.")

    email = payload.get("email")
    res = await db.execute(select(User).where(User.email == email))
    user = res.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı.")

    # Validate that the token matches the one stored in DB (prevents replay attacks)
    if not user.password_reset_token or not hmac.compare_digest(str(user.password_reset_token), str(data.token)):
        raise bad_request("Bu sıfırlama bağlantısı zaten kullanılmış.")

    user.password_hash = hash_password(data.new_password)
    user.password_reset_token = None
    await db.commit()
    return {"detail": "Şifreniz başarıyla güoncellendi."}














































async def _build_system_digest_email_content(db: AsyncSession, config: SystemEmailDigestConfig) -> tuple[str, str, int, int]:
    today = datetime.now(timezone.utc).date()
    now_year = datetime.now(timezone.utc).year

    logo_data_uri: Optional[str] = None
    logo_path = Path(__file__).with_name("logo1.png")
    if logo_path.exists():
        logo_data_uri = f"data:image/png;base64,{base64.b64encode(logo_path.read_bytes()).decode('ascii')}"

    hero_svg = """
    <svg xmlns='http://www.w3.org/2000/svg' width='1200' height='420' viewBox='0 0 1200 420' role='img' aria-label='HeptaCert community update banner'>
        <defs>
            <linearGradient id='g' x1='0' x2='1' y1='0' y2='1'>
                <stop offset='0%' stop-color='#0f172a'/>
                <stop offset='52%' stop-color='#0f766e'/>
                <stop offset='100%' stop-color='#06b6d4'/>
            </linearGradient>
            <radialGradient id='r' cx='30%' cy='25%' r='80%'>
                <stop offset='0%' stop-color='#ffffff' stop-opacity='0.28'/>
                <stop offset='100%' stop-color='#ffffff' stop-opacity='0'/>
            </radialGradient>
        </defs>
        <rect width='1200' height='420' fill='url(#g)'/>
        <circle cx='180' cy='110' r='170' fill='url(#r)'/>
        <circle cx='1020' cy='80' r='210' fill='#ffffff' fill-opacity='0.08'/>
        <rect x='78' y='84' width='180' height='180' rx='38' fill='#ffffff' fill-opacity='0.12' stroke='#ffffff' stroke-opacity='0.18'/>
        <path d='M138 126h60l42 42-42 42h-60l42-42z' fill='#ffffff' fill-opacity='0.96'/>
        <path d='M152 160h132v20H152zM152 192h96v20h-96z' fill='#ffffff' fill-opacity='0.72'/>
        <text x='308' y='155' fill='#ecfeff' font-size='58' font-weight='700' font-family='Arial, Helvetica, sans-serif'>HeptaCert topluluk özeti</text>
        <text x='308' y='215' fill='#cffafe' font-size='30' font-weight='400' font-family='Arial, Helvetica, sans-serif'>Yeni etkinlikler ve topluluk paylaşımları tek e-postada</text>
        <rect x='308' y='252' width='230' height='52' rx='26' fill='#ffffff' fill-opacity='0.14' stroke='#ffffff' stroke-opacity='0.24'/>
        <text x='338' y='286' fill='#ffffff' font-size='22' font-weight='700' font-family='Arial, Helvetica, sans-serif'>HeptaCert</text>
    </svg>
    """.strip()
    hero_data_uri = f"data:image/svg+xml;base64,{base64.b64encode(hero_svg.encode('utf-8')).decode('ascii')}"

    public_events_res = await db.execute(
        select(Event, Organization)
        .outerjoin(Organization, Organization.user_id == Event.admin_id)
        .order_by(Event.event_date.asc().nulls_last(), Event.created_at.desc())
        .limit(50)
    )
    public_event_rows = public_events_res.all()
    public_event_rows = [row for row in public_event_rows if _get_event_visibility(row[0]) == "public"]

    upcoming_events: list[dict[str, Any]] = []
    for event, org in public_event_rows:
        if event.event_date and event.event_date < today:
            continue
        upcoming_events.append(
            {
                "name": event.name,
                "date": event.event_date.isoformat() if event.event_date else None,
                "location": event.event_location or "Online",
                "organization": org.org_name if org else None,
                "link": f"{settings.frontend_base_url.rstrip('/')}/events/{_get_public_event_identifier(event)}",
            }
        )
        if len(upcoming_events) >= int(config.max_events or 3):
            break

    if len(upcoming_events) < int(config.max_events or 3):
        for event, org in public_event_rows:
            candidate = {
                "name": event.name,
                "date": event.event_date.isoformat() if event.event_date else None,
                "location": event.event_location or "Online",
                "organization": org.org_name if org else None,
                "link": f"{settings.frontend_base_url.rstrip('/')}/events/{_get_public_event_identifier(event)}",
            }
            if candidate not in upcoming_events:
                upcoming_events.append(candidate)
            if len(upcoming_events) >= int(config.max_events or 3):
                break

    try:
        posts_res = await db.execute(
            select(CommunityPost, PublicMember, Organization)
            .outerjoin(PublicMember, PublicMember.id == CommunityPost.author_public_member_id)
            .outerjoin(Organization, Organization.id == CommunityPost.org_id)
            .where(CommunityPost.status == "visible")
            .order_by(CommunityPost.created_at.desc())
            .limit(max(int(config.max_posts or 3) * 2, int(config.max_posts or 3)))
        )
        posts_rows = posts_res.all()
    except sa.exc.ProgrammingError:
        try:
            await db.rollback()
        except Exception:
            pass
        posts_res = await db.execute(
            select(CommunityPost, Organization)
            .outerjoin(Organization, Organization.id == CommunityPost.org_id)
            .where(CommunityPost.status == "visible")
            .order_by(CommunityPost.created_at.desc())
            .limit(max(int(config.max_posts or 3) * 2, int(config.max_posts or 3)))
        )
        posts_rows = [(row[0], None, row[1]) for row in posts_res.all()]

    recent_posts: list[dict[str, Any]] = []
    for post, member, org in posts_rows:
        if org is not None:
            author = org.org_name
        elif member is not None:
            author = member.display_name
        else:
            author = "HeptaCert"
        recent_posts.append(
            {
                "author": author,
                "snippet": (post.body or "").strip()[:180],
                "link": f"{settings.frontend_base_url.rstrip('/')}/post/{post.public_id}",
            }
        )
        if len(recent_posts) >= int(config.max_posts or 3):
            break

    subject = "HeptaCert'te bu hafta"
    events_count = len(upcoming_events)
    posts_count = len(recent_posts)

    apple_events_html = "".join(
        f"""
        <tr>
            <td style="padding:20px 0;border-top:1px solid #e5e5ea;">
                <div style="font-size:17px;line-height:1.35;font-weight:700;color:#1d1d1f;margin:0 0 6px 0;">{escape(item['name'])}</div>
                <div style="font-size:14px;line-height:1.5;color:#6e6e73;margin:0 0 12px 0;">{escape(item['date'] or 'Tarih yakında')}{' · ' + escape(item['location']) if item['location'] else ''}{' · ' + escape(item['organization']) if item['organization'] else ''}</div>
                <a href="{escape(item['link'])}" style="font-size:14px;line-height:1.4;color:#0071e3;text-decoration:none;font-weight:600;">Etkinliği aç →</a>
            </td>
        </tr>
        """
        for item in upcoming_events
    ) or """
        <tr>
            <td style="padding:20px 0;border-top:1px solid #e5e5ea;color:#6e6e73;font-size:14px;line-height:1.6;">
                Şu an öne çıkan yeni etkinlik yok.
            </td>
        </tr>
    """

    apple_posts_html = "".join(
        f"""
        <tr>
            <td style="padding:20px 0;border-top:1px solid #e5e5ea;">
                <div style="font-size:15px;line-height:1.4;font-weight:700;color:#1d1d1f;margin:0 0 7px 0;">{escape(item['author'])}</div>
                <div style="font-size:14px;line-height:1.6;color:#6e6e73;margin:0 0 12px 0;">{escape(item['snippet'] or 'Gönderi')}</div>
                <a href="{escape(item['link'])}" style="font-size:14px;line-height:1.4;color:#0071e3;text-decoration:none;font-weight:600;">Paylaşımı aç →</a>
            </td>
        </tr>
        """
        for item in recent_posts
    ) or """
        <tr>
            <td style="padding:20px 0;border-top:1px solid #e5e5ea;color:#6e6e73;font-size:14px;line-height:1.6;">
                Şu an gösterilecek yeni paylaşım yok.
            </td>
        </tr>
    """

    logo_html = (
        '<div style="display:inline-block;border-radius:14px;background:#f5f5f7;'
        'padding:10px 12px;margin:0 0 18px 0;color:#1d1d1f;'
        'font-size:15px;font-weight:800;letter-spacing:-.01em;">HeptaCert</div>'
    )

    events_url = settings.frontend_base_url.rstrip("/") + "/events"
    body_html = f"""
    <!doctype html>
    <html lang="tr">
    <head>
        <meta charset="utf-8" />
        <meta name="viewport" content="width=device-width,initial-scale=1.0" />
        <title>{escape(subject)}</title>
    </head>
    <body style="margin:0;padding:0;background:#f5f5f7;-webkit-font-smoothing:antialiased;">
        <div style="display:none;max-height:0;overflow:hidden;opacity:0;color:transparent;">
            HeptaCert'te öne çıkan {events_count} etkinlik ve {posts_count} topluluk paylaşımı.
        </div>
        <table role="presentation" width="100%" cellspacing="0" cellpadding="0" border="0" style="background:#f5f5f7;padding:28px 12px;">
            <tr>
                <td align="center">
                    <table role="presentation" width="100%" cellspacing="0" cellpadding="0" border="0" style="max-width:680px;background:#ffffff;border-radius:28px;overflow:hidden;">
                        <tr>
                            <td style="padding:44px 44px 28px 44px;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Helvetica,Arial,sans-serif;">
                                {logo_html}
                                <div style="font-size:12px;line-height:1.4;letter-spacing:.12em;text-transform:uppercase;color:#86868b;font-weight:700;margin:0 0 12px 0;">HeptaCert</div>
                                <h1 style="margin:0;color:#1d1d1f;font-size:36px;line-height:1.12;font-weight:800;letter-spacing:-.02em;">Bu haftanın kısa özeti.</h1>
                                <p style="margin:16px 0 0 0;color:#6e6e73;font-size:17px;line-height:1.55;">İlginizi çekebilecek etkinlikleri ve topluluktan son paylaşımları sade bir özet halinde hazırladık.</p>
                            </td>
                        </tr>
                        <tr>
                            <td style="padding:0 44px 28px 44px;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Helvetica,Arial,sans-serif;">
                                <table role="presentation" width="100%" cellspacing="0" cellpadding="0" border="0">
                                    <tr>
                                        <td style="width:50%;padding:0 8px 0 0;">
                                            <div style="background:#f5f5f7;border-radius:20px;padding:18px;">
                                                <div style="font-size:12px;line-height:1.4;color:#86868b;font-weight:700;text-transform:uppercase;letter-spacing:.08em;">Etkinlik</div>
                                                <div style="font-size:32px;line-height:1;font-weight:800;color:#1d1d1f;margin-top:10px;">{events_count}</div>
                                            </div>
                                        </td>
                                        <td style="width:50%;padding:0 0 0 8px;">
                                            <div style="background:#f5f5f7;border-radius:20px;padding:18px;">
                                                <div style="font-size:12px;line-height:1.4;color:#86868b;font-weight:700;text-transform:uppercase;letter-spacing:.08em;">Topluluk</div>
                                                <div style="font-size:32px;line-height:1;font-weight:800;color:#1d1d1f;margin-top:10px;">{posts_count}</div>
                                            </div>
                                        </td>
                                    </tr>
                                </table>
                            </td>
                        </tr>
                        <tr>
                            <td style="padding:0 44px 8px 44px;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Helvetica,Arial,sans-serif;">
                                <h2 style="margin:0;color:#1d1d1f;font-size:22px;line-height:1.25;font-weight:800;letter-spacing:-.01em;">Öne çıkan etkinlikler</h2>
                                <table role="presentation" width="100%" cellspacing="0" cellpadding="0" border="0">{apple_events_html}</table>
                            </td>
                        </tr>
                        <tr>
                            <td style="padding:18px 44px 8px 44px;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Helvetica,Arial,sans-serif;">
                                <h2 style="margin:0;color:#1d1d1f;font-size:22px;line-height:1.25;font-weight:800;letter-spacing:-.01em;">Son topluluk paylaşımları</h2>
                                <table role="presentation" width="100%" cellspacing="0" cellpadding="0" border="0">{apple_posts_html}</table>
                            </td>
                        </tr>
                        <tr>
                            <td style="padding:24px 44px 44px 44px;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Helvetica,Arial,sans-serif;">
                                <a href="{events_url}" style="display:inline-block;background:#0071e3;color:#ffffff;text-decoration:none;border-radius:999px;padding:13px 22px;font-size:15px;line-height:1.2;font-weight:700;">Etkinlikleri keşfet</a>
                                <p style="margin:28px 0 0 0;color:#86868b;font-size:12px;line-height:1.7;">
                                    Bu e-posta, HeptaCert topluluk güncellemeleri kapsamında gönderilmiştir.
                                    <br />
                                    Bu tür e-postaları almak istemiyorsanız <a href="{{{{ unsubscribe_url }}}}" style="color:#0071e3;text-decoration:none;font-weight:600;">buradan abonelikten çıkabilirsiniz</a>.
                                </p>
                                <p style="margin:14px 0 0 0;color:#a1a1a6;font-size:11px;line-height:1.5;">© {now_year} HeptaCert</p>
                            </td>
                        </tr>
                    </table>
                </td>
            </tr>
        </table>
    </body>
    </html>
    """.strip()
    return subject, body_html, events_count, posts_count

    if upcoming_events:
        events_html = "".join(
            f"""
            <tr>
                <td style='padding:0 0 14px 0;'>
                    <table role='presentation' width='100%' cellspacing='0' cellpadding='0' style='border-collapse:collapse;'>
                        <tr>
                            <td style='padding:18px 0;border-top:1px solid #e5e5ea;'>
                                <div style='font-size:17px;line-height:1.35;font-weight:700;color:#1d1d1f;margin:0 0 6px 0;'>{escape(item['name'])}</div>
                                <div style='font-size:14px;color:#475569;margin-bottom:12px;'>{escape(item['date'] or 'Tarih yakında')}{' - ' + escape(item['location']) if item['location'] else ''}{' - ' + escape(item['organization']) if item['organization'] else ''}</div>
                                <a href='{escape(item['link'])}' style='display:inline-block;padding:10px 14px;background:#0f766e;color:#ffffff;text-decoration:none;border-radius:999px;font-size:13px;font-weight:700;'>Etkinliği aç</a>
                            </td>
                        </tr>
                    </table>
                </td>
            </tr>
            """
            for item in upcoming_events
        )
    else:
        events_html = """
        <tr>
            <td style='padding:0 0 14px 0;'>
                <table role='presentation' width='100%' cellspacing='0' cellpadding='0' style='border:1px dashed #d1d5db;border-radius:18px;background:#f8fafc;'>
                    <tr>
                        <td style='padding:18px;color:#64748b;'>Şu an öne çıkan yeni etkinlik yok.</td>
                    </tr>
                </table>
            </td>
        </tr>
        """.strip()

    if recent_posts:
        posts_html = "".join(
            f"""
            <tr>
                <td style='padding:0 0 14px 0;'>
                    <table role='presentation' width='100%' cellspacing='0' cellpadding='0' style='border:1px solid #e5e7eb;border-radius:18px;overflow:hidden;background:#ffffff;'>
                        <tr>
                            <td style='padding:18px 18px 14px 18px;'>
                                <div style='display:inline-block;padding:4px 10px;border-radius:999px;background:#eef2ff;color:#4338ca;font-size:12px;font-weight:700;margin-bottom:10px;'>Topluluk</div>
                                <div style='font-size:18px;font-weight:700;color:#0f172a;margin-bottom:8px;'>{escape(item['author'])}</div>
                                <div style='font-size:14px;color:#475569;margin-bottom:12px;line-height:1.55;'>{escape(item['snippet'] or 'Gönderi')}</div>
                                <a href='{escape(item['link'])}' style='display:inline-block;padding:10px 14px;background:#4338ca;color:#ffffff;text-decoration:none;border-radius:999px;font-size:13px;font-weight:700;'>Paylaşımı aç</a>
                            </td>
                        </tr>
                    </table>
                </td>
            </tr>
            """
            for item in recent_posts
        )
    else:
        posts_html = """
        <tr>
            <td style='padding:0 0 14px 0;'>
                <table role='presentation' width='100%' cellspacing='0' cellpadding='0' style='border:1px dashed #d1d5db;border-radius:18px;background:#f8fafc;'>
                    <tr>
                        <td style='padding:18px;color:#64748b;'>Şu an gösterilecek yeni paylaşım yok.</td>
                    </tr>
                </table>
            </td>
        </tr>
        """.strip()

    body_html = f"""
    <!doctype html>
    <html lang='tr'>
    <head>
        <meta charset='utf-8' />
        <meta name='viewport' content='width=device-width,initial-scale=1.0' />
        <meta http-equiv='x-ua-compatible' content='ie=edge' />
        <title>{escape(subject)}</title>
    </head>
    <body style='margin:0;padding:0;background:#eef2ff;'>
        <div style='display:none;max-height:0;overflow:hidden;opacity:0;color:transparent;mso-hide:all;'>
            HeptaCert: {events_count} etkinlik ve {posts_count} topluluk içeriği.
        </div>
        <table role='presentation' width='100%' cellspacing='0' cellpadding='0' border='0' style='background:linear-gradient(180deg,#eef2ff 0%,#f8fafc 100%);padding:24px 12px;'>
            <tr>
                <td align='center'>
                    <table role='presentation' width='100%' cellspacing='0' cellpadding='0' border='0' style='max-width:760px;background:#ffffff;border-radius:28px;overflow:hidden;box-shadow:0 18px 60px rgba(15,23,42,0.12);'>
                        <tr>
                            <td>
                                <img src='{hero_data_uri}' alt='HeptaCert topluluk özeti' style='display:block;width:100%;max-width:760px;height:auto;border:0;'/>
                            </td>
                        </tr>
                        <tr>
                            <td style='padding:28px 28px 8px 28px;'>
                                {'<img src="' + logo_data_uri + '" alt="HeptaCert" style="display:block;width:56px;height:56px;border-radius:16px;margin-bottom:14px;"/>' if logo_data_uri else ''}
                                <div style='font-size:12px;letter-spacing:.08em;text-transform:uppercase;color:#0f766e;font-weight:800;margin-bottom:8px;'>HeptaCert</div>
                                <h1 style='margin:0 0 12px 0;font-size:28px;line-height:1.2;color:#0f172a;'>Yeni etkinlikler ve topluluk paylaşımları</h1>
                                <p style='margin:0;color:#475569;font-size:15px;line-height:1.7;max-width:620px;'>İlginizi çekebilecek etkinlikleri ve topluluktan son paylaşımları kısa bir özet halinde hazırladık.</p>
                            </td>
                        </tr>
                        <tr>
                            <td style='padding:16px 28px 10px 28px;'>
                                <table role='presentation' width='100%' cellspacing='0' cellpadding='0' border='0'>
                                    <tr>
                                        <td style='padding:0 10px 10px 0;width:50%;'>
                                            <div style='border-radius:18px;background:#f8fafc;border:1px solid #e5e7eb;padding:18px;'>
                                                <div style='font-size:12px;color:#64748b;text-transform:uppercase;letter-spacing:.08em;font-weight:700;'>Etkinlik</div>
                                                <div style='font-size:30px;font-weight:800;color:#0f172a;margin-top:8px;'>{events_count}</div>
                                                <div style='font-size:13px;color:#64748b;margin-top:4px;'>öne çıkan içerik</div>
                                            </div>
                                        </td>
                                        <td style='padding:0 0 10px 10px;width:50%;'>
                                            <div style='border-radius:18px;background:#f8fafc;border:1px solid #e5e7eb;padding:18px;'>
                                                <div style='font-size:12px;color:#64748b;text-transform:uppercase;letter-spacing:.08em;font-weight:700;'>Topluluk</div>
                                                <div style='font-size:30px;font-weight:800;color:#0f172a;margin-top:8px;'>{posts_count}</div>
                                                <div style='font-size:13px;color:#64748b;margin-top:4px;'>son paylaşım</div>
                                            </div>
                                        </td>
                                    </tr>
                                </table>
                            </td>
                        </tr>
                        <tr>
                            <td style='padding:10px 28px 0 28px;'>
                                <h2 style='margin:0 0 14px 0;font-size:20px;color:#0f172a;'>Öne çıkan etkinlikler</h2>
                            </td>
                        </tr>
                        <tr>
                            <td style='padding:0 28px;'>
                                <table role='presentation' width='100%' cellspacing='0' cellpadding='0' border='0'>
                                    {events_html}
                                </table>
                            </td>
                        </tr>
                        <tr>
                            <td style='padding:8px 28px 0 28px;'>
                                <h2 style='margin:0 0 14px 0;font-size:20px;color:#0f172a;'>Son topluluk paylaşımları</h2>
                            </td>
                        </tr>
                        <tr>
                            <td style='padding:0 28px;'>
                                <table role='presentation' width='100%' cellspacing='0' cellpadding='0' border='0'>
                                    {posts_html}
                                </table>
                            </td>
                        </tr>
                        <tr>
                            <td style='padding:14px 28px 30px 28px;'>
                                <table role='presentation' width='100%' cellspacing='0' cellpadding='0' border='0' style='background:linear-gradient(135deg,#0f766e,#4338ca);border-radius:22px;'>
                                    <tr>
                                        <td style='padding:22px 22px 20px 22px;color:#ffffff;'>
                                            <div style='font-size:18px;font-weight:800;margin-bottom:6px;'>HeptaCert'te kesfetmeye devam edin</div>
                                            <div style='font-size:14px;line-height:1.7;opacity:0.95;margin-bottom:16px;'>Etkinlikleri inceleyin, topluluğu takip edin ve size uygun içeriklere ulaşın.</div>
                                            <a href='{settings.frontend_base_url.rstrip("/")}/events' style='display:inline-block;padding:12px 18px;background:#ffffff;color:#0f172a;text-decoration:none;border-radius:999px;font-size:14px;font-weight:800;'>Etkinlikleri keşfet</a>
                                        </td>
                                    </tr>
                                </table>
                            </td>
                        </tr>
                        <tr>
                            <td style='padding:0 28px 28px 28px;'>
                                <div style='font-size:12px;color:#64748b;line-height:1.7;'>
                                    Bu e-posta, HeptaCert topluluk güncellemeleri kapsamında gönderilmiştir.
                                    <br />
                                    Bu tür e-postaları almak istemiyorsanız <a href='{{{{ unsubscribe_url }}}}' style='color:#0f766e;text-decoration:none;font-weight:700;'>buradan abonelikten çıkabilirsiniz</a>.
                                </div>
                                <div style='margin-top:14px;font-size:11px;color:#94a3b8;'>&copy; {now_year} HeptaCert</div>
                            </td>
                        </tr>
                    </table>
                </td>
            </tr>
        </table>
    </body>
    </html>
    """.strip()
    return subject, body_html, events_count, posts_count

async def _create_system_digest_job(
    db: AsyncSession,
    created_by: int,
    subject: str,
    body_html: str,
) -> SuperadminBulkEmailJob:
    recipient_emails = await _resolve_system_digest_recipient_emails(db)
    if not recipient_emails:
        raise HTTPException(status_code=400, detail="Hedef alici bulunamadi.")

    job = SuperadminBulkEmailJob(
        created_by=created_by,
        source="public_members",
        job_kind="system_digest",
        subject=subject,
        body_html=body_html,
        total_targets=len(recipient_emails),
        status="pending",
        cancel_requested=False,
    )
    db.add(job)
    await db.commit()
    await db.refresh(job)
    await _enqueue_superadmin_bulk_email_job(job.id)
    return job


async def _process_system_digest_emails() -> None:
    now = datetime.now(timezone.utc)
    try:
        async with SessionLocal() as db:
            config = await _ensure_system_digest_config(db)
            if not _system_digest_is_due(config, now):
                return

            superadmin_id = await _get_default_superadmin_id(db)
            if superadmin_id is None:
                logger.warning("System digest skipped: no superadmin found")
                return

            subject, body_html, _, _ = await _build_system_digest_email_content(db, config)
            job = await _create_system_digest_job(db, superadmin_id, subject, body_html)
            config.last_sent_at = now
            config.updated_by = superadmin_id
            db.add(config)
            await db.commit()
            logger.info("System digest job queued: %s", job.id)
    except HTTPException as exc:
        logger.info("System digest skipped: %s", exc.detail)
    except Exception as exc:
        logger.warning("System digest scheduler failed: %s", exc)


async def _run_superadmin_bulk_email_job(job_id: int) -> None:
    try:
        async with SessionLocal() as db_job:
            res = await db_job.execute(
                select(SuperadminBulkEmailJob).where(SuperadminBulkEmailJob.id == job_id)
            )
            job = res.scalar_one_or_none()
            if not job or job.status not in {"pending", "sending"}:
                return

            if job.cancel_requested:
                job.status = "cancelled"
                job.completed_at = datetime.now(timezone.utc)
                db_job.add(job)
                await db_job.commit()
                return

            job.status = "sending"
            if not job.started_at:
                job.started_at = datetime.now(timezone.utc)
            job.error_message = None
            db_job.add(job)
            await db_job.commit()

            if job.job_kind == "system_digest":
                recipients = await _resolve_system_digest_recipient_emails(db_job)
            else:
                recipients = await _resolve_superadmin_recipient_emails(db_job, job.source)
            job.total_targets = len(recipients)
            db_job.add(job)
            await db_job.commit()

            if not recipients:
                job.status = "failed"
                job.error_message = "Hedef alici bulunamadi."
                job.completed_at = datetime.now(timezone.utc)
                db_job.add(job)
                await db_job.commit()
                return

            batch_size = max(1, int(settings.email_batch_size or 10))
            sent_count = int(job.sent_count or 0)
            failed_count = int(job.failed_count or 0)
            start_index = max(0, min(sent_count + failed_count, len(recipients)))
            batch_recipients = recipients[start_index:start_index + batch_size]

            for idx, recipient in enumerate(batch_recipients):
                check_res = await db_job.execute(
                    select(SuperadminBulkEmailJob).where(SuperadminBulkEmailJob.id == job_id)
                )
                fresh_job = check_res.scalar_one_or_none()
                if not fresh_job:
                    return
                if fresh_job.cancel_requested:
                    fresh_job.sent_count = sent_count
                    fresh_job.failed_count = failed_count
                    fresh_job.status = "cancelled"
                    fresh_job.completed_at = datetime.now(timezone.utc)
                    db_job.add(fresh_job)
                    await db_job.commit()
                    return

                try:
                    sender_user_id = None if fresh_job.job_kind == "system_digest" else fresh_job.created_by
                    await send_email_async(
                        to=recipient,
                        subject=fresh_job.subject,
                        html_body=fresh_job.body_html,
                        raise_on_error=True,
                        sender_user_id=sender_user_id,
                    )
                    sent_count += 1
                except Exception:
                    failed_count += 1

                if (idx + 1) % batch_size == 0 or start_index + idx == len(recipients) - 1:
                    fresh_job.sent_count = sent_count
                    fresh_job.failed_count = failed_count
                    db_job.add(fresh_job)
                    await db_job.commit()

            if sent_count + failed_count < len(recipients):
                resume_res = await db_job.execute(
                    select(SuperadminBulkEmailJob).where(SuperadminBulkEmailJob.id == job_id)
                )
                resume_job = resume_res.scalar_one_or_none()
                if resume_job:
                    resume_job.sent_count = sent_count
                    resume_job.failed_count = failed_count
                    resume_job.status = "pending"
                    db_job.add(resume_job)
                    await db_job.commit()

                    async def _resume_superadmin_bulk_email() -> None:
                        await asyncio.sleep(float(settings.email_batch_pause_seconds or 0))
                        await _enqueue_superadmin_bulk_email_job(job_id)

                    asyncio.create_task(_resume_superadmin_bulk_email())
                return

            finish_res = await db_job.execute(
                select(SuperadminBulkEmailJob).where(SuperadminBulkEmailJob.id == job_id)
            )
            finish_job = finish_res.scalar_one_or_none()
            if not finish_job:
                return
            finish_job.sent_count = sent_count
            finish_job.failed_count = failed_count
            finish_job.status = "completed"
            finish_job.completed_at = datetime.now(timezone.utc)
            db_job.add(finish_job)
            await db_job.commit()

    except Exception as exc:
        async with SessionLocal() as db_err:
            res = await db_err.execute(
                select(SuperadminBulkEmailJob).where(SuperadminBulkEmailJob.id == job_id)
            )
            job = res.scalar_one_or_none()
            if job:
                job.status = "failed"
                job.error_message = str(exc)[:2000]
                job.completed_at = datetime.now(timezone.utc)
                db_err.add(job)
                await db_err.commit()


async def _enqueue_superadmin_bulk_email_job(job_id: int) -> None:
    async with superadmin_bulk_email_tasks_lock:
        current_task = superadmin_bulk_email_tasks.get(job_id)
        if current_task and not current_task.done():
            return
        task = asyncio.create_task(_run_superadmin_bulk_email_job(job_id))
        superadmin_bulk_email_tasks[job_id] = task

        def _cleanup(_: asyncio.Task) -> None:
            superadmin_bulk_email_tasks.pop(job_id, None)

        task.add_done_callback(_cleanup)


# Ã¢â€â‚¬Ã¢â€â‚¬ Admin: Email Templates Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬

# Ã¢â€â‚¬Ã¢â€â‚¬ Admin: Certificate Templates Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬

@app.get(
    "/api/system/cert-templates",
    response_model=list[CertificateTemplateOut],
)
async def list_cert_templates(db: AsyncSession = Depends(get_db)):
    """Get all available certificate templates."""
    res = await db.execute(
        select(CertificateTemplate)
        .where(CertificateTemplate.is_default == True)
        .order_by(CertificateTemplate.order_index.asc())
    )
    return res.scalars().all()


@app.post(
    "/api/admin/events/{event_id}/apply-cert-template",
    response_model=EventOut,
    dependencies=[Depends(require_role(Role.admin, Role.superadmin)), Depends(require_email_system_access)],
)
async def apply_cert_template(
    event_id: int,
    payload: dict,  # { cert_template_id: int }
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Apply a certificate template to an event."""
    event = await _get_event_for_admin(event_id, me, db, "certificates:write")
    
    cert_template_id = payload.get("cert_template_id")
    if not cert_template_id:
        raise HTTPException(status_code=400, detail="cert_template_id gerekli")
    
    # Get certificate template
    ct_res = await db.execute(
        select(CertificateTemplate).where(CertificateTemplate.id == cert_template_id)
    )
    cert_template = ct_res.scalar_one_or_none()
    if not cert_template:
        raise HTTPException(status_code=404, detail="Sertifika Ã…Å¸ablonu bulunamadÃ„Â±")
    
    # Update event with template image and config
    event.template_image_url = cert_template.template_image_url
    event.config = _merge_certificate_template_config(event.config, cert_template.config)
    db.add(event)
    await db.commit()
    await db.refresh(event)
    
    return _event_to_out(event)


# Ã¢â€â‚¬Ã¢â€â‚¬ Admin: Email Configuration (SMTP) Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬

# Ã¢â€â‚¬Ã¢â€â‚¬ Admin: Bulk Email Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬

@app.post(
    "/api/admin/webhooks",
    response_model=WebhookSubscriptionOut,
    status_code=201,
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def create_webhook_subscription(
    payload: WebhookSubscriptionIn,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Subscribe to email and registration events."""
    # Validate event type
    valid_events = ["email.sent", "email.failed", "email.bouonced", "email.opened", "attendee.register"]
    if payload.event_type not in valid_events:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid event_type. Must be one of: {', '.join(valid_events)}",
        )
    
    # Validate URL
    if not payload.url.startswith("https://"):
        raise HTTPException(
            status_code=400,
            detail="Webhook URL must be HTTPS for security",
        )
    
    webhook = WebhookSubscription(
        user_id=me.id,
        event_type=payload.event_type,
        url=payload.url,
        secret=payload.secret or secrets.token_urlsafe(32),
        is_active=True,
    )
    db.add(webhook)
    await db.commit()
    await db.refresh(webhook)
    return webhook


@app.get(
    "/api/admin/webhooks",
    response_model=list[WebhookSubscriptionOut],
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def list_webhooks(
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List all webhook subscriptions for the user."""
    res = await db.execute(
        select(WebhookSubscription).where(WebhookSubscription.user_id == me.id)
    )
    return res.scalars().all()


@app.patch(
    "/api/admin/webhooks/{webhook_id}",
    response_model=WebhookSubscriptionOut,
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def update_webhook(
    webhook_id: int,
    payload: dict,  # { is_active: bool, url: str }
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Update webhook subscription (enable/disable or change URL)."""
    res = await db.execute(
        select(WebhookSubscription).where(
            WebhookSubscription.id == webhook_id,
            WebhookSubscription.user_id == me.id,
        )
    )
    webhook = res.scalar_one_or_none()
    if not webhook:
        raise HTTPException(status_code=404, detail="Webhook not found")
    
    if "is_active" in payload:
        webhook.is_active = payload["is_active"]
    if "url" in payload:
        try:
            webhook.url = WebhookEndpointIn(url=str(payload["url"]), events=[]).url
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
    
    db.add(webhook)
    await db.commit()
    await db.refresh(webhook)
    return webhook


@app.delete(
    "/api/admin/webhooks/{webhook_id}",
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def delete_webhook(
    webhook_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Delete a webhook subscription."""
    res = await db.execute(
        select(WebhookSubscription).where(
            WebhookSubscription.id == webhook_id,
            WebhookSubscription.user_id == me.id,
        )
    )
    webhook = res.scalar_one_or_none()
    if not webhook:
        raise HTTPException(status_code=404, detail="Webhook not found")
    
    await db.delete(webhook)
    await db.commit()
    return {"message": "Webhook deleted"}


@app.post(
    "/api/admin/webhooks/{webhook_id}/test",
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def test_webhook(
    webhook_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Send a test webhook to verify endpoint is working."""
    res = await db.execute(
        select(WebhookSubscription).where(
            WebhookSubscription.id == webhook_id,
            WebhookSubscription.user_id == me.id,
        )
    )
    webhook = res.scalar_one_or_none()
    if not webhook:
        raise HTTPException(status_code=404, detail="Webhook not found")
    
    # Create test payload
    test_payload = {
        "event": webhook.event_type,
        "timestamp": datetime.utcnow().isoformat(),
        "test": True,
        "data": {
            "bulk_job_id": 0,
            "message": "Test webhook payload"
        }
    }
    
    # Send test webhook
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            response = await client.post(
                webhook.url,
                json=test_payload,
                headers={
                    "X-Heptacert-Event": webhook.event_type,
                    "X-Heptacert-Timestamp": str(int(datetime.utcnow().timestamp())),
                }
            )
        return {
            "status": "sent",
            "http_status": response.status_code,
            "message": f"Webhook sent, endpoint returned {response.status_code}",
        }
    except Exception as e:
        return {
            "status": "error",
            "message": f"Failed to send webhook: {str(e)}",
        }


@app.get("/api/superadmin/admins", response_model=list[AdminRowOut], dependencies=[Depends(require_role(Role.superadmin))])
async def list_admins(db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(User).where(User.role.in_([Role.admin, Role.superadmin])).order_by(User.id.asc()))
    users = res.scalars().all()
    return [
        AdminRowOut(
            id=u.id,
            email=u.email,
            role=u.role,                  # <-- kritik
            heptacoin_balance=u.heptacoin_balaonce
        )
        for u in users
    ]


@app.get("/api/superadmin/transactions", response_model=TxListOut, dependencies=[Depends(require_role(Role.superadmin))])
async def list_transactions(
    user_id: Optional[int] = None,
    page: int = 1,
    limit: int = 30,
    db: AsyncSession = Depends(get_db),
):
    if page < 1 or limit < 1 or limit > 200:
        raise bad_request("Invalid page/limit")

    q = select(Transaction)
    if user_id:
        q = q.where(Transaction.user_id == user_id)

    res_total = await db.execute(select(func.count()).select_from(q.subquery()))
    total = int(res_total.scalar_one())

    q = q.order_by(Transaction.timestamp.desc()).offset((page - 1) * limit).limit(limit)
    res = await db.execute(q)
    items = res.scalars().all()

    return TxListOut(
        items=[
            TxListItem(
                id=t.id,
                user_id=t.user_id,
                amount=t.amount,
                type=t.type,
                timestamp=t.timestamp,
            )
            for t in items
        ],
        total=total,
        page=page,
        limit=limit,
    )


@app.post("/api/superadmin/admins", dependencies=[Depends(require_role(Role.superadmin))])
async def create_admin(payload: CreateAdminIn, db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(User).where(User.email == str(payload.email)))
    if res.scalar_one_or_none():
        raise bad_request("Email already exists")

    admin = User(
        email=str(payload.email),
        password_hash=hash_password(payload.password),
        role=Role.admin,
        heptacoin_balaonce=0,
        is_verified=True,  # superadmin tarafÃ„Â±ndan oluÃ…Å¸turulan hesaplar otomatik doÃ„Å¸rulanÃ„Â±r
    )
    db.add(admin)
    await db.commit()
    await db.refresh(admin)
    return {"id": admin.id, "email": admin.email, "role": admin.role, "heptacoin_balance": admin.heptacoin_balaonce}




@app.delete("/api/superadmin/admins/{admin_id}", dependencies=[Depends(require_role(Role.superadmin))])
async def delete_admin(
    admin_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if admin_id == me.id:
        raise bad_request("Kendi hesabÃ„Â±nÃ„Â±zÃ„Â± silemezsiniz.")
    res = await db.execute(select(User).where(User.id == admin_id))
    user = res.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="Admin not found")
    await db.delete(user)
    await db.commit()
    return {"ok": True}


@app.patch("/api/superadmin/admins/{admin_id}/role", dependencies=[Depends(require_role(Role.superadmin))])
async def change_admin_role(
    admin_id: int,
    payload: AdminRoleIn,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if admin_id == me.id:
        raise bad_request("Kendi rolÃƒÂ¼nÃƒÂ¼zÃƒÂ¼ deÃ„Å¸iÃ…Å¸tiremezsiniz.")
    res = await db.execute(select(User).where(User.id == admin_id))
    user = res.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="Admin not found")
    user.role = Role(payload.role)
    await db.commit()
    return {"ok": True, "new_role": payload.role}


@app.get("/api/admin/transactions", dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def my_transactions(
    limit: int = Query(default=50, le=200),
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    res = await db.execute(
        select(Transaction)
        .where(Transaction.user_id == me.id)
        .order_by(Transaction.timestamp.desc())
        .limit(limit)
    )
    txs = res.scalars().all()
    return [{"id": t.id, "amount": t.amount, "type": t.type, "timestamp": t.timestamp} for t in txs]


@app.post("/api/superadmin/coins/credit", dependencies=[Depends(require_role(Role.superadmin))])
async def credit_coins(payload: CreditCoinsIn, db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(User).where(User.id == payload.admin_user_id))
    user = res.scalar_one_or_none()
    if not user or user.role not in (Role.admin, Role.superadmin):
        raise bad_request("Admin or superadmin user not found")

    user.heptacoin_balaonce += payload.amount
    db.add(Transaction(user_id=user.id, amount=payload.amount, type=TxType.credit))
    await db.commit()
    return {"admin_user_id": user.id, "new_balance": user.heptacoin_balaonce}


# Ã¢â€â‚¬Ã¢â€â‚¬ Waitlist Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬

@app.post("/api/waitlist", status_code=201)
@limiter.limit("5/minute")
async def join_waitlist(request: Request, data: WaitlistIn, db: AsyncSession = Depends(get_db)):
    """Public endpoint Ã¢â‚¬â€ anyone can join the waitlist."""
    existing = await db.execute(select(WaitlistEntry).where(WaitlistEntry.email == str(data.email)))
    if existing.scalar_one_or_none():
        return {"ok": True, "already_registered": True}
    entry = WaitlistEntry(
        name=data.name,
        email=str(data.email),
        phone=data.phone,
        plan_interest=data.plan_interest,
        note=data.note,
    )
    db.add(entry)
    await db.commit()
    return {"ok": True, "already_registered": False}


@app.get("/api/superadmin/waitlist", dependencies=[Depends(require_role(Role.superadmin))])
async def get_waitlist(db: AsyncSession = Depends(get_db)):
    """Superadmin: list all waitlist entries ordered newest first."""
    res = await db.execute(select(WaitlistEntry).order_by(WaitlistEntry.created_at.desc()))
    entries = res.scalars().all()
    return {
        "total": len(entries),
        "entries": [WaitlistEntryOut.model_validate(e) for e in entries],
    }


# Ã¢â€â‚¬Ã¢â€â‚¬ Pricing Config (public GET + superadmin PATCH) Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬

@app.get("/api/pricing/config", response_model=PricingConfigOut)
async def get_pricing_config(db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(SystemConfig).where(SystemConfig.key == "pricing"))
    cfg = res.scalar_one_or_none()
    if cfg is None or not cfg.value.get("tiers"):
        return PricingConfigOut(tiers=[PricingTier(**t) for t in DEFAULT_PRICING])
    return PricingConfigOut(tiers=[PricingTier(**t) for t in cfg.value["tiers"]])


@app.get("/api/superadmin/pricing", response_model=PricingConfigOut, dependencies=[Depends(require_role(Role.superadmin))])
async def get_pricing_config_admin(db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(SystemConfig).where(SystemConfig.key == "pricing"))
    cfg = res.scalar_one_or_none()
    if cfg is None or not cfg.value.get("tiers"):
        return PricingConfigOut(tiers=[PricingTier(**t) for t in DEFAULT_PRICING])
    return PricingConfigOut(tiers=[PricingTier(**t) for t in cfg.value["tiers"]])


@app.patch("/api/superadmin/pricing", response_model=PricingConfigOut, dependencies=[Depends(require_role(Role.superadmin))])
async def update_pricing_config(payload: PricingConfigOut, db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(SystemConfig).where(SystemConfig.key == "pricing"))
    cfg = res.scalar_one_or_none()
    data = {"tiers": [t.model_dump() for t in payload.tiers]}
    if cfg is None:
        cfg = SystemConfig(key="pricing", value=data)
        db.add(cfg)
    else:
        cfg.value = data
    await db.commit()
    return payload


# Ã¢â€â‚¬Ã¢â€â‚¬ Public Stats Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬

DEFAULT_STATS = {
    "active_members": "12.4K+",
    "hosted_events": "850+",
    "issued_certificates": "50.000+",
    # Backward compatibility keys
    "active_orgs": "500+",
    "certs_issued": "50.000+",
    "uptime_pct": "%100",
    "availability": "7/24",
}






@app.get("/api/stats", response_model=StatsOut)
async def get_public_stats(db: AsyncSession = Depends(get_db)):
    """Public stats endpoint Ã¢â‚¬â€ returns display values (overridden by superadmin or real DB counts)."""
    res = await db.execute(select(SystemConfig).where(SystemConfig.key == "stats"))
    cfg = res.scalar_one_or_none()
    overrides: dict = cfg.value if cfg else {}

    if overrides.get("use_real_counts", True):
        # Real DB counts
        member_count_res = await db.execute(select(func.count()).select_from(PublicMember))
        member_count = member_count_res.scalar_one() or 0
        event_count_res = await db.execute(select(func.count()).select_from(Event))
        event_count = event_count_res.scalar_one() or 0
        org_count_res = await db.execute(select(func.count(func.distinct(Event.admin_id))))
        org_count = org_count_res.scalar_one() or 0
        cert_count_res = await db.execute(
            select(func.count()).select_from(Certificate).where(Certificate.deleted_at.is_(None))
        )
        cert_count = cert_count_res.scalar_one() or 0

        return StatsOut(
            active_members=overrides.get("active_members") or f"{member_count:,}".replace(",", "."),
            hosted_events=overrides.get("hosted_events") or f"{event_count:,}".replace(",", "."),
            issued_certificates=overrides.get("issued_certificates") or f"{cert_count:,}".replace(",", "."),
            active_orgs=overrides.get("active_orgs") or f"{org_count:,}".replace(",", "."),
            certs_issued=overrides.get("certs_issued") or f"{cert_count:,}".replace(",", "."),
            uptime_pct=overrides.get("uptime_pct") or DEFAULT_STATS["uptime_pct"],
            availability=overrides.get("availability") or DEFAULT_STATS["availability"],
        )

    return StatsOut(
        active_members=overrides.get("active_members", DEFAULT_STATS["active_members"]),
        hosted_events=overrides.get("hosted_events", DEFAULT_STATS["hosted_events"]),
        issued_certificates=overrides.get("issued_certificates", DEFAULT_STATS["issued_certificates"]),
        active_orgs=overrides.get("active_orgs", DEFAULT_STATS["active_orgs"]),
        certs_issued=overrides.get("certs_issued", DEFAULT_STATS["certs_issued"]),
        uptime_pct=overrides.get("uptime_pct", DEFAULT_STATS["uptime_pct"]),
        availability=overrides.get("availability", DEFAULT_STATS["availability"]),
    )


@app.get("/api/superadmin/stats", response_model=dict, dependencies=[Depends(require_role(Role.superadmin))])
async def get_stats_config(db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(SystemConfig).where(SystemConfig.key == "stats"))
    cfg = res.scalar_one_or_none()
    return cfg.value if cfg else {**DEFAULT_STATS, "use_real_counts": True}


@app.patch("/api/superadmin/stats", response_model=dict, dependencies=[Depends(require_role(Role.superadmin))])
async def update_stats_config(payload: StatsIn, db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(SystemConfig).where(SystemConfig.key == "stats"))
    cfg = res.scalar_one_or_none()
    data = payload.model_dump(exclude_none=False)
    if cfg is None:
        cfg = SystemConfig(key="stats", value=data)
        db.add(cfg)
    else:
        cfg.value = data
    await db.commit()
    return data


# Ã¢â€â‚¬Ã¢â€â‚¬ Billing / Payment endpoints Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬

from .payments import get_provider, PaymentRequest  # noqa: E402






@app.get("/api/billing/status")
async def billing_status():
    """Returns whether payment is enabled and which provider is active."""
    return {
        "enabled": settings.payment_enabled,
        "provider": settings.active_payment_provider if settings.payment_enabled else None,
        "stripe_publishable_key": settings.stripe_publishable_key if settings.payment_enabled and settings.active_payment_provider == "stripe" else None,
    }


@app.post("/api/billing/create-payment")
async def create_payment(
    payload: CreatePaymentIn,
    request: Request,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if not settings.payment_enabled:
        raise HTTPException(status_code=503, detail="Payment system is not yet enabled.")

    provider = get_provider(settings)
    if provider is None:
        raise HTTPException(status_code=503, detail="No payment provider configured.")

    # Lookup pricing tier for amount
    cfg_res = await db.execute(select(SystemConfig).where(SystemConfig.key == "pricing"))
    pricing_cfg = cfg_res.scalar_one_or_none()
    tiers = pricing_cfg.value.get("tiers", DEFAULT_PRICING) if pricing_cfg else DEFAULT_PRICING

    tier = next((t for t in tiers if t.get("id") == payload.plan_id), None)
    if tier is None:
        raise HTTPException(status_code=404, detail="Plan not found.")

    if payload.billing_period == "annual":
        price = tier.get("price_annual")
    else:
        price = tier.get("price_monthly")

    if price is None:
        raise HTTPException(status_code=400, detail="This plan requires custom pricing. Contact sales.")

    amount_cents = int(float(price) * 100)

    # Create Order record
    order = Order(
        user_id=me.id,
        plan_id=payload.plan_id,
        amount_cents=amount_cents,
        currency="TRY",
        provider=provider.name,
        status=OrderStatus.pending,
        meta={"billing_period": payload.billing_period},
    )
    db.add(order)
    await db.commit()
    await db.refresh(order)

    client_ip = request.client.host if request.client else "127.0.0.1"
    frontend = settings.frontend_base_url.rstrip("/")

    pay_req = PaymentRequest(
        order_id=str(order.id),
        amount_cents=amount_cents,
        currency="TRY",
        description=f"HeptaCert {tier.get('name_tr', payload.plan_id)} - {payload.billing_period}",
        customer_email=me.email,
        customer_name=me.email.split("@")[0],
        customer_ip=client_ip,
        success_url=f"{frontend}/checkout/success?order_id={order.id}",
        cancel_url=f"{frontend}/checkout/cancel?order_id={order.id}",
        webhook_url=f"{settings.public_base_url}/api/billing/webhook/{provider.name}",
    )

    result = await provider.create_payment(pay_req)

    if not result.success:
        order.status = OrderStatus.failed
        await db.commit()
        raise HTTPException(status_code=502, detail=result.error or "Payment initiation failed.")

    order.provider_ref = result.provider_ref
    await db.commit()

    return {
        "order_id": order.id,
        "checkout_url": result.checkout_url,
        "checkout_html": result.checkout_html,
        "provider": provider.name,
    }


@app.post("/api/billing/webhook/{provider_name}")
async def payment_webhook(
    provider_name: str,
    request: Request,
    db: AsyncSession = Depends(get_db),
):
    """Server-to-server payment notification from provider."""
    if not settings.payment_enabled:
        return {"ok": True}

    provider = get_provider(settings)
    if provider is None or provider.name != provider_name:
        raise HTTPException(status_code=400, detail="Unknown provider")

    raw_body = await request.body()
    headers = dict(request.headers)

    try:
        notification = await provider.verify_webhook(raw_body, headers)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    order_id_raw = notification.get("order_id")
    provider_ref = str(notification.get("provider_ref") or "").strip()
    order_id: Optional[int] = None
    if order_id_raw:
        try:
            order_id = int(order_id_raw)
        except (ValueError, TypeError):
            order_id = None
    if order_id is not None:
        res = await db.execute(select(Order).where(Order.id == order_id).with_for_update())
    elif provider_ref:
        res = await db.execute(select(Order).where(Order.provider_ref == provider_ref).with_for_update())
    else:
        return {"ok": True}
    order = res.scalar_one_or_none()
    if order is None:
        return {"ok": True}
    if order.provider != provider.name:
        raise HTTPException(status_code=400, detail="Payment provider does not match order.")
    if provider.name in {"iyzico", "stripe"} and order.provider_ref and not hmac.compare_digest(order.provider_ref, provider_ref):
        raise HTTPException(status_code=400, detail="Payment reference does not match order.")

    verified_amount = notification.get("amount_cents")
    verified_currency = str(notification.get("currency") or "").upper()
    if verified_amount is not None and int(verified_amount) != order.amount_cents:
        raise HTTPException(status_code=400, detail="Payment amount does not match order.")
    if verified_currency and verified_currency != order.currency.upper():
        raise HTTPException(status_code=400, detail="Payment currency does not match order.")

    status = notification.get("status", "failed")
    order.provider_ref = provider_ref or order.provider_ref

    if status == "paid" and order.status != OrderStatus.paid:
        order.status = OrderStatus.paid
        order.paid_at = datetime.now(timezone.utc)
        # Create or extend subscription
        sub_res = await db.execute(
            select(Subscription).where(Subscription.user_id == order.user_id, Subscription.plan_id == order.plan_id, Subscription.is_active == True)
        )
        existing_sub = sub_res.scalar_one_or_none()
        period = (order.meta or {}).get("billing_period", "monthly")
        delta = timedelta(days=365 if period == "annual" else 31)
        hc_quota_pay = _get_hc_quota(order.plan_id)
        if existing_sub:
            now = datetime.now(timezone.utc)
            base = max(existing_sub.expires_at or now, now)
            existing_sub.expires_at = base + delta
            existing_sub.last_hc_credited_at = now
        else:
            now = datetime.now(timezone.utc)
            db.add(Subscription(
                user_id=order.user_id,
                plan_id=order.plan_id,
                order_id=order.id,
                expires_at=now + delta,
                is_active=True,
                last_hc_credited_at=now,
            ))
        # Credit HC quota on every successful payment period
        if hc_quota_pay:
            usr_pay_res = await db.execute(select(User).where(User.id == order.user_id))
            usr_pay = usr_pay_res.scalar_one_or_none()
            if usr_pay:
                usr_pay.heptacoin_balaonce += hc_quota_pay
                db.add(Transaction(
                    user_id=usr_pay.id, amount=hc_quota_pay, type=TxType.credit,
                    description=f"Plan {('aktivasyonu' if not existing_sub else 'yenileme')}: {order.plan_id} ({period})",
                ))
    elif status == "failed" and order.status == OrderStatus.pending:
        order.status = OrderStatus.failed
    elif status == "refunded":
        order.status = OrderStatus.refunded

    await db.commit()
    return {"ok": True}


@app.get("/api/billing/orders", response_model=List[OrderOut])
async def list_orders(me: CurrentUser = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(Order).where(Order.user_id == me.id).order_by(Order.created_at.desc()).limit(50)
    )
    orders = res.scalars().all()
    return [OrderOut(
        id=o.id, plan_id=o.plan_id, amount_cents=o.amount_cents, currency=o.currency,
        provider=o.provider, status=o.status, created_at=o.created_at, paid_at=o.paid_at,
    ) for o in orders]


@app.get("/api/billing/subscription")
async def my_subscription(me: CurrentUser = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(Subscription).where(Subscription.user_id == me.id, Subscription.is_active == True)
        .order_by(Subscription.expires_at.desc()).limit(1)
    )
    sub = res.scalar_one_or_none()
    if sub is None:
        return {"active": False, "plan_id": None, "expires_at": None, "role": me.role.value}
    now = datetime.now(timezone.utc)
    is_valid = sub.expires_at is None or sub.expires_at > now
    return {
        "active": is_valid,
        "plan_id": sub.plan_id,
        "expires_at": sub.expires_at.isoformat() if sub.expires_at else None,
        "role": me.role.value,
    }


# Superadmin: payment config management


@app.get("/api/superadmin/payment-config", dependencies=[Depends(require_role(Role.superadmin))])
async def get_payment_config(db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(SystemConfig).where(SystemConfig.key == "payment_config"))
    cfg = res.scalar_one_or_none()
    data = cfg.value if cfg else {}
    # Merge DB overrides with current env settings
    return {
        "enabled": data.get("enabled", settings.payment_enabled),
        "active_provider": data.get("active_provider", settings.active_payment_provider),
        "iyzico_api_key": data.get("iyzico_api_key", settings.iyzico_api_key),
        "iyzico_secret_key": data.get("iyzico_secret_key", settings.iyzico_secret_key),
        "iyzico_base_url": data.get("iyzico_base_url", settings.iyzico_base_url),
        "paytr_merchant_id": data.get("paytr_merchant_id", settings.paytr_merchant_id),
        "paytr_merchant_key": data.get("paytr_merchant_key", settings.paytr_merchant_key),
        "paytr_merchant_salt": data.get("paytr_merchant_salt", settings.paytr_merchant_salt),
        "stripe_secret_key": data.get("stripe_secret_key", settings.stripe_secret_key),
        "stripe_webhook_secret": data.get("stripe_webhook_secret", settings.stripe_webhook_secret),
        "stripe_publishable_key": data.get("stripe_publishable_key", settings.stripe_publishable_key),
    }


@app.patch("/api/superadmin/payment-config", dependencies=[Depends(require_role(Role.superadmin))])
async def update_payment_config(payload: PaymentConfigOut, db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(SystemConfig).where(SystemConfig.key == "payment_config"))
    cfg = res.scalar_one_or_none()
    data = payload.model_dump()
    if cfg is None:
        cfg = SystemConfig(key="payment_config", value=data)
        db.add(cfg)
    else:
        cfg.value = data
    await db.commit()
    return data







@app.get("/api/me", response_model=MeOut, dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def me(me: CurrentUser = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(User).where(User.id == me.id))
    u = res.scalar_one()
    return MeOut(id=u.id, email=u.email, role=u.role, heptacoin_balance=u.heptacoin_balaonce)


@app.get("/api/me/export", dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def export_admin_data(
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """KVKK md. 11 / GDPR Art. 20 — data portability export for admin accounts."""
    import json as _json
    from fastapi.responses import Response as _Response

    user_res = await db.execute(select(User).where(User.id == me.id))
    user = user_res.scalar_one()

    events_res = await db.execute(select(Event).where(Event.admin_id == me.id))
    events = events_res.scalars().all()

    event_ids = [e.id for e in events]
    certs_res = await db.execute(select(Certificate).where(Certificate.event_id.in_(event_ids), Certificate.deleted_at.is_(None))) if event_ids else None
    certs = certs_res.scalars().all() if certs_res else []

    payload = {
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "account": {
            "id": user.id,
            "email": user.email,
            "role": str(user.role),
            "created_at": user.created_at.isoformat() if user.created_at else None,
        },
        "events": [
            {
                "id": e.id,
                "name": e.name,
                "event_date": e.event_date.isoformat() if e.event_date else None,
                "created_at": e.created_at.isoformat() if e.created_at else None,
            }
            for e in events
        ],
        "certificates_issued": len(certs),
    }
    await write_audit_log(
        db,
        user_id=me.id,
        action="user.data_export",
        resource_type="user",
        resource_id=str(me.id),
    )
    await db.commit()
    body = _json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    return _Response(
        content=body,
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="heptacert-data-export-{me.id}.json"'},
    )


@app.get("/api/public/me/export")
async def export_public_member_data(
    member: CurrentPublicMember = Depends(get_current_public_member),
    db: AsyncSession = Depends(get_db),
):
    """KVKK md. 11 / GDPR Art. 20 — data portability export for public members."""
    import json as _json
    from fastapi.responses import Response as _Response

    member_res = await db.execute(select(PublicMember).where(PublicMember.id == member.id))
    db_member = member_res.scalar_one_or_none()
    if not db_member:
        raise HTTPException(status_code=404, detail="Member not found.")

    attendees_res = await db.execute(
        select(Attendee, Event)
        .join(Event, Event.id == Attendee.event_id)
        .where(func.lower(func.trim(Attendee.email)) == (db_member.email or "").strip().lower())
        .order_by(Attendee.registered_at.desc())
        .limit(500)
    )
    attendee_rows = attendees_res.all()

    payload = {
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "profile": {
            "email": db_member.email,
            "display_name": db_member.display_name,
            "bio": db_member.bio,
            "location": db_member.location,
            "created_at": db_member.created_at.isoformat() if db_member.created_at else None,
        },
        "event_registrations": [
            {
                "event_id": event.id,
                "event_name": event.name,
                "event_date": event.event_date.isoformat() if event.event_date else None,
                "registered_at": attendee.registered_at.isoformat() if attendee.registered_at else None,
                "email_verified": bool(attendee.email_verified),
                "unsubscribed": attendee.unsubscribed_at is not None,
            }
            for attendee, event in attendee_rows
        ],
    }
    body = _json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    return _Response(
        content=body,
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="heptacert-member-data-export.json"'},
    )


@app.get("/api/public/me", response_model=PublicMemberMeOut)
async def public_me(
    member: CurrentPublicMember = Depends(get_current_public_member),
    db: AsyncSession = Depends(get_db),
):
    res = await db.execute(select(PublicMember).where(PublicMember.id == member.id))
    db_member = res.scalar_one_or_none()
    if not db_member:
        raise HTTPException(status_code=404, detail="Member not found.")
    return PublicMemberMeOut(
        id=db_member.id,
        public_id=db_member.public_id,
        email=db_member.email,
        display_name=db_member.display_name,
        bio=db_member.bio,
        avatar_url=db_member.avatar_url,
        headline=db_member.headline,
        location=db_member.location,
        website_url=db_member.website_url,
        contact_email=db_member.contact_email,
        created_at=db_member.created_at,
    )


@app.get("/api/public/me/email-preferences", response_model=PublicMemberEmailPrefereoncesOut)
@app.get("/api/public/me/email-prefereonces", response_model=PublicMemberEmailPrefereoncesOut)  # geriye donuk uyumluluk (eski yazim hatasi)
async def get_public_member_email_prefereonces(
    member: CurrentPublicMember = Depends(get_current_public_member),
    db: AsyncSession = Depends(get_db),
):
    res = await db.execute(select(PublicMember).where(PublicMember.id == member.id))
    db_member = res.scalar_one_or_none()
    if not db_member:
        raise HTTPException(status_code=404, detail="Member not found.")
    return PublicMemberEmailPrefereoncesOut(digest_opt_in=bool(getattr(db_member, "digest_opt_in", True)))


@app.patch("/api/public/me/email-preferences", response_model=PublicMemberEmailPrefereoncesOut)
@app.patch("/api/public/me/email-prefereonces", response_model=PublicMemberEmailPrefereoncesOut)  # geriye donuk uyumluluk (eski yazim hatasi)
async def update_public_member_email_prefereonces(
    data: PublicMemberEmailPrefereoncesIn,
    member: CurrentPublicMember = Depends(get_current_public_member),
    db: AsyncSession = Depends(get_db),
):
    res = await db.execute(select(PublicMember).where(PublicMember.id == member.id))
    db_member = res.scalar_one_or_none()
    if not db_member:
        raise HTTPException(status_code=404, detail="Member not found.")
    db_member.digest_opt_in = bool(data.digest_opt_in)
    await db.commit()
    await db.refresh(db_member)
    return PublicMemberEmailPrefereoncesOut(digest_opt_in=bool(db_member.digest_opt_in))


@app.patch("/api/public/me", response_model=PublicMemberMeOut)
async def update_public_me(
    data: PublicMemberProfileUpdateIn,
    member: CurrentPublicMember = Depends(get_current_public_member),
    db: AsyncSession = Depends(get_db),
):
    res = await db.execute(select(PublicMember).where(PublicMember.id == member.id))
    db_member = res.scalar_one_or_none()
    if not db_member:
        raise HTTPException(status_code=404, detail="Member not found.")

    display_name = data.display_name.strip()
    if not display_name:
        raise bad_request("Display name is required.")

    bio = (data.bio or "").strip()
    headline = (data.headline or "").strip()
    location = (data.location or "").strip()
    website_url = (data.website_url or "").strip()
    contact_email = (str(data.contact_email).strip().lower() if data.contact_email else "")
    db_member.display_name = display_name
    db_member.bio = bio or None
    db_member.headline = headline or None
    db_member.location = location or None
    db_member.website_url = website_url or None
    db_member.contact_email = contact_email or None
    await db.commit()
    await db.refresh(db_member)
    return PublicMemberMeOut(
        id=db_member.id,
        public_id=db_member.public_id,
        email=db_member.email,
        display_name=db_member.display_name,
        bio=db_member.bio,
        avatar_url=db_member.avatar_url,
        headline=db_member.headline,
        location=db_member.location,
        website_url=db_member.website_url,
        contact_email=db_member.contact_email,
        created_at=db_member.created_at,
    )


@app.post("/api/public/me/avatar", response_model=PublicMemberMeOut)
async def upload_public_member_avatar(
    file: UploadFile = File(...),
    member: CurrentPublicMember = Depends(get_current_public_member),
    db: AsyncSession = Depends(get_db),
):
    res = await db.execute(select(PublicMember).where(PublicMember.id == member.id))
    db_member = res.scalar_one_or_none()
    if not db_member:
        raise HTTPException(status_code=404, detail="Member not found.")
    data, ext = await _read_safe_raster_upload(file)
    safe_name = f"member-avatars/member_{member.id}/avatar{ext}"
    dest = Path(settings.local_storage_dir) / safe_name
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(data)

    db_member.avatar_url = f"{settings.public_base_url}/api/files/{safe_name}"
    await db.commit()
    await db.refresh(db_member)
    return PublicMemberMeOut(
        id=db_member.id,
        public_id=db_member.public_id,
        email=db_member.email,
        display_name=db_member.display_name,
        bio=db_member.bio,
        avatar_url=db_member.avatar_url,
        headline=db_member.headline,
        location=db_member.location,
        website_url=db_member.website_url,
        contact_email=db_member.contact_email,
        created_at=db_member.created_at,
    )


@app.get("/api/public/members/{member_public_id}", response_model=PublicMemberProfileOut)
async def get_public_member_profile(
    member_public_id: str,
    db: AsyncSession = Depends(get_db),
    viewer: Optional[CurrentPublicMember] = Depends(get_optional_public_member),
):
    res = await db.execute(select(PublicMember).where(PublicMember.public_id == member_public_id))
    db_member = res.scalar_one_or_none()
    if not db_member:
        raise HTTPException(status_code=404, detail="Member not found.")

    event_count_res = await db.execute(
        select(func.count(func.distinct(Attendee.event_id))).where(Attendee.public_member_id == db_member.id)
    )
    comment_count_res = await db.execute(
        select(func.count(EventComment.id)).where(EventComment.public_member_id == db_member.id, EventComment.status == "visible")
    )
    from .member_certificates_api import can_view_member_certificate_wallet, list_public_member_certificates

    can_view_certificates = await can_view_member_certificate_wallet(db, db_member, viewer)
    certificates = await list_public_member_certificates(db, db_member.id) if can_view_certificates else []

    # Fetch earned badges across all events for this member
    attendee_ids_res = await db.execute(
        select(Attendee.id, Attendee.event_id).where(Attendee.public_member_id == db_member.id)
    )
    attendee_rows = attendee_ids_res.all()
    badges_out: list[PublicMemberBadgeOut] = []
    if attendee_rows:
        att_id_list = [row.id for row in attendee_rows]
        badges_res = await db.execute(
            select(ParticipantBadge, Event.name.label("event_name"))
            .join(Event, ParticipantBadge.event_id == Event.id)
            .where(ParticipantBadge.attendee_id.in_(att_id_list))
            .order_by(ParticipantBadge.awarded_at.desc())
            .limit(50)
        )
        for badge_row, event_name in badges_res.all():
            meta = badge_row.badge_metadata or {}
            badges_out.append(PublicMemberBadgeOut(
                badge_id=f"{badge_row.event_id}:{badge_row.badge_type}",
                name=meta.get("name") or badge_row.badge_type.replace("_", " ").title(),
                icon=meta.get("icon"),
                color=meta.get("color"),
                event_id=badge_row.event_id,
                event_name=event_name,
                awarded_at=badge_row.awarded_at,
            ))

    return PublicMemberProfileOut(
        public_id=db_member.public_id,
        display_name=db_member.display_name,
        bio=db_member.bio,
        avatar_url=db_member.avatar_url,
        headline=db_member.headline,
        location=db_member.location,
        website_url=db_member.website_url,
        contact_email=db_member.contact_email,
        created_at=db_member.created_at,
        event_count=int(event_count_res.scalar_one() or 0),
        comment_count=int(comment_count_res.scalar_one() or 0),
        certificates=certificates,
        certificates_hidden=not can_view_certificates,
        badges=badges_out,
    )


@app.patch("/api/public/me/password")
async def change_public_member_password(
    data: PublicMemberChangePasswordIn,
    member: CurrentPublicMember = Depends(get_current_public_member),
    db: AsyncSession = Depends(get_db),
):
    res = await db.execute(select(PublicMember).where(PublicMember.id == member.id))
    db_member = res.scalar_one_or_none()
    if not db_member:
        raise HTTPException(status_code=404, detail="Member not found.")
    if not verify_password(data.current_password, db_member.password_hash):
        raise HTTPException(status_code=400, detail="Current password is incorrect.")
    if verify_password(data.new_password, db_member.password_hash):
        raise HTTPException(status_code=400, detail="New password must be different from the current password.")

    db_member.password_hash = hash_password(data.new_password)
    db_member.password_reset_token = None
    await db.commit()
    return {"detail": "Password updated successfully."}


@app.delete("/api/public/me")
async def delete_public_member_account(
    data: DeleteAccountIn,
    member: CurrentPublicMember = Depends(get_current_public_member),
    db: AsyncSession = Depends(get_db),
):
    res = await db.execute(select(PublicMember).where(PublicMember.id == member.id))
    db_member = res.scalar_one_or_none()
    if not db_member:
        raise HTTPException(status_code=404, detail="Member not found.")
    if not verify_password(data.current_password, db_member.password_hash):
        raise HTTPException(status_code=400, detail="Current password is incorrect.")

    now = datetime.now(timezone.utc)
    db_member.deleted_at = now
    db_member.display_name = "Silinmiş Üye"
    db_member.bio = None
    db_member.avatar_url = None
    db_member.headline = None
    db_member.location = None
    db_member.website_url = None
    db_member.contact_email = None
    db_member.password_hash = hash_password(secrets.token_hex(32))
    db_member.password_reset_token = None
    db_member.verification_token = None
    await db.execute(delete(OrganizationFollower).where(OrganizationFollower.public_member_id == db_member.id))
    await db.execute(delete(CommunityPostLike).where(CommunityPostLike.public_member_id == db_member.id))
    await db.commit()
    return {"detail": "Hesabınız ve kişisel verileriniz silinmiştir. Yasal saklama süresi dolan veriler 30 gün içinde kalıcı olarak temizlenecektir."}


@app.get("/api/public/my-events", response_model=list[PublicMemberEventOut])
async def public_member_events(
    member: CurrentPublicMember = Depends(get_current_public_member),
    db: AsyncSession = Depends(get_db),
):
    legacy_res = await db.execute(
        select(Attendee).where(
            func.lower(Attendee.email) == member.email.lower(),
            Attendee.public_member_id.is_(None),
        )
    )
    legacy_attendees = legacy_res.scalars().all()
    if legacy_attendees:
        for attendee in legacy_attendees:
            attendee.public_member_id = member.id
        await db.commit()

    attendee_res = await db.execute(
        select(Attendee)
        .options(selectinload(Attendee.event))
        .where(Attendee.public_member_id == member.id)
        .order_by(Attendee.registered_at.desc())
    )
    attendees = attendee_res.scalars().all()
    if not attendees:
        return []

    attendaonce_counts_res = await db.execute(
        select(AttendaonceRecord.attendee_id, func.count().label("cnt"))
        .where(AttendaonceRecord.attendee_id.in_([attendee.id for attendee in attendees]))
        .group_by(AttendaonceRecord.attendee_id)
    )
    attendaonce_counts = {int(row.attendee_id): int(row.cnt or 0) for row in attendaonce_counts_res.all()}

    return [
        PublicMemberEventOut(
            attendee_id=attendee.id,
            event_id=attendee.event_id,
            event_name=attendee.event.name,
            event_date=attendee.event.event_date.isoformat() if attendee.event.event_date else None,
            event_location=attendee.event.event_location,
            event_banner_url=attendee.event.event_banner_url,
            registered_at=attendee.registered_at,
            email_verified=attendee.email_verified,
            sessions_attended=attendaonce_counts.get(attendee.id, 0),
            min_sessions_required=attendee.event.min_sessions_required,
            status_url=build_public_status_url(
                event_id=_get_public_event_identifier(attendee.event),
                attendee_id=attendee.id,
                email=attendee.email,
            ),
        )
        for attendee in attendees
    ]


@app.patch("/api/me/password", dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def change_password(
    data: ChangePasswordIn,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    res = await db.execute(select(User).where(User.id == me.id))
    user = res.scalar_one()
    if not verify_password(data.current_password, user.password_hash):
        raise bad_request("Mevcut Ã…Å¸ifre yanlÃ„Â±Ã…Å¸.")
    user.password_hash = hash_password(data.new_password)
    await db.commit()
    return {"detail": "Ã…Âifre baÃ…Å¸arÃ„Â±yla gÃƒÂ¼oncellendi."}


@app.patch("/api/me/email", dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def change_email(
    data: ChangeEmailIn,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    res = await db.execute(select(User).where(User.id == me.id))
    user = res.scalar_one()
    if not verify_password(data.current_password, user.password_hash):
        raise bad_request("Mevcut Ã…Å¸ifre yanlÃ„Â±Ã…Å¸.")
    exists = await db.execute(select(User).where(User.email == str(data.new_email)))
    if exists.scalar_one_or_none():
        raise bad_request("Bu e-posta adresi zaten kullanÃ„Â±mda.")
    user.email = str(data.new_email)
    await db.commit()
    return {"detail": "E-posta baÃ…Å¸arÃ„Â±yla gÃƒÂ¼oncellendi."}


@app.delete("/api/me", dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def delete_admin_account(
    data: DeleteAccountIn,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    res = await db.execute(select(User).where(User.id == me.id))
    user = res.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="Kullanici bulunamadi.")
    if user.role == Role.superadmin:
        raise HTTPException(status_code=400, detail="Superadmin hesabi panelden silinemez.")
    if not verify_password(data.current_password, user.password_hash):
        raise HTTPException(status_code=400, detail="Mevcut şifre hatalı.")

    now = datetime.now(timezone.utc)
    user.deleted_at = now
    user.password_hash = hash_password(secrets.token_hex(32))
    user.verification_token = None
    user.password_reset_token = None
    user.magic_link_token = None
    from .cache import cache
    await cache.delete(f"user:{me.id}")
    await write_audit_log(
        db,
        user_id=me.id,
        action="user.soft_deleted",
        resource_type="user",
        resource_id=str(me.id),
        extra={"email": user.email},
    )
    await db.commit()
    return {"detail": "Hesabınız silinmiştir. Yasal saklama süresi dolan veriler 30 gün içinde kalıcı olarak temizlenecektir."}


@app.get("/api/admin/events", response_model=list[EventOut], dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def list_events(request: Request, me: CurrentUser = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    from .organization_access_api import OrganizationMember, member_allows
    from .organization_access_api import get_organization_for_access, organization_id_from_request

    normalized_email = str(me.email).strip().lower()
    selected_organization_id = organization_id_from_request(request)
    if selected_organization_id is not None:
        selected_org = await get_organization_for_access(db, me, "events:manage", selected_organization_id)
        res = await db.execute(select(Event).where(Event.admin_id == selected_org.user_id).order_by(Event.created_at.desc()))
        return [_event_to_out(e) for e in res.scalars().all()]
    org_memberships = await db.execute(
        select(Organization.user_id, OrganizationMember)
        .join(OrganizationMember, OrganizationMember.organization_id == Organization.id)
        .where(
            OrganizationMember.status == "active",
            or_(
                OrganizationMember.user_id == me.id,
                func.lower(func.trim(OrganizationMember.email)) == normalized_email,
            ),
        )
    )
    managed_owner_ids = [
        owner_id for owner_id, membership in org_memberships.all()
        if member_allows(membership, "events:manage")
    ]
    res = await db.execute(
        select(Event)
        .outerjoin(EventTeamMember, EventTeamMember.event_id == Event.id)
        .where(
            or_(
                Event.admin_id == me.id,
                Event.admin_id.in_(managed_owner_ids),
                and_(
                    EventTeamMember.status == "active",
                    or_(
                        EventTeamMember.user_id == me.id,
                        func.lower(func.trim(EventTeamMember.email)) == normalized_email,
                    ),
                ),
            )
        )
        .distinct()
        .order_by(Event.created_at.desc())
    )
    items = res.scalars().all()
    return [_event_to_out(e) for e in items]






async def _sync_event_venue_reservation(
    db: AsyncSession,
    ev: Event,
    me: CurrentUser,
    organization_id: Optional[int],
    venue_id: Optional[int],
    auto_reserve: Optional[bool],
    start_at: Optional[datetime],
    end_at: Optional[datetime],
) -> None:
    if auto_reserve is False:
        return
    config = dict(ev.config or {})
    if venue_id is None:
        if "organization_venue_id" in config:
            config.pop("organization_venue_id", None)
            config.pop("venue_reservation_start_at", None)
            config.pop("venue_reservation_end_at", None)
            reservation_id = config.pop("venue_reservation_id", None)
            if reservation_id:
                from .venue_reservations_api import VenueReservation
                reservation = await db.get(VenueReservation, int(reservation_id))
                if reservation:
                    reservation.status = "cancelled"
                    reservation.updated_by = me.id
            ev.config = config
        return
    from .organization_access_api import get_organization_for_access
    from .venue_reservations_api import ReservationIn, VenueReservation, _ensure_no_conflict, _venue_for_organization

    organization = await get_organization_for_access(db, me, "reservations:write", organization_id)
    if organization.user_id != ev.admin_id:
        raise HTTPException(status_code=403, detail="Venue must belong to the event organization")
    await _venue_for_organization(db, organization.id, venue_id)
    start, end = _event_reservation_window(ev, start_at, end_at)
    payload = ReservationIn(
        venue_id=venue_id,
        title=ev.name,
        description=f"Etkinlik rezervasyonu: {ev.name}",
        start_at=start,
        end_at=end,
    )
    existing_id = config.get("venue_reservation_id")
    reservation: Optional[VenueReservation] = None
    if existing_id:
        reservation = await db.get(VenueReservation, int(existing_id))
        if reservation and reservation.organization_id != organization.id:
            reservation = None
    await _ensure_no_conflict(db, organization.id, payload, exclude_id=reservation.id if reservation else None)
    if reservation:
        reservation.venue_id = venue_id
        reservation.title = ev.name
        reservation.description = payload.description
        reservation.start_at = start
        reservation.end_at = end
        reservation.status = "confirmed"
        reservation.updated_by = me.id
    else:
        reservation = VenueReservation(
            organization_id=organization.id,
            venue_id=venue_id,
            title=ev.name,
            description=payload.description,
            start_at=start,
            end_at=end,
            created_by=me.id,
        )
        db.add(reservation)
        await db.flush()
    config["organization_venue_id"] = venue_id
    config["venue_reservation_id"] = reservation.id
    config["venue_reservation_start_at"] = start.isoformat()
    config["venue_reservation_end_at"] = end.isoformat()
    ev.config = config


@app.post("/api/admin/events", response_model=EventOut, status_code=201, dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def create_event(
    payload: EventCreateIn,
    background_tasks: BackgroundTasks,
    request: Request,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from .organization_access_api import get_organization_for_access, organization_id_from_request

    selected_organization_id = organization_id_from_request(request)
    organization = await get_organization_for_access(db, me, "events:manage", selected_organization_id)
    next_config = dict(payload.config or {})
    if "registration_fields" in next_config:
        next_config["registration_fields"] = _validate_registration_fields_for_write(next_config.get("registration_fields"))
    next_config["visibility"] = _normalize_event_visibility(next_config.get("visibility"))
    # New events should require KVKK consent by default.
    next_config.setdefault("kvkk_consent_required", True)
    public_id = await _generate_event_public_id(db)
    ev = Event(
        public_id=public_id,
        admin_id=organization.user_id,
        name=payload.name,
        template_image_url=payload.template_image_url or "placeholder",
        config=next_config,
        event_type=normalize_event_type(payload.event_type),
        certificate_enabled=normalize_feature_bool(payload.certificate_enabled, default=FEATURE_DEFAULTS["certificate_enabled"]),
        checkin_enabled=normalize_feature_bool(payload.checkin_enabled, default=FEATURE_DEFAULTS["checkin_enabled"]),
        ticketing_enabled=normalize_feature_bool(payload.ticketing_enabled, default=FEATURE_DEFAULTS["ticketing_enabled"]),
        registration_enabled=normalize_feature_bool(payload.registration_enabled, default=FEATURE_DEFAULTS["registration_enabled"]),
        raffles_enabled=normalize_feature_bool(payload.raffles_enabled, default=FEATURE_DEFAULTS["raffles_enabled"]),
        gamification_enabled=normalize_feature_bool(payload.gamification_enabled, default=FEATURE_DEFAULTS["gamification_enabled"]),
        requires_approval=normalize_feature_bool(payload.requires_approval, default=FEATURE_DEFAULTS["requires_approval"]),
        quiz_enabled=normalize_feature_bool(payload.quiz_enabled, default=FEATURE_DEFAULTS["quiz_enabled"]),
        cpd_enabled=normalize_feature_bool(payload.cpd_enabled, default=FEATURE_DEFAULTS["cpd_enabled"]),
    )
    db.add(ev)
    await db.flush()
    await _sync_event_venue_reservation(
        db,
        ev,
        me,
        organization.id,
        payload.organization_venue_id,
        payload.auto_reserve_venue,
        payload.venue_reservation_start_at,
        payload.venue_reservation_end_at,
    )
    await db.commit()
    await db.refresh(ev)
    if _get_event_visibility(ev) == "public":
        from .community_notifications import send_public_event_announcement_to_followers

        background_tasks.add_task(send_public_event_announcement_to_followers, ev.id)
    return _event_to_out(ev)


def _event_to_out(ev: Event) -> EventOut:
    config = ev.config if isinstance(ev.config, dict) else {}
    return EventOut(
        id=ev.id,
        public_id=ev.public_id,
        name=ev.name,
        template_image_url=ev.template_image_url or "placeholder",
        config=config,
        event_date=ev.event_date.isoformat() if ev.event_date else None,
        event_description=sanitize_event_description_html(ev.event_description),
        event_location=ev.event_location,
        min_sessions_required=int(ev.min_sessions_required or 1),
        event_banner_url=ev.event_banner_url,
        auto_email_on_cert=bool(ev.auto_email_on_cert),
        cert_email_template_id=ev.cert_email_template_id,
        registration_closed=_is_event_registration_closed(ev),
        visibility=_get_event_visibility(ev),
        require_email_verification=_get_event_email_verification_required(ev),
        registration_quota=_get_event_registration_quota(ev),
        registration_quota_enabled=_is_event_registration_quota_enabled(ev),
        event_type=normalize_event_type(getattr(ev, "event_type", None)),
        certificate_enabled=normalize_feature_bool(getattr(ev, "certificate_enabled", None), default=FEATURE_DEFAULTS["certificate_enabled"]),
        checkin_enabled=normalize_feature_bool(getattr(ev, "checkin_enabled", None), default=FEATURE_DEFAULTS["checkin_enabled"]),
        ticketing_enabled=normalize_feature_bool(getattr(ev, "ticketing_enabled", None), default=FEATURE_DEFAULTS["ticketing_enabled"]),
        registration_enabled=normalize_feature_bool(getattr(ev, "registration_enabled", None), default=FEATURE_DEFAULTS["registration_enabled"]),
        raffles_enabled=normalize_feature_bool(getattr(ev, "raffles_enabled", None), default=FEATURE_DEFAULTS["raffles_enabled"]),
        gamification_enabled=normalize_feature_bool(getattr(ev, "gamification_enabled", None), default=FEATURE_DEFAULTS["gamification_enabled"]),
        requires_approval=normalize_feature_bool(getattr(ev, "requires_approval", None), default=FEATURE_DEFAULTS["requires_approval"]),
        organization_venue_id=config.get("organization_venue_id"),
        venue_reservation_id=config.get("venue_reservation_id"),
        venue_reservation_start_at=config.get("venue_reservation_start_at"),
        venue_reservation_end_at=config.get("venue_reservation_end_at"),
    )




_AI_SYSTEM_PROMPT = (
    "You are HeptaCert's admin AI assistant. HeptaCert is an event management and certificate issuance platform. "
    "Help event organizers in Turkish or English (match the user's language). "
    "You can suggest event descriptions, email copy, survey questions, registration field configs, "
    "certificate/check-in settings, session agendas, and LMS course outlines. "
    "Never claim changes were saved — you only suggest. Keep answers concise and practical. "
    "When useful, structure your response with clear sections for: event_update suggestions, "
    "registration_fields, and sessions."
)






async def _call_openai_ai_assistant(payload: AIAssistantIn, event: Optional[Event]) -> AIAssistantOut:
    input_text = json.dumps(
        {
            "language": payload.language,
            "event_context": _build_ai_event_context(event),
            "history": [m.model_dump() for m in payload.history[-6:]],
            "user_message": payload.message,
        },
        ensure_ascii=False,
    )
    body = {
        "model": settings.openai_model,
        "instructions": _AI_SYSTEM_PROMPT,
        "input": input_text,
        "max_output_tokens": 900,
    }
    try:
        async with httpx.AsyncClient(timeout=20.0) as client:
            resp = await client.post(
                "https://api.openai.com/v1/responses",
                headers={
                    "Authorization": f"Bearer {settings.openai_api_key}",
                    "Content-Type": "application/json",
                },
                json=body,
            )
    except httpx.HTTPError as exc:
        logger.warning("OpenAI assistant request failed: %s", exc)
        return _build_local_ai_assistant_response(payload, event)
    if resp.status_code >= 400:
        logger.warning("OpenAI assistant call failed status=%s body=%s", resp.status_code, resp.text[:500])
        return _build_local_ai_assistant_response(payload, event)
    try:
        answer = _extract_openai_response_text(resp.json())
    except ValueError:
        logger.warning("OpenAI assistant returned invalid JSON")
        return _build_local_ai_assistant_response(payload, event)
    if not answer:
        return _build_local_ai_assistant_response(payload, event)
    return AIAssistantOut(answer=answer, provider="openai", suggestions={})


async def _call_claude_ai_assistant(payload: AIAssistantIn, event: Optional[Event]) -> AIAssistantOut:
    event_context = _build_ai_event_context(event)
    messages: List[Dict[str, Any]] = []
    for m in payload.history[-6:]:
        role = "user" if m.role in ("user", "human") else "assistant"
        messages.append({"role": role, "content": m.message})

    user_content = payload.message
    if event_context:
        user_content = (
            f"[Event context: {json.dumps(event_context, ensure_ascii=False)}]\n\n"
            f"Language: {payload.language}\n\n"
            f"{payload.message}"
        )
    elif payload.language:
        user_content = f"Language: {payload.language}\n\n{payload.message}"

    messages.append({"role": "user", "content": user_content})

    body = {
        "model": settings.anthropic_model,
        "system": _AI_SYSTEM_PROMPT,
        "messages": messages,
        "max_tokens": 1024,
    }
    try:
        async with httpx.AsyncClient(timeout=25.0) as client:
            resp = await client.post(
                "https://api.anthropic.com/v1/messages",
                headers={
                    "x-api-key": settings.anthropic_api_key,
                    "anthropic-version": "2023-06-01",
                    "Content-Type": "application/json",
                },
                json=body,
            )
    except httpx.HTTPError as exc:
        logger.warning("Anthropic assistant request failed: %s", exc)
        return _build_local_ai_assistant_response(payload, event)
    if resp.status_code >= 400:
        logger.warning("Anthropic assistant call failed status=%s body=%s", resp.status_code, resp.text[:500])
        return _build_local_ai_assistant_response(payload, event)
    try:
        data = resp.json()
        parts = [
            block.get("text", "")
            for block in (data.get("content") or [])
            if isinstance(block, dict) and block.get("type") == "text"
        ]
        answer = "\n".join(parts).strip()
    except (ValueError, AttributeError):
        logger.warning("Anthropic assistant returned invalid JSON")
        return _build_local_ai_assistant_response(payload, event)
    if not answer:
        return _build_local_ai_assistant_response(payload, event)
    return AIAssistantOut(answer=answer, provider="claude", suggestions={})


async def _get_optional_ai_event_context(event_id: Optional[int], me: CurrentUser, db: AsyncSession) -> Optional[Event]:
    if not event_id:
        return None
    res = await db.execute(select(Event).where(Event.id == event_id))
    event = res.scalar_one_or_none()
    if not event:
        return None
    if me.role == Role.superadmin or event.admin_id == me.id:
        return event
    normalized_email = str(me.email).strip().lower()
    member_res = await db.execute(
        select(EventTeamMember).where(
            EventTeamMember.event_id == event.id,
            EventTeamMember.status == "active",
            or_(
                EventTeamMember.user_id == me.id,
                func.lower(func.trim(EventTeamMember.email)) == normalized_email,
            ),
        )
    )
    return event if member_res.scalar_one_or_none() else None


@app.post(
    "/api/admin/ai/event-assistant",
    response_model=AIAssistantOut,
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
@limiter.limit("20/minute")
async def ai_event_assistant(
    request: Request,
    payload: AIAssistantIn,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    event = await _get_optional_ai_event_context(payload.event_id, me, db)
    # Provider priority: Claude → OpenAI → local fallback
    if settings.anthropic_api_key.strip():
        return await _call_claude_ai_assistant(payload, event)
    if settings.openai_api_key.strip():
        return await _call_openai_ai_assistant(payload, event)
    return _build_local_ai_assistant_response(payload, event)


# ── MCP Agent endpoints ────────────────────────────────────────────────────────




@app.get(
    "/api/admin/mcp/me",
    response_model=AgentMeOut,
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def mcp_me(
    Authorization: Optional[str] = FastAPIHeader(default=None),
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Return current user identity and API key scopes. Used by MCP server to validate scope."""
    scopes: List[str] = []
    key_prefix: Optional[str] = None
    token = (Authorization or "").split(" ", 1)[-1].strip()
    if token.startswith("hc_live_"):
        key_hash = _hash_api_key(token)
        res = await db.execute(select(ApiKey).where(ApiKey.key_hash == key_hash, ApiKey.is_active.is_(True)))
        api_key = res.scalar_one_or_none()
        if api_key:
            scopes = list(api_key.scopes or [])
            key_prefix = api_key.key_prefix
    return AgentMeOut(user_id=me.id, email=str(me.email), api_key_prefix=key_prefix, scopes=scopes)




@app.post(
    "/api/admin/mcp/agent-log",
    status_code=201,
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def mcp_agent_log(
    payload: AgentLogIn,
    request: Request,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Record an agent action in the audit trail."""
    log = AgentActionLog(
        user_id=me.id,
        api_key_prefix=payload.api_key_prefix,
        tool_name=payload.tool_name,
        event_id=payload.event_id,
        payload=payload.payload,
        result_summary=payload.result_summary,
        ip_address=_client_ip_for_rate_limit(request),
    )
    db.add(log)
    await db.commit()
    return {"ok": True}


@app.get(
    "/api/admin/mcp/agent-logs",
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def list_agent_logs(
    event_id: Optional[int] = None,
    tool_name: Optional[str] = None,
    limit: int = 50,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List recent agent (MCP) actions for this account."""
    q = select(AgentActionLog).where(AgentActionLog.user_id == me.id)
    if event_id:
        q = q.where(AgentActionLog.event_id == event_id)
    if tool_name:
        q = q.where(AgentActionLog.tool_name == tool_name)
    q = q.order_by(AgentActionLog.created_at.desc()).limit(min(limit, 200))
    rows = (await db.execute(q)).scalars().all()
    return [
        {
            "id": r.id,
            "tool_name": r.tool_name,
            "event_id": r.event_id,
            "api_key_prefix": r.api_key_prefix,
            "payload": r.payload,
            "result_summary": r.result_summary,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        }
        for r in rows
    ]


# ── Attendee CRUD (for agent use) ──────────────────────────────────────────────




@app.patch(
    "/api/admin/events/{event_id}/attendees/{attendee_id}",
    dependencies=[Depends(require_role(Role.admin, Role.superadmin)), Depends(require_paid_plan)],
)
async def update_attendee(
    event_id: int,
    attendee_id: int,
    payload: AttendeeUpdateIn,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _get_event_for_admin(event_id, me, db, "attendees:write")
    res = await db.execute(select(Attendee).where(Attendee.id == attendee_id, Attendee.event_id == event_id))
    attendee = res.scalar_one_or_none()
    if not attendee:
        raise HTTPException(status_code=404, detail="Attendee not found")
    if payload.first_name is not None or payload.last_name is not None:
        first = payload.first_name or (attendee.name or "").split(" ", 1)[0]
        last = payload.last_name or (" ".join((attendee.name or "").split(" ")[1:]) or "")
        attendee.name = f"{first} {last}".strip()
    if payload.email is not None:
        attendee.email = str(payload.email).strip().lower()
    await db.commit()
    await db.refresh(attendee)
    return {"id": attendee.id, "name": attendee.name, "email": attendee.email}



# ── Certificate revoke (for agent use) ─────────────────────────────────────────


@app.post(
    "/api/admin/certificates/{cert_id}/revoke",
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def revoke_certificate(
    cert_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    res = await db.execute(select(Certificate).where(Certificate.id == cert_id))
    cert = res.scalar_one_or_none()
    if not cert:
        raise HTTPException(status_code=404, detail="Certificate not found")
    ev = await _get_event_for_admin(cert.event_id, me, db, "certificates:write")
    if cert.status == CertStatus.revoked:
        raise HTTPException(status_code=409, detail="Certificate is already revoked")
    cert.status = CertStatus.revoked
    await db.commit()
    return {"id": cert.id, "status": "revoked"}




async def _send_event_team_invite_email(
    *,
    event: Event,
    member: EventTeamMember,
    invited_by: CurrentUser,
) -> None:
    token = make_email_token(
        {
            "action": "event_team_invite",
            "event_id": event.id,
            "email": member.email.strip().lower(),
            "role": member.role,
        }
    )
    invite_url = _event_team_invite_url(token)
    subject = f"HeptaCert etkinlik daveti: {event.name}"
    role_label = member.role or "team"
    html_body = f"""
    <div style="font-family:Inter,Arial,sans-serif;line-height:1.6;color:#1f2937">
      <h2 style="margin:0 0 12px">Etkinlik ekibine davet edildiniz</h2>
      <p><strong>{escape(event.name)}</strong> etkinliği için <strong>{escape(role_label)}</strong> rolüyle ekip erişimi tanımlandı.</p>
      <p>Davet eden: {escape(invited_by.email)}</p>
      <p style="margin:24px 0">
        <a href="{escape(invite_url, quote=True)}" style="background:#7c3aed;color:#fff;padding:12px 18px;border-radius:10px;text-decoration:none;font-weight:700">
          Daveti kabul et
        </a>
      </p>
      <p style="font-size:13px;color:#6b7280">Buton çalışmazsa bu bağlantıyı tarayıcınıza yapıştırın:<br>{escape(invite_url)}</p>
    </div>
    """
    await send_email_async(
        to=member.email,
        subject=subject,
        html_body=html_body,
        sender_user_id=event.admin_id,
    )


@app.get("/api/admin/events/{event_id}", response_model=EventOut, dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def get_event(event_id: int, me: CurrentUser = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    ev = await _get_event_for_admin(event_id, me, db, "event:view")
    return _event_to_out(ev)

@app.get(
    "/api/admin/events/{event_id}/health",
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def get_event_health(
    event_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    ev = await _get_event_for_admin(event_id, me, db, "event:view")
    attendee_count = int(
        (await db.execute(select(func.count()).select_from(Attendee).where(Attendee.event_id == event_id))).scalar_one()
        or 0
    )
    session_count = int(
        (await db.execute(select(func.count()).select_from(EventSession).where(EventSession.event_id == event_id))).scalar_one()
        or 0
    )
    attendance_count = int(
        (
            await db.execute(
                select(func.count())
                .select_from(AttendaonceRecord)
                .join(EventSession, AttendaonceRecord.session_id == EventSession.id)
                .where(EventSession.event_id == event_id)
            )
        ).scalar_one()
        or 0
    )
    ticket_count = int(
        (await db.execute(select(func.count()).select_from(EventTicket).where(EventTicket.event_id == event_id))).scalar_one()
        or 0
    )
    used_ticket_count = int(
        (
            await db.execute(
                select(func.count())
                .select_from(EventTicket)
                .where(EventTicket.event_id == event_id, EventTicket.status == "used")
            )
        ).scalar_one()
        or 0
    )

    cert_rows = (
        await db.execute(
            select(Certificate.status, func.count(Certificate.id))
            .where(Certificate.event_id == event_id, Certificate.deleted_at.is_(None))
            .group_by(Certificate.status)
        )
    ).all()
    certificate_counts = {"active": 0, "expired": 0, "revoked": 0}
    for status, count in cert_rows:
        key = status.value if hasattr(status, "value") else str(status)
        certificate_counts[key] = int(count or 0)
    certificate_total = sum(certificate_counts.values())

    latest_cert_job = (
        await db.execute(
            select(BulkCertificateJob)
            .where(BulkCertificateJob.event_id == event_id)
            .order_by(BulkCertificateJob.created_at.desc())
            .limit(1)
        )
    ).scalar_one_or_none()
    latest_email_job = (
        await db.execute(
            select(BulkEmailJob)
            .where(BulkEmailJob.event_id == event_id)
            .order_by(BulkEmailJob.created_at.desc())
            .limit(1)
        )
    ).scalar_one_or_none()
    sheets = await _event_sheets_status_payload(db, ev)

    def job_payload(job: Any) -> Optional[Dict[str, Any]]:
        if not job:
            return None
        return {
            "id": job.id,
            "status": job.status,
            "created_at": job.created_at.isoformat() if job.created_at else None,
            "completed_at": job.completed_at.isoformat() if getattr(job, "completed_at", None) else None,
            "failed_count": int(getattr(job, "failed_count", 0) or 0),
            "error_message": getattr(job, "error_message", None),
        }

    latest_jobs = {
        "certificate": job_payload(latest_cert_job),
        "email": job_payload(latest_email_job),
    }
    failed_job = next((job for job in latest_jobs.values() if job and job["status"] == "failed"), None)
    running_job = next(
        (job for job in latest_jobs.values() if job and job["status"] in {"pending", "processing", "sending", "scheduled"}),
        None,
    )

    checks = [
        {
            "key": "registration",
            "label": "Kayit",
            "status": "ok" if is_public_registration_enabled(ev) else "idle",
            "detail": f"{attendee_count} katılımcı",
        },
        {
            "key": "attendance",
            "label": "Yoklama",
            "status": "idle"
            if not is_checkin_enabled(ev)
            else "warning"
            if session_count == 0 or attendance_count == 0
            else "ok",
            "detail": f"{attendance_count} kayıt / {session_count} oturum",
        },
        {
            "key": "certificates",
            "label": "Sertifika",
            "status": "idle"
            if not is_certificate_enabled(ev)
            else "warning"
            if certificate_total == 0
            else "ok",
            "detail": f"{certificate_counts['active']} aktif / {certificate_total} toplam",
        },
        {
            "key": "sheets",
            "label": "Google Sheets",
            "status": "idle"
            if not sheets.enabled
            else "ok"
            if sheets.spreadsheet_id and sheets.last_synced_at
            else "warning",
            "detail": sheets.google_email or sheets.sheet_name or "Bagli degil",
        },
        {
            "key": "jobs",
            "label": "Isler",
            "status": "error" if failed_job else "warning" if running_job else "ok",
            "detail": failed_job["error_message"] if failed_job and failed_job["error_message"] else (running_job["status"] if running_job else "Bekleyen is yok"),
        },
    ]

    return {
        "event_id": ev.id,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "overview": {
            "attendees": attendee_count,
            "sessions": session_count,
            "attendance_records": attendance_count,
            "tickets": ticket_count,
            "used_tickets": used_ticket_count,
            "certificates": certificate_total,
            "active_certificates": certificate_counts["active"],
            "expired_certificates": certificate_counts["expired"],
            "revoked_certificates": certificate_counts["revoked"],
        },
        "sheets": sheets.model_dump(mode="json"),
        "latest_jobs": latest_jobs,
        "checks": checks,
    }


@app.get(
    "/api/admin/events/{event_id}/team",
    response_model=list[EventTeamMemberOut],
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def list_event_team_members(
    event_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await require_event_owner_premium_for_teams(event_id, db, me)
    await _get_event_for_admin(event_id, me, db, "team:manage")
    res = await db.execute(
        select(EventTeamMember)
        .where(EventTeamMember.event_id == event_id)
        .order_by(EventTeamMember.created_at.asc(), EventTeamMember.id.asc())
    )
    return [_event_team_member_to_out(member) for member in res.scalars().all()]


@app.get(
    "/api/admin/events/{event_id}/access",
    response_model=EventAccessOut,
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def get_event_access(
    event_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    event = await _get_event_for_admin(event_id, me, db, "event:view")
    all_permissions = sorted(EVENT_TEAM_PERMISSION_LABELS.keys())
    if me.role == Role.superadmin or event.admin_id == me.id:
        return EventAccessOut(
            event_id=event_id,
            is_owner=True,
            role="owner",
            permissions=all_permissions,
            permission_labels=EVENT_TEAM_PERMISSION_LABELS,
        )
    membership = await _get_event_team_membership(event_id, me, db)
    if membership:
        permissions = _effective_event_team_permissions(membership)
        return EventAccessOut(
            event_id=event_id,
            is_owner=False,
            role=membership.role,
            permissions=permissions,
            permission_labels={key: EVENT_TEAM_PERMISSION_LABELS[key] for key in permissions if key in EVENT_TEAM_PERMISSION_LABELS},
        )
    from .organization_access_api import user_can_manage_owner_organization
    if await user_can_manage_owner_organization(db, me, event.admin_id, "events:manage"):
        permissions = sorted(EVENT_TEAM_ROLE_PERMISSIONS["manager"] - {"team:manage"})
        return EventAccessOut(
            event_id=event_id,
            is_owner=False,
            role="organization_event_manager",
            permissions=permissions,
            permission_labels={key: EVENT_TEAM_PERMISSION_LABELS[key] for key in permissions if key in EVENT_TEAM_PERMISSION_LABELS},
        )
    return EventAccessOut(
        event_id=event_id,
        is_owner=False,
        role="none",
        permissions=[],
        permission_labels={},
    )


@app.post(
    "/api/admin/events/{event_id}/team",
    response_model=EventTeamMemberOut,
    status_code=201,
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def add_event_team_member(
    event_id: int,
    payload: EventTeamMemberIn,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await require_event_owner_premium_for_teams(event_id, db, me)
    event = await _get_event_for_admin(event_id, me, db, "team:manage")
    normalized_email = str(payload.email).strip().lower()
    if normalized_email == str(me.email).strip().lower() and me.role != Role.superadmin:
        raise HTTPException(status_code=400, detail="Event owner is already the full-access owner")
    owner_res = await db.execute(select(User).where(User.id == event.admin_id))
    owner = owner_res.scalar_one_or_none()
    if owner and normalized_email == owner.email.strip().lower():
        raise HTTPException(status_code=400, detail="Event owner is already the full-access owner")

    user_res = await db.execute(select(User).where(func.lower(func.trim(User.email)) == normalized_email))
    user = user_res.scalar_one_or_none()
    if user and user.role not in (Role.admin, Role.superadmin):
        user.role = Role.admin

    existing_res = await db.execute(
        select(EventTeamMember).where(
            EventTeamMember.event_id == event_id,
            func.lower(func.trim(EventTeamMember.email)) == normalized_email,
        )
    )
    member = existing_res.scalar_one_or_none()
    if member:
        member.role = payload.role
        member.permissions = payload.permissions
        member.status = "pending"
        member.user_id = user.id if user else member.user_id
        member.invited_by = me.id
    else:
        member = EventTeamMember(
            event_id=event_id,
            user_id=user.id if user else None,
            email=normalized_email,
            role=payload.role,
            permissions=payload.permissions,
            status="pending",
            invited_by=me.id,
        )
    db.add(member)
    await db.flush()
    await write_audit_log(
        db,
        user_id=me.id,
        action="team.member.invited",
        resource_type="event_team_member",
        resource_id=str(member.id) if member.id else None,
        extra={"event_id": event_id, "member_email": normalized_email, "role": payload.role},
    )
    await db.commit()
    await db.refresh(member)
    await _send_event_team_invite_email(event=event, member=member, invited_by=me)
    return _event_team_member_to_out(member)


@app.post("/api/event-team/invitations/accept", response_model=EventTeamInviteAcceptOut)
async def accept_event_team_invite(
    payload: EventTeamInviteAcceptIn,
    db: AsyncSession = Depends(get_db),
):
    try:
        invite = verify_email_token(payload.token, max_age=60 * 60 * 24 * 14)
    except Exception:
        raise HTTPException(status_code=400, detail="Davet baglantisi gecersiz veya suresi dolmus.")
    if invite.get("action") != "event_team_invite":
        raise HTTPException(status_code=400, detail="Gecersiz davet baglantisi.")

    event_id = int(invite.get("event_id") or 0)
    normalized_email = str(invite.get("email") or "").strip().lower()
    if not event_id or not normalized_email:
        raise HTTPException(status_code=400, detail="Davet bilgisi eksik.")

    res = await db.execute(select(EventTeamMember).where(
        EventTeamMember.event_id == event_id,
        func.lower(func.trim(EventTeamMember.email)) == normalized_email,
    ))
    member = res.scalar_one_or_none()
    if not member:
        raise HTTPException(status_code=404, detail="Davet bulunamadi.")

    event_res = await db.execute(select(Event).where(Event.id == event_id))
    event = event_res.scalar_one_or_none()
    if not event:
        raise HTTPException(status_code=404, detail="Etkinlik bulunamadi.")

    user_res = await db.execute(select(User).where(func.lower(func.trim(User.email)) == normalized_email))
    user = user_res.scalar_one_or_none()
    if user:
        member.user_id = user.id
        if user.role not in (Role.admin, Role.superadmin):
            user.role = Role.admin
    member.status = "active"
    db.add(member)
    await write_audit_log(
        db,
        user_id=user.id if user else member.invited_by,
        action="team.member.accepted",
        resource_type="event_team_member",
        resource_id=str(member.id),
        extra={"event_id": event_id, "member_email": normalized_email, "role": member.role},
    )
    await db.commit()
    return EventTeamInviteAcceptOut(
        ok=True,
        event_id=event.id,
        event_name=event.name,
        email=member.email,
        status=member.status,
        message="Davet kabul edildi.",
    )


@app.patch(
    "/api/admin/events/{event_id}/team/{member_id}",
    response_model=EventTeamMemberOut,
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def update_event_team_member(
    event_id: int,
    member_id: int,
    payload: EventTeamMemberUpdateIn,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await require_event_owner_premium_for_teams(event_id, db, me)
    await _get_event_for_admin(event_id, me, db, "team:manage")
    res = await db.execute(select(EventTeamMember).where(EventTeamMember.id == member_id, EventTeamMember.event_id == event_id))
    member = res.scalar_one_or_none()
    if not member:
        raise HTTPException(status_code=404, detail="Team member not found")
    if payload.role is not None:
        member.role = payload.role
    if "permissions" in payload.model_fields_set:
        member.permissions = payload.permissions
    if payload.status is not None:
        member.status = payload.status
    db.add(member)
    await write_audit_log(
        db,
        user_id=me.id,
        action="team.member.updated",
        resource_type="event_team_member",
        resource_id=str(member.id),
        extra={"event_id": event_id, "member_email": member.email, "role": member.role, "status": member.status},
    )
    await db.commit()
    await db.refresh(member)
    return _event_team_member_to_out(member)


@app.delete(
    "/api/admin/events/{event_id}/team/{member_id}",
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def delete_event_team_member(
    event_id: int,
    member_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await require_event_owner_premium_for_teams(event_id, db, me)
    await _get_event_for_admin(event_id, me, db, "team:manage")
    res = await db.execute(select(EventTeamMember).where(EventTeamMember.id == member_id, EventTeamMember.event_id == event_id))
    member = res.scalar_one_or_none()
    if not member:
        raise HTTPException(status_code=404, detail="Team member not found")
    removed_email = member.email
    await db.delete(member)
    await write_audit_log(
        db,
        user_id=me.id,
        action="team.member.removed",
        resource_type="event_team_member",
        resource_id=str(member_id),
        extra={"event_id": event_id, "member_email": removed_email},
    )
    await db.commit()
    return {"ok": True}






@app.get(
    "/api/admin/events/{event_id}/team/activity",
    response_model=list[EventTeamActivityOut],
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def list_event_team_activity(
    event_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await require_event_owner_premium_for_teams(event_id, db, me)
    event = await _get_event_for_admin(event_id, me, db, "team:manage")
    team_res = await db.execute(select(EventTeamMember).where(EventTeamMember.event_id == event_id))
    team_members = team_res.scalars().all()
    user_ids = {event.admin_id, me.id}
    emails = set()
    for member in team_members:
        if member.user_id:
            user_ids.add(member.user_id)
        emails.add(member.email.strip().lower())

    users_res = await db.execute(select(User).where(or_(User.id.in_(user_ids), func.lower(func.trim(User.email)).in_(emails))))
    users = users_res.scalars().all()
    user_email_by_id = {user.id: user.email for user in users}
    for user in users:
        user_ids.add(user.id)

    logs_res = await db.execute(
        select(AuditLog)
        .where(AuditLog.user_id.in_(user_ids))
        .order_by(AuditLog.created_at.desc())
        .limit(300)
    )
    items: List[EventTeamActivityOut] = []
    for log in logs_res.scalars().all():
        extra = log.extra if isinstance(log.extra, dict) else {}
        if str(extra.get("event_id")) != str(event_id):
            continue
        title, fallback_detail = _event_activity_label(log.action)
        actor_email = user_email_by_id.get(log.user_id) if log.user_id else None
        actor_label = "Organizator" if log.user_id == event.admin_id else (actor_email or "Ekip uyesi")
        items.append(
            EventTeamActivityOut(
                id=log.id,
                actor_email=actor_email,
                actor_label=actor_label,
                action=log.action,
                action_label=title,
                detail=_event_activity_detail(log) or fallback_detail,
                created_at=log.created_at,
            )
        )
        if len(items) >= 80:
            break
    return items


async def _event_sheets_status_payload(db: AsyncSession, event: Event) -> EventSheetsStatusOut:
    integration = await _get_user_google_integration(db, event.admin_id)
    scopes = _normalize_google_scopes(integration.scopes if integration else [])
    missing_scopes = _google_sheets_missing_scopes(scopes) if integration else GOOGLE_SHEETS_SCOPES
    sheets_config = _get_event_google_sheets_config(event)
    return EventSheetsStatusOut(
        google_configured=bool(settings.google_oauth_client_id and settings.google_oauth_client_secret),
        google_connected=bool(integration and integration.refresh_token and not missing_scopes),
        google_email=integration.google_email if integration else None,
        spreadsheet_id=sheets_config.get("spreadsheet_id"),
        spreadsheet_url=sheets_config.get("spreadsheet_url"),
        sheet_name=sheets_config.get("sheet_name"),
        enabled=bool(sheets_config.get("enabled")),
        last_synced_at=sheets_config.get("last_synced_at"),
        missing_scopes=missing_scopes,
    )


@app.get(
    "/api/admin/events/{event_id}/sheets",
    response_model=EventSheetsStatusOut,
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def get_event_sheets_status(
    event_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    event = await _get_event_for_admin(event_id, me, db, "attendees:read")
    return await _event_sheets_status_payload(db, event)


@app.post(
    "/api/admin/events/{event_id}/sheets/connect",
    response_model=EventSheetsStatusOut,
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def connect_event_google_sheet(
    event_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    event = await _get_event_for_admin(event_id, me, db, "settings:write")
    await _write_event_attendees_to_google_sheet(db, event, create_if_missing=True)
    return await _event_sheets_status_payload(db, event)


@app.post(
    "/api/admin/events/{event_id}/sheets/sync",
    response_model=EventSheetsStatusOut,
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def sync_event_google_sheet(
    event_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    event = await _get_event_for_admin(event_id, me, db, "settings:write")
    sheets_config = _get_event_google_sheets_config(event)
    if not sheets_config.get("enabled") or not sheets_config.get("spreadsheet_id"):
        raise HTTPException(status_code=409, detail="No Google Sheet is connected for this event.")
    await _write_event_attendees_to_google_sheet(db, event)
    return await _event_sheets_status_payload(db, event)


@app.delete(
    "/api/admin/events/{event_id}/sheets",
    response_model=EventSheetsStatusOut,
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def disconnect_event_google_sheet(
    event_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    event = await _get_event_for_admin(event_id, me, db, "settings:write")
    _set_event_google_sheets_config(event, None)
    db.add(event)
    await db.commit()
    await db.refresh(event)
    return await _event_sheets_status_payload(db, event)


async def _event_ms365_excel_status_payload(db: AsyncSession, event: Event) -> EventMicrosoftExcelStatusOut:
    integration = await _get_user_ms365_integration(db, event.admin_id)
    scopes = _normalize_ms365_scopes(integration.scopes if integration else [])
    missing_scopes = _ms365_excel_missing_scopes(scopes) if integration else MS365_EXCEL_REQUIRED_SCOPES
    excel_config = _get_event_ms365_excel_config(event)
    return EventMicrosoftExcelStatusOut(
        ms365_configured=bool(settings.ms365_oauth_client_id and settings.ms365_oauth_client_secret),
        ms365_connected=bool(integration and integration.refresh_token and not missing_scopes),
        microsoft_email=integration.microsoft_email if integration else None,
        workbook_id=excel_config.get("workbook_id"),
        workbook_url=excel_config.get("workbook_url"),
        workbook_name=excel_config.get("workbook_name"),
        sheet_name=excel_config.get("sheet_name"),
        enabled=bool(excel_config.get("enabled")),
        last_synced_at=excel_config.get("last_synced_at"),
        missing_scopes=missing_scopes,
    )




@app.patch("/api/admin/events/{event_id}", response_model=EventOut, dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def rename_event(
    event_id: int,
    payload: EventRenameIn,
    background_tasks: BackgroundTasks,
    request: Request,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from .organization_access_api import organization_id_from_request

    ev = await _get_event_for_admin(event_id, me, db, "settings:write")
    was_public = _get_event_visibility(ev) == "public"
    existing_registration_fields = _get_event_registration_fields(ev)
    ev.name = payload.name
    if payload.event_date is not None:
        from datetime import date as _date
        ev.event_date = _date.fromisoformat(payload.event_date) if payload.event_date else None
    if payload.event_description is not None:
        ev.event_description = sanitize_event_description_html(payload.event_description)
    if payload.event_location is not None:
        ev.event_location = payload.event_location
    if payload.min_sessions_required is not None:
        ev.min_sessions_required = payload.min_sessions_required
    if payload.event_banner_url is not None:
        ev.event_banner_url = payload.event_banner_url if payload.event_banner_url else None
    next_config = dict(ev.config or {})
    config_dirty = False
    if "registration_fields" in payload.model_fields_set:
        next_config["registration_fields"] = _validate_registration_fields_for_write(
            payload.registration_fields,
            existing_fields=existing_registration_fields,
        )
        config_dirty = True
    if "visibility" in payload.model_fields_set:
        next_config["visibility"] = _normalize_event_visibility(payload.visibility)
        config_dirty = True
    if "require_email_verification" in payload.model_fields_set:
        next_config["require_email_verification"] = bool(payload.require_email_verification)
        config_dirty = True
    if "registration_closed" in payload.model_fields_set:
        next_config["registration_closed"] = bool(payload.registration_closed)
        config_dirty = True
    if "registration_quota" in payload.model_fields_set:
        if payload.registration_quota is None:
            next_config.pop("registration_quota", None)
        else:
            next_config["registration_quota"] = int(payload.registration_quota)
        config_dirty = True
    if "registration_quota_enabled" in payload.model_fields_set:
        next_config["registration_quota_enabled"] = bool(payload.registration_quota_enabled)
        config_dirty = True
    if "event_type" in payload.model_fields_set:
        ev.event_type = normalize_event_type(payload.event_type)
    if "certificate_enabled" in payload.model_fields_set:
        ev.certificate_enabled = normalize_feature_bool(payload.certificate_enabled, default=FEATURE_DEFAULTS["certificate_enabled"])
    if "checkin_enabled" in payload.model_fields_set:
        ev.checkin_enabled = normalize_feature_bool(payload.checkin_enabled, default=FEATURE_DEFAULTS["checkin_enabled"])
    if "ticketing_enabled" in payload.model_fields_set:
        ev.ticketing_enabled = normalize_feature_bool(payload.ticketing_enabled, default=FEATURE_DEFAULTS["ticketing_enabled"])
    if "registration_enabled" in payload.model_fields_set:
        ev.registration_enabled = normalize_feature_bool(payload.registration_enabled, default=FEATURE_DEFAULTS["registration_enabled"])
    if "raffles_enabled" in payload.model_fields_set:
        ev.raffles_enabled = normalize_feature_bool(payload.raffles_enabled, default=FEATURE_DEFAULTS["raffles_enabled"])
    if "gamification_enabled" in payload.model_fields_set:
        ev.gamification_enabled = normalize_feature_bool(payload.gamification_enabled, default=FEATURE_DEFAULTS["gamification_enabled"])
    if "requires_approval" in payload.model_fields_set:
        ev.requires_approval = normalize_feature_bool(payload.requires_approval, default=FEATURE_DEFAULTS["requires_approval"])
    if "quiz_enabled" in payload.model_fields_set:
        ev.quiz_enabled = normalize_feature_bool(payload.quiz_enabled, default=FEATURE_DEFAULTS["quiz_enabled"])
    if "cpd_enabled" in payload.model_fields_set:
        ev.cpd_enabled = normalize_feature_bool(payload.cpd_enabled, default=FEATURE_DEFAULTS["cpd_enabled"])
    if "organizer_privacy_notice_enabled" in payload.model_fields_set:
        next_config["organizer_privacy_notice_enabled"] = bool(payload.organizer_privacy_notice_enabled)
        config_dirty = True
    if "organizer_privacy_notice_text" in payload.model_fields_set:
        if payload.organizer_privacy_notice_text is None:
            next_config.pop("organizer_privacy_notice_text", None)
        else:
            next_config["organizer_privacy_notice_text"] = sanitize_event_description_html(payload.organizer_privacy_notice_text)
        config_dirty = True
    if "show_cross_border_transfer_notice" in payload.model_fields_set:
        next_config["show_cross_border_transfer_notice"] = bool(payload.show_cross_border_transfer_notice)
        config_dirty = True
    if "require_cross_border_transfer_consent" in payload.model_fields_set:
        next_config["require_cross_border_transfer_consent"] = bool(payload.require_cross_border_transfer_consent)
        config_dirty = True
    if "data_controller_name" in payload.model_fields_set:
        if payload.data_controller_name is None:
            next_config.pop("data_controller_name", None)
        else:
            next_config["data_controller_name"] = payload.data_controller_name.strip()
        config_dirty = True
    if "data_controller_contact_email" in payload.model_fields_set:
        if payload.data_controller_contact_email is None:
            next_config.pop("data_controller_contact_email", None)
        else:
            next_config["data_controller_contact_email"] = str(payload.data_controller_contact_email).strip()
        config_dirty = True
    if "data_retention_note" in payload.model_fields_set:
        if payload.data_retention_note is None:
            next_config.pop("data_retention_note", None)
        else:
            next_config["data_retention_note"] = payload.data_retention_note.strip()
        config_dirty = True
    if payload.registration_quota is not None and bool(next_config.get("registration_quota_enabled")):
        attendee_count_res = await db.execute(
            select(func.count()).where(Attendee.event_id == ev.id)
        )
        attendee_count = int(attendee_count_res.scalar_one() or 0)
        if attendee_count >= int(payload.registration_quota):
            next_config["registration_closed"] = True
            config_dirty = True
    if config_dirty:
        ev.config = _stamp_event_legal_versions(next_config)
        if "registration_fields" in payload.model_fields_set:
            await _sync_registration_option_capacities(db, ev)
    venue_fields_present = any(
        field in payload.model_fields_set
        for field in (
            "organization_venue_id",
            "auto_reserve_venue",
            "venue_reservation_start_at",
            "venue_reservation_end_at",
        )
    )
    current_config = dict(ev.config or {})
    if venue_fields_present or current_config.get("organization_venue_id"):
        venue_id = (
            payload.organization_venue_id
            if "organization_venue_id" in payload.model_fields_set
            else current_config.get("organization_venue_id")
        )
        start_at = (
            payload.venue_reservation_start_at
            if "venue_reservation_start_at" in payload.model_fields_set
            else _parse_event_reservation_datetime(current_config.get("venue_reservation_start_at"))
        )
        end_at = (
            payload.venue_reservation_end_at
            if "venue_reservation_end_at" in payload.model_fields_set
            else _parse_event_reservation_datetime(current_config.get("venue_reservation_end_at"))
        )
        await _sync_event_venue_reservation(
            db,
            ev,
            me,
            organization_id_from_request(request),
            int(venue_id) if venue_id is not None else None,
            payload.auto_reserve_venue,
            start_at,
            end_at,
        )
    if "auto_email_on_cert" in payload.model_fields_set:
        ev.auto_email_on_cert = bool(payload.auto_email_on_cert)
    if "cert_email_template_id" in payload.model_fields_set:
        template_id = payload.cert_email_template_id
        if template_id is None:
            ev.cert_email_template_id = None
        else:
            template_res = await db.execute(
                select(EmailTemplate).where(
                    EmailTemplate.id == template_id,
                    or_(EmailTemplate.event_id == event_id, EmailTemplate.template_type == "system"),
                )
            )
            template = template_res.scalar_one_or_none()
            if not template:
                raise HTTPException(status_code=404, detail="Email template not found")
            ev.cert_email_template_id = template.id
    await db.commit()
    if not was_public and _get_event_visibility(ev) == "public":
        from .community_notifications import send_public_event_announcement_to_followers

        background_tasks.add_task(send_public_event_announcement_to_followers, ev.id)
    return _event_to_out(ev)


@app.delete("/api/admin/events/{event_id}", dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def delete_event(
    event_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    # Superadmin can delete any event; admin can only delete their own
    if me.role == Role.superadmin:
        res = await db.execute(select(Event).where(Event.id == event_id))
    else:
        res = await db.execute(select(Event).where(Event.id == event_id, Event.admin_id == me.id))
    ev = res.scalar_one_or_none()
    if not ev:
        raise HTTPException(status_code=404, detail="Event not found")
    await db.delete(ev)
    await db.commit()
    return {"ok": True}


@app.post("/api/admin/events/{event_id}/template-upload", dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def upload_template(
    event_id: int,
    file: UploadFile = File(...),
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    ev = await _get_event_for_admin(event_id, me, db, "certificates:write")
    _ensure_certificate_feature_enabled(ev)

    data, ext = await _read_safe_raster_upload(file)
    safe_name = f"templates/event_{event_id}_{secrets.token_hex(8)}{ext}"
    dest = Path(settings.local_storage_dir) / safe_name

    dest.parent.mkdir(parents=True, exist_ok=True)

    # Remove old template file from disk before overwriting
    if ev.template_image_url and ev.template_image_url not in ("placeholder", ""):
        try:
            old_path = local_path_from_url(ev.template_image_url)
            if old_path.exists() and "templates/" in str(old_path):
                old_path.unlink(missing_ok=True)
        except Exception as exc:
            logger.warning("Old template cleanup failed for event %s: %s", event_id, exc)

    dest.write_bytes(data)

    # Get image dimensions for the editor
    img_w, img_h = 1240, 877
    if PILImage is not None:
        try:
            img_obj = PILImage.open(io.BytesIO(data))
            img_w, img_h = img_obj.size
        except Exception:
            pass

    # Save template snapshot before overwriting
    snap = EventTemplateSnapshot(
        event_id=event_id,
        template_image_url=ev.template_image_url,
        config=ev.config,
        created_by=me.id,
    )
    db.add(snap)

    ev.template_image_url = safe_name
    await db.commit()
    pub_url = f"{settings.public_base_url}/api/files/{safe_name}"
    return {"template_image_url": safe_name, "url": pub_url, "width": img_w, "height": img_h}


@app.post("/api/admin/events/{event_id}/banner-upload", dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def upload_event_banner(
    event_id: int,
    file: UploadFile = File(...),
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    ev = await _get_event_for_admin(event_id, me, db, "settings:write")
    data, ext = await _read_safe_raster_upload(file)
    safe_name = f"banners/event_{event_id}/banner{ext}"
    dest = Path(settings.local_storage_dir) / safe_name
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(data)
    pub_url = f"{settings.public_base_url}/api/files/{safe_name}"
    ev.event_banner_url = pub_url
    await db.commit()
    return {"event_banner_url": pub_url}


@app.put("/api/admin/events/{event_id}/config", dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def save_event_config(
    event_id: int,
    payload: Dict[str, Any] = Body(...),
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    ev = await _get_event_for_admin(event_id, me, db, "certificates:write")

    if not isinstance(payload, dict):
        raise bad_request("config payload must be an object.")

    existing_registration_fields = _get_event_registration_fields(ev)
    next_config = dict(ev.config or {})
    if "registration_fields" in payload:
        next_config["registration_fields"] = _validate_registration_fields_for_write(
            payload.get("registration_fields"),
            existing_fields=existing_registration_fields,
        )

    for key, value in payload.items():
        if key == "registration_fields":
            continue
        if key == "organizer_privacy_notice_text":
            next_config[key] = sanitize_event_description_html(str(value)) if value is not None else None
        else:
            next_config[key] = value

    # Save config snapshot before merging the new payload
    snap = EventTemplateSnapshot(
        event_id=event_id,
        template_image_url=ev.template_image_url,
        config=ev.config,
        created_by=me.id,
    )
    db.add(snap)

    ev.config = next_config
    await db.commit()
    return {"event_id": ev.id, "config": ev.config}


#



@app.get("/api/verify/{uuid}", response_model=VerifyOut)
async def verify(uuid: str, request: Request, db: AsyncSession = Depends(get_db)):
    accept = (request.headers.get("accept") or "").lower()
    host = (request.headers.get("host") or "").split(":")[0].strip().lower()
    scheme = request.headers.get("x-forwarded-proto") or request.url.scheme or "https"

    org = None
    branding: Optional[Dict[str, Any]] = None
    verification_path: Optional[str] = None

    if host:
        org_res = await db.execute(
            select(Organization).where(Organization.custom_domain == host)
        )
        org = org_res.scalar_one_or_none()
        if org:
            branding = {
                "org_name": org.org_name,
                "brand_logo": org.brand_logo,
                "brand_color": org.brand_color,
            }
            org_settings = getattr(org, "settings", {}) or {}
            if isinstance(org_settings, dict):
                verification_path = org_settings.get("verification_path")

    if "text/html" in accept:
        return RedirectResponse(
            url=build_certificate_verify_url(
                uuid,
                host=host or None,
                scheme=scheme,
                verification_path=verification_path,
            ),
            status_code=307,
        )

    res = await db.execute(
        select(Certificate, Event)
        .join(Event, Certificate.event_id == Event.id)
        .where(Certificate.uuid == uuid, Certificate.deleted_at.is_(None))
    )
    row = res.first()
    if not row:
        raise HTTPException(status_code=404, detail="Certificate not found")

    cert, ev = row

    owner_org_res = await db.execute(
        select(Organization).where(Organization.user_id == ev.admin_id)
    )
    owner_org = owner_org_res.scalar_one_or_none()
    organizer_name = owner_org.org_name if owner_org else None
    organizer_logo = owner_org.brand_logo if owner_org else None
    organizer_public_id = owner_org.public_id if owner_org else None

    now = datetime.now(timezone.utc)
    if cert.hosting_ends_at and cert.hosting_ends_at < now and cert.status == CertStatus.active:
        cert.status = CertStatus.expired
        await db.commit()

    pdf_url = cert.pdf_url if cert.status == CertStatus.active else None

    # Ã¢â€â‚¬Ã¢â€â‚¬ Record verification hit Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬
    client_ip = request.client.host if request.client else None
    user_agent = request.headers.get("user-agent", "")
    referer = request.headers.get("referer", "")
    db.add(VerificationHit(
        cert_uuid=uuid,
        ip_address=client_ip,
        user_agent=user_agent[:512],
        referer=referer[:512],
    ))

    # Ã¢â€â‚¬Ã¢â€â‚¬ View count Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬
    count_res = await db.execute(
        select(func.count()).select_from(
            select(VerificationHit).where(VerificationHit.cert_uuid == uuid).subquery()
        )
    )
    view_count = int(count_res.scalar_one() or 0) + 1  # +1 for current hit

    # Ã¢â€â‚¬Ã¢â€â‚¬ Organization branding (match Host header to custom_domain) Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬
    branding: Optional[Dict[str, Any]] = None
    host = request.headers.get("host", "")
    if host:
        domain = host.split(":")[0]
        org_res = await db.execute(
            select(Organization).where(Organization.custom_domain == domain)
        )
        org = org_res.scalar_one_or_none()
        if org:
            branding = {
                "org_name": org.org_name,
                "brand_logo": org.brand_logo,
                "brand_color": org.brand_color,
            }

    # Ã¢â€â‚¬Ã¢â€â‚¬ LinkedIn URL Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬
    canonical_verify_url = build_certificate_verify_url(uuid)
    linkedin_share_url: Optional[str] = None
    linkedin_add_url: Optional[str] = None
    linkedin_url: Optional[str] = None
    if cert.status == CertStatus.active:
        from urllib.parse import urlencode
        issue_date = ensure_utc(getattr(cert, "issued_at", None)) or ensure_utc(getattr(cert, "created_at", None)) or now
        organization_name = organizer_name or (branding or {}).get("org_name") or "HeptaCert"
        share_text = f"{cert.student_name} - {ev.name} certificate"
        linkedin_share_url = build_linkedin_share_url(canonical_verify_url, share_text)
        params = {
            "startTask": "CERTIFICATION_NAME",
            "name": ev.name,
            "organizationName": organization_name,
            "issueYear": str(issue_date.year),
            "issueMonth": str(issue_date.month),
            "certId": cert.public_id or cert.uuid,
            "certUrl": canonical_verify_url,
        }
        linkedin_add_url = f"https://www.linkedin.com/profile/add?{urlencode(params)}"
        linkedin_url = linkedin_share_url

    await db.commit()

    # Ã¢â€â‚¬Ã¢â€â‚¬ Watermarked PNG URL (only if the file was generated) Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬
    png_url: Optional[str] = None
    rel_png_path = f"pngs/event_{ev.id}/{cert.uuid}.png"
    abs_png_path = Path(settings.local_storage_dir) / rel_png_path
    if abs_png_path.exists() and cert.status == CertStatus.active:
        png_url = build_public_pdf_url(rel_png_path)  # same /api/files/ base

    return VerifyOut(
        uuid=cert.uuid,
        public_id=cert.public_id,
        student_name=cert.student_name,
        event_name=ev.name,
        event_date=ev.event_date.isoformat() if ev.event_date else None,
        organizer_name=organizer_name,
        organizer_logo=organizer_logo,
        organizer_public_id=organizer_public_id,
        status=cert.status,
        pdf_url=pdf_url,
        png_url=png_url,
        issued_at=getattr(cert, "issued_at", None),
        hosting_ends_at=cert.hosting_ends_at,
        view_count=view_count,
        linkedin_url=linkedin_url,
        linkedin_share_url=linkedin_share_url,
        linkedin_add_url=linkedin_add_url,
        branding=branding,
    )




@app.post("/api/verify-watermark", response_model=WatermarkVerifyOut)
async def verify_watermark(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """
    Public endpoint: upload a HeptaCert certificate image (PNG recommended).
    If a valid steganographic watermark is detected, the corresponding
    certificate details are returned.
    """
    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Sadece görsel dosyaları kabul edilir (PNG veya JPEG).")
        raise HTTPException(status_code=400, detail="Sadece gÃƒÂ¶rsel dosyalarÃ„Â± kabul edilir (PNG, JPEG, Ã¢â‚¬Â¦)")

    img_bytes = await file.read()
    if len(img_bytes) > 30 * 1024 * 1024:  # 30 MB guard
        raise HTTPException(status_code=413, detail="Dosya 30 MB sinirini asiyor.")
        raise HTTPException(status_code=413, detail="Dosya 30 MB sÃ„Â±nÃ„Â±rÃ„Â±nÃ„Â± aÃ…Å¸Ã„Â±yor")

    # Extract watermark
    try:
        from .watermark import extract_watermark
        payload = await asyncio.to_thread(extract_watermark, img_bytes)
    except Exception:
        payload = None

    if not payload:
        return WatermarkVerifyOut(
            valid=False,
            message="Bu gorselde HeptaCert dijital damgasi bulunamadi. Mumkunse sertifikanin sistemden indirilen orijinal PNG dosyasini yukleyin.",
        )
        return WatermarkVerifyOut(
            valid=False,
            message="Bu gÃƒÂ¶rselde HeptaCert damgasÃ„Â± bulunamadÃ„Â±. Orijinal PNG dosyasÃ„Â±nÃ„Â± yÃƒÂ¼kleyin.",
        )

    payload = payload.strip()
    looks_like_public_id = re.fullmatch(r"EV\d+-\d{6,}", payload) is not None
    looks_like_uuid = re.fullmatch(
        r"[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}",
        payload,
    ) is not None
    if not looks_like_public_id and not looks_like_uuid:
        return WatermarkVerifyOut(
            valid=False,
            message="Dijital damga okunamadi veya gecersiz bir sertifika kodu iceriyor. Lutfen orijinal PNG dosyasini yukleyin.",
        )

    # Look up certificate by public_id
    res = await db.execute(
        select(Certificate, Event)
        .join(Event, Certificate.event_id == Event.id)
        .where(or_(Certificate.public_id == payload, Certificate.uuid == payload), Certificate.deleted_at.is_(None))
    )
    row = res.first()
    if not row:
        return WatermarkVerifyOut(
            valid=False,
            message=f"Dijital damga okundu ({payload}) ancak veritabaninda eslesen sertifika bulunamadi.",
            public_id=payload if looks_like_public_id else None,
            cert_uuid=payload if looks_like_uuid else None,
        )
        return WatermarkVerifyOut(
            valid=False,
            message=f"Damga gÃƒÂ¶rÃƒÂ¼ldÃƒÂ¼ ({payload}) ancak veritabanÃ„Â±nda eÃ…Å¸leÃ…Å¸en sertifika bulunamadÃ„Â±.",
            public_id=payload,
        )

    cert, ev = row
    issued_str = cert.issued_at.isoformat() if getattr(cert, "issued_at", None) else None

    return WatermarkVerifyOut(
        valid=cert.status == CertStatus.active,
        message=(
            "Bu görsel geçerli bir HeptaCert sertifika kaydına ait."
            if cert.status == CertStatus.active
            else f"Sertifika bulundu ancak durumu: {cert.status.value}."
        ),
        public_id=cert.public_id,
        cert_uuid=cert.uuid,
        student_name=cert.student_name,
        event_name=ev.name,
        issued_at=issued_str,
        status=cert.status.value,
    )

    return WatermarkVerifyOut(
        valid=cert.status == CertStatus.active,
        message=(
            f"Bu gÃƒÂ¶rsel geÃƒÂ§erli bir HeptaCert sertifika kaydÃ„Â±na ait."
            if cert.status == CertStatus.active
            else f"Sertifika bulundu ancak durumu: {cert.status.value}."
        ),
        public_id=cert.public_id,
        cert_uuid=cert.uuid,
        student_name=cert.student_name,
        event_name=ev.name,
        issued_at=issued_str,
        status=cert.status.value,
    )


@app.get("/api/files/{path:path}")
async def serve_file(path: str):
    # Sanitise: strip leading slashes, reject path traversal components
    path = path.lstrip("/")
    if ".." in path or path.startswith("/") or "\\" in path:
        raise HTTPException(status_code=400, detail="Invalid path")
    if path.split("/", 1)[0].lower() in {"registration_docs", "zips"}:
        raise HTTPException(status_code=404, detail="File not found")
    storage_root = Path(settings.local_storage_dir).resolve()
    abs_path = (storage_root / path).resolve()
    # Ensure the resolved path is still within the storage directory
    if not abs_path.is_relative_to(storage_root):
        raise HTTPException(status_code=403, detail="Access denied")
    if not abs_path.exists() or not abs_path.is_file():
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(abs_path)





@app.get("/api/branding")
async def get_branding(request: Request, db: AsyncSession = Depends(get_db)):
    """Public endpoint: returns organization branding for the current Host header (if any).

    This is intended for the frontend to fetch host-specific branding information.
    """
    host = (request.headers.get("host") or "").split(":")[0].strip().lower()
    if not host:
        return {"org_name": None, "brand_logo": None, "brand_color": None, "settings": {}}

    def host_local_file_url(value: Optional[str]) -> Optional[str]:
        if not value:
            return value
        marker = "/api/files/"
        if marker not in value:
            return value
        suffix = value.split(marker, 1)[1]
        return f"https://{host}{marker}{suffix}"

    try:
        org = await _organization_for_request_host(request, db)
        if not org:
            return {"org_name": None, "brand_logo": None, "brand_color": None, "settings": {}}
        return {
            "public_id": org.public_id,
            "org_name": org.org_name,
            "brand_logo": host_local_file_url(org.brand_logo),
            "brand_color": org.brand_color,
            "custom_domain": org.custom_domain,
            "settings": getattr(org, "settings", {}) or {},
        }
    except Exception:
        return {"org_name": None, "brand_logo": None, "brand_color": None, "settings": {}}






@app.get("/api/admin/organization/settings", dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def get_admin_organization_settings(request: Request, me: CurrentUser = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    from .organization_access_api import get_organization_for_access, organization_id_from_request
    org = await get_organization_for_access(db, me, "organization:view", organization_id_from_request(request))
    return _serialize_admin_organization(org)


@app.patch("/api/admin/organization/settings", dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def update_admin_organization_settings(payload: dict[str, Any], request: Request, me: CurrentUser = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    from .organization_access_api import get_organization_for_access, organization_id_from_request
    org = await get_organization_for_access(db, me, "organization:profile_write", organization_id_from_request(request))

    settings_data = dict(getattr(org, "settings", {}) or {})
    for key in (
        "verification_path",
        "certificate_footer",
        "hide_heptacert_home",
        "public_bio",
        "public_website_url",
        "public_linkedin_url",
        "public_github_url",
        "public_x_url",
        "public_instagram_url",
        "lms_portal_title",
        "lms_support_email",
        "lms_welcome_text",
    ):
        if key in payload:
            value = payload.get(key)
            if key == "hide_heptacert_home":
                settings_data[key] = bool(value)
            else:
                settings_data[key] = str(value).strip() if value is not None else ""

    org_name = str(payload.get("org_name") or "").strip()
    if org_name:
        org.org_name = org_name[:200]

    brand_color = str(payload.get("brand_color") or "").strip()
    if not brand_color:
        brand_color = org.brand_color or "#6366f1"
    if not re.fullmatch(r"^#[0-9a-fA-F]{6}$", brand_color):
        raise HTTPException(status_code=400, detail="Invalid brand color")
    org.brand_color = brand_color
    org.settings = settings_data

    await db.commit()
    await db.refresh(org)
    return _serialize_admin_organization(org)


@app.post("/api/admin/organization/logo", dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def upload_admin_organization_logo(
    request: Request,
    file: UploadFile = File(...),
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from .organization_access_api import get_organization_for_access, organization_id_from_request
    org = await get_organization_for_access(db, me, "organization:profile_write", organization_id_from_request(request))
    data, ext = await _read_safe_raster_upload(file)
    safe_name = f"org-logos/org_{org.id}/logo{ext}"
    dest = Path(settings.local_storage_dir) / safe_name
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(data)

    org.brand_logo = f"{settings.public_base_url}/api/files/{safe_name}"
    await db.commit()
    await db.refresh(org)
    return {"brand_logo": org.brand_logo}


@app.get("/api/admin/api-keys", response_model=list[ApiKeyOut], dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def list_api_keys(me: CurrentUser = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(ApiKey).where(ApiKey.user_id == me.id).order_by(ApiKey.created_at.desc()))
    api_keys = res.scalars().all()
    return [
        ApiKeyOut(
            id=api_key.id,
            name=api_key.name,
            key_prefix=api_key.key_prefix,
            is_active=api_key.is_active,
            scopes=list(api_key.scopes or []),
            last_used_at=api_key.last_used_at,
            expires_at=api_key.expires_at,
            created_at=api_key.created_at,
            rate_limit_per_min=api_key.rate_limit_per_min,
        )
        for api_key in api_keys
    ]


@app.post("/api/admin/api-keys", response_model=ApiKeyCreateOut, status_code=201, dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def create_api_key(payload: ApiKeyCreateIn, me: CurrentUser = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    full_key = f"hc_live_{secrets.token_urlsafe(32)}"
    key_prefix = full_key[:8]
    expires_at = None
    if payload.expires_days is not None:
        expires_at = datetime.now(timezone.utc) + timedelta(days=payload.expires_days)

    api_key = ApiKey(
        user_id=me.id,
        name=payload.name.strip()[:200],
        key_prefix=key_prefix,
        key_hash=_hash_api_key(full_key),
        scopes=payload.scopes,
        is_active=True,
        expires_at=expires_at,
        rate_limit_per_min=payload.rate_limit_per_min,
    )
    db.add(api_key)
    await db.commit()
    await db.refresh(api_key)
    return ApiKeyCreateOut(
        id=api_key.id,
        name=api_key.name,
        key_prefix=api_key.key_prefix,
        is_active=api_key.is_active,
        scopes=list(api_key.scopes or []),
        last_used_at=api_key.last_used_at,
        expires_at=api_key.expires_at,
        created_at=api_key.created_at,
        full_key=full_key,
    )


@app.delete("/api/admin/api-keys/{key_id}", dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def delete_api_key(key_id: int, me: CurrentUser = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(ApiKey).where(ApiKey.id == key_id, ApiKey.user_id == me.id))
    api_key = res.scalar_one_or_none()
    if not api_key:
        raise HTTPException(status_code=404, detail="API key not found")
    api_key.is_active = False
    await db.commit()
    return {"ok": True}


_DEFINED_SCOPES = [
    {"value": "events:read",        "label": "View events"},
    {"value": "events:write",       "label": "Create and manage events"},
    {"value": "attendees:read",     "label": "View attendees"},
    {"value": "attendees:write",    "label": "Add and manage attendees"},
    {"value": "certificates:read",  "label": "View certificates"},
    {"value": "certificates:write", "label": "Issue and revoke certificates"},
    {"value": "sessions:read",      "label": "View sessions"},
    {"value": "sessions:write",     "label": "Create and manage sessions"},
    {"value": "checkin:write",      "label": "Perform check-ins"},
    {"value": "analytics:read",     "label": "Access analytics and reports"},
    {"value": "automations:read",   "label": "View automation rules"},
    {"value": "automations:write",  "label": "Create and manage automation rules"},
]




@app.get("/api/admin/api-keys/scopes", dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def list_api_key_scopes():
    return _DEFINED_SCOPES


@app.get("/api/admin/api-keys/v2", response_model=list[ApiKeyOut], dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def list_api_keys_v2(me: CurrentUser = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(ApiKey).where(ApiKey.user_id == me.id).order_by(ApiKey.created_at.desc()))
    return [_api_key_to_out(k) for k in res.scalars().all()]


@app.post("/api/admin/api-keys/v2", response_model=ApiKeyCreateOut, status_code=201, dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def create_api_key_v2(payload: ApiKeyCreateIn, me: CurrentUser = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    full_key = f"hc_live_{secrets.token_urlsafe(32)}"
    key_prefix = full_key[:8]
    expires_at = None
    if payload.expires_days is not None:
        expires_at = datetime.now(timezone.utc) + timedelta(days=payload.expires_days)

    api_key = ApiKey(
        user_id=me.id,
        name=payload.name.strip()[:200],
        key_prefix=key_prefix,
        key_hash=_hash_api_key(full_key),
        scopes=payload.scopes,
        is_active=True,
        expires_at=expires_at,
        rate_limit_per_min=payload.rate_limit_per_min,
    )
    db.add(api_key)
    await db.commit()
    await db.refresh(api_key)
    out = _api_key_to_out(api_key)
    return ApiKeyCreateOut(**out.model_dump(), full_key=full_key)


@app.patch("/api/admin/api-keys/{key_id}/scopes", response_model=ApiKeyOut, dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def update_api_key_scopes(key_id: int, payload: ApiKeyScopePatchIn, me: CurrentUser = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(ApiKey).where(ApiKey.id == key_id, ApiKey.user_id == me.id))
    api_key = res.scalar_one_or_none()
    if not api_key:
        raise HTTPException(status_code=404, detail="API key not found")
    if payload.name is not None:
        api_key.name = payload.name.strip()[:200]
    if payload.scopes is not None:
        api_key.scopes = payload.scopes
    if payload.is_active is not None:
        api_key.is_active = payload.is_active
    await db.commit()
    await db.refresh(api_key)
    return _api_key_to_out(api_key)


@app.get("/public/branding")
async def get_public_branding(request: Request, db: AsyncSession = Depends(get_db)):
    host = request.headers.get("host")

    if not host:
        return JSONResponse(content={}, status_code=200)

    # localhost vs port temizle
    host = host.split(":")[0]

    # organization bul
    result = await db.execute(
        select(Organization).where(Organization.custom_domain == host)
    )
    org = result.scalar_one_or_none()

    # EÃ„Å¸er custom domain yoksa fallback
    if not org:
        return {
            "org_name": "HeptaCert",
            "brand_logo": None,
            "brand_color": "#7c3aed",
            "settings": {
                "hide_heptacert_home": False
            }
        }

    # settings ÃƒÂ§ek (senin DBÃ¢â‚¬â„¢de JSONB)
    settings_data = {}

    if hasattr(org, "settings") and org.settings:
        settings_data = org.settings

    return {
        "org_name": org.org_name,
        "brand_logo": org.brand_logo,
        "brand_color": org.brand_color,
        "settings": settings_data or {
            "hide_heptacert_home": False
        }
    }

@app.get(
    "/api/admin/events/{event_id}/certificates",
    response_model=CertificateListOut,
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def list_certificates(
    event_id: int,
    search: str = "",
    status: Optional[CertStatus] = None,
    page: int = 1,
    limit: int = 20,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if page < 1 or limit < 1 or limit > 200:
        raise bad_request("Invalid page/limit")

    ev = await _get_event_for_admin(event_id, me, db, "certificates:write")
    _ensure_certificate_feature_enabled(ev)

    q = select(Certificate).where(
        Certificate.event_id == event_id,
        Certificate.deleted_at.is_(None),
    )

    if search:
        s = search.strip()
        from sqlalchemy import or_
        if len(s) >= 2:
            # Sanitise: keep only alphanumeric, spaces, Turkish chars
            import re as _re
            _safe_words = [w for w in _re.sub(r"[^\w\sÃƒÂ§Ã„Å¸Ã„Â±ÃƒÂ¶Ã…Å¸ÃƒÂ¼Ãƒâ€¡Ã„ÂÃ„Â°Ãƒâ€“Ã…ÂÃƒÅ“]", "", s).split() if w]
            if _safe_words:
                tsq_str = " & ".join([w + ":*" for w in _safe_words])
                try:
                    q = q.where(
                        or_(
                            Certificate.student_name.ilike(f"%{s}%"),
                            func.to_tsvector("simple", Certificate.student_name).op("@@")(
                                func.to_tsquery("simple", tsq_str)
                            ),
                        )
                    )
                except Exception:
                    q = q.where(Certificate.student_name.ilike(f"%{s}%"))
            else:
                q = q.where(Certificate.student_name.ilike(f"%{s}%"))
        else:
            q = q.where(Certificate.student_name.ilike(f"%{s}%"))

    if status:
        q = q.where(Certificate.status == status)

    # total
    res_total = await db.execute(select(func.count()).select_from(q.subquery()))
    total = int(res_total.scalar_one())

    q = q.order_by(Certificate.created_at.desc()).offset((page - 1) * limit).limit(limit)
    res = await db.execute(q)
    items = res.scalars().all()

    def to_out(c: Certificate) -> CertificateOut:
        # expired/revoked -> pdf kapalÃ„Â± (X)
        pdf_url = c.pdf_url if c.status == CertStatus.active else None
        return CertificateOut(
            id=c.id,
            uuid=c.uuid,
            public_id=c.public_id,
            student_name=c.student_name,
            event_id=c.event_id,
            status=c.status,
            issued_at=getattr(c, "issued_at", None),
            hosting_term=getattr(c, "hosting_term", None),
            hosting_ends_at=getattr(c, "hosting_ends_at", None),
            pdf_url=pdf_url,
        )

    return CertificateListOut(
        items=[certificate_to_out(x) for x in items],
        total=total,
        page=page,
        limit=limit,
    )


@app.get(
    "/api/admin/events/{event_id}/certificates/cost-estimate",
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def estimate_certificate_cost(
    event_id: int,
    count: int = Query(default=1, ge=1, le=100000),
    asset_size_bytes: int = Query(default=0, ge=0),
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    ev = await _get_event_for_admin(event_id, me, db, "certificates:write")
    _ensure_certificate_feature_enabled(ev)
    monthly_units = hosting_units("monthly", asset_size_bytes)
    yearly_units = hosting_units("yearly", asset_size_bytes)
    return {
        "count": count,
        "asset_size_bytes": asset_size_bytes,
        "issue_units_per_certificate": ISSUE_UNITS_PER_CERT,
        "monthly_hosting_units_per_certificate": monthly_units,
        "yearly_hosting_units_per_certificate": yearly_units,
        "monthly_total_units": count * (ISSUE_UNITS_PER_CERT + monthly_units),
        "yearly_total_units": count * (ISSUE_UNITS_PER_CERT + yearly_units),
        "monthly_renewal_units": count * monthly_units,
        "yearly_renewal_units": count * yearly_units,
        "mb_per_coin_month": MB_PER_COIN_MONTH,
    }




@app.post(
    "/api/admin/events/{event_id}/certificates",
    response_model=CertificateOut,
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def issue_certificate(
    event_id: int,
    payload: IssueCertificateIn,
    background_tasks: BackgroundTasks,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    ev = await _get_event_for_admin(event_id, me, db, "certificates:write")
    _ensure_certificate_feature_enabled(ev)

    if not ev.config:
        raise bad_request("Event config missing. Save coordinates in editor first.")
    try:
        cfg = editor_config_to_template_config(ev.config)
    except Exception as e:
        raise bad_request(f"Invalid event config: {e}")

    billing_user_id = ev.admin_id

    # Enforce hologram: only Growth/Enterprise can disable it for the organizer account
    if not cfg.show_hologram and me.role != Role.superadmin:
        _sub_h = await db.execute(
            select(Subscription)
            .where(Subscription.user_id == billing_user_id, Subscription.is_active == True)
            .order_by(Subscription.expires_at.desc()).limit(1)
        )
        _sub_h_row = _sub_h.scalar_one_or_none()
        _now_h = datetime.now(timezone.utc)
        if not _sub_h_row or _sub_h_row.plan_id not in ("growth", "enterprise") or \
                (_sub_h_row.expires_at and _sub_h_row.expires_at < _now_h):
            cfg.show_hologram = True

    res_u = await db.execute(select(User).where(User.id == billing_user_id))
    user = res_u.scalar_one()

    template_path = local_path_from_url(ev.template_image_url)
    if not template_path.exists():
        raise bad_request("Template image not found on server. Upload template or fix template_image_url.")
    template_bytes = template_path.read_bytes()

    org_res2 = await db.execute(select(Organization).where(Organization.user_id == billing_user_id))
    org2 = org_res2.scalar_one_or_none()
    single_brand_logo_bytes: Optional[bytes] = None
    if org2 and org2.brand_logo:
        try:
            logo_path2 = local_path_from_url(org2.brand_logo)
            if logo_path2.exists():
                single_brand_logo_bytes = logo_path2.read_bytes()
        except Exception:
            pass
    certificate_footer: Optional[str] = None
    try:
        certificate_footer = (org2.settings or {}).get("certificate_footer") if org2 else None
    except Exception:
        certificate_footer = None

    res_lock = await db.execute(select(Event).where(Event.id == ev.id).with_for_update())
    ev = res_lock.scalar_one()

    term = payload.hosting_term

    cert_uuid = new_certificate_uuid()
    ev.cert_seq += 1
    public_id = f"EV{ev.id}-{ev.cert_seq:06d}"
    verify_url = build_certificate_verify_url(cert_uuid)

    pdf_bytes = render_certificate_pdf(
        template_image_bytes=template_bytes,
        student_name=payload.student_name,
        verify_url=verify_url,
        config=cfg,
        public_id=public_id,
        brand_logo_bytes=single_brand_logo_bytes,
        certificate_footer=certificate_footer,
    )

    rel_pdf_path = f"pdfs/event_{ev.id}/{cert_uuid}.pdf"
    abs_pdf_path = Path(settings.local_storage_dir) / rel_pdf_path
    abs_pdf_path.parent.mkdir(parents=True, exist_ok=True)
    abs_pdf_path.write_bytes(pdf_bytes)
    asset_size_bytes = abs_pdf_path.stat().st_size

    rel_png_path: Optional[str] = None
    try:
        png_bytes = await asyncio.to_thread(
            render_certificate_png_watermarked,
            template_image_bytes=template_bytes,
            student_name=payload.student_name,
            verify_url=verify_url,
            config=cfg,
            public_id=public_id,
            brand_logo_bytes=single_brand_logo_bytes,
            certificate_footer=certificate_footer,
        )
        rel_png_path = f"pngs/event_{ev.id}/{cert_uuid}.png"
        abs_png_path = Path(settings.local_storage_dir) / rel_png_path
        abs_png_path.parent.mkdir(parents=True, exist_ok=True)
        abs_png_path.write_bytes(png_bytes)
    except Exception:
        pass

    hosting_spend = hosting_units(term, asset_size_bytes)
    spend_units = ISSUE_UNITS_PER_CERT + hosting_spend

    if user.heptacoin_balaonce < spend_units:
        raise HTTPException(
            status_code=402,
            detail=f"Insufficient HeptaCoin. NeededUnits={spend_units}, balanceUnits={user.heptacoin_balaonce}",
        )

    pdf_url = build_public_pdf_url(rel_pdf_path)
    hosting_ends_at = compute_hosting_ends(term)

    cert = Certificate(
        uuid=cert_uuid,
        public_id=public_id,
        student_name=payload.student_name,
        event_id=ev.id,
        pdf_url=pdf_url,
        status=CertStatus.active,
        hosting_term=term,
        hosting_ends_at=hosting_ends_at,
        asset_size_bytes=asset_size_bytes,
    )
    db.add(cert)

    user.heptacoin_balaonce -= spend_units
    db.add(Transaction(user_id=user.id, amount=spend_units, type=TxType.spend))

    await write_audit_log(
        db,
        user_id=me.id,
        action="certificate.issued",
        ip_address=None,
        user_agent=None,
        extra={"event_id": ev.id, "public_id": public_id, "student_name": payload.student_name},
    )
    from .crm_snapshot_hooks import auto_tag_certified_for_attendee_email, refresh_crm_snapshots_for_certificate_name
    await refresh_crm_snapshots_for_certificate_name(db, event_id=ev.id, student_name=payload.student_name)
    attendee_email_res = await db.execute(
        select(Attendee.email).where(
            Attendee.event_id == ev.id,
            func.lower(func.trim(Attendee.name)) == payload.student_name.strip().lower(),
            Attendee.email.is_not(None),
            func.trim(Attendee.email) != "",
        ).limit(1)
    )
    _cert_email = attendee_email_res.scalar_one_or_none()
    if _cert_email:
        _org_res = await db.execute(select(Organization).where(Organization.user_id == ev.admin_id).limit(1))
        _cert_org = _org_res.scalar_one_or_none()
        if _cert_org:
            await auto_tag_certified_for_attendee_email(db, org_id=_cert_org.id, email=_cert_email)

    await db.commit()
    await db.refresh(cert)

    try:
        await _maybe_log_cpd(db, ev.id, cert.id, student_name=payload.student_name)
        await db.commit()
    except Exception:
        pass

    attendee_res = await db.execute(
        select(Attendee)
        .where(
            Attendee.event_id == ev.id,
            func.lower(Attendee.name) == (payload.student_name or "").strip().lower(),
        )
        .order_by(Attendee.id.asc())
        .limit(1)
    )
    attendee_for_email = attendee_res.scalar_one_or_none()
    if attendee_for_email and attendee_for_email.email:
        background_tasks.add_task(
            send_certificate_delivery_email_task,
            event_id=ev.id,
            cert_uuid=cert.uuid,
            recipient_name=attendee_for_email.name,
            recipient_email=attendee_for_email.email,
        )

    from .webhooks import WebhookEvent
    background_tasks.add_task(
        deliver_webhook_task, billing_user_id, WebhookEvent.cert_issued.value,
        {"uuid": cert.uuid, "public_id": cert.public_id, "student_name": cert.student_name, "event_id": ev.id},
    )

    return certificate_to_out(cert)


@app.patch(
    "/api/admin/certificates/{cert_id}",
    response_model=CertificateOut,
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def update_certificate_status(
    cert_id: int,
    payload: UpdateCertificateStatusIn,
    background_tasks: BackgroundTasks,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    cert_res = await db.execute(
        select(Certificate).where(Certificate.id == cert_id, Certificate.deleted_at.is_(None))
    )
    cert = cert_res.scalar_one_or_none()
    if not cert:
        raise HTTPException(status_code=404, detail="Certificate not found")

    ev = await _get_event_for_admin(cert.event_id, me, db, "certificates:write")
    _ensure_certificate_feature_enabled(ev)

    previous_status = cert.status
    previous_auto_renew = bool(getattr(cert, "auto_renew_enabled", False))
    if payload.status is None and payload.auto_renew_enabled is None:
        raise HTTPException(status_code=400, detail="No certificate update provided")

    if payload.status is not None:
        cert.status = payload.status
        if payload.status == CertStatus.active:
            hosting_ends_at = ensure_utc(getattr(cert, "hosting_ends_at", None))
            if hosting_ends_at is None or hosting_ends_at <= datetime.now(timezone.utc):
                cert.hosting_ends_at = compute_hosting_ends(getattr(cert, "hosting_term", None) or "yearly")

    if payload.auto_renew_enabled is not None:
        cert.auto_renew_enabled = payload.auto_renew_enabled

    await write_audit_log(
        db,
        user_id=me.id,
        action="certificate.status.updated",
        resource_type="certificate",
        resource_id=str(cert.id),
        extra={
            "event_id": cert.event_id,
            "public_id": cert.public_id,
            "from": previous_status.value if previous_status else None,
            "to": payload.status.value if payload.status else previous_status.value,
            "auto_renew_from": previous_auto_renew,
            "auto_renew_to": bool(getattr(cert, "auto_renew_enabled", False)),
        },
    )
    await db.commit()
    await db.refresh(cert)

    if payload.status == CertStatus.revoked and background_tasks:
        from .webhooks import WebhookEvent
        background_tasks.add_task(
            deliver_webhook_task,
            ev.admin_id,
            WebhookEvent.cert_revoked.value,
            {"uuid": cert.uuid, "public_id": cert.public_id, "event_id": cert.event_id},
        )

    return certificate_to_out(cert)


@app.delete(
    "/api/admin/certificates/{cert_id}",
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def delete_certificate(
    cert_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    cert_res = await db.execute(
        select(Certificate).where(Certificate.id == cert_id, Certificate.deleted_at.is_(None))
    )
    cert = cert_res.scalar_one_or_none()
    if not cert:
        raise HTTPException(status_code=404, detail="Certificate not found")

    await _get_event_for_admin(cert.event_id, me, db, "certificates:write")
    cert.deleted_at = datetime.now(timezone.utc)

    await write_audit_log(
        db,
        user_id=me.id,
        action="certificate.deleted",
        resource_type="certificate",
        resource_id=str(cert.id),
        extra={"event_id": cert.event_id, "public_id": cert.public_id},
    )
    await db.commit()
    return {"ok": True}


@app.post(
    "/api/admin/events/{event_id}/certificates/bulk-action",
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def bulk_certificate_action(
    event_id: int,
    payload: BulkActionIn,
    background_tasks: BackgroundTasks,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    ev = await _get_event_for_admin(event_id, me, db, "certificates:write")
    _ensure_certificate_feature_enabled(ev)

    res_certs = await db.execute(
        select(Certificate).where(
            Certificate.id.in_(payload.cert_ids),
            Certificate.event_id == event_id,
            Certificate.deleted_at.is_(None),
        )
    )
    certs = res_certs.scalars().all()
    if not certs:
        raise HTTPException(status_code=404, detail="No certificates found")

    processed = 0
    for cert in certs:
        if payload.action == "delete":
            cert.deleted_at = datetime.now(timezone.utc)
            if cert.pdf_url:
                try:
                    pdf_path = local_path_from_url(cert.pdf_url)
                    if pdf_path.exists():
                        pdf_path.unlink(missing_ok=True)
                except Exception as exc:
                    logger.warning("Bulk delete PDF cleanup failed for cert %s: %s", cert.id, exc)
        elif payload.action == "revoke":
            cert.status = CertStatus.revoked
            if cert.pdf_url:
                try:
                    pdf_path = local_path_from_url(cert.pdf_url)
                    if pdf_path.exists():
                        pdf_path.unlink(missing_ok=True)
                except Exception as exc:
                    logger.warning("Bulk revoke PDF cleanup failed for cert %s: %s", cert.id, exc)
        elif payload.action == "expire":
            cert.status = CertStatus.expired
        elif payload.action == "enable_auto_renew":
            cert.auto_renew_enabled = True
        elif payload.action == "disable_auto_renew":
            cert.auto_renew_enabled = False
        processed += 1

    await write_audit_log(
        db,
        user_id=me.id,
        action=f"certificate.bulk.{payload.action}",
        ip_address=None,
        user_agent=None,
        extra={"event_id": event_id, "count": processed},
    )
    await db.commit()

    if payload.action == "revoke" and background_tasks:
        from .webhooks import WebhookEvent
        background_tasks.add_task(
            deliver_webhook_task, ev.admin_id, WebhookEvent.cert_bulk_completed.value,
            {"event_id": event_id, "action": "revoke", "count": processed},
        )

    return {"ok": True, "processed": processed, "action": payload.action}


# Ã¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢Â
# Certificate Export (CSV / XLSX)
# Ã¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢Â

@app.get(
    "/api/admin/events/{event_id}/certificates/export",
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def export_certificates(
    event_id: int,
    format: str = Query(default="csv", pattern="^(csv|xlsx)$"),
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    ev = await _get_event_for_admin(event_id, me, db, "certificates:write")
    _ensure_certificate_feature_enabled(ev)

    res = await db.execute(
        select(Certificate)
        .where(Certificate.event_id == event_id, Certificate.deleted_at.is_(None))
        .order_by(Certificate.created_at.asc())
    )
    certs = res.scalars().all()

    columns = ["public_id", "student_name", "status", "hosting_term", "issued_at", "hosting_ends_at", "uuid"]

    if format == "xlsx":
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Certificates"
        ws.append(columns)
        for c in certs:
            ws.append([
                c.public_id or "",
                c.student_name,
                c.status.value if c.status else "",
                getattr(c, "hosting_term", "") or "",
                getattr(c, "issued_at", None).isoformat() if getattr(c, "issued_at", None) else "",
                getattr(c, "hosting_ends_at", None).isoformat() if getattr(c, "hosting_ends_at", None) else "",
                c.uuid,
            ])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=certificates-event-{event_id}.xlsx"},
        )
    else:
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(columns)
        for c in certs:
            writer.writerow([
                c.public_id or "",
                c.student_name,
                c.status.value if c.status else "",
                getattr(c, "hosting_term", "") or "",
                getattr(c, "issued_at", None).isoformat() if getattr(c, "issued_at", None) else "",
                getattr(c, "hosting_ends_at", None).isoformat() if getattr(c, "hosting_ends_at", None) else "",
                c.uuid,
            ])
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename=certificates-event-{event_id}.csv"},
        )


# Ã¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢Â
# Dashboard Analytics
# Ã¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢Â

@app.get(
    "/api/admin/dashboard/stats",
    response_model=DashboardStatsOut,
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def dashboard_stats(
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    # Get events for this user
    res_events = await db.execute(
        select(Event).where(Event.admin_id == me.id).order_by(Event.created_at.desc())
    )
    events = res_events.scalars().all()
    event_ids = [e.id for e in events]

    if not event_ids:
        return DashboardStatsOut(
            total_events=0, total_certs=0, active_certs=0, revoked_certs=0,
            expired_certs=0, total_spent_hc=0, events_with_stats=[],
        )

    # Aggregate cert counts by status per event
    from sqlalchemy import case as sa_case
    agg_res = await db.execute(
        select(
            Certificate.event_id,
            func.count(Certificate.id).label("total"),
            func.sum(sa_case((Certificate.status == CertStatus.active, 1), else_=0)).label("active"),
            func.sum(sa_case((Certificate.status == CertStatus.revoked, 1), else_=0)).label("revoked"),
            func.sum(sa_case((Certificate.status == CertStatus.expired, 1), else_=0)).label("expired"),
        )
        .where(Certificate.event_id.in_(event_ids), Certificate.deleted_at.is_(None))
        .group_by(Certificate.event_id)
    )
    agg_rows = agg_res.all()
    event_stats_map: Dict[int, Dict] = {r.event_id: r._asdict() for r in agg_rows}

    # Total HC spent
    tx_res = await db.execute(
        select(func.coalesce(func.sum(Transaction.amount), 0))
        .where(Transaction.user_id == me.id, Transaction.type == TxType.spend)
    )
    total_spent = int(tx_res.scalar_one() or 0)

    total_certs = sum(r["total"] for r in event_stats_map.values())
    active_certs = sum(int(r["active"] or 0) for r in event_stats_map.values())
    revoked_certs = sum(int(r["revoked"] or 0) for r in event_stats_map.values())
    expired_certs = sum(int(r["expired"] or 0) for r in event_stats_map.values())

    event_name_map = {e.id: e.name for e in events}
    events_with_stats = [
        {
            "event_id": ev_id,
            "event_name": event_name_map.get(ev_id, ""),
            "name": event_name_map.get(ev_id, ""),
            "total": int(stats["total"] or 0),
            "active": int(stats["active"] or 0),
            "revoked": int(stats["revoked"] or 0),
            "expired": int(stats["expired"] or 0),
            "cert_count": int(stats["total"] or 0),
            "active_count": int(stats["active"] or 0),
            "revoked_count": int(stats["revoked"] or 0),
            "expired_count": int(stats["expired"] or 0),
        }
        for ev_id, stats in event_stats_map.items()
    ]

    return DashboardStatsOut(
        total_events=len(events),
        total_certs=total_certs,
        active_certs=active_certs,
        revoked_certs=revoked_certs,
        expired_certs=expired_certs,
        total_spent_hc=total_spent,
        events_with_stats=events_with_stats,
    )


# Ã¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢Â
# Superadmin System Health
# Ã¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢Â

@app.get(
    "/api/superadmin/system-health",
    dependencies=[Depends(require_role(Role.superadmin))],
)
async def system_health(db: AsyncSession = Depends(get_db)):
    import shutil
    from sqlalchemy import text as sa_text

    # Disk usage
    try:
        disk = shutil.disk_usage(settings.local_storage_dir)
        disk_info = {
            "total_bytes": disk.total,
            "used_bytes": disk.used,
            "free_bytes": disk.free,
            "used_pct": round(disk.used / disk.total * 100, 1) if disk.total else 0,
        }
    except Exception:
        disk_info = {}

    # DB size
    try:
        db_size_res = await db.execute(sa_text("SELECT pg_database_size(current_database())"))
        db_size_bytes = int(db_size_res.scalar_one() or 0)
    except Exception:
        db_size_bytes = 0

    # Active DB connections
    try:
        conn_res = await db.execute(sa_text("SELECT count(*) FROM pg_stat_activity"))
        active_connections = int(conn_res.scalar_one() or 0)
    except Exception:
        active_connections = 0

    # Recent activity (last 24h audit log count)
    try:
        activity_res = await db.execute(
            select(func.count()).select_from(AuditLog)
            .where(AuditLog.created_at >= datetime.now(timezone.utc) - timedelta(hours=24))
        )
        recent_actions = int(activity_res.scalar_one() or 0)
    except Exception:
        recent_actions = 0

    # Totals
    try:
        user_res = await db.execute(select(func.count()).select_from(User))
        total_users = int(user_res.scalar_one() or 0)
        cert_res = await db.execute(
            select(func.count()).select_from(Certificate).where(Certificate.deleted_at.is_(None))
        )
        total_certs = int(cert_res.scalar_one() or 0)
    except Exception:
        total_users = 0
        total_certs = 0

    return {
        "uptime_seconds": round(time.time() - _startup_time),
        "disk_total_gb":  round(disk_info.get("total_bytes", 0) / (1024 ** 3), 2),
        "disk_used_gb":   round(disk_info.get("used_bytes",  0) / (1024 ** 3), 2),
        "disk_free_gb":   round(disk_info.get("free_bytes",  0) / (1024 ** 3), 2),
        "disk_percent":   disk_info.get("used_pct", 0),
        "db_size_mb":     round(db_size_bytes / (1024 ** 2), 2),
        "db_active_connections": active_connections,
        "recent_24h_actions": recent_actions,
        "total_users": total_users,
        "total_certs": total_certs,
    }


# Ã¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢Â
# Superadmin Subscription Management
# Ã¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢Â



@app.get(
    "/api/superadmin/subscriptions",
    dependencies=[Depends(require_role(Role.superadmin))],
)
async def list_all_subscriptions(db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(Subscription, User.email)
        .join(User, User.id == Subscription.user_id)
        .order_by(Subscription.started_at.desc())
        .limit(500)
    )
    rows = [
        {
            "id": sub.id,
            "target_type": "admin",
            "user_id": sub.user_id,
            "user_email": email,
            "plan_id": sub.plan_id,
            "order_id": sub.order_id,
            "started_at": sub.started_at.isoformat() if sub.started_at else None,
            "expires_at": sub.expires_at.isoformat() if sub.expires_at else None,
            "is_active": sub.is_active,
        }
        for sub, email in res.all()
    ]
    return rows


@app.post(
    "/api/superadmin/subscriptions/grant",
    dependencies=[Depends(require_role(Role.superadmin))],
    status_code=201,
)
async def grant_subscription(payload: GrantSubscriptionIn, db: AsyncSession = Depends(get_db)):
    # Find user by email
    user_res = await db.execute(select(User).where(User.email == payload.user_email))
    user = user_res.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="KullanÃ„Â±cÃ„Â± bulunamadÃ„Â±.")

    # Validate plan_id
    valid_plans = [t["id"] for t in DEFAULT_PRICING]
    if payload.plan_id not in valid_plans:
        raise HTTPException(status_code=400, detail=f"GeÃ§ersiz plan. GeÃ§erli planlar: {', '.join(valid_plans)}")

    # Deactivate all existing active subscriptions for this user
    await db.execute(
        update(Subscription)
        .where(Subscription.user_id == user.id, Subscription.is_active == True)
        .values(is_active=False)
    )

    now = datetime.now(timezone.utc)
    new_sub = Subscription(
        user_id=user.id,
        plan_id=payload.plan_id,
        order_id=None,
        started_at=now,
        expires_at=now + timedelta(days=payload.days),
        is_active=True,
    )
    new_sub.last_hc_credited_at = datetime.now(timezone.utc)
    db.add(new_sub)
    # Credit initial HC quota immediately
    hc_quota_grant = _get_hc_quota(payload.plan_id)
    if hc_quota_grant:
        user.heptacoin_balaonce += hc_quota_grant
        db.add(Transaction(
            user_id=user.id, amount=hc_quota_grant, type=TxType.credit,
            description=f"Superadmin plan aktivasyonu: {payload.plan_id}",
        ))
    await db.commit()
    await db.refresh(new_sub)

    return {
        "id": new_sub.id,
        "target_type": "admin",
        "user_id": new_sub.user_id,
        "user_email": user.email,
        "plan_id": new_sub.plan_id,
        "started_at": new_sub.started_at.isoformat() if new_sub.started_at else None,
        "expires_at": new_sub.expires_at.isoformat() if new_sub.expires_at else None,
        "is_active": new_sub.is_active,
    }


@app.delete(
    "/api/superadmin/subscriptions/{sub_id}",
    dependencies=[Depends(require_role(Role.superadmin))],
)
async def revoke_subscription(sub_id: int, db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(Subscription).where(Subscription.id == sub_id))
    sub = res.scalar_one_or_none()
    if not sub:
        raise HTTPException(status_code=404, detail="Abonelik bulunamadÃ„Â±.")
    sub.is_active = False
    await db.commit()
    return {"detail": "Abonelik iptal edildi."}


# Ã¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢Â
# Magic Link Authentication
# Ã¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢Â

@app.post("/api/auth/magic-link")
@limiter.limit("3/minute")
async def request_magic_link(
    request: Request,
    data: MagicLinkIn,
    db: AsyncSession = Depends(get_db),
):
    res = await db.execute(select(User).where(User.email == str(data.email)))
    user = res.scalar_one_or_none()
    # Always 200 to avoid enumeration
    if user and user.is_verified:
        token = make_email_token({"email": str(data.email), "action": "magic_link"})
        user.magic_link_token = token
        await db.commit()

        verify_link = f"{settings.frontend_base_url}/admin/magic-verify?token={token}"
        await send_email_async(
            to=str(data.email),
            subject="HeptaCert - Giriş Bağlantısı",
            html_body=f"""
            <p>Merhaba,</p>
            <p>HeptaCert'e giriş yapmak icin aşağıdaki bağlantıya tıklayın:</p>
            <p><a href="{verify_link}" style="background:#6366f1;color:#fff;padding:10px 20px;border-radius:6px;text-decoration:none;display:inline-block">Giriş Yap</a></p>
            <p>Bu bağlantı 15 dakika gecerlidir.</p>
            <p>Eğer bu isteği siz yapmadıysanız, bu e-postayı gormezden gelebilirsiniz.</p>
            """,
        )
    return {"detail": "Giriş bağlantısı e-posta adresinize gonderildi."}


@app.get("/api/auth/magic-link/verify")
async def verify_magic_link(
    token: str = Query(...),
    db: AsyncSession = Depends(get_db),
):
    try:
        payload = verify_email_token(token, max_age=900)  # 15 minutes
    except SignatureExpired:
        raise bad_request("Giriş bağlantısının suresi dolmuş. Lutfen yeni bir bağlantı isteyin.")
    except (BadSignature, Exception):
        raise bad_request("Gecersiz giriş bağlantısı.")

    if payload.get("action") != "magic_link":
        raise bad_request("Gecersiz token turu.")

    email = payload.get("email")
    res = await db.execute(select(User).where(User.email == email))
    user = res.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı.")
    if not user.is_verified:
        raise bad_request("Hesabınız henuz doğrulanmamış.")

    # Invalidate token after use
    user.magic_link_token = None
    await db.commit()

    return TokenOut(
        access_token=create_access_token(user_id=user.id, role=user.role),
        token_type="bearer",
    )


# Ã¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢Â
# Template History
# Ã¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢Â



# Ã¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢Â
# Attendaonce Management
# Ã¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢Â

# Ã¢â€â‚¬Ã¢â€â‚¬ Pydantic schemas for attendaonce Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬















LEGAL_DOCUMENT_VERSION = "2026-05-30"








async def _write_attendee_consent_audit(
    db: AsyncSession,
    *,
    event: Event,
    attendee: Attendee,
    payload: SelfRegisterIn,
    member: Optional[CurrentPublicMember],
    client_ip: Optional[str],
    user_agent: Optional[str],
    device_id: str,
    result: str,
) -> None:
    await write_audit_log(
        db,
        user_id=None,
        action="legal.consent.accept",
        resource_type="legal_consent",
        resource_id=str(attendee.id),
        ip_address=client_ip,
        user_agent=user_agent,
        extra={
            "event_id": event.id,
            "event_public_id": _get_public_event_identifier(event),
            "attendee_id": attendee.id,
            "email": attendee.email.lower(),
            "public_member_id": member.id if member else None,
            "device_id": device_id,
            "result": result,
            "legal_document_version": LEGAL_DOCUMENT_VERSION,
            "kvkk_required": _is_event_kvkk_consent_required(event),
            "kvkk_accepted": bool(payload.kvkk_accepted),
            "kvkk_text_hash": _legal_text_hash(_get_event_kvkk_consent_text(event)),
            "organizer_notice_required": _is_event_organizer_privacy_notice_enabled(event),
            "organizer_notice_accepted": bool(payload.organizer_notice_accepted),
            "organizer_notice_text_hash": _legal_text_hash(_get_event_organizer_privacy_notice_text(event)),
            "cross_border_notice_required": _is_event_cross_border_transfer_notice_enabled(event),
            "cross_border_notice_read": bool(payload.cross_border_notice_read),
            "cross_border_notice_text_hash": _legal_text_hash(_cross_border_notice_text_for_audit()),
            "cross_border_transfer_consent_required": _is_event_cross_border_transfer_consent_required(event),
            "cross_border_transfer_consent": bool(payload.cross_border_transfer_consent),
        },
    )




def _build_attendee_register_webhook_payload(
    event: Event,
    attendee: Attendee,
    *,
    result: str,
) -> Dict[str, Any]:
    answers = attendee.registration_answers or {}
    registration_fields = _get_event_registration_fields(event)
    labeled_answers: Dict[str, Any] = {}
    for field in registration_fields:
        field_id = str(field.get("id") or "")
        if not field_id:
            continue
        label = str(field.get("label") or field_id)
        labeled_answers[label] = answers.get(field_id, "")

    return {
        "event_id": event.id,
        "event_public_id": _get_public_event_identifier(event),
        "event_name": event.name,
        "result": result,
        "attendee": {
            "id": attendee.id,
            "name": attendee.name,
            "email": attendee.email,
            "source": attendee.source,
            "registered_at": attendee.registered_at.isoformat() if attendee.registered_at else None,
            "email_verified": bool(attendee.email_verified),
        },
        "registration_answers": answers,
        "registration_answer_labels": labeled_answers,
    }


















# Ã¢â€â‚¬Ã¢â€â‚¬ Helper: resolve event + ownership Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬





async def _get_event_team_membership(event_id: int, me: CurrentUser, db: AsyncSession) -> Optional[EventTeamMember]:
    normalized_email = (me.email or "").strip().lower()
    res = await db.execute(
        select(EventTeamMember).where(
            EventTeamMember.event_id == event_id,
            EventTeamMember.status == "active",
            or_(
                EventTeamMember.user_id == me.id,
                func.lower(func.trim(EventTeamMember.email)) == normalized_email,
            ),
        )
    )
    return res.scalar_one_or_none()


async def _get_event_for_admin(
    event_id: int,
    me: CurrentUser,
    db: AsyncSession,
    required_permission: Optional[str] = "team:manage",
) -> Event:
    res = await db.execute(select(Event).where(Event.id == event_id))
    ev = res.scalar_one_or_none()
    if not ev:
        raise HTTPException(status_code=404, detail="Event not found")
    if me.role == Role.superadmin or ev.admin_id == me.id:
        return ev
    if not await _event_owner_has_enterprise_plan(event_id, db):
        raise HTTPException(status_code=404, detail="Event not found")
    membership = await _get_event_team_membership(event_id, me, db)
    if membership and _event_team_member_allows(membership, required_permission):
        return ev
    from .organization_access_api import user_can_manage_owner_organization
    if await user_can_manage_owner_organization(db, me, ev.admin_id, "events:manage"):
        return ev
    raise HTTPException(status_code=404, detail="Event not found")


async def _get_event_for_owner(event_id: int, me: CurrentUser, db: AsyncSession) -> Event:
    res = await db.execute(select(Event).where(Event.id == event_id))
    ev = res.scalar_one_or_none()
    if not ev:
        raise HTTPException(status_code=404, detail="Event not found")
    if me.role != Role.superadmin and ev.admin_id != me.id:
        raise HTTPException(status_code=403, detail="Only the event owner can manage the team")
    return ev



















































# Ã¢â€â‚¬Ã¢â€â‚¬ Public: event info & self-register Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬



def _build_public_event_detail(
    event: Event,
    sessions: List[EventSession],
    survey: Optional["EventSurvey"],
    organization: Optional[Organization] = None,
    has_active_quiz: bool = False,
) -> PublicEventDetailOut:
    return PublicEventDetailOut(
        id=event.id,
        public_id=_get_public_event_identifier(event),
        name=event.name,
        organization_public_id=organization.public_id if organization else None,
        organization_name=organization.org_name if organization else None,
        organization_logo=organization.brand_logo if organization else None,
        event_date=event.event_date.isoformat() if event.event_date else None,
        event_description=sanitize_event_description_html(event.event_description),
        event_location=event.event_location,
        min_sessions_required=int(event.min_sessions_required or 1),
        registration_closed=_is_event_registration_closed(event),
        event_banner_url=event.event_banner_url,
        registration_fields=_get_event_registration_fields(event),
        survey=_build_public_survey_info(survey),
        visibility=_get_event_visibility(event),
        require_email_verification=_get_event_email_verification_required(event),
        registration_quota=_get_event_registration_quota(event),
        registration_quota_enabled=_is_event_registration_quota_enabled(event),
        event_type=normalize_event_type(getattr(event, "event_type", None)),
        certificate_enabled=is_certificate_enabled(event),
        checkin_enabled=is_checkin_enabled(event),
        ticketing_enabled=is_ticketing_enabled(event),
        registration_enabled=is_public_registration_enabled(event),
        raffles_enabled=is_raffles_enabled(event),
        gamification_enabled=is_gamification_enabled(event),
        requires_approval=normalize_feature_bool(getattr(event, "requires_approval", None), default=FEATURE_DEFAULTS["requires_approval"]),
        quiz_enabled=is_quiz_enabled(event),
        cpd_enabled=is_cpd_enabled(event),
        kvkk_consent_required=_is_event_kvkk_consent_required(event),
        kvkk_consent_text=_get_event_kvkk_consent_text(event),
        organizer_privacy_notice_enabled=_is_event_organizer_privacy_notice_enabled(event),
        organizer_privacy_notice_text=_get_event_organizer_privacy_notice_text(event) if _is_event_organizer_privacy_notice_enabled(event) else None,
        show_cross_border_transfer_notice=_is_event_cross_border_transfer_notice_enabled(event),
        require_cross_border_transfer_consent=_is_event_cross_border_transfer_consent_required(event),
        data_controller_name=_get_event_data_controller_name(event),
        data_controller_contact_email=_get_event_data_controller_contact_email(event),
        data_retention_note=_get_event_data_retention_note(event),
        has_active_quiz=has_active_quiz,
        sessions=[
            {
                "id": s.id,
                "name": s.name,
                "session_date": s.session_date.isoformat() if s.session_date else None,
                "session_start": s.session_start.strftime("%H:%M") if s.session_start else None,
                "session_location": s.session_location,
            }
            for s in sessions
        ],
    )




@app.get("/api/public/events", response_model=list[PublicEventListItemOut])
async def list_public_events(
    request: Request,
    scope: str = Query(default="all", pattern="^(all|upcoming|past)$"),
    limit: int = Query(default=24, ge=1, le=100),
    offset: int = Query(default=0, ge=0),
    search: Optional[str] = Query(default=None, min_length=1, max_length=120),
    db: AsyncSession = Depends(get_db),
):
    bind = db.get_bind()
    dialect_name = bind.dialect.name if bind is not None else ""

    if dialect_name == "sqlite":
        visibility_clause = func.lower(func.coalesce(func.json_extract(Event.config, "$.visibility"), "private")) == "public"
    elif dialect_name == "postgresql":
        visibility_clause = (
            func.lower(
                func.coalesce(
                    func.jsonb_extract_path_text(Event.config, "visibility"),
                    "private",
                )
            )
            == "public"
        )
    else:
        visibility_clause = None

    host_org = await _organization_for_request_host(request, db)
    stmt = select(Event, Organization).outerjoin(Organization, Organization.user_id == Event.admin_id)
    if host_org:
        stmt = stmt.where(Event.admin_id == host_org.user_id)

    if visibility_clause is not None:
        stmt = stmt.where(visibility_clause)

    today = datetime.now(timezone.utc).date()
    if scope == "upcoming":
        stmt = stmt.where(or_(Event.event_date.is_(None), Event.event_date >= today))
        stmt = stmt.order_by(Event.event_date.asc().nulls_last(), Event.created_at.desc())
    elif scope == "past":
        stmt = stmt.where(Event.event_date.is_not(None), Event.event_date < today)
        stmt = stmt.order_by(Event.event_date.desc(), Event.created_at.desc())
    else:
        stmt = stmt.order_by(Event.created_at.desc())

    normalized_search = (search or "").strip().lower()
    if normalized_search:
        like_term = f"%{normalized_search}%"
        stmt = stmt.where(
            or_(
                func.lower(func.coalesce(Event.name, "")).like(like_term),
                func.lower(func.coalesce(Event.event_location, "")).like(like_term),
                func.lower(func.coalesce(Event.event_description, "")).like(like_term),
                func.lower(func.coalesce(Organization.org_name, "")).like(like_term),
            )
        )

    stmt = stmt.offset(offset).limit(limit)
    events_res = await db.execute(stmt)
    event_rows = events_res.all()
    if visibility_clause is None:
        event_rows = [(event, org) for event, org in event_rows if _get_event_visibility(event) == "public"]

    if not event_rows:
        return []

    visible_events = [event for event, _org in event_rows]
    orgs_by_user_id: Dict[int, Organization] = {
        org.user_id: org for _event, org in event_rows if org is not None
    }
    event_ids = [event.id for event in visible_events]
    session_counts_res = await db.execute(
        select(EventSession.event_id, func.count(EventSession.id).label("cnt"))
        .where(EventSession.event_id.in_(event_ids))
        .group_by(EventSession.event_id)
    )
    session_counts = {int(row.event_id): int(row.cnt or 0) for row in session_counts_res.all()}

    return [
        PublicEventListItemOut(
            id=event.id,
            public_id=_get_public_event_identifier(event),
            name=event.name,
            organization_public_id=orgs_by_user_id.get(event.admin_id).public_id if orgs_by_user_id.get(event.admin_id) else None,
            organization_name=orgs_by_user_id.get(event.admin_id).org_name if orgs_by_user_id.get(event.admin_id) else None,
            organization_logo=orgs_by_user_id.get(event.admin_id).brand_logo if orgs_by_user_id.get(event.admin_id) else None,
            event_date=event.event_date.isoformat() if event.event_date else None,
            event_description=sanitize_event_description_html(event.event_description),
            event_location=event.event_location,
            event_banner_url=event.event_banner_url,
            min_sessions_required=int(event.min_sessions_required or 1),
            registration_closed=_is_event_registration_closed(event),
            visibility=_get_event_visibility(event),
            session_count=session_counts.get(event.id, 0),
            event_type=normalize_event_type(getattr(event, "event_type", None)),
            certificate_enabled=is_certificate_enabled(event),
            checkin_enabled=is_checkin_enabled(event),
            ticketing_enabled=is_ticketing_enabled(event),
            registration_enabled=is_public_registration_enabled(event),
            raffles_enabled=is_raffles_enabled(event),
            gamification_enabled=is_gamification_enabled(event),
            quiz_enabled=is_quiz_enabled(event),
            cpd_enabled=is_cpd_enabled(event),
        )
        for event in visible_events
    ]


@app.get("/api/public/events/{event_id}", response_model=PublicEventDetailOut)
async def get_public_event_detail(event_id: str, request: Request, db: AsyncSession = Depends(get_db)):
    event = await _resolve_public_event(db, event_id)
    if not event or _get_event_visibility(event) == "private":
        raise HTTPException(status_code=404, detail="Event not found")
    await _ensure_event_allowed_for_request_host(request, db, event)

    sessions_res = await db.execute(
        select(EventSession).where(EventSession.event_id == event.id).order_by(EventSession.session_date, EventSession.session_start)
    )
    survey_res = await db.execute(select(EventSurvey).where(EventSurvey.event_id == event.id))
    org_res = await db.execute(select(Organization).where(Organization.user_id == event.admin_id))

    from .quiz_models import Quiz as _Quiz
    quiz_count_res = await db.execute(
        select(func.count(_Quiz.id)).where(_Quiz.event_id == event.id, _Quiz.is_active == True)
    )
    has_active_quiz = (quiz_count_res.scalar_one() or 0) > 0

    return _build_public_event_detail(event, sessions_res.scalars().all(), survey_res.scalar_one_or_none(), org_res.scalar_one_or_none(), has_active_quiz=has_active_quiz)


@app.get("/api/public/events/{event_id}/comments", response_model=list[PublicEventCommentOut])
async def list_public_event_comments(event_id: str, request: Request, db: AsyncSession = Depends(get_db)):
    event = await _resolve_public_event(db, event_id)
    if not event or _get_event_visibility(event) == "private":
        raise HTTPException(status_code=404, detail="Event not found")
    await _ensure_event_allowed_for_request_host(request, db, event)

    comments_res = await db.execute(
        select(EventComment)
        .options(selectinload(EventComment.public_member))
        .where(EventComment.event_id == event.id, EventComment.status == "visible")
        .order_by(EventComment.created_at.desc())
    )
    return [_event_comment_to_out(comment) for comment in comments_res.scalars().all()]


@app.post("/api/public/events/{event_id}/comments", response_model=PublicEventCommentOut, status_code=201)
@limiter.limit("5/minute")
async def create_public_event_comment(
    request: Request,
    event_id: str,
    payload: PublicEventCommentCreateIn,
    member: CurrentPublicMember = Depends(get_current_public_member),
    db: AsyncSession = Depends(get_db),
):
    event = await _resolve_public_event(db, event_id)
    if not event or _get_event_visibility(event) == "private":
        raise HTTPException(status_code=404, detail="Event not found")
    await _ensure_event_allowed_for_request_host(request, db, event)
    body = moderate_public_text(payload.body)

    comment = EventComment(
        event_id=event.id,
        public_member_id=member.id,
        body=body,
        status="visible",
    )
    db.add(comment)
    await db.commit()
    await db.refresh(comment)
    await db.refresh(comment, attribute_names=["public_member"])

    await write_audit_log(
        db,
        user_id=None,
        action="public.comment.create",
        resource_type="event_comment",
        resource_id=str(comment.id),
        ip_address=_client_ip_for_rate_limit(request),
        user_agent=request.headers.get("User-Agent"),
        extra={"event_id": event.id, "public_member_id": member.id},
    )
    await db.commit()
    return _event_comment_to_out(comment)


@app.post("/api/public/events/{event_id}/comments/{comment_id}/report")
@limiter.limit("10/minute")
async def report_public_event_comment(
    request: Request,
    event_id: str,
    comment_id: int,
    member: CurrentPublicMember = Depends(get_current_public_member),
    db: AsyncSession = Depends(get_db),
):
    event = await _resolve_public_event(db, event_id)
    if not event or _get_event_visibility(event) == "private":
        raise HTTPException(status_code=404, detail="Event not found")
    await _ensure_event_allowed_for_request_host(request, db, event)
    comment_res = await db.execute(
        select(EventComment).where(EventComment.id == comment_id, EventComment.event_id == event.id)
    )
    comment = comment_res.scalar_one_or_none()
    if not comment:
        raise HTTPException(status_code=404, detail="Comment not found")

    comment.report_count += 1
    if comment.status == "visible":
        comment.status = "reported"
    await write_audit_log(
        db,
        user_id=None,
        action="public.comment.report",
        resource_type="event_comment",
        resource_id=str(comment.id),
        ip_address=_client_ip_for_rate_limit(request),
        user_agent=request.headers.get("User-Agent"),
        extra={"event_id": event.id, "public_member_id": member.id},
    )
    await db.commit()
    return {"ok": True, "status": comment.status, "report_count": comment.report_count}


@app.get("/api/events/{event_id}/info")
async def public_event_info(event_id: str, db: AsyncSession = Depends(get_db)):
    ev = await _resolve_public_event(db, event_id)
    if not ev:
        raise HTTPException(status_code=404, detail="Event not found")
    sess_res = await db.execute(
        select(EventSession).where(EventSession.event_id == ev.id).order_by(EventSession.session_date, EventSession.session_start)
    )
    survey_res = await db.execute(select(EventSurvey).where(EventSurvey.event_id == ev.id))
    return _build_public_event_detail(ev, sess_res.scalars().all(), survey_res.scalar_one_or_none()).model_dump()


@app.post("/api/legal/document-events")
@limiter.limit("60/minute")
async def log_legal_document_event(
    request: Request,
    payload: LegalDocumentEventIn,
    db: AsyncSession = Depends(get_db),
    member: Optional[CurrentPublicMember] = Depends(get_optional_public_member),
):
    event_db_id: Optional[int] = None
    event_public_id: Optional[str] = None
    if payload.event_id:
        ev = await _resolve_public_event(db, payload.event_id)
        if ev:
            event_db_id = ev.id
            event_public_id = _get_public_event_identifier(ev)

    await write_audit_log(
        db,
        user_id=None,
        action=f"legal.document.{payload.event_type}",
        resource_type="legal_document",
        resource_id=payload.document,
        ip_address=_client_ip_for_rate_limit(request),
        user_agent=request.headers.get("User-Agent"),
        extra={
            "document": payload.document,
            "event_id": event_db_id,
            "event_public_id": event_public_id,
            "public_member_id": member.id if member else None,
            "context": payload.context,
            "source_path": payload.source_path,
            "legal_document_version": LEGAL_DOCUMENT_VERSION,
        },
    )
    await db.commit()
    return {"ok": True}


async def _scan_upload_with_clamav(raw: bytes) -> None:
    if not settings.clamav_enabled:
        return
    try:
        reader, writer = await asyncio.open_connection(settings.clamav_host, settings.clamav_port)
        writer.write(b"zINSTREAM\0")
        await writer.drain()
        chunk_size = 1024 * 1024
        for idx in range(0, len(raw), chunk_size):
            chunk = raw[idx:idx + chunk_size]
            writer.write(len(chunk).to_bytes(4, "big") + chunk)
            await writer.drain()
        writer.write((0).to_bytes(4, "big"))
        await writer.drain()
        response = await asyncio.wait_for(reader.read(4096), timeout=15)
        writer.close()
        await writer.wait_closed()
        verdict = response.decode("utf-8", errors="ignore")
        if "FOUND" in verdict:
            raise HTTPException(status_code=400, detail="Uploaded file failed antivirus scan")
        if "OK" not in verdict:
            raise HTTPException(status_code=502, detail="Antivirus scan did not return a clean result")
    except HTTPException:
        raise
    except Exception as exc:
        logger.warning("ClamAV scan failed: %s", exc)
        if settings.require_clamav:
            raise HTTPException(status_code=503, detail="Antivirus scan is required but temporarily unavailable. Please try again.")
        logger.warning("ClamAV unavailable — REQUIRE_CLAMAV=false so upload proceeds without scan")


@app.post("/api/events/{event_id}/registration-document")
@limiter.limit("10/minute")
async def upload_public_registration_document(
    request: Request,
    event_id: str,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    ev = await _resolve_public_event(db, event_id)
    if not ev:
        raise HTTPException(status_code=404, detail="Event not found")
    if not is_public_registration_enabled(ev):
        raise HTTPException(status_code=403, detail="Registration is disabled for this event.")
    if _is_event_registration_closed(ev):
        raise HTTPException(status_code=403, detail="Registration is closed for this event.")

    allowed_content_types = {
        "application/pdf": ".pdf",
        "image/jpeg": ".jpg",
        "image/png": ".png",
        "image/webp": ".webp",
    }
    ctype = str(file.content_type or "").lower().strip()
    if ctype not in allowed_content_types:
        raise bad_request("Only PDF, JPG, PNG or WEBP files are allowed")

    raw = await file.read()
    if not raw:
        raise bad_request("Document file is empty")
    max_size = 2 * 1024 * 1024
    if len(raw) > max_size:
        raise HTTPException(status_code=413, detail="Document exceeds 2 MB limit")

    ext = allowed_content_types[ctype]
    if ctype == "application/pdf":
        if not raw.startswith(b"%PDF-"):
            raise bad_request("Invalid PDF file")
    else:
        signatures = {
            "image/jpeg": (b"\xff\xd8\xff",),
            "image/png": (b"\x89PNG\r\n\x1a\n",),
            "image/webp": (b"RIFF",),
        }
        if not any(raw.startswith(prefix) for prefix in signatures.get(ctype, ())):
            raise bad_request("Invalid image file")
        if ctype == "image/webp" and raw[8:12] != b"WEBP":
            raise bad_request("Invalid WEBP file")
        if PILImage is not None:
            try:
                image = PILImage.open(io.BytesIO(raw))
                image.verify()
            except Exception:
                raise bad_request("Invalid image file")
    await _scan_upload_with_clamav(raw)

    safe_name = f"registration_docs/event_{ev.id}/{secrets.token_hex(16)}{ext}"
    dest = Path(settings.local_storage_dir) / safe_name
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(raw)

    return {
        "path": safe_name,
        "name": Path(file.filename or f"document{ext}").name[:200],
        "content_type": ctype,
        "size_bytes": len(raw),
        "sha256": hashlib.sha256(raw).hexdigest(),
    }


@app.post("/api/events/{event_id}/register", status_code=201)
@limiter.limit("10/minute")
async def public_event_register(
    request: Request,
    event_id: str,
    payload: SelfRegisterIn,
    db: AsyncSession = Depends(get_db),
    member: Optional[CurrentPublicMember] = Depends(get_optional_public_member),
):
    ev = await _resolve_public_event(db, event_id)
    if not ev:
        raise HTTPException(status_code=404, detail="Event not found")
    if not is_public_registration_enabled(ev):
        raise HTTPException(status_code=403, detail="Registration is disabled for this event.")
    if _is_event_registration_closed(ev):
        raise HTTPException(status_code=403, detail="Registration is closed for this event.")
    if _is_event_kvkk_consent_required(ev) and not payload.kvkk_accepted:
        raise bad_request("KVKK consent is required for this event.")
    if _is_event_organizer_privacy_notice_enabled(ev) and not payload.organizer_notice_accepted:
        raise bad_request("Organizer privacy notice acceptaonce is required for this event.")
    if _is_event_cross_border_transfer_notice_enabled(ev) and not payload.cross_border_notice_read:
        raise bad_request("Cross-border transfer notice must be acknowledged for this event.")
    if _is_event_cross_border_transfer_consent_required(ev) and not payload.cross_border_transfer_consent:
        raise bad_request("Cross-border transfer consent is required for this event.")
    registration_quota = _get_event_registration_quota(ev)
    quota_enabled = _is_event_registration_quota_enabled(ev)
    if quota_enabled and registration_quota is not None:
        attendee_count_res = await db.execute(
            select(func.count()).where(Attendee.event_id == ev.id)
        )
        attendee_count = int(attendee_count_res.scalar_one() or 0)
        if attendee_count >= registration_quota:
            next_config = dict(ev.config or {})
            next_config["registration_closed"] = True
            ev.config = next_config
            await db.commit()
            raise HTTPException(status_code=403, detail="Registration quota reached for this event.")
    event_db_id = ev.id
    require_email_verification = _get_event_email_verification_required(ev)

    normalized_email = payload.email.lower()
    if member and normalized_email != member.email.lower():
        raise bad_request("Signed-in members must register with their own email address.")

    client_ip = _client_ip_for_rate_limit(request)
    user_agent = request.headers.get("User-Agent")
    device_id, should_set_device_cookie = _get_registration_device_id(request)
    registration_fields = _get_event_registration_fields(ev)
    file_fields = [field for field in registration_fields if field.get("type") == "file"]
    file_field_ids = {str(field.get("id")) for field in file_fields if field.get("id")}
    registration_answers = _normalize_registration_answers(
        registration_fields,
        payload.registration_answers,
    )
    reservations_to_attempt = _collect_registration_option_reservations(ev, registration_answers)

    # Per-option capacity is only consumed immediately when no email verification is required.
    # Otherwise it is consumed oonce the attendee verifies their email.
    if not require_email_verification:
        for (ev_id, fid, label, cap) in reservations_to_attempt:
            ok = await _reserve_option_capacity(db, ev_id, fid, label, cap)
            if not ok:
                raise bad_request(f'"{fid}" alanındaki "{label}" seeneği icin kontenjan doldu.')
    required_file_field_ids = {
        str(field.get("id"))
        for field in file_fields
        if field.get("id") and (
            bool(field.get("required"))
            or _is_registration_field_condition_met(field, registration_answers)
        )
    }
    if payload.registration_documents:
        if not file_field_ids:
            raise bad_request("Document upload is not enabled for this event")
        sanitized_docs: List[Dict[str, Any]] = []
        storage_root = Path(settings.local_storage_dir).resolve()
        expected_prefix = f"registration_docs/event_{event_db_id}/"
        for doc in payload.registration_documents[:10]:
            field_id = str((doc or {}).get("field_id") or "").strip()
            if not field_id or field_id not in file_field_ids:
                raise bad_request("Invalid registration document field")
            rel_path = str((doc or {}).get("path") or "").strip().lstrip("/")
            if not rel_path or not rel_path.startswith(expected_prefix):
                raise bad_request("Invalid registration document path")
            abs_path = (storage_root / rel_path).resolve()
            if not abs_path.is_relative_to(storage_root) or not abs_path.exists() or not abs_path.is_file():
                raise bad_request("Registration document not found")
            stat_info = abs_path.stat()
            if stat_info.st_size > 2 * 1024 * 1024:
                raise bad_request("Registration document exceeds size limit")
            sanitized_docs.append(
                {
                    "field_id": field_id,
                    "path": rel_path,
                    "name": str((doc or {}).get("name") or Path(rel_path).name)[:200],
                    "content_type": str((doc or {}).get("content_type") or "application/octet-stream")[:120],
                    "size_bytes": int((doc or {}).get("size_bytes") or stat_info.st_size),
                    "sha256": str((doc or {}).get("sha256") or "")[:128],
                    "uploaded_at": datetime.now(timezone.utc).isoformat(),
                }
            )
        if sanitized_docs:
            uploaded_field_ids = {
                str(doc.get("field_id"))
                for doc in sanitized_docs
                if str(doc.get("field_id") or "").strip()
            }
            missing_required_fields = required_file_field_ids - uploaded_field_ids
            if missing_required_fields:
                raise bad_request("Required document fields are missing")
            registration_answers["__documents"] = sanitized_docs
    elif required_file_field_ids:
        raise bad_request("Required document fields are missing")
    registration_answers["__kvkk"] = {
        "accepted": bool(payload.kvkk_accepted),
        "accepted_at": datetime.now(timezone.utc).isoformat(),
        "ip_address": client_ip,
        "user_agent": user_agent,
        "organizer_notice_accepted": bool(payload.organizer_notice_accepted),
        "cross_border_notice_read": bool(payload.cross_border_notice_read),
        "cross_border_transfer_consent": bool(payload.cross_border_transfer_consent),
    }

    await _enforce_registration_risk_controls(
        db,
        event_id=event_db_id,
        email=normalized_email,
        ip_address=client_ip,
        user_agent=user_agent,
        device_id=device_id,
    )

    existing = await db.execute(
        select(Attendee).where(Attendee.event_id == event_db_id, func.lower(Attendee.email) == normalized_email)
    )
    existing_attendee = existing.scalar_one_or_none()
    if existing_attendee:
        existing_attendee.name = payload.name
        existing_attendee.registration_answers = registration_answers
        if member and not existing_attendee.public_member_id:
            existing_attendee.public_member_id = member.id

        if not existing_attendee.email_verified and require_email_verification:
            existing_attendee.email_verification_token = make_email_token(
                {
                    "action": "attendee_verify",
                    "attendee_id": existing_attendee.id,
                    "event_id": event_db_id,
                    "email": existing_attendee.email.lower(),
                }
            )
            await _write_attendee_consent_audit(
                db,
                event=ev,
                attendee=existing_attendee,
                payload=payload,
                member=member,
                client_ip=client_ip,
                user_agent=user_agent,
                device_id=device_id,
                result="existing_unverified",
            )
            await write_audit_log(
                db,
                user_id=None,
                action="attendee.register",
                resource_type="attendee",
                resource_id=str(existing_attendee.id),
                ip_address=client_ip,
                user_agent=user_agent,
                extra={
                    "event_id": event_id,
                    "event_public_id": _get_public_event_identifier(ev),
                    "email": normalized_email,
                    "device_id": device_id,
                    "public_member_id": member.id if member else None,
                    "result": "existing_unverified",
                    "verification_required": True,
                },
            )
            await db.commit()
            await trigger_webhooks(
                ev.admin_id,
                "attendee.register",
                _build_attendee_register_webhook_payload(ev, existing_attendee, result="existing_unverified"),
            )
            await _sync_google_sheet_if_enabled(db, ev)
            await _sync_ms365_excel_if_enabled(db, ev)
            await send_attendee_verification_email(attendee=existing_attendee, event=ev)
            response = JSONResponse(
                status_code=200,
                content={
                    "ok": True,
                    "already_registered": True,
                    "email_verified": False,
                    "verification_required": True,
                    "message": "Bu e-posta ile kayıt bulundu. Devam etmek için doğrulama e-postasını onaylayın.",
                    "attendee_id": existing_attendee.id,
                    "attendee_name": existing_attendee.name,
                    "attendee_email": existing_attendee.email,
                },
            )
            if should_set_device_cookie:
                response.set_cookie(
                    key=REGISTRATION_DEVICE_COOKIE,
                    value=device_id,
                    max_age=60 * 60 * 24 * 365,
                    httponly=True,
                    samesite="Lax",
                )
            return response

        if not existing_attendee.email_verified and not require_email_verification:
            existing_attendee.email_verified = True
            existing_attendee.email_verification_token = None
            existing_attendee.email_verified_at = datetime.now(timezone.utc)

        survey_token = make_survey_access_token(
            attendee_id=existing_attendee.id,
            event_id=event_db_id,
            email=existing_attendee.email,
        )
        ticket = await _issue_event_ticket_if_needed(db, ev, existing_attendee)
        consent_result = "existing_verified" if existing_attendee.email_verified else "existing_auto_verified"
        await _write_attendee_consent_audit(
            db,
            event=ev,
            attendee=existing_attendee,
            payload=payload,
            member=member,
            client_ip=client_ip,
            user_agent=user_agent,
            device_id=device_id,
            result=consent_result,
        )
        await write_audit_log(
            db,
            user_id=None,
            action="attendee.register",
            resource_type="attendee",
            resource_id=str(existing_attendee.id),
            ip_address=client_ip,
            user_agent=user_agent,
            extra={
                "event_id": event_db_id,
                "email": normalized_email,
                "device_id": device_id,
                "public_member_id": member.id if member else None,
                "result": "existing_verified" if existing_attendee.email_verified else "existing_auto_verified",
                "verification_required": require_email_verification,
            },
        )
        await db.commit()
        await trigger_webhooks(
            ev.admin_id,
            "attendee.register",
            _build_attendee_register_webhook_payload(
                ev,
                existing_attendee,
                result="existing_verified" if existing_attendee.email_verified else "existing_auto_verified",
            ),
        )
        await _sync_google_sheet_if_enabled(db, ev)
        await _sync_ms365_excel_if_enabled(db, ev)
        response = JSONResponse(
            status_code=200,
            content={
                "ok": True,
                "already_registered": True,
                "email_verified": bool(existing_attendee.email_verified),
                "verification_required": require_email_verification,
                "message": "Bu e-posta ile zaten kayitlisiniz.",
                "attendee_id": existing_attendee.id,
                "attendee_name": existing_attendee.name,
                "attendee_email": existing_attendee.email,
                "survey_token": survey_token,
                "survey_url": build_public_survey_url(
                    event_id=_get_public_event_identifier(ev),
                    attendee_id=existing_attendee.id,
                    email=existing_attendee.email,
                ),
                "status_url": build_public_status_url(
                    event_id=_get_public_event_identifier(ev),
                    attendee_id=existing_attendee.id,
                    email=existing_attendee.email,
                ),
                "ticket": _ticket_response_payload(ticket),
            },
        )
        if should_set_device_cookie:
            response.set_cookie(
                key=REGISTRATION_DEVICE_COOKIE,
                value=device_id,
                max_age=60 * 60 * 24 * 365,
                httponly=True,
                samesite="Lax",
            )
        return response

    attendee = Attendee(
        event_id=event_db_id,
        name=payload.name,
        email=normalized_email,
        source="self_register",
        email_verified=not require_email_verification,
        public_member_id=member.id if member else None,
        registration_answers=registration_answers,
    )
    db.add(attendee)
    await db.flush()
    if require_email_verification:
        attendee.email_verification_token = make_email_token(
            {
                "action": "attendee_verify",
                "attendee_id": attendee.id,
                "event_id": event_db_id,
                "email": attendee.email.lower(),
            }
        )
    else:
        attendee.email_verification_token = None
        attendee.email_verified_at = datetime.now(timezone.utc)
        await _issue_event_ticket_if_needed(db, ev, attendee)
    from .crm_snapshot_hooks import refresh_crm_snapshot_for_attendee
    await refresh_crm_snapshot_for_attendee(db, attendee)
    consent_result = "created_unverified" if require_email_verification else "created_verified"
    await _write_attendee_consent_audit(
        db,
        event=ev,
        attendee=attendee,
        payload=payload,
        member=member,
        client_ip=client_ip,
        user_agent=user_agent,
        device_id=device_id,
        result=consent_result,
    )
    await write_audit_log(
        db,
        user_id=None,
        action="attendee.register",
        resource_type="attendee",
        resource_id=str(attendee.id),
        ip_address=client_ip,
        user_agent=user_agent,
        extra={
            "event_id": event_db_id,
            "email": normalized_email,
            "device_id": device_id,
            "public_member_id": member.id if member else None,
            "result": "created_unverified" if require_email_verification else "created_verified",
            "verification_required": require_email_verification,
        },
    )
    if quota_enabled and registration_quota is not None and not require_email_verification:
        attendee_count_res = await db.execute(
            select(func.count()).where(Attendee.event_id == event_db_id)
        )
        attendee_count = int(attendee_count_res.scalar_one() or 0)
        if attendee_count >= registration_quota:
            next_config = dict(ev.config or {})
            next_config["registration_closed"] = True
            ev.config = next_config
    await db.commit()
    await trigger_webhooks(
        ev.admin_id,
        "attendee.register",
        _build_attendee_register_webhook_payload(
            ev,
            attendee,
            result="created_unverified" if require_email_verification else "created_verified",
        ),
    )
    await _append_attendee_to_google_sheet_if_enabled(db, ev, attendee)
    await _sync_ms365_excel_if_enabled(db, ev)
    ticket_res = await db.execute(
        select(EventTicket).where(EventTicket.event_id == event_db_id, EventTicket.attendee_id == attendee.id)
    )
    issued_ticket = ticket_res.scalar_one_or_none()
    if require_email_verification:
        await send_attendee_verification_email(attendee=attendee, event=ev)
    survey_token = make_survey_access_token(
        attendee_id=attendee.id,
        event_id=event_db_id,
        email=attendee.email,
    )
    response = JSONResponse(
        status_code=201,
        content={
            "ok": True,
            "already_registered": False,
            "email_verified": bool(attendee.email_verified),
            "verification_required": require_email_verification,
            "message": "Kayit alindi. Etkinlige katilimi tamamlamak icin e-posta dogrulamasi gerekiyor." if require_email_verification else "Kayit alindi.",
            "attendee_id": attendee.id,
            "attendee_name": attendee.name,
            "attendee_email": attendee.email,
            "survey_token": survey_token if attendee.email_verified else None,
            "survey_url": build_public_survey_url(
                event_id=_get_public_event_identifier(ev),
                attendee_id=attendee.id,
                email=attendee.email,
            ) if attendee.email_verified else None,
            "status_url": build_public_status_url(
                event_id=_get_public_event_identifier(ev),
                attendee_id=attendee.id,
                email=attendee.email,
            ) if attendee.email_verified else None,
            "ticket": _ticket_response_payload(issued_ticket) if attendee.email_verified else None,
        },
    )
    if should_set_device_cookie:
        response.set_cookie(
            key=REGISTRATION_DEVICE_COOKIE,
            value=device_id,
            max_age=60 * 60 * 24 * 365,
            httponly=True,
            samesite="Lax",
        )
    return response


@app.get("/api/events/{event_id}/verify-email")
async def verify_attendee_email(event_id: str, token: str = Query(...), db: AsyncSession = Depends(get_db)):
    event = await _resolve_public_event(db, event_id)
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    try:
        payload = verify_email_token(token, max_age=86400)
    except SignatureExpired:
        raise bad_request("Dogrulama baglantisinin suresi dolmus.")
    except (BadSignature, Exception):
        raise bad_request("Geçersiz doğrulama bağlantısı.")

    if payload.get("action") != "attendee_verify":
        raise bad_request("Gecersiz token turu.")
    if int(payload.get("event_id") or 0) != event.id:
        raise bad_request("Etkinlik doğrulama bilgisi eşleşmiyor.")

    attendee_id = int(payload.get("attendee_id") or 0)
    email = str(payload.get("email") or "").lower()
    res = await db.execute(
        select(Attendee)
        .where(Attendee.id == attendee_id, Attendee.event_id == event.id)
        .with_for_update()
    )
    attendee = res.scalar_one_or_none()
    if not attendee or attendee.email.lower() != email:
        raise HTTPException(status_code=404, detail="Katılımcı bulunamadı.")

    if attendee.email_verified:
        return {
            "detail": "E-posta zaten dogrulanmis.",
            "attendee_id": attendee.id,
            "event_id": event.id,
            "status_url": build_public_status_url(event_id=_get_public_event_identifier(event), attendee_id=attendee.id, email=attendee.email),
        }

    registration_quota = _get_event_registration_quota(event)
    quota_enabled = _is_event_registration_quota_enabled(event)
    if quota_enabled and registration_quota is not None:
        verified_count_res = await db.execute(
            select(func.count()).where(Attendee.event_id == event.id, Attendee.email_verified.is_(True))
        )
        verified_count = int(verified_count_res.scalar_one() or 0)
        if verified_count >= registration_quota:
            next_config = dict(event.config or {})
            next_config["registration_closed"] = True
            event.config = next_config
            await db.commit()
            raise HTTPException(status_code=403, detail="Registration quota reached for this event.")

    if _get_event_email_verification_required(event):
        reservation_requests = _collect_registration_option_reservations(event, attendee.registration_answers or {})
        for (ev_id, fid, label, cap) in reservation_requests:
            ok = await _reserve_option_capacity(db, ev_id, fid, label, cap)
            if not ok:
                raise bad_request(f'"{fid}" alanındaki "{label}" seeneği icin kontenjan doldu.')

    attendee.email_verified = True
    attendee.email_verification_token = None
    attendee.email_verified_at = datetime.now(timezone.utc)

    if quota_enabled and registration_quota is not None:
        verified_count_res = await db.execute(
            select(func.count()).where(Attendee.event_id == event.id, Attendee.email_verified.is_(True))
        )
        verified_count = int(verified_count_res.scalar_one() or 0)
        if verified_count + 1 >= registration_quota:
            next_config = dict(event.config or {})
            next_config["registration_closed"] = True
            event.config = next_config

    await _issue_event_ticket_if_needed(db, event, attendee)
    await db.commit()
    await _sync_google_sheet_if_enabled(db, event)
    await _sync_ms365_excel_if_enabled(db, event)

    return {
        "detail": "E-posta dogrulandi.",
        "attendee_id": attendee.id,
        "event_id": event.id,
        "status_url": build_public_status_url(event_id=_get_public_event_identifier(event), attendee_id=attendee.id, email=attendee.email),
    }


# Ã¢â€â‚¬Ã¢â€â‚¬ Public: QR check-in Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬
@app.post("/api/events/{event_id}/resend-verification")
@limiter.limit("3/hour")
async def resend_attendee_verification_email(
    request: Request,
    event_id: str,
    data: ForgotPasswordIn,
    db: AsyncSession = Depends(get_db),
):
    event = await _resolve_public_event(db, event_id)
    if not event:
        raise HTTPException(status_code=404, detail="Etkinlik bulunamadı.")

    email = str(data.email).strip().lower()
    res = await db.execute(
        select(Attendee).where(
            Attendee.event_id == event.id,
            func.lower(func.trim(Attendee.email)) == email,
        )
    )
    attendee = res.scalar_one_or_none()
    if attendee and not attendee.email_verified:
        attendee.email_verification_token = make_email_token(
            {
                "action": "attendee_verify",
                "attendee_id": attendee.id,
                "event_id": event.id,
                "email": attendee.email.lower(),
            }
        )
        db.add(attendee)
        await db.commit()
        await send_attendee_verification_email(attendee=attendee, event=event)

    return {"detail": "Eğer doğrulanmamış bir kayıt varsa doğrulama e-postası yeniden gönderildi."}


API_AUDIT_SKIP_PREFIXES_EXTENDED = ("/api/attend/", "/api/events/")

@app.get("/api/attend/{checkin_token}")
async def get_session_by_token(checkin_token: str, db: AsyncSession = Depends(get_db)):
    ctx = await _get_checkin_context_by_token(checkin_token, db)
    if not ctx:
        raise HTTPException(status_code=404, detail="Gecersiz QR kodu")
    count_res = await db.execute(
        select(func.count()).where(AttendaonceRecord.session_id == ctx["session_id"])
    )
    count = int(count_res.scalar_one() or 0)
    return {
        "session_id": ctx["session_id"],
        "session_name": ctx["session_name"],
        "session_date": ctx["session_date"].isoformat() if ctx["session_date"] else None,
        "session_start": ctx["session_start"].strftime("%H:%M") if ctx["session_start"] else None,
        "session_location": ctx["session_location"],
        "is_active": ctx["is_active"],
        "event_id": ctx["event_id"],
        "event_public_id": ctx["event_public_id"],
        "event_name": ctx["event_name"],
        "event_date": ctx["event_date"].isoformat() if ctx["event_date"] else None,
        "min_sessions_required": ctx["min_sessions_required"],
        "attendaonce_count": count,
    }


@app.post("/api/attend/{checkin_token}", response_model=CheckinOut)
@limiter.limit("600/minute")
async def self_checkin(
    checkin_token: str,
    payload: CheckinIn,
    request: Request,
    db: AsyncSession = Depends(get_db),
):
    from .checkin_ops_api import record_checkin_activity

    ctx = await _get_checkin_context_by_token(checkin_token, db)
    if not ctx:
        raise HTTPException(status_code=404, detail="Gecersiz QR kodu")
    if not ctx.get("checkin_enabled", True):
        raise HTTPException(status_code=403, detail="Check-in is disabled for this event.")
    if not ctx["is_active"]:
        raise HTTPException(status_code=403, detail="Bu oturum icin check-in kapalı")

    att_res = await db.execute(
        select(Attendee).where(
            Attendee.event_id == ctx["event_id"],
            func.lower(Attendee.email) == payload.email.lower(),
        )
    )
    attendee = att_res.scalar_one_or_none()
    event_res = await db.execute(select(Event).where(Event.id == ctx["event_id"]))
    event = event_res.scalar_one_or_none()
    require_email_verification = _get_event_email_verification_required(event) if event else True
    if attendee and require_email_verification and not attendee.email_verified:
        raise HTTPException(status_code=403, detail="Check-in icin oonce e-posta dogrulamasi yapmalisiniz.")
    if not attendee:
        raise HTTPException(
            status_code=404,
            detail="Bu e-posta ile etkinlikte kayıtlı değilsiniz. Lutfen once kayıt olun.",
        )

    ip = _client_ip_for_rate_limit(request)
    insert_stmt = (
        _pg_insert(AttendaonceRecord.__table__)
        .values(
            attendee_id=attendee.id,
            session_id=ctx["session_id"],
            ip_address=ip,
        )
        .on_conflict_do_nothing(index_elements=["attendee_id", "session_id"])
        .returning(AttendaonceRecord.id)
    )
    inserted_res = await db.execute(insert_stmt)
    inserted_id = inserted_res.scalar_one_or_none()
    if inserted_id is not None:
        from .crm_snapshot_hooks import refresh_crm_snapshot_for_attendee
        await refresh_crm_snapshot_for_attendee(db, attendee)
    await record_checkin_activity(
        db,
        event_id=ctx["event_id"],
        session_id=ctx["session_id"],
        attendee_id=attendee.id,
        actor_user_id=None,
        method="self",
        source="public",
        success=inserted_id is not None,
        message="Self check-in successful" if inserted_id is not None else "Duplicate self check-in",
        ip_address=ip,
    )
    await db.commit()

    if inserted_id is None:
        existing_res = await db.execute(
            select(AttendaonceRecord.id, AttendaonceRecord.checked_in_at).where(
                AttendaonceRecord.attendee_id == attendee.id,
                AttendaonceRecord.session_id == session_id,
            )
        )
        existing = existing_res.first()
        return {
            "ok": False,
            "duplicate": True,
            "record_id": existing.id if existing else None,
            "checked_in_at": existing.checked_in_at.isoformat() if existing and existing.checked_in_at else None,
            "attendee_id": attendee.id,
            "attendee_name": attendee.name,
            "attendee_email": attendee.email,
            "session_id": session.id,
            "session_name": session.name,
            "message": "Bu katılımcı bu oturum için zaten check-in yapmış.",
        }

    return {
        "ok": True,
        "duplicate": False,
        "record_id": inserted_id,
        "attendee_id": attendee.id,
        "attendee_name": attendee.name,
        "attendee_email": attendee.email,
        "session_id": session.id,
        "session_name": session.name,
        "message": f"Check-in başarılı: {attendee.name}",
    }

    attended_res = await db.execute(
        select(func.count()).where(AttendaonceRecord.attendee_id == attendee.id)
    )
    attended_count = int(attended_res.scalar_one() or 0)
    total_sessions = await _get_event_total_sessions_cached(ctx["event_id"], db)

    if inserted_id is None:
        return CheckinOut(
            success=False,
            message="Bu oturuma zaten check-in yaptınız.",
            attendee_name=attendee.name,
            sessions_attended=attended_count,
            sessions_required=ctx["min_sessions_required"],
            total_sessions=total_sessions,
        )

    return CheckinOut(
        success=True,
        message=f"Check-in başarılı! Hoş geldiniz, {attendee.name}.",
        attendee_name=attendee.name,
        sessions_attended=attended_count,
        sessions_required=ctx["min_sessions_required"],
        total_sessions=total_sessions,
    )


# ============================================================================
# SUPPORT TICKETS - AI Assistant Escalation
# ============================================================================

async def send_ticket_notification_email(ticket: SupportTicket, user_email: str):
    """Send email notification to superadmins when a support ticket is created"""
    try:
        async with SessionLocal() as notification_db:
            superadmins_res = await notification_db.execute(
                select(User).where(User.role == Role.superadmin)
            )
            superadmins = superadmins_res.scalars().all()

            org_res = await notification_db.execute(
                select(Organization).where(Organization.id == ticket.organization_id)
            )
            org = org_res.scalar_one_or_none()
        
        for superadmin in superadmins:
            # Create email
            msg = MIMEMultipart("alternative")
            msg["Subject"] = str(EmailHeader(f"Yeni Destek Talebineği: {ticket.subject}", charset="utf-8"))
            msg["From"] = formataddr(("HeptaCert Sistem", settings.smtp_from_email))
            msg["To"] = superadmin.email
            
            html = f"""
            <html>
                <body style="font-family: Arial, sans-serif;">
                    <h2>Yeni Destek Talebi</h2>
                    <p><strong>Organizasyon:</strong> {escape(org.org_name) if org else 'Bilinmiyor'}</p>
                    <p><strong>Kullanıcı:</strong> {escape(user_email)}</p>
                    <p><strong>Konu:</strong> {escape(ticket.subject)}</p>
                    <p><strong>Durum:</strong> {escape(ticket.status)}</p>
                    <hr>
                    <p><a href="{escape(settings.frontend_base_url)}/admin/superadmin/support-tickets/{ticket.id}">Talebi Görüntüle</a></p>
                </body>
            </html>
            """
            msg.attach(MIMEText(html, "html", "utf-8"))
            
            # Send email (non-blocking)
            try:
                await aiosmtplib.send_message(
                    msg,
                    hostname=settings.smtp_server,
                    port=settings.smtp_port,
                    use_tls=True,
                    username=settings.smtp_username,
                    password=settings.smtp_password,
                    timeout=10,
                )
            except Exception as e:
                logger.error(f"Failed to send support ticket email: {e}")
    except Exception as e:
        logger.error(f"Error in send_ticket_notification_email: {e}")


@app.post("/api/admin/support-tickets", response_model=SupportTicketOut)
async def create_support_ticket(
    req: SupportTicketCreateIn,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Create a support ticket - Called when AI assistant can't help"""
    # User must be an admin/superadmin
    if me.role not in (Role.admin, Role.superadmin):
        raise HTTPException(status_code=403, detail="Only admins can create support tickets")
    
    # New admins can need support before they have visited organization settings.
    org = await _get_or_create_admin_organization(db, me.id)
    
    # Create ticket
    ticket = SupportTicket(
        organization_id=org.id,
        user_id=me.id,
        subject=req.subject,
        messages=[{
            "role": "user",
            "message": req.message,
            "timestamp": datetime.now(timezone.utc).isoformat()
        }],
        status="open"
    )
    
    db.add(ticket)
    await db.commit()
    await db.refresh(ticket)
    
    # Send notification email to superadmins (non-blocking)
    asyncio.create_task(send_ticket_notification_email(ticket, me.email))
    
    return ticket


@app.get("/api/superadmin/support-tickets", response_model=list[SupportTicketOut], dependencies=[Depends(require_role(Role.superadmin))])
async def list_support_tickets(
    status: Optional[str] = Query(default=None),
    db: AsyncSession = Depends(get_db),
):
    """List all support tickets (superadmin only)"""
    query = select(SupportTicket).order_by(SupportTicket.created_at.desc())
    
    if status:
        query = query.where(SupportTicket.status == status)
    
    res = await db.execute(query)
    tickets = res.scalars().all()
    return tickets


@app.get("/api/superadmin/support-tickets/{ticket_id}", response_model=SupportTicketOut, dependencies=[Depends(require_role(Role.superadmin))])
async def get_support_ticket(
    ticket_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Get a single support ticket"""
    ticket_res = await db.execute(
        select(SupportTicket).where(SupportTicket.id == ticket_id)
    )
    ticket = ticket_res.scalar_one_or_none()
    if not ticket:
        raise HTTPException(status_code=404, detail="Support ticket not found")
    
    return ticket


@app.patch("/api/superadmin/support-tickets/{ticket_id}", response_model=SupportTicketOut, dependencies=[Depends(require_role(Role.superadmin))])
async def update_support_ticket(
    ticket_id: int,
    req: SupportTicketUpdateIn,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Update support ticket status or add admin reply"""
    ticket_res = await db.execute(
        select(SupportTicket).where(SupportTicket.id == ticket_id)
    )
    ticket = ticket_res.scalar_one_or_none()
    if not ticket:
        raise HTTPException(status_code=404, detail="Support ticket not found")
    
    # Update status if provided
    if req.status:
        ticket.status = req.status
    
    # Add admin reply if provided
    if req.admin_reply:
        ticket.messages.append({
            "role": "admin",
            "message": req.admin_reply,
            "timestamp": datetime.now(timezone.utc).isoformat()
        })
    
    ticket.updated_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(ticket)
    
    return ticket


# Ã¢â€â‚¬Ã¢â€â‚¬ Admin: Sessions Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬

@app.get("/api/admin/events/{event_id}/sessions", dependencies=[Depends(require_role(Role.admin, Role.superadmin)), Depends(require_paid_plan)])
async def list_sessions(
    event_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    ev = await _get_event_for_admin(event_id, me, db, "checkin:write")
    _ensure_checkin_feature_enabled(ev)
    res = await db.execute(
        select(EventSession).where(EventSession.event_id == event_id).order_by(EventSession.session_date, EventSession.session_start, EventSession.id)
    )
    sessions = res.scalars().all()
    results = []
    for s in sessions:
        cnt_res = await db.execute(select(func.count()).where(AttendaonceRecord.session_id == s.id))
        cnt = int(cnt_res.scalar_one() or 0)
        results.append(_session_to_out(s, cnt))
    return results


@app.post("/api/admin/events/{event_id}/sessions", status_code=201, dependencies=[Depends(require_role(Role.admin, Role.superadmin)), Depends(require_paid_plan)])
async def create_session(
    event_id: int,
    payload: SessionCreateIn,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    ev = await _get_event_for_admin(event_id, me, db, "checkin:write")
    _ensure_checkin_feature_enabled(ev)
    from datetime import date as _date, time as _time
    sd = _date.fromisoformat(payload.session_date) if payload.session_date else None
    st = None
    if payload.session_start:
        parts = payload.session_start.split(":")
        st = _time(int(parts[0]), int(parts[1]))
    session = EventSession(
        event_id=event_id,
        name=payload.name,
        session_date=sd,
        session_start=st,
        session_location=payload.session_location,
        checkin_token=str(_uuid_module.uuid4()).replace("-", ""),
        is_active=False,
    )
    db.add(session)
    await db.commit()
    return _session_to_out(session, 0)


@app.patch("/api/admin/events/{event_id}/sessions/{session_id}", dependencies=[Depends(require_role(Role.admin, Role.superadmin)), Depends(require_paid_plan)])
async def update_session(
    event_id: int,
    session_id: int,
    payload: SessionCreateIn,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    ev = await _get_event_for_admin(event_id, me, db, "checkin:write")
    _ensure_checkin_feature_enabled(ev)
    res = await db.execute(select(EventSession).where(EventSession.id == session_id, EventSession.event_id == event_id))
    session = res.scalar_one_or_none()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    from datetime import date as _date, time as _time
    session.name = payload.name
    if payload.session_date is not None:
        session.session_date = _date.fromisoformat(payload.session_date) if payload.session_date else None
    if payload.session_start is not None:
        parts = payload.session_start.split(":")
        session.session_start = _time(int(parts[0]), int(parts[1])) if payload.session_start else None
    if payload.session_location is not None:
        session.session_location = payload.session_location
    await db.commit()
    cnt_res = await db.execute(select(func.count()).where(AttendaonceRecord.session_id == session.id))
    cnt = int(cnt_res.scalar_one() or 0)
    return _session_to_out(session, cnt)


@app.delete("/api/admin/events/{event_id}/sessions/{session_id}", dependencies=[Depends(require_role(Role.admin, Role.superadmin)), Depends(require_paid_plan)])
async def delete_session(
    event_id: int,
    session_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    ev = await _get_event_for_admin(event_id, me, db, "checkin:write")
    _ensure_checkin_feature_enabled(ev)
    res = await db.execute(select(EventSession).where(EventSession.id == session_id, EventSession.event_id == event_id))
    session = res.scalar_one_or_none()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    await db.delete(session)
    await db.commit()
    return {"ok": True}


@app.patch("/api/admin/events/{event_id}/sessions/{session_id}/toggle", dependencies=[Depends(require_role(Role.admin, Role.superadmin)), Depends(require_paid_plan)])
async def toggle_session_checkin(
    event_id: int,
    session_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    ev = await _get_event_for_admin(event_id, me, db, "checkin:write")
    _ensure_checkin_feature_enabled(ev)
    res = await db.execute(select(EventSession).where(EventSession.id == session_id, EventSession.event_id == event_id))
    session = res.scalar_one_or_none()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    session.is_active = not session.is_active
    await db.commit()
    cnt_res = await db.execute(select(func.count()).where(AttendaonceRecord.session_id == session.id))
    cnt = int(cnt_res.scalar_one() or 0)
    return _session_to_out(session, cnt)


@app.get(
    "/api/admin/events/{event_id}/sessions/{session_id}/qr",
    dependencies=[Depends(require_role(Role.admin, Role.superadmin)), Depends(require_paid_plan)],
)
async def get_session_qr(
    event_id: int,
    session_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    ev = await _get_event_for_admin(event_id, me, db, "checkin:write")
    _ensure_checkin_feature_enabled(ev)

    res = await db.execute(
        select(EventSession).where(
            EventSession.id == session_id,
            EventSession.event_id == event_id,
        )
    )
    session = res.scalar_one_or_none()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    # Event sahibinin organization/domain bilgisini Ã§ek
    org_res = await db.execute(
        select(Organization).where(Organization.user_id == ev.admin_id)
    )
    org = org_res.scalar_one_or_none()

    # Ã–oncelik: organization custom domain
    if org and org.custom_domain:
        checkin_url = f"https://{org.custom_domain}/attend/{session.checkin_token}"
    else:
        checkin_url = f"{settings.frontend_base_url.rstrip('/')}/attend/{session.checkin_token}"

    qr = qrcode.QRCode(
        version=2,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=4,
    )
    qr.add_data(checkin_url)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)

    from fastapi.responses import Response
    return Response(
        content=buf.getvalue(),
        media_type="image/png",
        headers={"X-Checkin-URL": checkin_url},
    )
# Ã¢â€â‚¬Ã¢â€â‚¬ Admin: Attendees Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬

@app.get("/api/admin/events/{event_id}/attendees", dependencies=[Depends(require_role(Role.admin, Role.superadmin)), Depends(require_paid_plan)])
async def list_attendees(
    event_id: int,
    page: int = Query(default=1, ge=1),
    limit: int = Query(default=50, le=500),
    search: str = Query(default=""),
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    ev = await _get_event_for_admin(event_id, me, db, "attendees:read")
    q = select(Attendee).options(selectinload(Attendee.public_member)).where(Attendee.event_id == event_id)
    if search:
        like = f"%{search.lower()}%"
        from sqlalchemy import or_
        q = q.where(or_(func.lower(Attendee.name).like(like), func.lower(Attendee.email).like(like)))
    total_res = await db.execute(select(func.count()).select_from(q.subquery()))
    total = int(total_res.scalar_one() or 0)
    q = q.order_by(Attendee.registered_at.desc()).offset((page - 1) * limit).limit(limit)
    res = await db.execute(q)
    attendees = res.scalars().all()

    # Get session counts per attendee
    results = []
    for a in attendees:
        cnt_res = await db.execute(select(func.count()).where(AttendaonceRecord.attendee_id == a.id))
        cnt = int(cnt_res.scalar_one() or 0)
        # Check if has certificate
        cert_res = await db.execute(
            select(Certificate).where(
                Certificate.event_id == event_id,
                Certificate.student_name == a.name,
                Certificate.deleted_at.is_(None),
            )
        )
        has_cert = cert_res.scalar_one_or_none() is not None
        results.append(AttendeeOut(
            id=a.id,
            event_id=a.event_id,
            name=a.name,
            email=a.email,
            source=a.source,
            registered_at=a.registered_at,
            sessions_attended=cnt,
            has_certificate=has_cert,
            public_member_id=a.public_member_id,
            public_member_name=a.public_member.display_name if a.public_member else None,
            public_member_email=a.public_member.email if a.public_member else None,
            registration_answers=(a.registration_answers or {}),
        ))
    return {"items": results, "total": total, "page": page, "limit": limit}


@app.get(
    "/api/admin/events/{event_id}/attendaonce/export",
    dependencies=[Depends(require_role(Role.admin, Role.superadmin)), Depends(require_paid_plan)],
)
async def export_attendaonce(
    event_id: int,
    fmt: str = Query(default="xlsx", pattern="^(csv|xlsx)$"),
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export attendees list as CSV or XLSX file."""
    ev = await _get_event_for_admin(event_id, me, db, "attendees:read")
    
    # Get registration fields mapping (field_id -> field_label)
    registration_fields = _get_event_registration_fields(ev)
    field_id_to_label = {field.get("id"): field.get("label", field.get("id")) for field in registration_fields}
    
    # Fetch all attendees without pagination
    res = await db.execute(
        select(Attendee)
        .where(Attendee.event_id == event_id)
        .order_by(Attendee.registered_at.desc())
    )
    attendees = res.scalars().all()
    
    # Build data for export
    data = []
    all_answer_keys = set()  # Track all registration answer keys
    
    for a in attendees:
        # Get session count
        cnt_res = await db.execute(
            select(func.count()).where(AttendaonceRecord.attendee_id == a.id)
        )
        sessions_attended = int(cnt_res.scalar_one() or 0)
        
        # Check if has certificate
        cert_res = await db.execute(
            select(Certificate).where(
                Certificate.event_id == event_id,
                Certificate.student_name == a.name,
                Certificate.deleted_at.is_(None),
            )
        )
        has_certificate = cert_res.scalar_one_or_none() is not None
        
        row_data = {
            "İsim": a.name,
            "Email": a.email,
            "Kayıt Tarihi": _format_export_datetime(a.registered_at),
            "Katıldığı Oturumlar": sessions_attended,
            "Sertifika": "Evet" if has_certificate else "Hayır",
            "Kaynak": a.source or "",
        }
        
        # Add registration answers as individual columns
        if a.registration_answers:
            for key, value in a.registration_answers.items():
                # Convert field ID to field label using the mapping
                field_label = field_id_to_label.get(key, key)
                all_answer_keys.add(field_label)
                row_data[field_label] = str(value) if value is not None else ""
        
        data.append(row_data)
    
    if fmt == "xlsx":
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl library not available")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Katılımcılar"
        
        # Build columns: base columns + dynamic answer keys
        base_columns = ["İsim", "Email", "Kayıt Tarihi", "Katıldığı Oturumlar", "Sertifika", "Kaynak"]
        answer_columns = sorted(list(all_answer_keys))  # Sort for consistent ordering
        columns = base_columns + answer_columns
        
        ws.append(columns)
        
        for row in data:
            ws.append([row.get(col, "") for col in columns])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=katilimcilar-event-{event_id}.xlsx"},
        )
    else:
        # CSV export
        buf = io.StringIO()
        base_columns = ["İsim", "Email", "Kayıt Tarihi", "Katıldığı Oturumlar", "Sertifika", "Kaynak"]
        answer_columns = sorted(list(all_answer_keys))  # Sort for consistent ordering
        columns = base_columns + answer_columns
        
        writer = csv.DictWriter(buf, fieldnames=columns)
        writer.writeheader()
        
        for row in data:
            writer.writerow({col: row.get(col, "") for col in columns})
        
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename=katilimcilar-event-{event_id}.csv"},
        )


@app.get(
    "/api/admin/events/{event_id}/registration-documents/file",
    dependencies=[Depends(require_role(Role.admin, Role.superadmin)), Depends(require_paid_plan)],
)
async def download_registration_document(
    event_id: int,
    path: str,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _get_event_for_admin(event_id, me, db, "attendees:read")
    rel_path = path.strip().lstrip("/")
    expected_prefix = f"registration_docs/event_{event_id}/"
    if not rel_path.startswith(expected_prefix):
        raise HTTPException(status_code=404, detail="Document not found")
    storage_root = Path(settings.local_storage_dir).resolve()
    abs_path = (storage_root / rel_path).resolve()
    if not abs_path.is_relative_to(storage_root) or not abs_path.is_file():
        raise HTTPException(status_code=404, detail="Document not found")
    filename = _safe_registration_document_name(abs_path.name)
    return FileResponse(
        abs_path,
        filename=filename,
        media_type="application/octet-stream",
    )


@app.get(
    "/api/admin/events/{event_id}/registration-documents/export",
    dependencies=[Depends(require_role(Role.admin, Role.superadmin)), Depends(require_paid_plan)],
)
async def export_registration_documents_grouped(
    event_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    event = await _get_event_for_admin(event_id, me, db, "attendees:read")
    registration_fields = _get_event_registration_fields(event)
    file_fields = [field for field in registration_fields if field.get("type") == "file"]
    field_id_to_label = {
        str(field.get("id")): str(field.get("label") or field.get("id") or "Belge")
        for field in file_fields
        if field.get("id")
    }

    res = await db.execute(
        select(Attendee)
        .where(Attendee.event_id == event_id)
        .order_by(Attendee.registered_at.asc())
    )
    attendees = res.scalars().all()

    storage_root = Path(settings.local_storage_dir).resolve()
    expected_prefix = f"registration_docs/event_{event_id}/"

    zip_buffer = io.BytesIO()
    used_paths: set[str] = set()
    manifest_rows: List[Dict[str, Any]] = []
    added_files = 0

    with zipfile.ZipFile(zip_buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zip_file:
        for attendee in attendees:
            answers = attendee.registration_answers or {}
            docs_raw = answers.get("__documents")
            if not isinstance(docs_raw, list):
                continue

            attendee_folder = _sanitize_zip_path_part(
                f"{attendee.name}_{attendee.email}_{attendee.id}",
                fallback=f"attendee_{attendee.id}",
                max_len=120,
            )

            for doc_index, doc in enumerate(docs_raw, start=1):
                if not isinstance(doc, dict):
                    continue
                rel_path = str(doc.get("path") or "").strip().lstrip("/")
                if not rel_path or not rel_path.startswith(expected_prefix):
                    continue

                abs_path = (storage_root / rel_path).resolve()
                if not abs_path.is_relative_to(storage_root) or not abs_path.exists() or not abs_path.is_file():
                    continue

                field_id = str(doc.get("field_id") or "").strip()
                field_label = field_id_to_label.get(field_id) or (field_id if field_id else "Diger")
                field_folder = _sanitize_zip_path_part(field_label, fallback="Diger", max_len=100)

                original_name = str(doc.get("name") or Path(rel_path).name)
                safe_name = _safe_registration_document_name(original_name, fallback=f"document_{doc_index}")
                base_arcname = f"{field_folder}/{attendee_folder}/{safe_name}"
                arcname = base_arcname
                duplicate_no = 2
                while arcname in used_paths:
                    alt_name = _safe_registration_document_name(
                        f"{Path(safe_name).stem}_{duplicate_no}{Path(safe_name).suffix}",
                        fallback=f"document_{doc_index}_{duplicate_no}",
                    )
                    arcname = f"{field_folder}/{attendee_folder}/{alt_name}"
                    duplicate_no += 1

                zip_file.write(abs_path, arcname=arcname)
                used_paths.add(arcname)
                added_files += 1

                manifest_rows.append(
                    {
                        "attendee_id": attendee.id,
                        "attendee_name": attendee.name,
                        "attendee_email": attendee.email,
                        "field_id": field_id,
                        "field_label": field_label,
                        "original_name": original_name,
                        "zip_path": arcname,
                        "size_bytes": int(doc.get("size_bytes") or abs_path.stat().st_size),
                    }
                )

        if manifest_rows:
            manifest_buf = io.StringIO()
            writer = csv.DictWriter(
                manifest_buf,
                fieldnames=[
                    "attendee_id",
                    "attendee_name",
                    "attendee_email",
                    "field_id",
                    "field_label",
                    "original_name",
                    "zip_path",
                    "size_bytes",
                ],
            )
            writer.writeheader()
            writer.writerows(manifest_rows)
            zip_file.writestr("manifest.csv", manifest_buf.getvalue().encode("utf-8-sig"))

        if added_files == 0:
            zip_file.writestr(
                "README.txt",
                "Bu etkinlikte indirilebilir kayıt belgesi bulunamadı.",
            )

    zip_buffer.seek(0)
    return StreamingResponse(
        iter([zip_buffer.getvalue()]),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="registration-documents-event-{event_id}.zip"',
        },
    )


@app.get(
    "/api/admin/events/{event_id}/attendees/{attendee_id}/survey-link",
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def get_attendee_survey_link(
    event_id: int,
    attendee_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    event = await _get_event_for_admin(event_id, me, db, "attendees:read")
    res = await db.execute(
        select(Attendee).where(Attendee.id == attendee_id, Attendee.event_id == event_id)
    )
    attendee = res.scalar_one_or_none()
    if not attendee:
        raise HTTPException(status_code=404, detail="Katılımcı bulunamadı")

    survey_token = make_survey_access_token(
        attendee_id=attendee.id,
        event_id=event_id,
        email=attendee.email,
    )
    return {
        "attendee_id": attendee.id,
        "attendee_name": attendee.name,
        "attendee_email": attendee.email,
        "survey_token": survey_token,
        "survey_url": build_public_survey_url(
            event_id=_get_public_event_identifier(event),
            attendee_id=attendee.id,
            email=attendee.email,
        ),
    }


@app.get("/api/admin/events/{event_id}/comments", response_model=list[PublicEventCommentOut], dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def list_admin_event_comments(
    event_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    ev = await _get_event_for_admin(event_id, me, db, "attendees:read")
    comments_res = await db.execute(
        select(EventComment)
        .options(selectinload(EventComment.public_member))
        .where(EventComment.event_id == event_id)
        .order_by(EventComment.created_at.desc())
    )
    return [_event_comment_to_out(comment) for comment in comments_res.scalars().all()]


@app.patch("/api/admin/events/{event_id}/comments/{comment_id}", response_model=PublicEventCommentOut, dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def update_admin_event_comment(
    event_id: int,
    comment_id: int,
    payload: AdminEventCommentUpdateIn,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _get_event_for_admin(event_id, me, db, "settings:write")
    comment_res = await db.execute(
        select(EventComment)
        .options(selectinload(EventComment.public_member))
        .where(EventComment.id == comment_id, EventComment.event_id == event_id)
    )
    comment = comment_res.scalar_one_or_none()
    if not comment:
        raise HTTPException(status_code=404, detail="Comment not found")

    comment.status = payload.status
    await write_audit_log(
        db,
        user_id=me.id,
        action="admin.comment.update",
        resource_type="event_comment",
        resource_id=str(comment.id),
        extra={"event_id": event_id, "status": payload.status},
    )
    await db.commit()
    await db.refresh(comment, attribute_names=["public_member"])
    return _event_comment_to_out(comment)


@app.get(
    "/api/admin/events/{event_id}/attendees/filter-for-email",
    dependencies=[Depends(require_role(Role.admin, Role.superadmin)), Depends(require_paid_plan)],
)
async def filter_attendees_for_email(
    event_id: int,
    filter_type: str = Query(default="all"),  # all | with_email | unsubscribed | certified
    subscribed_only: bool = Query(default=True),
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Get filtered attendees for email targeting.
    
    Filter types:
    - all: all attendees
    - with_email: attendees with valid emails
    - unsubscribed: attendees who unsubscribed
    - certified: attendees with certificates
    """
    ev = await _get_event_for_admin(event_id, me, db, "email:write")
    
    q = select(Attendee).where(Attendee.event_id == event_id)
    
    # Apply filters
    if filter_type == "with_email":
        q = q.where(Attendee.email.isnot(None), Attendee.email != "")
    elif filter_type == "unsubscribed":
        q = q.where(Attendee.unsubscribed_at.isnot(None))
    elif filter_type == "certified":
        # Join with certificates
        from sqlalchemy import exists
        cert_subq = select(Certificate).where(
            Certificate.event_id == event_id,
            Certificate.deleted_at.is_(None),
            Certificate.student_name == Attendee.name
        )
        q = q.where(exists(cert_subq))
    
    if subscribed_only and filter_type != "unsubscribed":
        q = q.where(Attendee.unsubscribed_at.is_(None))
    
    res = await db.execute(q.order_by(Attendee.registered_at.desc()))
    attendees = res.scalars().all()
    
    return {
        "filter_type": filter_type,
        "count": len(attendees),
        "attendees": [
            {
                "id": a.id,
                "name": a.name,
                "email": a.email,
                "unsubscribed": a.unsubscribed_at is not None,
            }
            for a in attendees
        ],
    }


@app.post("/api/admin/events/{event_id}/attendees/import", dependencies=[Depends(require_role(Role.admin, Role.superadmin)), Depends(require_paid_plan)])
async def import_attendees(
    event_id: int,
    file: UploadFile = File(...),
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _get_event_for_admin(event_id, me, db, "attendees:write")
    content = await file.read()
    try:
        if file.filename and file.filename.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(content))
        else:
            df = pd.read_excel(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Dosya okunamadÃ„Â±: {e}")

    # Find name and email columns (case-insensitive)
    col_map = {c.lower(): c for c in df.columns}
    name_col = col_map.get("name") or col_map.get("ad") or col_map.get("isim")
    email_col = col_map.get("email") or col_map.get("e-posta") or col_map.get("eposta")
    if not name_col or not email_col:
        raise HTTPException(status_code=400, detail="'name' ve 'email' kolonlarÃ„Â± gerekli")

    added = 0
    skipped = 0
    errors = []
    for _, row in df.iterrows():
        raw_name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ""
        raw_email = str(row[email_col]).strip().lower() if pd.notna(row[email_col]) else ""
        if not raw_name or not raw_email or "@" not in raw_email:
            skipped += 1
            continue
        existing = await db.execute(
            select(Attendee).where(Attendee.event_id == event_id, func.lower(Attendee.email) == raw_email)
        )
        if existing.scalar_one_or_none():
            skipped += 1
            continue
        attendee = Attendee(event_id=event_id, name=raw_name, email=raw_email, source="import")
        db.add(attendee)
        await db.flush()
        from .crm_snapshot_hooks import refresh_crm_snapshot_for_attendee
        await refresh_crm_snapshot_for_attendee(db, attendee)
        added += 1
        if added % 100 == 0:
            await db.flush()

    await db.commit()
    return {"added": added, "skipped": skipped}


@app.post(
    "/api/admin/events/{event_id}/attendees",
    response_model=AttendeeOut,
    dependencies=[Depends(require_role(Role.admin, Role.superadmin)), Depends(require_paid_plan)],
)
async def create_manual_attendee(
    event_id: int,
    payload: ManualAttendeeCreateIn,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    event = await _get_event_for_admin(event_id, me, db, "attendees:write")

    normalized_email = str(payload.email).strip().lower()
    first_name = " ".join(payload.first_name.strip().split())
    last_name = " ".join(payload.last_name.strip().split())
    full_name = f"{first_name} {last_name}".strip()

    if not full_name:
        raise bad_request("Ad ve soyad gerekli")

    existing_attendee_res = await db.execute(
        select(Attendee).where(
            Attendee.event_id == event_id,
            func.lower(Attendee.email) == normalized_email,
        )
    )
    if existing_attendee_res.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Bu e-posta ile katılımcı zaten kayıtlı")

    public_member_res = await db.execute(
        select(PublicMember).where(func.lower(PublicMember.email) == normalized_email)
    )
    public_member = public_member_res.scalar_one_or_none()

    attendee = Attendee(
        event_id=event_id,
        name=full_name,
        email=normalized_email,
        source="import",
        email_verified=True,
        email_verified_at=datetime.now(timezone.utc),
        public_member_id=public_member.id if public_member else None,
        registration_answers={},
    )
    db.add(attendee)
    await db.flush()
    await _issue_event_ticket_if_needed(db, event, attendee)
    from .crm_snapshot_hooks import refresh_crm_snapshot_for_attendee
    await refresh_crm_snapshot_for_attendee(db, attendee)

    await write_audit_log(
        db,
        user_id=me.id,
        action="attendee.manual_add",
        resource_type="attendee",
        resource_id=str(attendee.id),
        extra={
            "event_id": event_id,
            "attendee_email": attendee.email,
            "source": "admin_manual",
        },
    )

    await db.commit()
    await db.refresh(attendee)

    return AttendeeOut(
        id=attendee.id,
        event_id=attendee.event_id,
        name=attendee.name,
        email=attendee.email,
        source=attendee.source,
        registered_at=attendee.registered_at,
        sessions_attended=0,
        has_certificate=False,
        public_member_id=attendee.public_member_id,
        public_member_name=public_member.display_name if public_member else None,
        public_member_email=public_member.email if public_member else None,
        registration_answers=attendee.registration_answers or {},
    )


@app.delete(
    "/api/admin/events/{event_id}/attendees/{attendee_id}",
    dependencies=[Depends(require_role(Role.admin, Role.superadmin)), Depends(require_paid_plan)],
)
async def delete_attendee(
    event_id: int,
    attendee_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _get_event_for_admin(event_id, me, db, "attendees:write")
    res = await db.execute(select(Attendee).where(Attendee.id == attendee_id, Attendee.event_id == event_id))
    att = res.scalar_one_or_none()
    if not att:
        raise HTTPException(status_code=404, detail="Attendee not found")
    snapshot_email = att.email
    await db.delete(att)
    await db.flush()
    from .crm_snapshot_hooks import refresh_crm_snapshot_for_event_email
    await refresh_crm_snapshot_for_event_email(db, event_id=event_id, email=snapshot_email)
    await db.commit()
    return {"ok": True}


# Ã¢â€â‚¬Ã¢â€â‚¬ Admin: Attendaonce matrix & manual check-in Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬

@app.post(
    "/api/admin/events/{event_id}/sessions/{session_id}/checkin",
    dependencies=[Depends(require_role(Role.admin, Role.superadmin)), Depends(require_paid_plan)],
)
async def admin_manual_checkin(
    event_id: int,
    session_id: int,
    payload: CheckinIn,
    request: Request,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from .checkin_ops_api import record_checkin_activity

    ev = await _get_event_for_admin(event_id, me, db, "checkin:write")
    _ensure_checkin_feature_enabled(ev)
    ip = _client_ip_for_rate_limit(request)

    session_res = await db.execute(
        select(EventSession).where(EventSession.id == session_id, EventSession.event_id == event_id)
    )
    session = session_res.scalar_one_or_none()
    if not session:
        raise HTTPException(status_code=404, detail="Oturum bulunamadı")

    email = payload.email.strip().lower()
    attendee_res = await db.execute(
        select(Attendee).where(
            Attendee.event_id == event_id,
            func.lower(Attendee.email) == email,
        )
    )
    attendee = attendee_res.scalar_one_or_none()
    if not attendee:
        raise HTTPException(
            status_code=404,
            detail="Bu e-posta ile etkinlikte kayıtlı katılımcı bulunamadı.",
        )

    insert_stmt = (
        _pg_insert(AttendaonceRecord.__table__)
        .values(
            attendee_id=attendee.id,
            session_id=session_id,
            ip_address=ip,
        )
        .on_conflict_do_nothing(index_elements=["attendee_id", "session_id"])
        .returning(AttendaonceRecord.id)
    )
    inserted_res = await db.execute(insert_stmt)
    inserted_id = inserted_res.scalar_one_or_none()
    if inserted_id is not None:
        from .crm_snapshot_hooks import refresh_crm_snapshot_for_attendee
        await refresh_crm_snapshot_for_attendee(db, attendee)
    await record_checkin_activity(
        db,
        event_id=event_id,
        session_id=session_id,
        attendee_id=attendee.id,
        actor_user_id=me.id,
        method="manual",
        source="admin",
        success=inserted_id is not None,
        message="Manual check-in successful" if inserted_id is not None else "Duplicate manual check-in",
        ip_address=ip,
    )
    await db.commit()

    if inserted_id is None:
        return {"ok": False, "message": "Bu katılımcı bu oturum için zaten check-in yapılmış."}

    try:
        from .checkin_ops_api import publish_checkin_event
        publish_checkin_event(event_id, {
            "type": "checkin",
            "attendee_id": attendee.id,
            "attendee_name": attendee.name,
            "session_id": session_id,
            "checked_in_at": datetime.now(timezone.utc).isoformat(),
            "method": "manual",
        })
    except Exception:
        pass

    return {"ok": True, "message": f"Check-in başarılı: {attendee.name}"}

@app.get(
    "/api/admin/events/{event_id}/operations",
    dependencies=[Depends(require_role(Role.admin, Role.superadmin)), Depends(require_paid_plan)],
)
async def get_event_operations(
    event_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    ev = await _get_event_for_admin(event_id, me, db, "checkin:write")
    _ensure_checkin_feature_enabled(ev)

    attendee_count = int(
        (await db.execute(select(func.count(Attendee.id)).where(Attendee.event_id == event_id))).scalar_one() or 0
    )
    session_count = int(
        (await db.execute(select(func.count(EventSession.id)).where(EventSession.event_id == event_id))).scalar_one() or 0
    )
    active_session_count = int(
        (
            await db.execute(
                select(func.count(EventSession.id)).where(
                    EventSession.event_id == event_id,
                    EventSession.is_active.is_(True),
                )
            )
        ).scalar_one()
        or 0
    )
    attendance_count = int(
        (
            await db.execute(
                select(func.count(AttendaonceRecord.id))
                .join(EventSession, EventSession.id == AttendaonceRecord.session_id)
                .where(EventSession.event_id == event_id)
            )
        ).scalar_one()
        or 0
    )

    session_rows = (
        await db.execute(
            select(EventSession, func.count(AttendaonceRecord.id).label("attendance_count"))
            .outerjoin(AttendaonceRecord, AttendaonceRecord.session_id == EventSession.id)
            .where(EventSession.event_id == event_id)
            .group_by(EventSession.id)
            .order_by(EventSession.session_date, EventSession.session_start, EventSession.id)
        )
    ).all()

    ticket_rows = (
        await db.execute(
            select(EventTicket.status, func.count(EventTicket.id))
            .where(EventTicket.event_id == event_id)
            .group_by(EventTicket.status)
        )
    ).all()
    tickets_by_status = {str(status or "unknown"): int(count or 0) for status, count in ticket_rows}
    ticket_total = sum(tickets_by_status.values())

    recent_rows = (
        await db.execute(
            select(AttendaonceRecord, Attendee, EventSession)
            .join(Attendee, Attendee.id == AttendaonceRecord.attendee_id)
            .join(EventSession, EventSession.id == AttendaonceRecord.session_id)
            .where(EventSession.event_id == event_id)
            .order_by(AttendaonceRecord.checked_in_at.desc(), AttendaonceRecord.id.desc())
            .limit(40)
        )
    ).all()

    return {
        "event_id": event_id,
        "event_name": ev.name,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "overview": {
            "attendees": attendee_count,
            "sessions": session_count,
            "active_sessions": active_session_count,
            "attendance_records": attendance_count,
            "tickets_total": ticket_total,
            "tickets_used": tickets_by_status.get("used", 0),
        },
        "tickets": {
            "total": ticket_total,
            "by_status": tickets_by_status,
        },
        "sessions": [
            {
                "id": session.id,
                "name": session.name,
                "is_active": session.is_active,
                "session_date": session.session_date.isoformat() if session.session_date else None,
                "session_start": session.session_start.isoformat() if session.session_start else None,
                "attendance_count": int(count or 0),
            }
            for session, count in session_rows
        ],
        "recent_checkins": [
            {
                "id": record.id,
                "attendee_id": attendee.id,
                "attendee_name": attendee.name,
                "attendee_email": attendee.email,
                "session_id": session.id,
                "session_name": session.name,
                "checked_in_at": record.checked_in_at.isoformat() if record.checked_in_at else None,
                "ip_address": record.ip_address,
            }
            for record, attendee, session in recent_rows
        ],
    }


@app.delete(
    "/api/admin/events/{event_id}/attendance-records/{record_id}",
    dependencies=[Depends(require_role(Role.admin, Role.superadmin)), Depends(require_paid_plan)],
)
async def undo_event_attendance_record(
    event_id: int,
    record_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    ev = await _get_event_for_admin(event_id, me, db, "checkin:write")
    _ensure_checkin_feature_enabled(ev)

    record_res = await db.execute(
        select(AttendaonceRecord, Attendee, EventSession)
        .join(Attendee, Attendee.id == AttendaonceRecord.attendee_id)
        .join(EventSession, EventSession.id == AttendaonceRecord.session_id)
        .where(AttendaonceRecord.id == record_id, EventSession.event_id == event_id)
    )
    row = record_res.first()
    if not row:
        raise HTTPException(status_code=404, detail="Check-in kaydı bulunamadı.")

    record, attendee, session = row
    await write_audit_log(
        db,
        user_id=me.id,
        action="attendance.undo",
        resource_type="attendance_record",
        resource_id=str(record.id),
        extra={
            "event_id": event_id,
            "attendee_id": attendee.id,
            "attendee_email": attendee.email,
            "session_id": session.id,
            "session_name": session.name,
        },
    )
    await db.delete(record)
    await db.commit()
    return {
        "ok": True,
        "deleted_id": record_id,
        "message": f"{attendee.name} için {session.name} check-in kaydı geri alındı.",
    }


async def _get_raffle_for_admin(
    event_id: int,
    raffle_id: int,
    me: CurrentUser,
    db: AsyncSession,
) -> EventRaffle:
    event = await _get_event_for_admin(event_id, me, db, "settings:write")
    if not is_raffles_enabled(event):
        raise HTTPException(status_code=403, detail="Raffle features are disabled for this event.")
    raffle_res = await db.execute(
        select(EventRaffle)
        .options(selectinload(EventRaffle.winners).selectinload(EventRaffleWinner.attendee))
        .where(EventRaffle.id == raffle_id, EventRaffle.event_id == event_id)
    )
    raffle = raffle_res.scalar_one_or_none()
    if not raffle:
        raise HTTPException(status_code=404, detail="Raffle not found")
    return raffle




@app.get("/api/admin/events/{event_id}/attendance", dependencies=[Depends(require_role(Role.admin, Role.superadmin)), Depends(require_paid_plan)])
@app.get("/api/admin/events/{event_id}/attendaonce", dependencies=[Depends(require_role(Role.admin, Role.superadmin)), Depends(require_paid_plan)])
@app.get("/api/admin/events/{event_id}/attendance/export", dependencies=[Depends(require_role(Role.admin, Role.superadmin)), Depends(require_paid_plan)])
async def get_attendaonce_matrix(
    event_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    fmt: str = Query(default="json", pattern="^(csv|xlsx|json)$"),
):
    ev = await _get_event_for_admin(event_id, me, db, "certificates:write")
    _ensure_checkin_feature_enabled(ev)
    sess_res = await db.execute(
        select(EventSession).where(EventSession.event_id == event_id).order_by(EventSession.session_date, EventSession.session_start, EventSession.id)
    )
    sessions = sess_res.scalars().all()
    att_res = await db.execute(
        select(Attendee)
        .where(Attendee.event_id == event_id)
        .order_by(Attendee.name)
    )
    attendees = att_res.scalars().all()

    # Build set of (attendee_id, session_id) for O(1) lookup
    rec_res = await db.execute(
        select(AttendaonceRecord.attendee_id, AttendaonceRecord.session_id, AttendaonceRecord.checked_in_at).where(
            AttendaonceRecord.attendee_id.in_([a.id for a in attendees])
        )
    )
    records = rec_res.all()
    rec_set: dict[tuple, str] = {(r.attendee_id, r.session_id): r.checked_in_at.isoformat() for r in records}

    # Build certificate lookup by student_name for this event
    cert_res = await db.execute(
        select(Certificate.student_name, Certificate.uuid)
        .where(Certificate.event_id == event_id)
    )
    certs = cert_res.all()
    cert_map: dict[str, str] = {c.student_name: c.uuid for c in certs}

    # Return JSON format if requested
    if fmt == "json":
        json_rows = []
        for a in attendees:
            checkins: dict[str, str | None] = {}
            for s in sessions:
                checkins[str(s.id)] = rec_set.get((a.id, s.id), None)
            
            sessions_attended = sum(1 for s in sessions if (a.id, s.id) in rec_set)
            has_cert = a.name in cert_map
            cert_uuid = cert_map.get(a.name)
            
            row = {
                "attendee_id": a.id,
                "name": a.name,
                "email": a.email,
                "source": a.source,
                "sessions_attended": sessions_attended,
                "meets_threshold": sessions_attended >= ev.min_sessions_required,
                "has_certificate": has_cert,
                "certificate_uuid": cert_uuid,
                "checkins": checkins,
            }
            json_rows.append(row)
        
        return {
            "event_id": event_id,
            "min_sessions_required": ev.min_sessions_required,
            "sessions": [{"id": s.id, "name": s.name, "session_date": s.session_date.isoformat() if s.session_date else None} for s in sessions],
            "rows": json_rows,
        }

    # Excel/CSV format
    rows = []
    for a in attendees:
        registered_at = a.registered_at
        if registered_at:
            if registered_at.tzinfo is None:
                registered_at = registered_at.replace(tzinfo=timezone.utc)
            registered_at_text = registered_at.astimezone(timezone.utc).strftime("%d.%m.%Y %H:%M:%S UTC")
        else:
            registered_at_text = ""
        row = {
            "Ad Soyad": a.name,
            "E-posta": a.email,
            "Kaynak": a.source,
            "Kayit Tarihi": registered_at_text,
        }
        for s in sessions:
            row[s.name] = "Evet" if (a.id, s.id) in rec_set else "Hayir"
        row["Katilinan Oturum"] = sum(1 for s in sessions if (a.id, s.id) in rec_set)
        row["Esigi Geciyor"] = "Evet" if row["Katilinan Oturum"] >= ev.min_sessions_required else "Hayir"
        answers = a.registration_answers or {}
        registration_fields = _get_event_registration_fields(ev)
        for field in registration_fields:
            row[field["label"]] = answers.get(field["id"], "")
        rows.append(row)

    df = pd.DataFrame(rows)
    buf = io.BytesIO()
    if fmt == "csv":
        df.to_csv(buf, index=False)
        buf.seek(0)
        return StreamingResponse(buf, media_type="text/csv", headers={"Content-Disposition": f'attachment; filename="attendance_{event_id}.csv"'})
    else:
        df.to_excel(buf, index=False)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="attendance_{event_id}.xlsx"'},
        )


# Ã¢â€â‚¬Ã¢â€â‚¬ Admin: Bulk certify Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬



@app.post(
    "/api/admin/events/{event_id}/bulk-certify",
    response_model=BulkCertifyOut,
    dependencies=[Depends(require_role(Role.admin, Role.superadmin)), Depends(require_paid_plan)],
)
async def bulk_certify_attendees(
    event_id: int,
    payload: BulkCertifyIn,
    background_tasks: BackgroundTasks,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from .generator import (
        TemplateConfig,
        new_certificate_uuid,
        render_certificate_pdf,
        render_certificate_png_watermarked,
    )

    ev = await _get_event_for_admin(event_id, me, db, "certificates:write")
    _ensure_certificate_feature_enabled(ev)
    if not ev.config or ev.template_image_url in ("", "placeholder"):
        raise HTTPException(status_code=400, detail="Etkinlik sablon yapilandirmasi eksik")

    # Fetch all attendees
    att_res = await db.execute(select(Attendee).where(Attendee.event_id == event_id))
    attendees = att_res.scalars().all()
    if not attendees:
        raise HTTPException(status_code=400, detail="Katılımcı listesi boş")

    # Determine hologram policy: only Growth/Enterprise can disable it
    billing_user_id = ev.admin_id
    _allow_no_hologram = me.role == Role.superadmin
    if not _allow_no_hologram:
        _sub_hb = await db.execute(
            select(Subscription)
            .where(Subscription.user_id == billing_user_id, Subscription.is_active == True)
            .order_by(Subscription.expires_at.desc()).limit(1)
        )
        _sub_hb_row = _sub_hb.scalar_one_or_none()
        _now_hb = datetime.now(timezone.utc)
        _allow_no_hologram = bool(
            _sub_hb_row and _sub_hb_row.plan_id in ("growth", "enterprise") and
            (not _sub_hb_row.expires_at or _sub_hb_row.expires_at > _now_hb)
        )

    # Fetch attendaonce counts
    rec_res = await db.execute(
        select(AttendaonceRecord.attendee_id, func.count().label("cnt"))
        .where(AttendaonceRecord.attendee_id.in_([a.id for a in attendees]))
        .group_by(AttendaonceRecord.attendee_id)
    )
    attend_counts: dict[int, int] = {r.attendee_id: r.cnt for r in rec_res.all()}

    # Eligible attendees
    eligible = [a for a in attendees if attend_counts.get(a.id, 0) >= ev.min_sessions_required]
    below_threshold = len(attendees) - len(eligible)

    if not eligible:
        return BulkCertifyOut(
            created=0, already_had_cert=0,
            below_threshold=below_threshold,
            total_attendees=len(attendees),
            spent_heptacoin=0
        )

    # Balance check
    user_res = await db.execute(select(User).where(User.id == billing_user_id))
    user = user_res.scalar_one()
    # Rough estimate: check at least 10 HC per cert available
    if user.heptacoin_balaonce < 10:
        raise HTTPException(status_code=402, detail="Yetersiz HeptaCoin")

    # Load org branding
    org_res = await db.execute(select(Organization).where(Organization.user_id == billing_user_id))
    org = org_res.scalar_one_or_none()
    brand_logo_path: Optional[Path] = None
    if org and org.brand_logo:
        brand_logo_path = local_path_from_url(org.brand_logo)
        if not brand_logo_path.exists():
            brand_logo_path = None

    template_path = local_path_from_url(ev.template_image_url)
    cfg_raw = ev.config or {}

    created = 0
    already_had_cert = 0
    total_spent = 0

    for attendee in eligible:
        # Check if cert already exists
        cert_check = await db.execute(
            select(Certificate).where(
                Certificate.event_id == event_id,
                Certificate.student_name == attendee.name,
                Certificate.deleted_at.is_(None),
            )
        )
        if cert_check.scalar_one_or_none():
            already_had_cert += 1
            continue

        # Re-check balance each iteration
        fresh_user = await db.execute(select(User).where(User.id == billing_user_id))
        user = fresh_user.scalar_one()
        if user.heptacoin_balaonce < 10:
            break  # stop if out of coins

        cert_uuid = new_certificate_uuid()
        # Atomic seq increment
        await db.execute(
            update(Event).where(Event.id == ev.id).values(cert_seq=Event.cert_seq + 1)
        )
        await db.flush()
        ev_fresh = await db.execute(select(Event).where(Event.id == ev.id))
        ev = ev_fresh.scalar_one()
        public_id = f"EV{ev.id}-{ev.cert_seq:06d}"

        hosting_term = payload.hosting_term
        ends_at = compute_hosting_ends(hosting_term)

        try:
            tc = editor_config_to_template_config(cfg_raw)
            # Override hologram policy based on plan
            if not _allow_no_hologram:
                tc = TemplateConfig(
                    **{k: v for k, v in tc.__dict__.items() if k != "show_hologram"},
                    show_hologram=True,
                )
        except Exception as e:
            logger.warning("BulkCertify: TemplateConfig error for %s: %s", attendee.name, e)
            continue

        # Load template image bytes
        try:
            template_bytes = template_path.read_bytes()
        except Exception as e:
            logger.error("BulkCertify: cannot read template for %s: %s", attendee.name, e)
            continue

        # Load brand logo bytes
        brand_logo_bytes: Optional[bytes] = None
        if brand_logo_path and brand_logo_path.exists():
            try:
                brand_logo_bytes = brand_logo_path.read_bytes()
            except Exception:
                pass

        rel_path = f"pdfs/event_{event_id}/{cert_uuid}.pdf"
        abs_path = Path(settings.local_storage_dir) / rel_path
        abs_path.parent.mkdir(parents=True, exist_ok=True)
        rel_png_path = _certificate_png_rel_path(event_id, cert_uuid)
        abs_png_path = Path(settings.local_storage_dir) / rel_png_path
        abs_png_path.parent.mkdir(parents=True, exist_ok=True)

        try:
            pdf_bytes = render_certificate_pdf(
                template_image_bytes=template_bytes,
                student_name=attendee.name,
                verify_url=build_certificate_verify_url(cert_uuid),
                config=tc,
                public_id=public_id,
                brand_logo_bytes=brand_logo_bytes,
                certificate_footer=(cfg_raw.get("certificate_footer") if isinstance(cfg_raw, dict) else None),
            )
            abs_path.write_bytes(pdf_bytes)
            try:
                png_bytes = render_certificate_png_watermarked(
                    template_image_bytes=template_bytes,
                    student_name=attendee.name,
                    verify_url=build_certificate_verify_url(cert_uuid),
                    config=tc,
                    public_id=public_id,
                    brand_logo_bytes=brand_logo_bytes,
                    certificate_footer=(cfg_raw.get("certificate_footer") if isinstance(cfg_raw, dict) else None),
                )
                abs_png_path.write_bytes(png_bytes)
            except Exception as png_error:
                logger.warning("BulkCertify PNG render error for %s: %s", attendee.name, png_error)
        except Exception as e:
            logger.error("BulkCertify render error for %s: %s", attendee.name, e)
            continue

        asset_size = abs_path.stat().st_size if abs_path.exists() else 0
        h_units = hosting_units(hosting_term, asset_size)
        cost = 10 + h_units

        cert = Certificate(
            uuid=cert_uuid,
            student_name=attendee.name,
            event_id=event_id,
            pdf_url=build_public_pdf_url(rel_path),
            status=CertStatus.active,
            public_id=public_id,
            hosting_term=hosting_term,
            hosting_ends_at=ends_at,
            asset_size_bytes=asset_size,
        )
        db.add(cert)
        user.heptacoin_balaonce -= cost
        db.add(Transaction(user_id=me.id, amount=cost, type=TxType.spend))
        created += 1
        total_spent += cost
        await db.flush()
        from .crm_snapshot_hooks import refresh_crm_snapshot_for_attendee
        await refresh_crm_snapshot_for_attendee(db, attendee)
        if attendee.email:
            background_tasks.add_task(
                send_certificate_delivery_email_task,
                event_id=event_id,
                cert_uuid=cert_uuid,
                recipient_name=attendee.name,
                recipient_email=attendee.email,
            )

    await db.commit()
    return BulkCertifyOut(
        created=created,
        already_had_cert=already_had_cert,
        below_threshold=below_threshold,
        total_attendees=len(attendees),
        spent_heptacoin=total_spent,
    )


@app.post(
    "/api/admin/events/{event_id}/bulk-certify-queue",
    response_model=BulkCertificateJobOut,
    dependencies=[Depends(require_role(Role.admin, Role.superadmin)), Depends(require_paid_plan)],
)
async def bulk_certify_attendees_queue(
    event_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Creates a background bulk certificate job for all eligible attendees.

    Uses the existing job queue (BulkCertificateJob) so the request returns
    immediately and heavy PDF rendering happens asynchronously, preventing
    HTTP timeouts for large attendee lists.
    """
    ev = await _get_event_for_admin(event_id, me, db, "certificates:write")
    _ensure_certificate_feature_enabled(ev)
    if not ev.config or ev.template_image_url in ("", "placeholder"):
        raise HTTPException(status_code=400, detail="Etkinlik sablon yapilandirmasi eksik")

    # Fetch all attendees
    att_res = await db.execute(select(Attendee).where(Attendee.event_id == event_id))
    attendees = att_res.scalars().all()
    if not attendees:
        raise HTTPException(status_code=400, detail="Katılımcı listesi boş")

    # Count attendaonce per attendee
    rec_res = await db.execute(
        select(AttendaonceRecord.attendee_id, func.count().label("cnt"))
        .where(AttendaonceRecord.attendee_id.in_([a.id for a in attendees]))
        .group_by(AttendaonceRecord.attendee_id)
    )
    attend_counts: dict[int, int] = {r.attendee_id: r.cnt for r in rec_res.all()}

    eligible = [a for a in attendees if attend_counts.get(a.id, 0) >= ev.min_sessions_required]
    if not eligible:
        raise HTTPException(status_code=400, detail="Eşiği geçen katılımcı bulunamadı")

    names = [a.name for a in eligible]

    # Early balance check
    res_u = await db.execute(select(User).where(User.id == ev.admin_id))
    user = res_u.scalar_one()
    ISSUE_UNITS_PER_CERT = 10
    HOSTING_ESTIMATE_UNITS = 20
    estimated_total = len(names) * (ISSUE_UNITS_PER_CERT + HOSTING_ESTIMATE_UNITS)
    if user.heptacoin_balaonce < estimated_total:
        raise HTTPException(
            status_code=402,
            detail=f"Yetersiz HeptaCoin. Tahmini Gereksinim={estimated_total}, Bakiye={user.heptacoin_balaonce}",
        )

    chunk_size = 5 if len(names) >= 500 else 10
    job = BulkCertificateJob(
        event_id=ev.id,
        created_by=ev.admin_id,
        names=names,
        chunk_size=chunk_size,
        total_count=len(names),
        status="pending",
    )
    db.add(job)
    await db.commit()
    await db.refresh(job)

    # Nudge the background processor to start quickly
    asyncio.create_task(_process_bulk_certificate_jobs())

    return job


# Ã¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢Â
# Admin: Transaction list (paginated)
# Ã¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢ÂÃ¢â€¢Â

@app.get(
    "/api/admin/transactions/list",
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
)
async def list_my_transactions_paginated(
    page: int = Query(default=1, ge=1),
    limit: int = Query(default=20, le=100),
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    q = select(Transaction).where(Transaction.user_id == me.id)
    total_res = await db.execute(select(func.count()).select_from(q.subquery()))
    total = int(total_res.scalar_one() or 0)

    q = q.order_by(Transaction.timestamp.desc()).offset((page - 1) * limit).limit(limit)
    res = await db.execute(q)
    txs = res.scalars().all()
    return {
        "items": [
            {"id": t.id, "amount": t.amount, "type": t.type, "timestamp": t.timestamp.isoformat() if t.timestamp else None}
            for t in txs
        ],
        "total": total,
        "page": page,
        "limit": limit,
    }


# Ã¢â€â‚¬Ã¢â€â‚¬ Public: Email Unsubscribe Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬Ã¢â€â‚¬

# ═════ SUPERADMIN: AUDIT LOGS ═════











@app.get("/api/superadmin/audit-logs", dependencies=[Depends(require_role(Role.superadmin))])
async def get_audit_logs(
    page: int = Query(default=1, ge=1),
    limit: int = Query(default=50, le=500),
    action: Optional[str] = Query(None),
    resource_type: Optional[str] = Query(None),
    category: Optional[str] = Query(default=None, pattern="^(legal|security)$"),
    from_date: Optional[datetime] = Query(default=None),
    to_date: Optional[datetime] = Query(default=None),
    db: AsyncSession = Depends(get_db),
):
    """Get paginated audit logs - superadmin only."""
    q = select(AuditLog, User.email).outerjoin(User, User.id == AuditLog.user_id)
    if action:
        q = q.where(AuditLog.action.ilike(f"%{action}%"))
    if resource_type:
        q = q.where(AuditLog.resource_type == resource_type)
    if from_date:
        q = q.where(AuditLog.created_at >= from_date)
    if to_date:
        q = q.where(AuditLog.created_at <= to_date)
    q = _audit_category_filter(q, category)
    
    total_res = await db.execute(select(func.count()).select_from(q.subquery()))
    total = int(total_res.scalar_one() or 0)
    
    q = q.order_by(AuditLog.created_at.desc()).offset((page - 1) * limit).limit(limit)
    res = await db.execute(q)
    rows = res.all()
    
    return {
        "total": total,
        "page": page,
        "limit": limit,
        "items": [_audit_row_payload(log, email) for log, email in rows],
    }


@app.get("/api/superadmin/audit-logs/export", dependencies=[Depends(require_role(Role.superadmin))])
async def export_audit_logs(
    format: str = Query(default="csv", pattern="^(csv|pdf)$"),
    category: Optional[str] = Query(default=None, pattern="^(legal|security)$"),
    action: Optional[str] = Query(None),
    resource_type: Optional[str] = Query(None),
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if format == "pdf":
        from .document_export_jobs import DocumentExportJob

        job = DocumentExportJob(
            export_type="audit_logs",
            export_format="pdf",
            requested_by=me.id,
            filters={"category": category, "action": action, "resource_type": resource_type},
            status="pending",
        )
        db.add(job)
        await db.flush()
        await write_audit_log(
            db,
            user_id=me.id,
            action="document_export.enqueue",
            resource_type="document_export_job",
            resource_id=str(job.id),
            extra={"export_type": "audit_logs", "format": "pdf", "legacy_endpoint": True},
        )
        await db.commit()
        return JSONResponse(
            {"id": job.id, "status": "pending", "message": "PDF export queued. You will receive an email when it is ready."},
            status_code=202,
        )

    q = select(AuditLog, User.email).outerjoin(User, User.id == AuditLog.user_id)
    if action:
        q = q.where(AuditLog.action.ilike(f"%{action}%"))
    if resource_type:
        q = q.where(AuditLog.resource_type == resource_type)
    q = _audit_category_filter(q, category)
    q = q.order_by(AuditLog.created_at.desc()).limit(5000)
    rows = [_audit_row_payload(log, email) for log, email in (await db.execute(q)).all()]
    suffix = category or "all"
    return _audit_csv_response(rows, f"audit-logs-{suffix}.csv")


@app.get("/api/admin/organization/legal-consents/export", dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def export_organization_legal_consents(
    request: Request,
    format: str = Query(default="csv", pattern="^(csv|pdf)$"),
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from .organization_access_api import get_organization_for_access, organization_id_from_request

    organization = await get_organization_for_access(db, me, "organization:view", organization_id_from_request(request))
    if format == "pdf":
        from .document_export_jobs import DocumentExportJob

        job = DocumentExportJob(
            export_type="organization_legal_consents",
            export_format="pdf",
            requested_by=me.id,
            organization_id=organization.id,
            filters={},
            status="pending",
        )
        db.add(job)
        await db.flush()
        await write_audit_log(
            db,
            user_id=me.id,
            action="document_export.enqueue",
            resource_type="document_export_job",
            resource_id=str(job.id),
            extra={"export_type": "organization_legal_consents", "format": "pdf", "legacy_endpoint": True},
        )
        await db.commit()
        return JSONResponse(
            {"id": job.id, "status": "pending", "message": "PDF export queued. You will receive an email when it is ready."},
            status_code=202,
        )

    event_ids = set((await db.execute(select(Event.id).where(Event.admin_id == organization.user_id))).scalars().all())
    rows: List[Dict[str, Any]] = []
    if event_ids:
        q = (
            select(AuditLog, User.email)
            .outerjoin(User, User.id == AuditLog.user_id)
            .where(AuditLog.action.in_(["legal.consent.accept", "legal.document.click", "legal.document.view"]))
            .order_by(AuditLog.created_at.desc())
            .limit(10000)
        )
        for log, email in (await db.execute(q)).all():
            extra = log.extra if isinstance(log.extra, dict) else {}
            try:
                log_event_id = int(extra.get("event_id") or 0)
            except (TypeError, ValueError):
                log_event_id = 0
            if log_event_id in event_ids:
                rows.append(_audit_row_payload(log, email))
    await write_audit_log(
        db,
        user_id=me.id,
        action="organization.legal_consents.export",
        resource_type="organization",
        resource_id=str(organization.id),
        extra={"format": format, "row_count": len(rows)},
    )
    await db.commit()
    return _audit_csv_response(rows, f"organization-consent-logs-{organization.id}.csv")


@app.get("/api/superadmin/security-events", dependencies=[Depends(require_role(Role.superadmin))])
async def get_security_events(db: AsyncSession = Depends(get_db)):
    cutoff = datetime.now(timezone.utc) - timedelta(hours=24)
    q = _audit_category_filter(select(AuditLog), "security").where(AuditLog.created_at >= cutoff)
    logs = (await db.execute(q.order_by(AuditLog.created_at.desc()).limit(500))).scalars().all()
    by_action: Dict[str, int] = {}
    by_ip: Dict[str, int] = {}
    for log in logs:
        by_action[log.action] = by_action.get(log.action, 0) + 1
        if log.ip_address:
            by_ip[str(log.ip_address)] = by_ip.get(str(log.ip_address), 0) + 1
    return {
        "total_24h": len(logs),
        "by_action": by_action,
        "suspicious_ips": [{"ip": ip, "count": count} for ip, count in sorted(by_ip.items(), key=lambda item: item[1], reverse=True)[:20] if count >= 3],
        "items": [_audit_row_payload(log) for log in logs[:100]],
    }


@app.get("/api/admin/organization/venue-reservations/google-calendar/status", dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def organization_google_calendar_status(
    request: Request,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from .organization_access_api import get_organization_for_access, organization_id_from_request

    await get_organization_for_access(db, me, "reservations:read", organization_id_from_request(request))
    integration = await _get_user_google_integration(db, me.id)
    scopes = _normalize_google_scopes(integration.scopes if integration else [])
    missing_scopes = _google_calendar_missing_scopes(scopes) if integration else GOOGLE_CALENDAR_REQUIRED_SCOPES
    return {
        "configured": bool(settings.google_oauth_client_id and settings.google_oauth_client_secret),
        "connected": bool(integration and integration.refresh_token and not missing_scopes),
        "google_email": integration.google_email if integration else None,
        "missing_scopes": missing_scopes,
    }


@app.get("/api/admin/organization/venue-reservations/google-calendar/start", dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def organization_google_calendar_start(
    request: Request,
    next: Optional[str] = Query(default="/admin/settings?tab=venues"),
    frontend_origin: Optional[str] = Query(default=None),
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from .organization_access_api import get_organization_for_access, organization_id_from_request

    await get_organization_for_access(db, me, "reservations:write", organization_id_from_request(request))
    if not settings.google_oauth_client_id or not settings.google_oauth_client_secret:
        raise HTTPException(status_code=503, detail="Google OAuth is not configured.")
    state = make_email_token({
        "action": "google_calendar_oauth",
        "user_id": me.id,
        "next": _normalize_oauth_next(next, "/admin/settings?tab=venues"),
        "frontend_origin": _normalize_oauth_frontend_origin(frontend_origin),
    })
    params = {
        "client_id": settings.google_oauth_client_id,
        "redirect_uri": _google_sheets_redirect_uri(),
        "response_type": "code",
        "scope": " ".join(GOOGLE_CALENDAR_SCOPES),
        "state": state,
        "access_type": "offline",
        "prompt": "consent select_account",
        "include_granted_scopes": "true",
    }
    return {"authorization_url": f"https://accounts.google.com/o/oauth2/v2/auth?{urlencode(params)}"}






async def _sync_organization_reservations_with_google_calendar(db: AsyncSession, organization_id: int, calendar_user_id: int) -> Dict[str, int]:
    from .venue_reservations_api import VenueReservation
    from .venues_api import OrganizationVenue

    access_token = await _get_google_access_token_for_calendar(db, calendar_user_id)
    res = await db.execute(
        select(VenueReservation, OrganizationVenue)
        .join(OrganizationVenue, OrganizationVenue.id == VenueReservation.venue_id)
        .where(VenueReservation.organization_id == organization_id, VenueReservation.status == "confirmed")
        .order_by(VenueReservation.start_at.asc())
    )
    rows = res.all()
    reservations_by_id = {reservation.id: reservation for reservation, _venue in rows}
    pulled = await _pull_google_calendar_reservations(db, access_token, organization_id, reservations_by_id)
    await db.flush()

    pushed = 0
    updated = 0
    for reservation, venue in rows:
        body = {
            "summary": reservation.title,
            "description": reservation.description or "",
            "location": f"{venue.name}{' - ' + venue.location if venue.location else ''}",
            "start": {"dateTime": ensure_utc(reservation.start_at).isoformat()},
            "end": {"dateTime": ensure_utc(reservation.end_at).isoformat()},
            "extendedProperties": {
                "private": {
                    "heptacert_source": "venue_reservation",
                    "heptacert_reservation_id": str(reservation.id),
                    "heptacert_organization_id": str(organization_id),
                }
            },
        }
        if reservation.external_event_id:
            await _google_json_request(
                access_token,
                "PATCH",
                f"https://www.googleapis.com/calendar/v3/calendars/primary/events/{quote(reservation.external_event_id, safe='')}",
                json_body=body,
            )
            updated += 1
        else:
            created = await _google_json_request(
                access_token,
                "POST",
                "https://www.googleapis.com/calendar/v3/calendars/primary/events",
                json_body=body,
            )
            reservation.calendar_provider = "google"
            reservation.external_event_id = str(created.get("id") or "")
            pushed += 1
    await db.commit()
    return {"pulled": pulled, "pushed": pushed, "updated": updated}


@app.post("/api/admin/organization/venue-reservations/google-calendar/sync", dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def organization_google_calendar_sync(
    request: Request,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from .organization_access_api import get_organization_for_access, organization_id_from_request

    organization = await get_organization_for_access(db, me, "reservations:write", organization_id_from_request(request))
    result = await _sync_organization_reservations_with_google_calendar(db, organization.id, me.id)
    await write_audit_log(
        db,
        user_id=me.id,
        action="organization.reservation.google_calendar_sync",
        resource_type="organization",
        resource_id=str(organization.id),
        extra=result,
    )
    await db.commit()
    return {"ok": True, **result}


# ═════ SUPERADMIN: ORGANIZATIONS ═════













@app.get("/api/superadmin/organizations", dependencies=[Depends(require_role(Role.superadmin))])
async def get_organizations(
    page: int = Query(default=1, ge=1),
    limit: int = Query(default=20, le=100),
    db: AsyncSession = Depends(get_db),
):
    """Get paginated organization records - superadmin only."""
    q = select(Organization)
    
    total_res = await db.execute(select(func.count()).select_from(q.subquery()))
    total = int(total_res.scalar_one() or 0)
    
    q = q.order_by(Organization.created_at.desc()).offset((page - 1) * limit).limit(limit)
    res = await db.execute(q)
    organizations = res.scalars().all()
    domains_by_name: dict[str, Any] = {}
    custom_domains = [org.custom_domain for org in organizations if org.custom_domain]
    if custom_domains:
        from .domains import Domain

        domain_res = await db.execute(select(Domain).where(Domain.domain.in_(custom_domains)))
        domains_by_name = {dom.domain: dom for dom in domain_res.scalars().all()}
    
    return {
        "total": total,
        "page": page,
        "limit": limit,
        "items": [_serialize_superadmin_org(org, domains_by_name.get(org.custom_domain or "")) for org in organizations],
    }


@app.post("/api/superadmin/organizations", dependencies=[Depends(require_role(Role.superadmin))])
async def create_organization(
    payload: OrgIn,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if not payload.user_id:
        raise HTTPException(status_code=400, detail="Admin user ID is required")

    user_res = await db.execute(select(User).where(User.id == payload.user_id))
    user = user_res.scalar_one_or_none()
    if not user or user.role not in {Role.admin, Role.superadmin}:
        raise HTTPException(status_code=404, detail="Admin user not found")

    existing_res = await db.execute(select(Organization).where(Organization.user_id == payload.user_id))
    if existing_res.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Organization already exists for this user")

    custom_domain = payload.custom_domain.strip().lower() if payload.custom_domain else None
    if custom_domain:
        domain_res = await db.execute(select(Organization.id).where(Organization.custom_domain == custom_domain))
        if domain_res.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="Custom domain is already in use")

    org = Organization(
        user_id=payload.user_id,
        public_id=await _generate_organization_public_id(db),
        org_name=payload.org_name.strip(),
        custom_domain=custom_domain,
        brand_logo=payload.brand_logo.strip() if payload.brand_logo else None,
        brand_color=payload.brand_color,
        settings={},
    )
    db.add(org)
    await db.flush()
    await _ensure_domain_row_for_org(db, org)
    await write_audit_log(db, user_id=me.id, action="organization.create", resource_type="organization", resource_id=str(payload.user_id))
    await db.commit()
    await db.refresh(org)
    return _serialize_superadmin_org(org, await _domain_row_for_org(db, org))


@app.patch("/api/superadmin/organizations/{org_id}", dependencies=[Depends(require_role(Role.superadmin))])
async def update_organization(
    org_id: int,
    payload: OrgIn,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    org_res = await db.execute(select(Organization).where(Organization.id == org_id))
    org = org_res.scalar_one_or_none()
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")

    if payload.user_id and payload.user_id != org.user_id:
        user_res = await db.execute(select(User).where(User.id == payload.user_id))
        user = user_res.scalar_one_or_none()
        if not user or user.role not in {Role.admin, Role.superadmin}:
            raise HTTPException(status_code=404, detail="Admin user not found")
        existing_res = await db.execute(
            select(Organization.id).where(Organization.user_id == payload.user_id, Organization.id != org_id)
        )
        if existing_res.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="Organization already exists for this user")
        org.user_id = payload.user_id

    old_domain = org.custom_domain
    custom_domain = payload.custom_domain.strip().lower() if payload.custom_domain else None
    if custom_domain:
        domain_res = await db.execute(
            select(Organization.id).where(Organization.custom_domain == custom_domain, Organization.id != org_id)
        )
        if domain_res.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="Custom domain is already in use")

    org.org_name = payload.org_name.strip()
    org.custom_domain = custom_domain
    org.brand_logo = payload.brand_logo.strip() if payload.brand_logo else None
    org.brand_color = payload.brand_color
    await _sync_superadmin_org_domain(db, org, old_domain)
    await write_audit_log(db, user_id=me.id, action="organization.update", resource_type="organization", resource_id=str(org_id))
    await db.commit()
    await db.refresh(org)
    return _serialize_superadmin_org(org, await _domain_row_for_org(db, org))


@app.post("/api/superadmin/organizations/{org_id}/domain/approve", dependencies=[Depends(require_role(Role.superadmin))])
async def approve_organization_domain(
    org_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from .domains import DomainStatus

    org_res = await db.execute(select(Organization).where(Organization.id == org_id))
    org = org_res.scalar_one_or_none()
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")
    if not org.custom_domain:
        raise HTTPException(status_code=400, detail="Organization has no custom domain")

    dom = await _ensure_domain_row_for_org(db, org)
    dom.status = DomainStatus.active
    db.add(dom)
    await write_audit_log(
        db,
        user_id=me.id,
        action="organization.domain.approve",
        resource_type="organization",
        resource_id=str(org.id),
        extra={"domain": org.custom_domain},
    )
    await db.commit()
    await db.refresh(org)
    return _serialize_superadmin_org(org, await _domain_row_for_org(db, org))


@app.post("/api/superadmin/organizations/{org_id}/domain/revoke", dependencies=[Depends(require_role(Role.superadmin))])
async def revoke_organization_domain(
    org_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from .domains import Domain, DomainStatus

    org_res = await db.execute(select(Organization).where(Organization.id == org_id))
    org = org_res.scalar_one_or_none()
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")

    old_domain = org.custom_domain
    if old_domain:
        dom = await Domain.get_by_domain(db, old_domain)
        if dom:
            dom.status = DomainStatus.revoked
            db.add(dom)
    org.custom_domain = None
    await write_audit_log(
        db,
        user_id=me.id,
        action="organization.domain.revoke",
        resource_type="organization",
        resource_id=str(org.id),
        extra={"domain": old_domain},
    )
    await db.commit()
    await db.refresh(org)
    return _serialize_superadmin_org(org, None)


@app.delete("/api/superadmin/organizations/{org_id}", dependencies=[Depends(require_role(Role.superadmin))])
async def delete_organization(
    org_id: int,
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    org_res = await db.execute(select(Organization).where(Organization.id == org_id))
    org = org_res.scalar_one_or_none()
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")
    if org.custom_domain:
        from .domains import Domain, DomainStatus

        dom = await Domain.get_by_domain(db, org.custom_domain)
        if dom:
            dom.status = DomainStatus.revoked
            db.add(dom)
    await db.delete(org)
    await write_audit_log(db, user_id=me.id, action="organization.delete", resource_type="organization", resource_id=str(org_id))
    await db.commit()
    return {"ok": True}


from . import email_api as _email_api
app.include_router(_email_api.router)

from . import auth_2fa_api as _auth_2fa_api
app.include_router(_auth_2fa_api.router)

from . import document_outputs_api as _document_outputs_api
app.include_router(_document_outputs_api.router)

from . import document_export_jobs as _document_export_jobs
app.include_router(_document_export_jobs.router)

from . import community_api as _community_api
app.include_router(_community_api.router)

from . import social_api as _social_api
app.include_router(_social_api.router)

from . import connections_api as _connections_api
app.include_router(_connections_api.router)

from . import member_certificates_api as _member_certificates_api
app.include_router(_member_certificates_api.router)

from . import certificate_templates_api as _certificate_templates_api
app.include_router(_certificate_templates_api.router)

from . import automation_api as _automation_api
app.include_router(_automation_api.router)

from . import audience_segments_api as _audience_segments_api
app.include_router(_audience_segments_api.router)

from . import oidc_sso_api as _oidc_sso_api
app.include_router(_oidc_sso_api.router)

from . import event_crm_api as _event_crm_api
app.include_router(_event_crm_api.router)

from . import crm_sequences_api as _crm_sequences_api
app.include_router(_crm_sequences_api.router)

from . import training_api as _training_api
app.include_router(_training_api.router)

from . import checkin_ops_api as _checkin_ops_api
app.include_router(_checkin_ops_api.router)

from . import platform_health_api as _platform_health_api
app.include_router(_platform_health_api.router)

from . import product_telemetry_api as _product_telemetry_api
app.include_router(_product_telemetry_api.router)

from . import qa_seed_api as _qa_seed_api
app.include_router(_qa_seed_api.router)

from .plan_policy import feature_policy_payload as _feature_policy_payload

@app.get("/api/feature-policies")
async def feature_policies():
    return _feature_policy_payload()

from .product_observability import install_product_observability as _install_product_observability
_install_product_observability(app)

# API module imports - these are loaded after all models are defined
from . import analytics_api as _analytics_api
app.include_router(_analytics_api.router)

from . import tickets_api as _tickets_api
app.include_router(_tickets_api.router)

from . import organization_access_api as _organization_access_api
app.include_router(_organization_access_api.router)

from . import venues_api as _venues_api
app.include_router(_venues_api.router)

from . import venue_reservations_api as _venue_reservations_api
app.include_router(_venue_reservations_api.router)

from . import notification_integrations_api as _notification_integrations_api
app.include_router(_notification_integrations_api.router)

from . import oauth_api as _oauth_api
app.include_router(_oauth_api.router)

from . import ai_content_api as _ai_content_api
app.include_router(_ai_content_api.router)

from . import ai_proactive_api as _ai_proactive_api
app.include_router(_ai_proactive_api.router)

from . import event_extras_api as _event_extras_api
app.include_router(_event_extras_api.router)


# ── Background job status dashboard ──────────────────────────────────────────

@app.get("/api/superadmin/job-status", dependencies=[Depends(require_role(Role.superadmin))])
async def get_job_status(db: AsyncSession = Depends(get_db)):
    """Real-time status of all background job queues — for ops/monitoring."""
    now = datetime.now(timezone.utc)
    since_hour = now - timedelta(hours=1)

    bulk_pending = (await db.execute(select(func.count(BulkEmailJob.id)).where(BulkEmailJob.status == "pending"))).scalar_one()
    bulk_processing = (await db.execute(select(func.count(BulkEmailJob.id)).where(BulkEmailJob.status == "processing"))).scalar_one()
    bulk_failed_recent = (await db.execute(select(func.count(BulkEmailJob.id)).where(BulkEmailJob.status == "failed", BulkEmailJob.created_at >= since_hour))).scalar_one()

    from .audience_segments_api import SegmentExportJob
    seg_pending = (await db.execute(select(func.count(SegmentExportJob.id)).where(SegmentExportJob.status == "pending"))).scalar_one()
    seg_processing = (await db.execute(select(func.count(SegmentExportJob.id)).where(SegmentExportJob.status == "processing"))).scalar_one()

    from .document_export_jobs import DocumentExportJob
    doc_pending = (await db.execute(select(func.count(DocumentExportJob.id)).where(DocumentExportJob.status == "pending"))).scalar_one()

    bulk_cert_pending = (await db.execute(select(func.count(BulkCertificateJob.id)).where(BulkCertificateJob.status == "pending"))).scalar_one()
    bulk_cert_processing = (await db.execute(select(func.count(BulkCertificateJob.id)).where(BulkCertificateJob.status == "processing"))).scalar_one()

    from .training_api import TrainingRenewalNotificationLog
    notif_failed_recent = (await db.execute(select(func.count(TrainingRenewalNotificationLog.id)).where(
        TrainingRenewalNotificationLog.status == "failed",
        TrainingRenewalNotificationLog.created_at >= since_hour,
    ))).scalar_one()

    return {
        "timestamp": now.isoformat(),
        "bulk_email": {"pending": int(bulk_pending), "processing": int(bulk_processing), "failed_last_hour": int(bulk_failed_recent)},
        "segment_export": {"pending": int(seg_pending), "processing": int(seg_processing)},
        "document_export": {"pending": int(doc_pending)},
        "certificate_bulk": {"pending": int(bulk_cert_pending), "processing": int(bulk_cert_processing)},
        "training_notifications": {"failed_last_hour": int(notif_failed_recent)},
        # scheduler_enabled reflects THIS instance only. In production the scheduler
        # runs exclusively in the backend_jobs container (ENABLE_SCHEDULER=true there).
        "scheduler_enabled_this_instance": settings.enable_scheduler,
        "scheduler_enabled": True,  # jobs always run somewhere in the deployment
    }


# ── Unified Job Center ────────────────────────────────────────────────────────

@app.get("/api/admin/jobs", dependencies=[Depends(require_role(Role.admin, Role.superadmin))])
async def list_my_jobs(
    me: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    limit: int = Query(default=40, ge=1, le=100),
):
    """Return recent + all active jobs across all types for the current user."""
    from .audience_segments_api import SegmentExportJob
    from .document_export_jobs import DocumentExportJob

    now = datetime.now(timezone.utc)
    since = now - timedelta(days=7)
    jobs: list[dict] = []

    # ── Bulk email jobs ──────────────────────────────────────────────────
    email_rows = (
        await db.execute(
            select(BulkEmailJob, Event)
            .join(Event, Event.id == BulkEmailJob.event_id)
            .where(
                BulkEmailJob.created_by == me.id,
                or_(
                    BulkEmailJob.status.in_(["pending", "sending", "in_progress"]),
                    BulkEmailJob.created_at >= since,
                ),
            )
            .order_by(BulkEmailJob.created_at.desc())
            .limit(limit)
        )
    ).all()
    for job, event in email_rows:
        total = int(job.recipients_count or 0)
        sent = int(job.sent_count or 0)
        failed = int(job.failed_count or 0)
        jobs.append({
            "id": job.id,
            "type": "bulk_email",
            "type_label": "Toplu E-posta",
            "event_id": job.event_id,
            "event_name": event.name if event else None,
            "status": job.status,
            "total": total,
            "done": sent + failed,
            "success": sent,
            "failed": failed,
            "progress_pct": round((sent + failed) / total * 100) if total else 0,
            "created_at": job.created_at.isoformat() if job.created_at else None,
            "started_at": job.started_at.isoformat() if job.started_at else None,
            "completed_at": job.completed_at.isoformat() if job.completed_at else None,
            "error_message": job.error_message,
            "detail_url": f"/admin/events/{job.event_id}/analytics/{job.id}",
            "can_cancel": False,
            "can_download": False,
        })

    # ── Bulk certificate jobs ────────────────────────────────────────────
    cert_rows = (
        await db.execute(
            select(BulkCertificateJob, Event)
            .join(Event, Event.id == BulkCertificateJob.event_id)
            .where(
                BulkCertificateJob.created_by == me.id,
                or_(
                    BulkCertificateJob.status.in_(["pending", "processing"]),
                    BulkCertificateJob.created_at >= since,
                ),
            )
            .order_by(BulkCertificateJob.created_at.desc())
            .limit(limit)
        )
    ).all()
    for job, event in cert_rows:
        total = int(job.total_count or 0)
        done = int(job.current_index or 0)
        jobs.append({
            "id": job.id,
            "type": "bulk_certificate",
            "type_label": "Toplu Sertifika",
            "event_id": job.event_id,
            "event_name": event.name if event else None,
            "status": job.status,
            "total": total,
            "done": done,
            "success": int(job.created_count or 0),
            "failed": int(job.failed_count or 0),
            "already_exists": int(job.already_exists_count or 0),
            "progress_pct": round(done / total * 100) if total else 0,
            "created_at": job.created_at.isoformat() if job.created_at else None,
            "started_at": job.started_at.isoformat() if job.started_at else None,
            "completed_at": job.completed_at.isoformat() if job.completed_at else None,
            "error_message": job.error_message,
            "detail_url": f"/admin/events/{job.event_id}/certificates",
            "download_url": f"/api/admin/events/{job.event_id}/bulk-generate-jobs/{job.id}/download" if job.status == "completed" and job.zip_file_path else None,
            "can_cancel": job.status in ("pending", "processing"),
            "can_download": job.status == "completed" and bool(job.zip_file_path),
        })

    # ── Segment export jobs ──────────────────────────────────────────────
    seg_rows = (
        await db.execute(
            select(SegmentExportJob, Event)
            .join(Event, Event.id == SegmentExportJob.event_id)
            .where(
                SegmentExportJob.created_by == me.id,
                or_(
                    SegmentExportJob.status.in_(["pending", "processing"]),
                    SegmentExportJob.created_at >= since,
                ),
            )
            .order_by(SegmentExportJob.created_at.desc())
            .limit(limit)
        )
    ).all()
    for job, event in seg_rows:
        jobs.append({
            "id": job.id,
            "type": "segment_export",
            "type_label": "Segment Export",
            "event_id": job.event_id,
            "event_name": event.name if event else None,
            "status": job.status,
            "total": int(job.row_count or 0),
            "done": int(job.row_count or 0) if job.status == "completed" else 0,
            "success": int(job.row_count or 0),
            "failed": 0,
            "progress_pct": 100 if job.status == "completed" else (50 if job.status == "processing" else 0),
            "created_at": job.created_at.isoformat() if job.created_at else None,
            "started_at": job.started_at.isoformat() if job.started_at else None,
            "completed_at": job.completed_at.isoformat() if job.completed_at else None,
            "error_message": job.error_message,
            "detail_url": f"/admin/events/{job.event_id}/segments",
            "download_url": f"/api/admin/events/{job.event_id}/segments/export-jobs/{job.id}/download" if job.status == "completed" and job.file_path else None,
            "can_cancel": False,
            "can_download": job.status == "completed" and bool(job.file_path),
        })

    jobs.sort(key=lambda j: j.get("created_at") or "", reverse=True)
    active_count = sum(1 for j in jobs if j["status"] in ("pending", "processing", "sending", "in_progress"))
    return {"jobs": jobs[:limit], "active_count": active_count}


# ── Built-in badge template gallery ──────────────────────────────────────────

@app.get(
    "/api/admin/badge-templates",
    dependencies=[Depends(require_role(Role.admin, Role.superadmin))],
    summary="List built-in badge template presets",
)
async def list_builtin_badge_templates(lang: str = "tr"):
    """
    Return the platform-level badge template gallery.
    Organizers pick from these to pre-fill their event's badge definitions.
    """
    from .badge_template_seeds import get_builtin_badge_templates  # noqa: PLC0415
    templates = await get_builtin_badge_templates(lang=lang if lang in ("tr", "en") else "tr")
    return {"templates": templates, "total": len(templates)}


# ── Feature module routers ────────────────────────────────────────────────────

from . import quiz_api as _quiz_api  # noqa: E402
app.include_router(_quiz_api.router)

from . import learning_path_api as _learning_path_api  # noqa: E402
app.include_router(_learning_path_api.router)

from . import crm_accounts_api as _crm_accounts_api  # noqa: E402
app.include_router(_crm_accounts_api.router)

from . import lead_forms_api as _lead_forms_api  # noqa: E402
app.include_router(_lead_forms_api.router)

# LMS sistemi devre disi birakildi — arsivlendi.
# API router'lari backend/_archive_lms/ klasorune tasindi (lms_api.py, lms_extended_api.py).
# Frontend sayfalari frontend/_archive_lms/ klasorunde. Yeniden aktive etmek icin
# o klasordeki README.md'ye bakin.
# Model import zorunlu: marketplace_api lms_models'i yukluyor, CourseEnrollment->CourseGradeSummary
# iliskisi cozumlenemezse SQLAlchemy startup'ta kiliyor — bu yuzden modeller src/ icinde kaliyor.
from . import lms_extended_models as _lms_extended_models  # noqa: E402, F401

from . import org_modules_api as _org_modules_api  # noqa: E402
app.include_router(_org_modules_api.router)

from . import org_staff_api as _org_staff_api  # noqa: E402
app.include_router(_org_staff_api.router)

from . import org_analytics_api as _org_analytics_api  # noqa: E402
app.include_router(_org_analytics_api.router)

from . import report_scheduler_api as _report_scheduler_api  # noqa: E402
app.include_router(_report_scheduler_api.router)

from . import marketplace_api as _marketplace_api  # noqa: E402
app.include_router(_marketplace_api.router)

from . import lti_api as _lti_api  # noqa: E402
app.include_router(_lti_api.router)

from . import sso_api as _sso_api  # noqa: E402
app.include_router(_sso_api.router)

from . import api_keys_ext_api as _api_keys_ext_api  # noqa: E402
app.include_router(_api_keys_ext_api.router)

from . import accreditation_api as _accreditation_api  # noqa: E402
app.include_router(_accreditation_api.router)

from . import raffles_api as _raffles_api  # routers refactor 4d
app.include_router(_raffles_api.router)

from . import event_sponsors_api as _event_sponsors_api  # routers refactor 4d
app.include_router(_event_sponsors_api.router)

from . import certificate_tiers_api as _certificate_tiers_api  # routers refactor 4d
app.include_router(_certificate_tiers_api.router)

from . import survey_config_api as _survey_config_api  # routers refactor 4d
app.include_router(_survey_config_api.router)

from . import template_history_api as _template_history_api  # routers refactor 4d
app.include_router(_template_history_api.router)

from . import ms_excel_api as _ms_excel_api  # 4d
app.include_router(_ms_excel_api.router)

from . import bulk_generate_api as _bulk_generate_api  # 4d
app.include_router(_bulk_generate_api.router)

# ── MCP hosted endpoint (/mcp) ─────────────────────────────────────────────────
# Agents connect to https://yourapp.com/mcp with Authorization: Bearer hc_live_...
# Each request is authenticated per-user via the MCP server's Context-based auth.
from . import mcp_server as _mcp_server_module  # noqa: E402

app.mount("/mcp", _mcp_server_module.mcp.streamable_http_app())



