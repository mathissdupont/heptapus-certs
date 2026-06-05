import csv
import io
from typing import Any, List, Tuple

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import distinct, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from .main import (
    AttendaonceRecord,
    Attendee,
    CertStatus,
    Certificate,
    Event,
    EventSession,
    EventTicket,
    ParticipantBadge,
    Role,
    User,
    VerificationHit,
    get_current_user,
    get_db,
    is_certificate_enabled,
    is_checkin_enabled,
    is_gamification_enabled,
    is_public_registration_enabled,
    is_raffles_enabled,
    is_ticketing_enabled,
    normalize_event_type,
)

router = APIRouter()


@router.get("/api/admin/events/{event_id}/analytics")
async def get_event_analytics(
    event_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Get aggregated analytics for an event (attendees, certs, sessions)."""
    e_res = await db.execute(select(Event).where(Event.id == event_id))
    event = e_res.scalar_one_or_none()
    if not event:
        raise HTTPException(status_code=404, detail="Etkinlik bulunamad\u0131")

    if event.admin_id != current_user.id and current_user.role != Role.superadmin:
        raise HTTPException(status_code=403, detail="Yetkisiz eri\u015fim")

    att_res = await db.execute(
        select(func.count(Attendee.id)).where(Attendee.event_id == event_id)
    )
    total_attendees = att_res.scalar() or 0

    cert_res = await db.execute(
        select(func.count(Certificate.id)).where(
            Certificate.event_id == event_id,
            Certificate.status == CertStatus.active,
        )
    )
    certified_count = cert_res.scalar() or 0
    pending_count = max(0, total_attendees - certified_count)

    sess_res = await db.execute(
        select(EventSession).where(EventSession.event_id == event_id).order_by(EventSession.id)
    )
    sessions = sess_res.scalars().all()

    session_data = []
    for session in sessions:
        arc_res = await db.execute(
            select(func.count(distinct(AttendaonceRecord.attendee_id))).where(
                AttendaonceRecord.session_id == session.id
            )
        )
        attended = arc_res.scalar() or 0
        rate = (attended / total_attendees) if total_attendees > 0 else 0.0
        session_data.append({"id": session.id, "name": session.name, "attendance_rate": rate})

    return {
        "event_id": event.id,
        "event_name": event.name,
        "total_attendees": total_attendees,
        "certified_count": certified_count,
        "pending_count": pending_count,
        "sessions": session_data,
    }


@router.get("/api/admin/events/{event_id}/analytics/engagement")
async def get_engagement_analytics(
    event_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Get engagement analytics for an event (attendance, surveys, badges)."""
    e_res = await db.execute(select(Event).where(Event.id == event_id))
    event = e_res.scalar_one_or_none()
    if not event:
        raise HTTPException(status_code=404, detail="Etkinlik bulunamad\u0131")

    if event.admin_id != current_user.id and current_user.role != Role.superadmin:
        raise HTTPException(status_code=403, detail="Yetkisiz eri\u015fim")

    att_count_res = await db.execute(
        select(func.count(Attendee.id)).where(Attendee.event_id == event_id)
    )
    total_attendees = att_count_res.scalar() or 0

    survey_completed_res = await db.execute(
        select(func.count(Attendee.id)).where(
            Attendee.event_id == event_id,
            Attendee.survey_completed_at.isnot(None),
        )
    )
    surveys_completed = survey_completed_res.scalar() or 0

    pb_res = await db.execute(
        select(func.count(ParticipantBadge.id)).where(
            ParticipantBadge.event_id == event_id
        )
    )
    total_badges = pb_res.scalar() or 0

    arc_res = await db.execute(
        select(func.count(distinct(AttendaonceRecord.attendee_id))).where(
            AttendaonceRecord.id.in_(
                select(AttendaonceRecord.id).join(
                    EventSession,
                    EventSession.id == AttendaonceRecord.session_id,
                ).where(EventSession.event_id == event_id)
            )
        )
    )
    attended_count = arc_res.scalar() or 0
    ticketing_enabled = is_ticketing_enabled(event)
    ticket_counts = {
        "total": 0,
        "active_total": 0,
        "issued": 0,
        "used": 0,
        "cancelled": 0,
        "revoked": 0,
        "no_show": 0,
        "no_show_rate": 0,
        "usage_rate": 0,
    }
    if ticketing_enabled:
        ticket_status_res = await db.execute(
            select(EventTicket.status, func.count(EventTicket.id))
            .where(EventTicket.event_id == event_id)
            .group_by(EventTicket.status)
        )
        by_status = {str(status): int(count or 0) for status, count in ticket_status_res.all()}
        ticket_counts.update(
            {
                "issued": by_status.get("issued", 0),
                "used": by_status.get("used", 0),
                "cancelled": by_status.get("cancelled", 0),
                "revoked": by_status.get("revoked", 0),
            }
        )
        ticket_counts["total"] = sum(int(ticket_counts[key]) for key in ("issued", "used", "cancelled", "revoked"))
        ticket_counts["active_total"] = int(ticket_counts["issued"]) + int(ticket_counts["used"])
        ticket_counts["no_show"] = int(ticket_counts["issued"])
        ticket_counts["usage_rate"] = (
            ticket_counts["used"] / ticket_counts["active_total"] * 100
            if ticket_counts["active_total"] > 0
            else 0
        )
        ticket_counts["no_show_rate"] = (
            ticket_counts["no_show"] / ticket_counts["active_total"] * 100
            if ticket_counts["active_total"] > 0
            else 0
        )

        ticket_attended_res = await db.execute(
            select(func.count(distinct(EventTicket.attendee_id))).where(
                EventTicket.event_id == event_id,
                EventTicket.status == "used",
            )
        )
        attended_count = max(int(attended_count or 0), int(ticket_attended_res.scalar() or 0))

    not_attended_count = max(0, int(total_attendees or 0) - int(attended_count or 0))

    return {
        "event_type": normalize_event_type(getattr(event, "event_type", None)),
        "certificate_enabled": is_certificate_enabled(event),
        "checkin_enabled": is_checkin_enabled(event),
        "ticketing_enabled": ticketing_enabled,
        "registration_enabled": is_public_registration_enabled(event),
        "raffles_enabled": is_raffles_enabled(event),
        "gamification_enabled": is_gamification_enabled(event),
        "total_attendees": total_attendees,
        "survey_completion": {
            "completed": surveys_completed,
            "pending": total_attendees - surveys_completed,
            "completion_rate": (surveys_completed / total_attendees * 100) if total_attendees > 0 else 0,
        },
        "badges": {
            "total_awarded": total_badges,
            "average_per_attendee": (total_badges / total_attendees) if total_attendees > 0 else 0,
        },
        "attendance": {
            "attended": attended_count,
            "not_attended": not_attended_count,
            "attendance_rate": (attended_count / total_attendees * 100) if total_attendees > 0 else 0,
            "no_show_rate": (not_attended_count / total_attendees * 100) if total_attendees > 0 else 0,
        },
        "tickets": ticket_counts,
    }


@router.get("/api/admin/events/{event_id}/analytics/badges")
async def get_badge_analytics(
    event_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Get badge distribution analytics."""
    e_res = await db.execute(select(Event).where(Event.id == event_id))
    event = e_res.scalar_one_or_none()
    if not event:
        raise HTTPException(status_code=404, detail="Etkinlik bulunamad\u0131")

    if event.admin_id != current_user.id and current_user.role != Role.superadmin:
        raise HTTPException(status_code=403, detail="Yetkisiz eri\u015fim")

    pb_res = await db.execute(
        select(ParticipantBadge.badge_type, func.count(ParticipantBadge.id))
        .where(ParticipantBadge.event_id == event_id)
        .group_by(ParticipantBadge.badge_type)
    )
    badge_counts = pb_res.all()

    auto_res = await db.execute(
        select(func.count(ParticipantBadge.id)).where(
            ParticipantBadge.event_id == event_id,
            ParticipantBadge.is_automatic == True,
        )
    )
    automatic = auto_res.scalar() or 0

    manual_res = await db.execute(
        select(func.count(ParticipantBadge.id)).where(
            ParticipantBadge.event_id == event_id,
            ParticipantBadge.is_automatic == False,
        )
    )
    manual = manual_res.scalar() or 0

    return {
        "by_type": {badge_type: count for badge_type, count in badge_counts},
        "by_award_method": {
            "automatic": automatic,
            "manual": manual,
        },
        "total_badges": automatic + manual,
    }


@router.get("/api/admin/events/{event_id}/analytics/tiers")
async def get_tier_analytics(
    event_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Get certificate tier distribution analytics."""
    e_res = await db.execute(select(Event).where(Event.id == event_id))
    event = e_res.scalar_one_or_none()
    if not event:
        raise HTTPException(status_code=404, detail="Etkinlik bulunamad\u0131")

    if event.admin_id != current_user.id and current_user.role != Role.superadmin:
        raise HTTPException(status_code=403, detail="Yetkisiz eri\u015fim")

    cert_res = await db.execute(
        select(Certificate.certificate_tier, func.count(Certificate.id))
        .where(
            Certificate.event_id == event_id,
            Certificate.status == CertStatus.active,
        )
        .group_by(Certificate.certificate_tier)
    )
    tier_counts = cert_res.all()

    total = sum(count for _, count in tier_counts)
    tier_dist = {}
    for tier_name, count in tier_counts:
        tier_key = tier_name or "Unassigned"
        percentage = (count / total * 100) if total > 0 else 0
        tier_dist[tier_key] = {
            "count": count,
            "percentage": round(percentage, 2),
        }

    verification_hits_res = await db.execute(
        select(func.count(VerificationHit.id))
        .join(Certificate, Certificate.uuid == VerificationHit.cert_uuid)
        .where(
            Certificate.event_id == event_id,
            Certificate.status == CertStatus.active,
        )
    )
    verification_hits = int(verification_hits_res.scalar() or 0)

    verified_certificates_res = await db.execute(
        select(func.count(distinct(Certificate.id)))
        .join(VerificationHit, Certificate.uuid == VerificationHit.cert_uuid)
        .where(
            Certificate.event_id == event_id,
            Certificate.status == CertStatus.active,
        )
    )
    verified_certificates = int(verified_certificates_res.scalar() or 0)

    return {
        "total_certificates": total,
        "tier_distribution": tier_dist,
        "unassigned_count": tier_dist.get("Unassigned", {}).get("count", 0),
        "verification_hits": verification_hits,
        "verified_certificates": verified_certificates,
        "verification_rate": (verified_certificates / total * 100) if total > 0 else 0,
    }


@router.get("/api/admin/events/{event_id}/analytics/timeline")
async def get_timeline_analytics(
    event_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Get timeline analytics (registrations, completions, downloads over time)."""
    e_res = await db.execute(select(Event).where(Event.id == event_id))
    event = e_res.scalar_one_or_none()
    if not event:
        raise HTTPException(status_code=404, detail="Etkinlik bulunamad\u0131")

    if event.admin_id != current_user.id and current_user.role != Role.superadmin:
        raise HTTPException(status_code=403, detail="Yetkisiz eri\u015fim")

    reg_res = await db.execute(
        select(
            func.date(Attendee.registered_at).label("date"),
            func.count(Attendee.id).label("count"),
        )
        .where(Attendee.event_id == event_id)
        .group_by(func.date(Attendee.registered_at))
        .order_by("date")
    )
    registrations = reg_res.all()

    surv_res = await db.execute(
        select(
            func.date(Attendee.survey_completed_at).label("date"),
            func.count(Attendee.id).label("count"),
        )
        .where(
            Attendee.event_id == event_id,
            Attendee.survey_completed_at.isnot(None),
        )
        .group_by(func.date(Attendee.survey_completed_at))
        .order_by("date")
    )
    surveys = surv_res.all()

    cert_res = await db.execute(
        select(
            func.date(Certificate.created_at).label("date"),
            func.count(Certificate.id).label("count"),
        )
        .where(
            Certificate.event_id == event_id,
            Certificate.status == CertStatus.active,
        )
        .group_by(func.date(Certificate.created_at))
        .order_by("date")
    )
    certificates = cert_res.all()

    ticket_checkins: List[Tuple[Any, int]] = []
    if is_ticketing_enabled(event):
        ticket_res = await db.execute(
            select(
                func.date(EventTicket.checked_in_at).label("date"),
                func.count(EventTicket.id).label("count"),
            )
            .where(
                EventTicket.event_id == event_id,
                EventTicket.status == "used",
                EventTicket.checked_in_at.isnot(None),
            )
            .group_by(func.date(EventTicket.checked_in_at))
            .order_by("date")
        )
        ticket_checkins = ticket_res.all()

    return {
        "event_type": normalize_event_type(getattr(event, "event_type", None)),
        "ticketing_enabled": is_ticketing_enabled(event),
        "certificate_enabled": is_certificate_enabled(event),
        "checkin_enabled": is_checkin_enabled(event),
        "registrations": [
            {"date": str(d), "count": c} for d, c in registrations
        ],
        "survey_completions": [
            {"date": str(d), "count": c} for d, c in surveys
        ],
        "certificate_creations": [
            {"date": str(d), "count": c} for d, c in certificates
        ],
        "ticket_checkins": [
            {"date": str(d), "count": c} for d, c in ticket_checkins
        ],
    }


@router.get("/api/admin/events/{event_id}/analytics/export.csv")
async def export_event_analytics_csv(
    event_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export full attendee analytics as CSV."""
    e_res = await db.execute(select(Event).where(Event.id == event_id))
    event = e_res.scalar_one_or_none()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    if event.admin_id != current_user.id and current_user.role != Role.superadmin:
        raise HTTPException(status_code=403, detail="Unauthorized")

    # Fetch attendees with their cert and attendance info
    atts_res = await db.execute(
        select(Attendee).where(Attendee.event_id == event_id).order_by(Attendee.registered_at.asc())
    )
    attendees = atts_res.scalars().all()

    # Attendee IDs with certs
    cert_res = await db.execute(
        select(Certificate.student_name, func.count(Certificate.id).label("cert_count"))
        .where(Certificate.event_id == event_id, Certificate.deleted_at.is_(None))
        .group_by(Certificate.student_name)
    )
    cert_map = {row.student_name.strip().lower(): row.cert_count for row in cert_res.all()}

    # Attendance records per attendee
    ar_res = await db.execute(
        select(AttendaonceRecord.attendee_id, func.count(AttendaonceRecord.id).label("sessions"))
        .where(AttendaonceRecord.attendee_id.in_([a.id for a in attendees]))
        .group_by(AttendaonceRecord.attendee_id)
    )
    attendance_map = {row.attendee_id: row.sessions for row in ar_res.all()}

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "ID", "Name", "Email", "Registered At", "Email Verified",
        "Survey Completed", "Sessions Attended", "Has Certificate", "Check-in",
    ])
    for att in attendees:
        writer.writerow([
            att.id,
            att.name or "",
            att.email,
            att.registered_at.isoformat() if att.registered_at else "",
            "yes" if att.email_verified else "no",
            "yes" if att.survey_completed_at else "no",
            attendance_map.get(att.id, 0),
            "yes" if cert_map.get((att.name or "").strip().lower()) else "no",
            "yes" if att.checked_in_at else "no",
        ])
    output.seek(0)
    safe_name = (event.name or f"event_{event_id}").replace(" ", "_")[:60]
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="analytics_{safe_name}.csv"'},
    )


@router.get("/api/admin/events/{event_id}/analytics/export.xlsx")
async def export_event_analytics_xlsx(
    event_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export full attendee analytics as Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    import io as _io

    e_res = await db.execute(select(Event).where(Event.id == event_id))
    event = e_res.scalar_one_or_none()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    if event.admin_id != current_user.id and current_user.role != Role.superadmin:
        raise HTTPException(status_code=403, detail="Unauthorized")

    atts_res = await db.execute(
        select(Attendee).where(Attendee.event_id == event_id).order_by(Attendee.registered_at.asc())
    )
    attendees = atts_res.scalars().all()

    cert_res = await db.execute(
        select(Certificate.student_name, func.count(Certificate.id).label("cert_count"))
        .where(Certificate.event_id == event_id, Certificate.deleted_at.is_(None))
        .group_by(Certificate.student_name)
    )
    cert_map = {row.student_name.strip().lower(): row.cert_count for row in cert_res.all()}

    ar_res = await db.execute(
        select(AttendaonceRecord.attendee_id, func.count(AttendaonceRecord.id).label("sessions"))
        .where(AttendaonceRecord.attendee_id.in_([a.id for a in attendees]))
        .group_by(AttendaonceRecord.attendee_id)
    )
    attendance_map = {row.attendee_id: row.sessions for row in ar_res.all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendees"
    headers = ["ID", "Name", "Email", "Registered At", "Email Verified", "Survey Completed", "Sessions Attended", "Has Certificate", "Check-in"]
    ws.append(headers)
    header_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for att in attendees:
        ws.append([
            att.id,
            att.name or "",
            att.email,
            att.registered_at.isoformat() if att.registered_at else "",
            "yes" if att.email_verified else "no",
            "yes" if att.survey_completed_at else "no",
            attendance_map.get(att.id, 0),
            "yes" if cert_map.get((att.name or "").strip().lower()) else "no",
            "yes" if att.checked_in_at else "no",
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(max(len(str(col[0].value or "")) + 4, 12), 40)

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = (event.name or f"event_{event_id}").replace(" ", "_")[:60]
    return StreamingResponse(
        iter([buf.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="analytics_{safe_name}.xlsx"'},
    )
